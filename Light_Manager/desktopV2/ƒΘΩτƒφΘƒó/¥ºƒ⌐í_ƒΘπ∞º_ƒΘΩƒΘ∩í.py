#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
نظام إدارة العهد المالية للمشاريع - متكامل ومترابط
يتضمن إدارة العهد المالية، دفعات العهد، ومصروفات العهد
مع دعم كامل للغة العربية والتخطيط RTL
"""

import sys
from datetime import datetime, date
from decimal import Decimal
import mysql.connector

# استيراد مكتبات PyQt6/PySide6
try:
    from PySide6.QtWidgets import *
    from PySide6.QtCore import *
    from PySide6.QtGui import *
except ImportError:
    from PyQt6.QtWidgets import *
    from PyQt6.QtCore import *
    from PyQt6.QtGui import *

# استيراد الإعدادات والأنماط
from الإعدادات_العامة import *
from ستايل import apply_stylesheet
import qtawesome as qta

# النظام الرئيسي لإدارة العهد المالية
class CustodyManagementSystem(QDialog):
    
    # init
    def __init__(self, parent=None, project_id=None, project_data=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.project_data = project_data or {}
        
        # إعداد النافذة
        self.setup_window()
        
        # إنشاء الواجهة
        self.create_interface()
        
        # تحميل البيانات
        self.load_data()
        
        # تطبيق الأنماط
        self.apply_styles()
        
    # إعداد النافذة الأساسية
    def setup_window(self):
        project_name = self.project_data.get('اسم_المشروع', 'مشروع جديد')
        self.setWindowTitle(f"إدارة العهد المالية - {project_name}")
        self.setGeometry(100, 100, 1400, 800)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setWindowFlags(Qt.Window | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
        
    # إنشاء واجهة المستخدم
    def create_interface(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)
        
        # إنشاء التابات
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # تاب العهد المالية
        self.create_custody_tab()
        
        # تاب دفعات العهد
        self.create_custody_payments_tab()
        
        # تاب مصروفات العهد
        self.create_custody_expenses_tab()
        
        # تاب التقارير والإحصائيات
        self.create_reports_tab()
        
        # ربط إشارة تغيير التاب
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
    # إنشاء تاب العهد المالية
    def create_custody_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # شريط الأدوات والفلاتر
        toolbar_layout = QHBoxLayout()
        
        # أزرار الإجراءات
        self.add_custody_btn = QPushButton("إضافة عهدة جديدة")
        self.add_custody_btn.setIcon(qta.icon('fa5s.plus', color='#27ae60'))
        self.add_custody_btn.clicked.connect(self.add_custody)
        
        self.edit_custody_btn = QPushButton("تعديل العهدة")
        self.edit_custody_btn.setIcon(qta.icon('fa5s.edit', color='#3498db'))
        self.edit_custody_btn.clicked.connect(self.edit_custody)
        
        self.close_custody_btn = QPushButton("إغلاق العهدة")
        self.close_custody_btn.setIcon(qta.icon('fa5s.lock', color='#e74c3c'))
        self.close_custody_btn.clicked.connect(self.close_custody)
        
        self.refresh_custody_btn = QPushButton("تحديث")
        self.refresh_custody_btn.setIcon(qta.icon('fa5s.sync', color='#95a5a6'))
        self.refresh_custody_btn.clicked.connect(self.load_custody_data)
        
        toolbar_layout.addWidget(self.add_custody_btn)
        toolbar_layout.addWidget(self.edit_custody_btn)
        toolbar_layout.addWidget(self.close_custody_btn)
        toolbar_layout.addWidget(self.refresh_custody_btn)
        toolbar_layout.addStretch()
        
        # فلاتر البحث
        search_layout = QHBoxLayout()
        
        self.custody_search_edit = QLineEdit()
        self.custody_search_edit.setPlaceholderText("البحث في العهد المالية...")
        self.custody_search_edit.textChanged.connect(self.filter_custody_data)
        
        self.custody_status_filter = QComboBox()
        self.custody_status_filter.addItems(["جميع الحالات", "مفتوحة", "مغلقة", "مرحلة"])
        self.custody_status_filter.currentTextChanged.connect(self.filter_custody_data)
        
        search_layout.addWidget(QLabel("البحث:"))
        search_layout.addWidget(self.custody_search_edit)
        search_layout.addWidget(QLabel("الحالة:"))
        search_layout.addWidget(self.custody_status_filter)
        search_layout.addStretch()
        
        layout.addLayout(toolbar_layout)
        layout.addLayout(search_layout)
        
        # جدول العهد المالية
        self.custody_table = QTableWidget()
        self.setup_custody_table()
        layout.addWidget(self.custody_table)
        
        # إحصائيات العهد المالية
        self.create_custody_statistics(layout)
        
        self.tab_widget.addTab(tab, qta.icon('fa5s.credit-card', color='#27ae60'), "العهد المالية")
        
    # إعداد جدول العهد المالية
    def setup_custody_table(self):
        headers = [
            "المعرف", "رقم العهدة", "وصف العهدة", "مبلغ العهدة", "نسبة المكتب %",
            "مبلغ المكتب", "المبلغ الصافي", "المصروف", "المتبقي", 
            "تاريخ الاستلام", "الحالة", "ملاحظات"
        ]
        
        self.custody_table.setColumnCount(len(headers))
        self.custody_table.setHorizontalHeaderLabels(headers)
        
        # إعداد خصائص الجدول
        self.custody_table.setAlternatingRowColors(True)
        self.custody_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_table.setSortingEnabled(True)
        
        # إخفاء عمود المعرف
        self.custody_table.setColumnHidden(0, True)
        
        # تعديل عرض الأعمدة
        header = self.custody_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        
        # ربط الأحداث
        self.custody_table.itemDoubleClicked.connect(self.edit_custody)
        self.custody_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_table.customContextMenuRequested.connect(self.show_custody_context_menu)
        
    # إنشاء قسم إحصائيات العهد المالية
    def create_custody_statistics(self, layout):
        stats_frame = QFrame()
        stats_frame.setFrameStyle(QFrame.StyledPanel)
        stats_layout = QHBoxLayout(stats_frame)
        
        # بطاقات الإحصائيات
        self.total_custody_card = self.create_stat_card("إجمالي العهد", "0", "#3498db")
        self.total_amount_card = self.create_stat_card("إجمالي المبالغ", f"0 {Currency_type}", "#27ae60")
        self.total_expenses_card = self.create_stat_card("إجمالي المصروفات",f"0 {Currency_type}", "#e74c3c")
        self.total_remaining_card = self.create_stat_card("إجمالي المتبقي", f"0 {Currency_type}", "#f39c12")
        
        stats_layout.addWidget(self.total_custody_card)
        stats_layout.addWidget(self.total_amount_card)
        stats_layout.addWidget(self.total_expenses_card)
        stats_layout.addWidget(self.total_remaining_card)
        
        layout.addWidget(stats_frame)
        
    # إنشاء بطاقة إحصائية
    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 10px;
                padding: 10px;
                margin: 5px;
            }}
            QLabel {{
                color: white;
                font-weight: bold;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)
        
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 12px;")
        
        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        value_label.setObjectName(f"{title}_value")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        return card
        
    # إنشاء تاب دفعات العهد
    def create_custody_payments_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # شريط الأدوات
        toolbar_layout = QHBoxLayout()
        
        self.add_payment_btn = QPushButton("إضافة دفعة جديدة")
        self.add_payment_btn.setIcon(qta.icon('fa5s.plus', color='#27ae60'))
        self.add_payment_btn.clicked.connect(self.add_custody_payment)
        
        self.edit_payment_btn = QPushButton("تعديل الدفعة")
        self.edit_payment_btn.setIcon(qta.icon('fa5s.edit', color='#3498db'))
        self.edit_payment_btn.clicked.connect(self.edit_custody_payment)
        
        self.delete_payment_btn = QPushButton("حذف الدفعة")
        self.delete_payment_btn.setIcon(qta.icon('fa5s.trash', color='#e74c3c'))
        self.delete_payment_btn.clicked.connect(self.delete_custody_payment)
        
        self.refresh_payments_btn = QPushButton("تحديث")
        self.refresh_payments_btn.setIcon(qta.icon('fa5s.sync', color='#95a5a6'))
        self.refresh_payments_btn.clicked.connect(self.load_custody_payments_data)
        
        toolbar_layout.addWidget(self.add_payment_btn)
        toolbar_layout.addWidget(self.edit_payment_btn)
        toolbar_layout.addWidget(self.delete_payment_btn)
        toolbar_layout.addWidget(self.refresh_payments_btn)
        toolbar_layout.addStretch()
        
        layout.addLayout(toolbar_layout)
        
        # جدول دفعات العهد
        self.custody_payments_table = QTableWidget()
        self.setup_custody_payments_table()
        layout.addWidget(self.custody_payments_table)
        
        self.tab_widget.addTab(tab, qta.icon('fa5s.money-bill', color='#2ecc71'), "دفعات العهد")
        
    # إعداد جدول دفعات العهد
    def setup_custody_payments_table(self):
        headers = [
            "المعرف", "رقم العهدة", "وصف الدفعة", "المبلغ", "تاريخ الدفعة",
            "طريقة الدفع", "المستلم", "ملاحظات"
        ]
        
        self.custody_payments_table.setColumnCount(len(headers))
        self.custody_payments_table.setHorizontalHeaderLabels(headers)
        
        # إعداد خصائص الجدول
        self.custody_payments_table.setAlternatingRowColors(True)
        self.custody_payments_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_payments_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_payments_table.setSortingEnabled(True)
        
        # إخفاء عمود المعرف
        self.custody_payments_table.setColumnHidden(0, True)
        
        # تعديل عرض الأعمدة
        header = self.custody_payments_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        
        # ربط الأحداث
        self.custody_payments_table.itemDoubleClicked.connect(self.edit_custody_payment)
        self.custody_payments_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_payments_table.customContextMenuRequested.connect(self.show_payments_context_menu)

    # إنشاء تاب مصروفات العهد
    def create_custody_expenses_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # شريط الأدوات
        toolbar_layout = QHBoxLayout()

        self.add_expense_btn = QPushButton("إضافة معاملة")
        self.add_expense_btn.setIcon(qta.icon('fa5s.plus', color='#27ae60'))
        self.add_expense_btn.clicked.connect(self.add_custody_expense)

        self.edit_expense_btn = QPushButton("تعديل المصروف")
        self.edit_expense_btn.setIcon(qta.icon('fa5s.edit', color='#3498db'))
        self.edit_expense_btn.clicked.connect(self.edit_custody_expense)

        self.delete_expense_btn = QPushButton("حذف المصروف")
        self.delete_expense_btn.setIcon(qta.icon('fa5s.trash', color='#e74c3c'))
        self.delete_expense_btn.clicked.connect(self.delete_custody_expense)

        self.refresh_expenses_btn = QPushButton("تحديث")
        self.refresh_expenses_btn.setIcon(qta.icon('fa5s.sync', color='#95a5a6'))
        self.refresh_expenses_btn.clicked.connect(self.load_custody_expenses_data)

        toolbar_layout.addWidget(self.add_expense_btn)
        toolbar_layout.addWidget(self.edit_expense_btn)
        toolbar_layout.addWidget(self.delete_expense_btn)
        toolbar_layout.addWidget(self.refresh_expenses_btn)
        toolbar_layout.addStretch()

        # فلاتر المصروفات
        filter_layout = QHBoxLayout()

        self.expense_type_filter = QComboBox()
        self.expense_type_filter.addItems(["جميع الأنواع", "مرتبط_بعهدة", "غير_مرتبط_بعهدة", "خسائر"])
        self.expense_type_filter.currentTextChanged.connect(self.filter_expenses_data)

        self.expense_custody_filter = QComboBox()
        self.expense_custody_filter.addItem("جميع العهد")
        self.expense_custody_filter.currentTextChanged.connect(self.filter_expenses_data)

        filter_layout.addWidget(QLabel("نوع المصروف:"))
        filter_layout.addWidget(self.expense_type_filter)
        filter_layout.addWidget(QLabel("العهدة:"))
        filter_layout.addWidget(self.expense_custody_filter)
        filter_layout.addStretch()

        layout.addLayout(toolbar_layout)
        layout.addLayout(filter_layout)

        # جدول مصروفات العهد
        self.custody_expenses_table = QTableWidget()
        self.setup_custody_expenses_table()
        layout.addWidget(self.custody_expenses_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.receipt', color='#e74c3c'), "مصروفات العهد")

    # إعداد جدول مصروفات العهد
    def setup_custody_expenses_table(self):
        headers = [
            "المعرف", "نوع المصروف", "رقم العهدة", "المسؤول", "وصف المصروف",
            "المبلغ", "تاريخ المصروف", "طريقة الدفع", "رقم الفاتورة",
            "المورد", "ملاحظات"
        ]

        self.custody_expenses_table.setColumnCount(len(headers))
        self.custody_expenses_table.setHorizontalHeaderLabels(headers)

        # إعداد خصائص الجدول
        self.custody_expenses_table.setAlternatingRowColors(True)
        self.custody_expenses_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_expenses_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_expenses_table.setSortingEnabled(True)

        # إخفاء عمود المعرف
        self.custody_expenses_table.setColumnHidden(0, True)

        # تعديل عرض الأعمدة
        header = self.custody_expenses_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # ربط الأحداث
        self.custody_expenses_table.itemDoubleClicked.connect(self.edit_custody_expense)
        self.custody_expenses_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_expenses_table.customContextMenuRequested.connect(self.show_expenses_context_menu)

    # إنشاء تاب التقارير والإحصائيات
    def create_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # شريط أدوات التقارير
        toolbar_layout = QHBoxLayout()

        self.generate_report_btn = QPushButton("إنشاء تقرير شامل")
        self.generate_report_btn.setIcon(qta.icon('fa5s.file-alt', color='#3498db'))
        self.generate_report_btn.clicked.connect(self.generate_comprehensive_report)

        self.export_excel_btn = QPushButton("تصدير إلى Excel")
        self.export_excel_btn.setIcon(qta.icon('fa5s.file-excel', color='#27ae60'))
        self.export_excel_btn.clicked.connect(self.export_to_excel)

        self.print_report_btn = QPushButton("طباعة التقرير")
        self.print_report_btn.setIcon(qta.icon('fa5s.print', color='#95a5a6'))
        self.print_report_btn.clicked.connect(self.print_report)

        toolbar_layout.addWidget(self.generate_report_btn)
        toolbar_layout.addWidget(self.export_excel_btn)
        toolbar_layout.addWidget(self.print_report_btn)
        toolbar_layout.addStretch()

        layout.addLayout(toolbar_layout)

        # منطقة عرض التقرير
        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)
        layout.addWidget(self.report_text)

        self.tab_widget.addTab(tab, qta.icon('fa5s.chart-bar', color='#9b59b6'), "التقارير والإحصائيات")

    # معالج تغيير التاب
    def on_tab_changed(self, index):
        if index == 0:  # تاب العهد المالية
            self.load_custody_data()
        elif index == 1:  # تاب دفعات العهد
            self.load_custody_payments_data()
        elif index == 2:  # تاب مصروفات العهد
            self.load_custody_expenses_data()
        elif index == 3:  # تاب التقارير
            self.update_reports()

    # تحميل جميع البيانات
    def load_data(self):
        self.load_custody_data()
        self.load_custody_payments_data()
        self.load_custody_expenses_data()
        self.update_statistics()

    # تحميل بيانات العهد المالية
    def load_custody_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT id, رقم_العهدة, وصف_العهدة, مبلغ_العهدة, نسبة_المكتب,
                       مبلغ_نسبة_المكتب, المبلغ_الصافي, المصروف, المتبقي,
                       تاريخ_الإستلام, حالة_العهدة, ملاحظات
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
                ORDER BY تاريخ_الإستلام DESC
            """

            cursor.execute(query, (self.project_id,))
            custody_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_table.setRowCount(len(custody_data))

            for row, data in enumerate(custody_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)

                    # تلوين الحالات
                    if col == 10:  # عمود الحالة
                        if value == "مفتوحة":
                            item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                        elif value == "مغلقة":
                            item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر
                        elif value == "مرحلة":
                            item.setForeground(QBrush(QColor(243, 156, 18)))  # برتقالي

                    self.custody_table.setItem(row, col, item)

            conn.close()
            self.update_statistics()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات العهد المالية: {str(e)}")

    # تحميل بيانات دفعات العهد
    def load_custody_payments_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT dp.id, c.رقم_العهدة, dp.وصف_الدفعة, dp.المبلغ, dp.تاريخ_الدفعة,
                       dp.طريقة_الدفع, dp.المستلم, dp.ملاحظات
                FROM المقاولات_دفعات_العهد dp
                JOIN المقاولات_العهد c ON dp.معرف_العهدة = c.id
                WHERE c.معرف_المشروع = %s
                ORDER BY dp.تاريخ_الدفعة DESC
            """

            cursor.execute(query, (self.project_id,))
            payments_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_payments_table.setRowCount(len(payments_data))

            for row, data in enumerate(payments_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.custody_payments_table.setItem(row, col, item)

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات دفعات العهد: {str(e)}")

    # تحميل بيانات مصروفات العهد
    def load_custody_expenses_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT me.id, me.نوع_المصروف, COALESCE(c.رقم_العهدة, 'غير مرتبط'),
                       me.المسؤول, me.وصف_المصروف, me.المبلغ, me.تاريخ_المصروف,
                       me.طريقة_الدفع, me.رقم_الفاتورة, me.المورد, me.ملاحظات
                FROM المقاولات_مصروفات_العهد me
                LEFT JOIN المقاولات_العهد c ON me.معرف_العهدة = c.id
                WHERE me.معرف_المشروع = %s
                ORDER BY me.تاريخ_المصروف DESC
            """

            cursor.execute(query, (self.project_id,))
            expenses_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_expenses_table.setRowCount(len(expenses_data))

            for row, data in enumerate(expenses_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)

                    # تلوين أنواع المصروفات
                    if col == 1:  # عمود نوع المصروف
                        if value == "مرتبط_بعهدة":
                            item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                        elif value == "غير_مرتبط_بعهدة":
                            item.setForeground(QBrush(QColor(52, 152, 219)))  # أزرق
                        elif value == "خسائر":
                            item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر

                    self.custody_expenses_table.setItem(row, col, item)

            # تحديث فلتر العهد
            self.update_custody_filter()
            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات مصروفات العهد: {str(e)}")

    # تحديث فلتر العهد في تاب المصروفات
    def update_custody_filter(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT رقم_العهدة FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND رقم_العهدة IS NOT NULL
                ORDER BY رقم_العهدة
            """, (self.project_id,))

            custodies = cursor.fetchall()

            # تحديث ComboBox
            current_text = self.expense_custody_filter.currentText()
            self.expense_custody_filter.clear()
            self.expense_custody_filter.addItem("جميع العهد")

            for custody in custodies:
                self.expense_custody_filter.addItem(custody[0])

            # استعادة الاختيار السابق إن أمكن
            index = self.expense_custody_filter.findText(current_text)
            if index >= 0:
                self.expense_custody_filter.setCurrentIndex(index)

            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث فلتر العهد: {e}")

    # تحديث الإحصائيات
    def update_statistics(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # إحصائيات العهد المالية
            cursor.execute("""
                SELECT
                    COUNT(*) as count_custody,
                    COALESCE(SUM(مبلغ_العهدة), 0) as total_amount,
                    COALESCE(SUM(المصروف), 0) as total_expenses,
                    COALESCE(SUM(المتبقي), 0) as total_remaining
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
            """, (self.project_id,))

            result = cursor.fetchone()
            if result:
                count_custody, total_amount, total_expenses, total_remaining = result

                # تحديث البطاقات
                self.total_custody_card.findChild(QLabel, "إجمالي العهد_value").setText(str(count_custody))
                self.total_amount_card.findChild(QLabel, "إجمالي المبالغ_value").setText(f"{float(total_amount):,.2f}  {Currency_type}")
                self.total_expenses_card.findChild(QLabel, "إجمالي المصروفات_value").setText(f"{float(total_expenses):,.2f}  {Currency_type}")
                self.total_remaining_card.findChild(QLabel, "إجمالي المتبقي_value").setText(f"{float(total_remaining):,.2f}  {Currency_type}")

            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث الإحصائيات: {e}")

    # فلترة بيانات العهد المالية
    def filter_custody_data(self):
        search_text = self.custody_search_edit.text().lower()
        status_filter = self.custody_status_filter.currentText()

        for row in range(self.custody_table.rowCount()):
            show_row = True

            # فلترة النص
            if search_text:
                row_text = ""
                for col in range(self.custody_table.columnCount()):
                    item = self.custody_table.item(row, col)
                    if item:
                        row_text += item.text().lower() + " "

                if search_text not in row_text:
                    show_row = False

            # فلترة الحالة
            if status_filter != "جميع الحالات":
                status_item = self.custody_table.item(row, 10)  # عمود الحالة
                if status_item and status_item.text() != status_filter:
                    show_row = False

            self.custody_table.setRowHidden(row, not show_row)

    # فلترة بيانات مصروفات العهد
    def filter_expenses_data(self):
        type_filter = self.expense_type_filter.currentText()
        custody_filter = self.expense_custody_filter.currentText()

        for row in range(self.custody_expenses_table.rowCount()):
            show_row = True

            # فلترة نوع المصروف
            if type_filter != "جميع الأنواع":
                type_item = self.custody_expenses_table.item(row, 1)  # عمود نوع المصروف
                if type_item and type_item.text() != type_filter:
                    show_row = False

            # فلترة العهدة
            if custody_filter != "جميع العهد":
                custody_item = self.custody_expenses_table.item(row, 2)  # عمود رقم العهدة
                if custody_item and custody_item.text() != custody_filter:
                    show_row = False

            self.custody_expenses_table.setRowHidden(row, not show_row)

    # عرض القائمة السياقية للعهد المالية
    def show_custody_context_menu(self, position):
        if self.custody_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        close_action = menu.addAction(qta.icon('fa5s.lock', color='#e74c3c'), "إغلاق العهدة")

        action = menu.exec_(self.custody_table.mapToGlobal(position))

        if action == view_action:
            self.view_custody_details()
        elif action == edit_action:
            self.edit_custody()
        elif action == close_action:
            self.close_custody()

    # عرض القائمة السياقية لدفعات العهد
    def show_payments_context_menu(self, position):
        if self.custody_payments_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        delete_action = menu.addAction(qta.icon('fa5s.trash', color='#e74c3c'), "حذف")

        action = menu.exec_(self.custody_payments_table.mapToGlobal(position))

        if action == view_action:
            self.view_payment_details()
        elif action == edit_action:
            self.edit_custody_payment()
        elif action == delete_action:
            self.delete_custody_payment()

    # عرض القائمة السياقية لمصروفات العهد
    def show_expenses_context_menu(self, position):
        if self.custody_expenses_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        delete_action = menu.addAction(qta.icon('fa5s.trash', color='#e74c3c'), "حذف")

        action = menu.exec_(self.custody_expenses_table.mapToGlobal(position))

        if action == view_action:
            self.view_expense_details()
        elif action == edit_action:
            self.edit_custody_expense()
        elif action == delete_action:
            self.delete_custody_expense()

    # ==================== دوال إدارة العهد المالية ====================

    # إضافة عهدة مالية جديدة
    def add_custody(self):
        dialog = CustodyDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_data()
            QMessageBox.information(self, "نجح", "تم إضافة العهدة المالية بنجاح")

    # تعديل عهدة مالية
    def edit_custody(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار عهدة للتعديل")
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        if not custody_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف العهدة")
            return

        custody_id = int(custody_id_item.text())
        dialog = CustodyDialog(self, project_id=self.project_id, custody_id=custody_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_data()
            QMessageBox.information(self, "نجح", "تم تعديل العهدة المالية بنجاح")

    # إغلاق عهدة مالية
    def close_custody(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار عهدة للإغلاق")
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        custody_number_item = self.custody_table.item(current_row, 1)

        if not custody_id_item or not custody_number_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات العهدة")
            return

        custody_id = int(custody_id_item.text())
        custody_number = custody_number_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الإغلاق",
            f"هل تريد إغلاق العهدة رقم {custody_number}؟\n"
            "لن تتمكن من إضافة مصروفات جديدة لهذه العهدة بعد الإغلاق.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE المقاولات_العهد
                    SET حالة_العهدة = 'مغلقة', تاريخ_الإغلاق = CURDATE()
                    WHERE id = %s
                """, (custody_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم إغلاق العهدة بنجاح")
                self.load_custody_data()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في إغلاق العهدة: {str(e)}")

    # عرض تفاصيل العهدة
    def view_custody_details(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        if not custody_id_item:
            return

        custody_id = int(custody_id_item.text())
        dialog = CustodyDetailsDialog(self, custody_id)
        dialog.exec_()

    # ==================== دوال إدارة دفعات العهد ====================

    # إضافة دفعة عهد جديدة
    def add_custody_payment(self):
        dialog = CustodyPaymentDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_payments_data()
            self.load_custody_data()  # تحديث العهد أيضاً
            QMessageBox.information(self, "نجح", "تم إضافة دفعة العهد بنجاح")

    # تعديل دفعة عهد
    def edit_custody_payment(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار دفعة للتعديل")
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        if not payment_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الدفعة")
            return

        payment_id = int(payment_id_item.text())
        dialog = CustodyPaymentDialog(self, project_id=self.project_id, payment_id=payment_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_payments_data()
            self.load_custody_data()  # تحديث العهد أيضاً
            QMessageBox.information(self, "نجح", "تم تعديل دفعة العهد بنجاح")

    # حذف دفعة عهد
    def delete_custody_payment(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار دفعة للحذف")
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        payment_desc_item = self.custody_payments_table.item(current_row, 2)

        if not payment_id_item or not payment_desc_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات الدفعة")
            return

        payment_id = int(payment_id_item.text())
        payment_desc = payment_desc_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل تريد حذف الدفعة '{payment_desc}'؟\n"
            "سيتم تحديث مبلغ العهدة تلقائياً.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("DELETE FROM المقاولات_دفعات_العهد WHERE id = %s", (payment_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم حذف دفعة العهد بنجاح")
                self.load_custody_payments_data()
                self.load_custody_data()  # تحديث العهد أيضاً

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف دفعة العهد: {str(e)}")

    # عرض تفاصيل دفعة العهد
    def view_payment_details(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        if not payment_id_item:
            return

        payment_id = int(payment_id_item.text())
        dialog = PaymentDetailsDialog(self, payment_id)
        dialog.exec_()

    # ==================== دوال إدارة مصروفات العهد ====================

    # إضافة مصروف عهد جديد
    def add_custody_expense(self):
        dialog = CustodyExpenseDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_expenses_data()
            self.load_custody_data()  # تحديث العهد أيضاً
            QMessageBox.information(self, "نجح", "تم إضافة مصروف العهد بنجاح")

    # تعديل مصروف عهد
    def edit_custody_expense(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار مصروف للتعديل")
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        if not expense_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المصروف")
            return

        expense_id = int(expense_id_item.text())
        dialog = CustodyExpenseDialog(self, project_id=self.project_id, expense_id=expense_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_custody_expenses_data()
            self.load_custody_data()  # تحديث العهد أيضاً
            QMessageBox.information(self, "نجح", "تم تعديل مصروف العهد بنجاح")

    # حذف مصروف عهد
    def delete_custody_expense(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار مصروف للحذف")
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        expense_desc_item = self.custody_expenses_table.item(current_row, 4)

        if not expense_id_item or not expense_desc_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات المصروف")
            return

        expense_id = int(expense_id_item.text())
        expense_desc = expense_desc_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل تريد حذف المصروف '{expense_desc}'؟\n"
            "سيتم تحديث مبلغ العهدة تلقائياً.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("DELETE FROM المقاولات_مصروفات_العهد WHERE id = %s", (expense_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم حذف مصروف العهد بنجاح")
                self.load_custody_expenses_data()
                self.load_custody_data()  # تحديث العهد أيضاً

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف مصروف العهد: {str(e)}")

    # عرض تفاصيل مصروف العهد
    def view_expense_details(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        if not expense_id_item:
            return

        expense_id = int(expense_id_item.text())
        dialog = ExpenseDetailsDialog(self, expense_id)
        dialog.exec_()

    # ==================== دوال التقارير والإحصائيات ====================

    # تحديث التقارير
    def update_reports(self):
        self.generate_comprehensive_report()

    # إنشاء تقرير شامل
    def generate_comprehensive_report(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            project_name = self.project_data.get('اسم_المشروع', 'غير محدد')
            client_name = self.project_data.get('اسم_العميل', 'غير محدد')

            report_html = f"""
            <html dir="rtl">
            <head>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: 'Arial', sans-serif; direction: rtl; }}
                    .header {{ background-color: #3498db; color: white; padding: 20px; text-align: center; }}
                    .section {{ margin: 20px 0; padding: 15px; border: 1px solid #ddd; }}
                    .table {{ width: 100%; border-collapse: collapse; }}
                    .table th, .table td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                    .table th {{ background-color: #f2f2f2; }}
                    .summary {{ background-color: #ecf0f1; padding: 15px; margin: 10px 0; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>تقرير العهد المالية الشامل</h1>
                    <h2>المشروع: {project_name}</h2>
                    <h3>العميل: {client_name}</h3>
                    <p>تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
                </div>
            """

            # ملخص العهد المالية
            cursor.execute("""
                SELECT
                    COUNT(*) as total_custody,
                    COUNT(CASE WHEN حالة_العهدة = 'مفتوحة' THEN 1 END) as open_custody,
                    COUNT(CASE WHEN حالة_العهدة = 'مغلقة' THEN 1 END) as closed_custody,
                    COALESCE(SUM(مبلغ_العهدة), 0) as total_amount,
                    COALESCE(SUM(المصروف), 0) as total_expenses,
                    COALESCE(SUM(المتبقي), 0) as total_remaining
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
            """, (self.project_id,))

            summary = cursor.fetchone()
            if summary:
                total_custody, open_custody, closed_custody, total_amount, total_expenses, total_remaining = summary

                report_html += f"""
                <div class="section">
                    <h2>ملخص العهد المالية</h2>
                    <div class="summary">
                        <p><strong>إجمالي العهد:</strong> {total_custody}</p>
                        <p><strong>العهد المفتوحة:</strong> {open_custody}</p>
                        <p><strong>العهد المغلقة:</strong> {closed_custody}</p>
                        <p><strong>إجمالي المبالغ:</strong> {float(total_amount):,.2f} {Currency_type}</p>
                        <p><strong>إجمالي المصروفات:</strong> {float(total_expenses):,.2f} {Currency_type}</p>
                        <p><strong>إجمالي المتبقي:</strong> {float(total_remaining):,.2f} {Currency_type}</p>
                    </div>
                </div>
                """

            # تفاصيل العهد المالية
            cursor.execute("""
                SELECT رقم_العهدة, وصف_العهدة, مبلغ_العهدة, المصروف, المتبقي,
                       تاريخ_الإستلام, حالة_العهدة
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
                ORDER BY تاريخ_الإستلام DESC
            """, (self.project_id,))

            custody_details = cursor.fetchall()

            report_html += """
            <div class="section">
                <h2>تفاصيل العهد المالية</h2>
                <table class="table">
                    <tr>
                        <th>رقم العهدة</th>
                        <th>الوصف</th>
                        <th>المبلغ</th>
                        <th>المصروف</th>
                        <th>المتبقي</th>
                        <th>تاريخ الاستلام</th>
                        <th>الحالة</th>
                    </tr>
            """

            for custody in custody_details:
                report_html += f"""
                    <tr>
                        <td>{custody[0] or ''}</td>
                        <td>{custody[1] or ''}</td>
                        <td>{float(custody[2]):,.2f}</td>
                        <td>{float(custody[3]):,.2f}</td>
                        <td>{float(custody[4]):,.2f}</td>
                        <td>{custody[5]}</td>
                        <td>{custody[6]}</td>
                    </tr>
                """

            report_html += """
                </table>
            </div>
            """

            report_html += """
            </body>
            </html>
            """

            self.report_text.setHtml(report_html)
            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في إنشاء التقرير: {str(e)}")

    # تصدير البيانات إلى Excel
    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ التقرير",
                f"تقرير_العهد_المالية_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # إنشاء ملف Excel جديد
            wb = Workbook()

            # إزالة الورقة الافتراضية
            wb.remove(wb.active)

            # ورقة العهد المالية
            ws_custody = wb.create_sheet("العهد المالية")
            self.export_custody_to_sheet(ws_custody)

            # ورقة دفعات العهد
            ws_payments = wb.create_sheet("دفعات العهد")
            self.export_payments_to_sheet(ws_payments)

            # ورقة مصروفات العهد
            ws_expenses = wb.create_sheet("مصروفات العهد")
            self.export_expenses_to_sheet(ws_expenses)

            # حفظ الملف
            wb.save(file_path)

            QMessageBox.information(self, "نجح", f"تم تصدير التقرير بنجاح إلى:\n{file_path}")

        except ImportError:
            QMessageBox.warning(self, "خطأ", "يتطلب تصدير Excel تثبيت مكتبات pandas و openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تصدير التقرير: {str(e)}")

    # تصدير بيانات العهد المالية إلى ورقة Excel
    def export_custody_to_sheet(self, worksheet):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # عناوين الأعمدة
            headers = ["رقم العهدة", "وصف العهدة", "مبلغ العهدة", "نسبة المكتب %",
                      "مبلغ المكتب", "المبلغ الصافي", "المصروف", "المتبقي",
                      "تاريخ الاستلام", "الحالة", "ملاحظات"]

            # كتابة العناوين
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # جلب البيانات
            cursor.execute("""
                SELECT رقم_العهدة, وصف_العهدة, مبلغ_العهدة, نسبة_المكتب,
                       مبلغ_نسبة_المكتب, المبلغ_الصافي, المصروف, المتبقي,
                       تاريخ_الإستلام, حالة_العهدة, ملاحظات
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
                ORDER BY تاريخ_الإستلام DESC
            """, (self.project_id,))

            data = cursor.fetchall()

            # كتابة البيانات
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

            conn.close()

        except Exception as e:
            print(f"خطأ في تصدير العهد المالية: {e}")

    # تصدير بيانات دفعات العهد إلى ورقة Excel
    def export_payments_to_sheet(self, worksheet):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # عناوين الأعمدة
            headers = ["رقم العهدة", "وصف الدفعة", "المبلغ", "تاريخ الدفعة",
                      "طريقة الدفع", "المستلم", "ملاحظات"]

            # كتابة العناوين
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # جلب البيانات
            cursor.execute("""
                SELECT c.رقم_العهدة, dp.وصف_الدفعة, dp.المبلغ, dp.تاريخ_الدفعة,
                       dp.طريقة_الدفع, dp.المستلم, dp.ملاحظات
                FROM المقاولات_دفعات_العهد dp
                JOIN المقاولات_العهد c ON dp.معرف_العهدة = c.id
                WHERE c.معرف_المشروع = %s
                ORDER BY dp.تاريخ_الدفعة DESC
            """, (self.project_id,))

            data = cursor.fetchall()

            # كتابة البيانات
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

            conn.close()

        except Exception as e:
            print(f"خطأ في تصدير دفعات العهد: {e}")

    # تصدير بيانات مصروفات العهد إلى ورقة Excel
    def export_expenses_to_sheet(self, worksheet):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # عناوين الأعمدة
            headers = ["نوع المصروف", "رقم العهدة", "المسؤول", "وصف المصروف",
                      "المبلغ", "تاريخ المصروف", "طريقة الدفع", "رقم الفاتورة",
                      "المورد", "ملاحظات"]

            # كتابة العناوين
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # جلب البيانات
            cursor.execute("""
                SELECT me.نوع_المصروف, COALESCE(c.رقم_العهدة, 'غير مرتبط'),
                       me.المسؤول, me.وصف_المصروف, me.المبلغ, me.تاريخ_المصروف,
                       me.طريقة_الدفع, me.رقم_الفاتورة, me.المورد, me.ملاحظات
                FROM المقاولات_مصروفات_العهد me
                LEFT JOIN المقاولات_العهد c ON me.معرف_العهدة = c.id
                WHERE me.معرف_المشروع = %s
                ORDER BY me.تاريخ_المصروف DESC
            """, (self.project_id,))

            data = cursor.fetchall()

            # كتابة البيانات
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center")

            conn.close()

        except Exception as e:
            print(f"خطأ في تصدير مصروفات العهد: {e}")

    # طباعة التقرير
    def print_report(self):
        try:
            from PySide6.QtPrintSupport import QPrinter, QPrintDialog

            printer = QPrinter()
            dialog = QPrintDialog(printer, self)

            if dialog.exec_() == QPrintDialog.Accepted:
                self.report_text.print_(printer)
                QMessageBox.information(self, "نجح", "تم إرسال التقرير للطباعة")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في طباعة التقرير: {str(e)}")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# ==================== كلاسات الحوارات ====================

# حوار إضافة/تعديل العهد المالية
class CustodyDialog(QDialog):

    # init
    def __init__(self, parent=None, project_id=None, custody_id=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.custody_id = custody_id
        self.is_edit_mode = custody_id is not None

        self.setup_dialog()
        self.create_form()
        self.setup_connections()

        if self.is_edit_mode:
            self.load_custody_data()

        self.apply_styles()

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل العهدة المالية" if self.is_edit_mode else "إضافة عهدة مالية جديدة"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 600, 500)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء النموذج
    def create_form(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # عنوان الحوار
        title_label = QLabel("تعديل العهدة المالية" if self.is_edit_mode else "إضافة عهدة مالية جديدة")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # النموذج الرئيسي
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFormAlignment(Qt.AlignRight)
        form_layout.setSpacing(10)

        # رقم العهدة
        self.custody_number_edit = QLineEdit()
        self.custody_number_edit.setPlaceholderText("أدخل رقم العهدة...")
        form_layout.addRow("رقم العهدة:", self.custody_number_edit)

        # وصف العهدة
        self.custody_description_edit = QLineEdit()
        self.custody_description_edit.setPlaceholderText("أدخل وصف العهدة...")
        form_layout.addRow("وصف العهدة:", self.custody_description_edit)

        # مبلغ العهدة الأولي
        self.initial_amount_spin = QDoubleSpinBox()
        self.initial_amount_spin.setRange(0, 999999999)
        self.initial_amount_spin.setDecimals(2)
        self.initial_amount_spin.setSuffix(f"  {Currency_type}")
        form_layout.addRow("المبلغ الأولي:", self.initial_amount_spin)

        # نسبة المكتب
        self.office_percentage_spin = QDoubleSpinBox()
        self.office_percentage_spin.setRange(0, 100)
        self.office_percentage_spin.setDecimals(2)
        self.office_percentage_spin.setSuffix(" %")
        self.office_percentage_spin.valueChanged.connect(self.calculate_amounts)
        form_layout.addRow("نسبة المكتب:", self.office_percentage_spin)

        # مبلغ المكتب (محسوب تلقائياً)
        self.office_amount_label = QLabel(f"0.00 {Currency_type}")
        self.office_amount_label.setStyleSheet("color: #27ae60; font-weight: bold;")
        form_layout.addRow("مبلغ المكتب:", self.office_amount_label)

        # المبلغ الصافي (محسوب تلقائياً)
        self.net_amount_label = QLabel(f"0.00 {Currency_type}")
        self.net_amount_label.setStyleSheet("color: #3498db; font-weight: bold;")
        form_layout.addRow("المبلغ الصافي:", self.net_amount_label)

        # تاريخ الاستلام
        self.receipt_date_edit = QDateEdit()
        self.receipt_date_edit.setDate(QDate.currentDate())
        self.receipt_date_edit.setCalendarPopup(True)
        form_layout.addRow("تاريخ الاستلام:", self.receipt_date_edit)

        # حالة العهدة
        self.custody_status_combo = QComboBox()
        self.custody_status_combo.addItems(["مفتوحة", "مغلقة", "مرحلة"])
        form_layout.addRow("حالة العهدة:", self.custody_status_combo)

        # ملاحظات
        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(80)
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات...")
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()

        self.save_btn = QPushButton("حفظ")
        self.save_btn.setIcon(qta.icon('fa5s.save', color='#27ae60'))
        self.save_btn.clicked.connect(self.save_custody)

        self.cancel_btn = QPushButton("إلغاء")
        self.cancel_btn.setIcon(qta.icon('fa5s.times', color='#e74c3c'))
        self.cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.cancel_btn)

        layout.addLayout(buttons_layout)

    # إعداد الاتصالات
    def setup_connections(self):
        self.initial_amount_spin.valueChanged.connect(self.calculate_amounts)

    # حساب المبالغ تلقائياً
    def calculate_amounts(self):
        initial_amount = self.initial_amount_spin.value()
        office_percentage = self.office_percentage_spin.value()

        office_amount = (initial_amount * office_percentage) / 100
        net_amount = initial_amount - office_amount

        self.office_amount_label.setText(f"{office_amount:,.2f}  {Currency_type}")
        self.net_amount_label.setText(f"{net_amount:,.2f}  {Currency_type}")

    # تحميل بيانات العهدة للتعديل
    def load_custody_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT رقم_العهدة, وصف_العهدة, مبلغ_العهدة, نسبة_المكتب,
                       تاريخ_الإستلام, حالة_العهدة, ملاحظات
                FROM المقاولات_العهد
                WHERE id = %s
            """, (self.custody_id,))

            result = cursor.fetchone()
            if result:
                custody_number, description, amount, percentage, receipt_date, status, notes = result

                self.custody_number_edit.setText(custody_number or "")
                self.custody_description_edit.setText(description or "")
                self.initial_amount_spin.setValue(float(amount) if amount else 0)
                self.office_percentage_spin.setValue(float(percentage) if percentage else 0)

                if receipt_date:
                    self.receipt_date_edit.setDate(QDate.fromString(str(receipt_date), "yyyy-MM-dd"))

                status_index = self.custody_status_combo.findText(status or "مفتوحة")
                if status_index >= 0:
                    self.custody_status_combo.setCurrentIndex(status_index)

                self.notes_edit.setPlainText(notes or "")

                # حساب المبالغ
                self.calculate_amounts()

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات العهدة: {str(e)}")

    # التحقق من صحة البيانات
    def validate_data(self):
        if not self.custody_number_edit.text().strip():
            QMessageBox.warning(self, "خطأ", "يرجى إدخال رقم العهدة")
            self.custody_number_edit.setFocus()
            return False

        if not self.custody_description_edit.text().strip():
            QMessageBox.warning(self, "خطأ", "يرجى إدخال وصف العهدة")
            self.custody_description_edit.setFocus()
            return False

        if self.initial_amount_spin.value() <= 0:
            QMessageBox.warning(self, "خطأ", "يرجى إدخال مبلغ أولي صحيح")
            self.initial_amount_spin.setFocus()
            return False

        return True

    # حفظ العهدة
    def save_custody(self):
        if not self.validate_data():
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            custody_data = {
                'رقم_العهدة': self.custody_number_edit.text().strip(),
                'وصف_العهدة': self.custody_description_edit.text().strip(),
                'مبلغ_العهدة': self.initial_amount_spin.value(),
                'نسبة_المكتب': self.office_percentage_spin.value(),
                'تاريخ_الإستلام': self.receipt_date_edit.date().toString("yyyy-MM-dd"),
                'حالة_العهدة': self.custody_status_combo.currentText(),
                'ملاحظات': self.notes_edit.toPlainText().strip(),
                'المستخدم': 'النظام',  # يمكن تحديث هذا ليكون المستخدم الحالي
                'السنة': datetime.now().year
            }

            if self.is_edit_mode:
                # تحديث العهدة الموجودة
                cursor.execute("""
                    UPDATE المقاولات_العهد
                    SET رقم_العهدة = %(رقم_العهدة)s, وصف_العهدة = %(وصف_العهدة)s,
                        مبلغ_العهدة = %(مبلغ_العهدة)s, نسبة_المكتب = %(نسبة_المكتب)s,
                        تاريخ_الإستلام = %(تاريخ_الإستلام)s, حالة_العهدة = %(حالة_العهدة)s,
                        ملاحظات = %(ملاحظات)s
                    WHERE id = %s
                """, {**custody_data, 'id': self.custody_id})
            else:
                # إضافة عهدة جديدة
                custody_data['معرف_المشروع'] = self.project_id
                cursor.execute("""
                    INSERT INTO المقاولات_العهد
                    (معرف_المشروع, رقم_العهدة, وصف_العهدة, مبلغ_العهدة, نسبة_المكتب,
                     تاريخ_الإستلام, حالة_العهدة, ملاحظات, المستخدم, السنة)
                    VALUES (%(معرف_المشروع)s, %(رقم_العهدة)s, %(وصف_العهدة)s, %(مبلغ_العهدة)s,
                            %(نسبة_المكتب)s, %(تاريخ_الإستلام)s, %(حالة_العهدة)s, %(ملاحظات)s,
                            %(المستخدم)s, %(السنة)s)
                """, custody_data)

            conn.commit()
            conn.close()

            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ العهدة: {str(e)}")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# حوار إضافة/تعديل دفعات العهد
class CustodyPaymentDialog(QDialog):

    # init
    def __init__(self, parent=None, project_id=None, payment_id=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.payment_id = payment_id
        self.is_edit_mode = payment_id is not None

        self.setup_dialog()
        self.create_form()
        self.load_custodies()

        if self.is_edit_mode:
            self.load_payment_data()

        self.apply_styles()

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل دفعة العهد" if self.is_edit_mode else "إضافة دفعة عهد جديدة"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 500, 400)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء النموذج
    def create_form(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # عنوان الحوار
        title_label = QLabel("تعديل دفعة العهد" if self.is_edit_mode else "إضافة دفعة عهد جديدة")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # النموذج الرئيسي
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFormAlignment(Qt.AlignRight)
        form_layout.setSpacing(10)

        # اختيار العهدة
        self.custody_combo = QComboBox()
        self.custody_combo.setEditable(False)
        form_layout.addRow("العهدة:", self.custody_combo)

        # وصف الدفعة
        self.payment_description_edit = QLineEdit()
        self.payment_description_edit.setPlaceholderText("أدخل وصف الدفعة...")
        form_layout.addRow("وصف الدفعة:", self.payment_description_edit)

        # مبلغ الدفعة
        self.payment_amount_spin = QDoubleSpinBox()
        self.payment_amount_spin.setRange(0, 999999999)
        self.payment_amount_spin.setDecimals(2)
        self.payment_amount_spin.setSuffix(f"  {Currency_type}")
        form_layout.addRow("مبلغ الدفعة:", self.payment_amount_spin)

        # تاريخ الدفعة
        self.payment_date_edit = QDateEdit()
        self.payment_date_edit.setDate(QDate.currentDate())
        self.payment_date_edit.setCalendarPopup(True)
        form_layout.addRow("تاريخ الدفعة:", self.payment_date_edit)

        # طريقة الدفع
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItems(["نقدي", "شيك", "تحويل بنكي"])
        form_layout.addRow("طريقة الدفع:", self.payment_method_combo)

        # المستلم
        self.receiver_edit = QLineEdit()
        self.receiver_edit.setPlaceholderText("أدخل اسم المستلم...")
        form_layout.addRow("المستلم:", self.receiver_edit)

        # ملاحظات
        self.notes_edit = QTextEdit()
        self.notes_edit.setMaximumHeight(80)
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات...")
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()

        self.save_btn = QPushButton("حفظ")
        self.save_btn.setIcon(qta.icon('fa5s.save', color='#27ae60'))
        self.save_btn.clicked.connect(self.save_payment)

        self.cancel_btn = QPushButton("إلغاء")
        self.cancel_btn.setIcon(qta.icon('fa5s.times', color='#e74c3c'))
        self.cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.cancel_btn)

        layout.addLayout(buttons_layout)

    # تحميل قائمة العهد المتاحة
    def load_custodies(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, رقم_العهدة, وصف_العهدة
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND حالة_العهدة = 'مفتوحة'
                ORDER BY تاريخ_الإستلام DESC
            """, (self.project_id,))

            custodies = cursor.fetchall()

            self.custody_combo.clear()
            for custody in custodies:
                custody_id, custody_number, custody_desc = custody
                display_text = f"{custody_number} - {custody_desc}" if custody_desc else custody_number
                self.custody_combo.addItem(display_text, custody_id)

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل قائمة العهد: {str(e)}")

    # تحميل بيانات الدفعة للتعديل
    def load_payment_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT معرف_العهدة, وصف_الدفعة, المبلغ, تاريخ_الدفعة,
                       طريقة_الدفع, المستلم, ملاحظات
                FROM المقاولات_دفعات_العهد
                WHERE id = %s
            """, (self.payment_id,))

            result = cursor.fetchone()
            if result:
                custody_id, description, amount, payment_date, method, receiver, notes = result

                # تحديد العهدة
                for i in range(self.custody_combo.count()):
                    if self.custody_combo.itemData(i) == custody_id:
                        self.custody_combo.setCurrentIndex(i)
                        break

                self.payment_description_edit.setText(description or "")
                self.payment_amount_spin.setValue(float(amount) if amount else 0)

                if payment_date:
                    self.payment_date_edit.setDate(QDate.fromString(str(payment_date), "yyyy-MM-dd"))

                method_index = self.payment_method_combo.findText(method or "نقدي")
                if method_index >= 0:
                    self.payment_method_combo.setCurrentIndex(method_index)

                self.receiver_edit.setText(receiver or "")
                self.notes_edit.setPlainText(notes or "")

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات الدفعة: {str(e)}")

    # التحقق من صحة البيانات
    def validate_data(self):
        if self.custody_combo.currentData() is None:
            QMessageBox.warning(self, "خطأ", "يرجى اختيار العهدة")
            self.custody_combo.setFocus()
            return False

        if not self.payment_description_edit.text().strip():
            QMessageBox.warning(self, "خطأ", "يرجى إدخال وصف الدفعة")
            self.payment_description_edit.setFocus()
            return False

        if self.payment_amount_spin.value() <= 0:
            QMessageBox.warning(self, "خطأ", "يرجى إدخال مبلغ صحيح")
            self.payment_amount_spin.setFocus()
            return False

        return True

    # حفظ دفعة العهد
    def save_payment(self):
        if not self.validate_data():
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            payment_data = {
                'معرف_العهدة': self.custody_combo.currentData(),
                'وصف_الدفعة': self.payment_description_edit.text().strip(),
                'المبلغ': self.payment_amount_spin.value(),
                'تاريخ_الدفعة': self.payment_date_edit.date().toString("yyyy-MM-dd"),
                'طريقة_الدفع': self.payment_method_combo.currentText(),
                'المستلم': self.receiver_edit.text().strip(),
                'ملاحظات': self.notes_edit.toPlainText().strip(),
                'المستخدم': 'النظام',  # يمكن تحديث هذا ليكون المستخدم الحالي
                'السنة': datetime.now().year
            }

            if self.is_edit_mode:
                # تحديث الدفعة الموجودة
                cursor.execute("""
                    UPDATE المقاولات_دفعات_العهد
                    SET معرف_العهدة = %(معرف_العهدة)s, وصف_الدفعة = %(وصف_الدفعة)s,
                        المبلغ = %(المبلغ)s, تاريخ_الدفعة = %(تاريخ_الدفعة)s,
                        طريقة_الدفع = %(طريقة_الدفع)s, المستلم = %(المستلم)s,
                        ملاحظات = %(ملاحظات)s
                    WHERE id = %s
                """, {**payment_data, 'id': self.payment_id})
            else:
                # إضافة دفعة جديدة
                cursor.execute("""
                    INSERT INTO المقاولات_دفعات_العهد
                    (معرف_العهدة, وصف_الدفعة, المبلغ, تاريخ_الدفعة, طريقة_الدفع,
                     المستلم, ملاحظات, المستخدم, السنة)
                    VALUES (%(معرف_العهدة)s, %(وصف_الدفعة)s, %(المبلغ)s, %(تاريخ_الدفعة)s,
                            %(طريقة_الدفع)s, %(المستلم)s, %(ملاحظات)s, %(المستخدم)s, %(السنة)s)
                """, payment_data)

            conn.commit()
            conn.close()

            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ دفعة العهد: {str(e)}")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# حوار إضافة/تعديل مصروفات العهد
class CustodyExpenseDialog(QDialog):

    # init
    def __init__(self, parent=None, project_id=None, expense_id=None, expense_type=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.expense_id = expense_id
        self.is_edit_mode = expense_id is not None
        self.preset_expense_type = expense_type  # نوع المصروف المحدد مسبقاً

        self.setup_dialog()
        self.create_form()
        self.load_custodies()
        self.setup_connections()

        # تحديد نوع المصروف مسبقاً إذا تم تمريره
        if self.preset_expense_type and not self.is_edit_mode:
            self.set_preset_expense_type()

        if self.is_edit_mode:
            self.load_expense_data()

        self.apply_styles()

    # تحديد نوع المصروف مسبقاً
    def set_preset_expense_type(self):
        if self.preset_expense_type in ["خسائر", "مردودات"]:
            # البحث عن الفهرس المناسب في ComboBox
            index = self.expense_type_combo.findText(self.preset_expense_type)
            if index >= 0:
                self.expense_type_combo.setCurrentIndex(index)
                # تحديث الواجهة حسب النوع المحدد
                self.on_expense_type_changed()

    # إعداد الحوار
    def setup_dialog(self):
        if self.is_edit_mode:
            title = "تعديل مصروف العهد"
        elif self.preset_expense_type == "خسائر":
            title = "إضافة خسارة جديدة"
        elif self.preset_expense_type == "مردودات":
            title = "إضافة مردود جديد"
        else:
            title = "إضافة مصروف عهد جديد"

        self.setWindowTitle(title)
        self.setGeometry(200, 200, 600, 500)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء النموذج
    def create_form(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # عنوان الحوار
        if self.is_edit_mode:
            title_text = "تعديل مصروف العهد"
        elif self.preset_expense_type == "خسائر":
            title_text = "إضافة خسارة جديدة"
        elif self.preset_expense_type == "مردودات":
            title_text = "إضافة مردود جديد"
        else:
            title_text = "إضافة مصروف عهد جديد"

        title_label = QLabel(title_text)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # النموذج الرئيسي
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFormAlignment(Qt.AlignRight)
        form_layout.setSpacing(10)

        # نوع المصروف
        self.expense_type_combo = QComboBox()
        self.expense_type_combo.addItems(["مرتبط_بعهدة", "خسائر", "مردودات"])
        self.expense_type_combo.currentTextChanged.connect(self.on_expense_type_changed)
        form_layout.addRow("نوع المصروف:", self.expense_type_combo)

        # اختيار العهدة (يظهر فقط للمصروفات المرتبطة بعهدة)
        self.custody_combo = QComboBox()
        self.custody_combo.setEditable(False)
        self.custody_label = QLabel("العهدة:")
        form_layout.addRow(self.custody_label, self.custody_combo)

        # متحمل الخسائر (يظهر فقط للخسائر)
        self.loss_bearer_combo = QComboBox()
        self.loss_bearer_combo.addItems(["الشركة", "مهندس", "مقاول", "عامل", "موظف"])
        self.loss_bearer_combo.currentTextChanged.connect(self.on_loss_bearer_changed)
        self.loss_bearer_label = QLabel("متحمل الخسائر:")
        form_layout.addRow(self.loss_bearer_label, self.loss_bearer_combo)

        # اختيار المسؤول (يظهر عند اختيار موظف كمتحمل للخسائر)
        self.responsible_combo = QComboBox()
        self.responsible_combo.setEditable(True)
        self.responsible_combo_label = QLabel("المسؤول:")
        form_layout.addRow(self.responsible_combo_label, self.responsible_combo)

        # اختيار العهدة للمردودات (يظهر فقط للمردودات)
        self.return_custody_combo = QComboBox()
        self.return_custody_combo.setEditable(False)
        self.return_custody_label = QLabel("العهدة المردودة:")
        form_layout.addRow(self.return_custody_label, self.return_custody_combo)

        # المسؤول (حقل نصي - سيتم إخفاؤه للخسائر)
        self.responsible_edit = QLineEdit()
        self.responsible_edit.setPlaceholderText("أدخل اسم المسؤول...")
        self.responsible_edit_label = QLabel("المسؤول:")
        form_layout.addRow(self.responsible_edit_label, self.responsible_edit)

        # وصف المصروف (سيتم تحديث التسمية حسب النوع)
        self.expense_description_edit = QLineEdit()
        self.expense_description_edit.setPlaceholderText("أدخل وصف المصروف...")
        self.expense_description_label = QLabel("وصف المصروف:")
        form_layout.addRow(self.expense_description_label, self.expense_description_edit)

        # مبلغ المصروف
        self.expense_amount_spin = QDoubleSpinBox()
        self.expense_amount_spin.setRange(0, 999999999)
        self.expense_amount_spin.setDecimals(2)
        self.expense_amount_spin.setSuffix(f"  {Currency_type}")
        form_layout.addRow("مبلغ المصروف:", self.expense_amount_spin)

        # تاريخ المصروف
        self.expense_date_edit = QDateEdit()
        self.expense_date_edit.setDate(QDate.currentDate())
        self.expense_date_edit.setCalendarPopup(True)
        form_layout.addRow("تاريخ المصروف:", self.expense_date_edit)

        # طريقة الدفع
        self.payment_method_edit = QComboBox()
        self.payment_method_edit.addItems([
            "",
            "نقدي",
            "اجل/دين",
            "تحويل بنكي",
            "شيك",
            "أخرى"
        ])
        form_layout.addRow("طريقة الدفع:", self.payment_method_edit)

        # رقم الفاتورة
        self.invoice_number_edit = QLineEdit()
        self.invoice_number_edit.setPlaceholderText("أدخل رقم الفاتورة...")
        form_layout.addRow("رقم الفاتورة:", self.invoice_number_edit)

        # المورد
        self.supplier_edit = QComboBox()
        self.supplier_edit.setEditable(True)
        self.supplier_edit.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.load_suppliers()
        form_layout.addRow("المورد:", self.supplier_edit)

        # ملاحظات
        self.notes_edit = QLineEdit()
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات...")
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()

        self.save_btn = QPushButton("حفظ")
        self.save_btn.setIcon(qta.icon('fa5s.save', color='#27ae60'))
        self.save_btn.clicked.connect(self.save_expense)

        self.cancel_btn = QPushButton("إلغاء")
        self.cancel_btn.setIcon(qta.icon('fa5s.times', color='#e74c3c'))
        self.cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.cancel_btn)

        layout.addLayout(buttons_layout)

    # إعداد الاتصالات
    def setup_connections(self):
        # إخفاء جميع الحقول الإضافية في البداية
        self.loss_bearer_combo.setVisible(False)
        self.loss_bearer_label.setVisible(False)
        self.responsible_combo.setVisible(False)
        self.responsible_combo_label.setVisible(False)
        self.return_custody_combo.setVisible(False)
        self.return_custody_label.setVisible(False)

        self.on_expense_type_changed()  # تحديد حالة العهدة الأولية

    # معالج تغيير نوع المصروف
    def on_expense_type_changed(self):
        expense_type = self.expense_type_combo.currentText()

        # تحديث التسميات حسب نوع المصروف
        self.update_labels_for_expense_type(expense_type)

        # إظهار/إخفاء حقول العهدة
        is_custody_related = expense_type == "مرتبط_بعهدة"
        self.custody_combo.setVisible(is_custody_related)
        self.custody_label.setVisible(is_custody_related)

        # إظهار/إخفاء حقول الخسائر
        is_loss = expense_type == "خسائر"
        self.loss_bearer_combo.setVisible(is_loss)
        self.loss_bearer_label.setVisible(is_loss)

        # إظهار/إخفاء حقول المردودات
        is_return = expense_type == "مردودات"
        self.return_custody_combo.setVisible(is_return)
        self.return_custody_label.setVisible(is_return)

        # إظهار/إخفاء حقل المسؤول النصي للخسائر
        if is_loss:
            self.responsible_edit.setVisible(False)
            self.responsible_edit_label.setVisible(False)
            self.on_loss_bearer_changed()
        else:
            self.responsible_edit.setVisible(True)
            self.responsible_edit_label.setVisible(True)
            self.responsible_combo.setVisible(False)
            self.responsible_combo_label.setVisible(False)

        # تحميل العهد للمردودات
        if is_return:
            self.load_custodies_for_returns()

    # تحديث تسميات الواجهة حسب نوع المصروف
    def update_labels_for_expense_type(self, expense_type):
        if expense_type == "خسائر":
            self.expense_description_label.setText("وصف الخسائر:")
            self.expense_description_edit.setPlaceholderText("أدخل وصف الخسائر...")
        elif expense_type == "مردودات":
            self.expense_description_label.setText("وصف المردودات:")
            self.expense_description_edit.setPlaceholderText("أدخل وصف المردودات...")
        else:
            self.expense_description_label.setText("وصف المصروف:")
            self.expense_description_edit.setPlaceholderText("أدخل وصف المصروف...")

    # معالج تغيير متحمل الخسائر
    def on_loss_bearer_changed(self):
        bearer = self.loss_bearer_combo.currentText()

        # إظهار/إخفاء كومبو بوكس المسؤول
        show_responsible = bearer in ["مهندس", "مقاول", "عامل", "موظف"]
        self.responsible_combo.setVisible(show_responsible)
        self.responsible_combo_label.setVisible(show_responsible)

        if show_responsible:
            self.load_employees_by_type(bearer)

    # تحميل الموظفين حسب النوع
    def load_employees_by_type(self, employee_type):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # تحديد نوع الحساب حسب الاختيار
            account_type_map = {
                "مهندس": "مهندس",
                "مقاول": "مقاول",
                "عامل": "عامل",
                "موظف": "موظف"
            }

            account_type = account_type_map.get(employee_type, employee_type)

            cursor.execute("""
                SELECT id, اسم_الموظف, الوظيفة
                FROM الموظفين
                WHERE التصنيف = %s OR الوظيفة LIKE %s
                ORDER BY اسم_الموظف
            """, (account_type, f"%{account_type}%"))

            employees = cursor.fetchall()

            self.responsible_combo.clear()
            self.responsible_combo.addItem("-- اختر المسؤول --", None)

            for employee in employees:
                emp_id, emp_name, emp_job = employee
                display_text = f"{emp_name} - {emp_job}" if emp_job else emp_name
                self.responsible_combo.addItem(display_text, emp_id)

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل قائمة الموظفين: {str(e)}")

    # تحميل قائمة الموردين من قاعدة البيانات
    def load_suppliers(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("SELECT اسم_المورد FROM الموردين ORDER BY اسم_المورد")
            suppliers = cursor.fetchall()

            # إضافة خيار فارغ أولاً
            self.supplier_edit.addItem("")

            # إضافة الموردين
            for supplier in suppliers:
                if supplier[0]:  # التأكد من أن اسم المورد ليس فارغاً
                    self.supplier_edit.addItem(supplier[0])

            conn.close()

        except Exception as e:
            print(f"خطأ في تحميل الموردين: {str(e)}")
            # في حالة الخطأ، إضافة خيار فارغ على الأقل
            self.supplier_edit.addItem("")

    # تحميل قائمة العهد للمردودات
    def load_custodies_for_returns(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, رقم_العهدة, وصف_العهدة, المبلغ_الصافي
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND حالة_العهدة = 'مفتوحة'
                ORDER BY تاريخ_الإستلام DESC
            """, (self.project_id,))

            custodies = cursor.fetchall()

            self.return_custody_combo.clear()
            self.return_custody_combo.addItem("-- اختر العهدة --", None)

            for custody in custodies:
                custody_id, custody_number, custody_desc, net_amount = custody
                display_text = f"{custody_number} - {custody_desc} ({net_amount} {Currency_type})" if custody_desc else f"{custody_number} ({net_amount} {Currency_type})"
                self.return_custody_combo.addItem(display_text, custody_id)

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل قائمة العهد: {str(e)}")

    # تحميل قائمة العهد المتاحة
    def load_custodies(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, رقم_العهدة, وصف_العهدة
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND حالة_العهدة = 'مفتوحة'
                ORDER BY تاريخ_الإستلام DESC
            """, (self.project_id,))

            custodies = cursor.fetchall()

            self.custody_combo.clear()
            self.custody_combo.addItem("-- اختر العهدة --", None)
            for custody in custodies:
                custody_id, custody_number, custody_desc = custody
                display_text = f"{custody_number} - {custody_desc}" if custody_desc else custody_number
                self.custody_combo.addItem(display_text, custody_id)

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل قائمة العهد: {str(e)}")

    # تحميل بيانات المصروف للتعديل
    def load_expense_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT نوع_المصروف, معرف_العهدة, المسؤول, وصف_المصروف, المبلغ,
                       تاريخ_المصروف, طريقة_الدفع, رقم_الفاتورة, المورد, ملاحظات,
                       متحمل_الخسائر, معرف_المتحمل, معرف_العهدة_المردودة
                FROM المقاولات_مصروفات_العهد
                WHERE id = %s
            """, (self.expense_id,))

            result = cursor.fetchone()
            if result:
                (expense_type, custody_id, responsible, description, amount,
                 expense_date, payment_method, invoice_number, supplier, notes,
                 loss_bearer, responsible_id, return_custody_id) = result

                # تحديد نوع المصروف
                type_index = self.expense_type_combo.findText(expense_type or "غير_مرتبط_بعهدة")
                if type_index >= 0:
                    self.expense_type_combo.setCurrentIndex(type_index)

                # تحديد العهدة إذا كان المصروف مرتبط بعهدة
                if custody_id:
                    for i in range(self.custody_combo.count()):
                        if self.custody_combo.itemData(i) == custody_id:
                            self.custody_combo.setCurrentIndex(i)
                            break

                self.responsible_edit.setText(responsible or "")
                self.expense_description_edit.setText(description or "")
                self.expense_amount_spin.setValue(float(amount) if amount else 0)

                if expense_date:
                    self.expense_date_edit.setDate(QDate.fromString(str(expense_date), "yyyy-MM-dd"))

                self.payment_method_edit.setCurrentText(payment_method or "")
                self.invoice_number_edit.setText(invoice_number or "")
                self.supplier_edit.setCurrentText(supplier or "")
                self.notes_edit.setText(notes or "")

                # تحديد متحمل الخسائر إذا كان نوع المصروف خسائر
                if expense_type == "خسائر" and loss_bearer:
                    bearer_index = self.loss_bearer_combo.findText(loss_bearer)
                    if bearer_index >= 0:
                        self.loss_bearer_combo.setCurrentIndex(bearer_index)

                    # تحديد المسؤول إذا كان موجود
                    if responsible_id and loss_bearer != "الشركة":
                        self.load_employees_by_type(loss_bearer)
                        for i in range(self.responsible_combo.count()):
                            if self.responsible_combo.itemData(i) == responsible_id:
                                self.responsible_combo.setCurrentIndex(i)
                                break

                # تحديد العهدة المردودة إذا كان نوع المصروف مردودات
                if expense_type == "مردودات" and return_custody_id:
                    self.load_custodies_for_returns()
                    for i in range(self.return_custody_combo.count()):
                        if self.return_custody_combo.itemData(i) == return_custody_id:
                            self.return_custody_combo.setCurrentIndex(i)
                            break

            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المصروف: {str(e)}")

    # التحقق من صحة البيانات
    def validate_data(self):
        expense_type = self.expense_type_combo.currentText()

        # تحديد نوع الوصف حسب نوع المصروف
        description_type = "وصف الخسائر" if expense_type == "خسائر" else \
                          "وصف المردودات" if expense_type == "مردودات" else "وصف المصروف"

        if not self.expense_description_edit.text().strip():
            QMessageBox.warning(self, "خطأ", f"يرجى إدخال {description_type}")
            self.expense_description_edit.setFocus()
            return False

        if self.expense_amount_spin.value() <= 0:
            QMessageBox.warning(self, "خطأ", "يرجى إدخال مبلغ صحيح")
            self.expense_amount_spin.setFocus()
            return False

        expense_type = self.expense_type_combo.currentText()

        # التحقق من اختيار العهدة للمصروفات المرتبطة بعهدة
        if expense_type == "مرتبط_بعهدة" and self.custody_combo.currentData() is None:
            QMessageBox.warning(self, "خطأ", "يرجى اختيار العهدة للمصروف المرتبط بعهدة")
            self.custody_combo.setFocus()
            return False

        # التحقق من اختيار المسؤول للخسائر عند اختيار موظف
        if expense_type == "خسائر":
            bearer = self.loss_bearer_combo.currentText()
            if bearer in ["مهندس", "مقاول", "عامل", "موظف"] and self.responsible_combo.currentData() is None:
                QMessageBox.warning(self, "خطأ", f"يرجى اختيار {bearer} المسؤول عن الخسائر")
                self.responsible_combo.setFocus()
                return False

        # التحقق من اختيار العهدة للمردودات
        if expense_type == "مردودات" and self.return_custody_combo.currentData() is None:
            QMessageBox.warning(self, "خطأ", "يرجى اختيار العهدة المراد إرجاع مردودات منها")
            self.return_custody_combo.setFocus()
            return False

        return True

    # عرض رسالة تأكيد مع تفاصيل العملية
    def show_confirmation_message(self):
        expense_type = self.expense_type_combo.currentText()
        amount = self.expense_amount_spin.value()
        description = self.expense_description_edit.text().strip()

        # إعداد رسالة التأكيد حسب نوع المصروف
        if expense_type == "خسائر":
            bearer = self.loss_bearer_combo.currentText()
            if bearer == "الشركة":
                message = f"سيتم إضافة مبلغ {amount:.2f} {Currency_type} كخسائر إلى مصروفات الشركة\n\n"
                message += f"التفاصيل:\n"
                message += f"• النوع: خسائر الشركة\n"
                message += f"• المبلغ: {amount:.2f} {Currency_type}\n"
                message += f"• الوصف: {description}\n"
                message += f"• سيتم تسجيلها في جدول الحسابات"
            else:
                responsible_name = self.responsible_combo.currentText().split(' - ')[0] if self.responsible_combo.currentData() else "غير محدد"
                message = f"سيتم خصم مبلغ {amount:.2f} {Currency_type} من رصيد {responsible_name}\n\n"
                message += f"التفاصيل:\n"
                message += f"• النوع: خسائر {bearer}\n"
                message += f"• المسؤول: {responsible_name}\n"
                message += f"• المبلغ: {amount:.2f} {Currency_type}\n"
                message += f"• الوصف: {description}\n"
                message += f"• سيتم خصمها من رصيد الموظف"

        elif expense_type == "مردودات":
            custody_text = self.return_custody_combo.currentText()
            message = f"سيتم إضافة مردود بمبلغ {amount:.2f} {Currency_type}\n\n"
            message += f"التفاصيل:\n"
            message += f"• النوع: مردودات\n"
            message += f"• العهدة: {custody_text}\n"
            message += f"• المبلغ: {amount:.2f} {Currency_type}\n"
            message += f"• الوصف: {description}\n"
            message += f"• سيتم طرح المبلغ من مصروفات العهدة"

        else:
            message = f"سيتم إضافة مصروف بمبلغ {amount:.2f} {Currency_type}\n\n"
            message += f"التفاصيل:\n"
            message += f"• النوع: {expense_type}\n"
            message += f"• المبلغ: {amount:.2f} {Currency_type}\n"
            message += f"• الوصف: {description}"

            if expense_type == "مرتبط_بعهدة":
                custody_text = self.custody_combo.currentText()
                message += f"\n• العهدة: {custody_text}"

        # عرض رسالة التأكيد
        reply = QMessageBox.question(
            self,
            "تأكيد العملية",
            message + "\n\nهل تريد المتابعة؟",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        return reply == QMessageBox.Yes

    # حفظ مصروف العهد
    def save_expense(self):
        if not self.validate_data():
            return

        # عرض رسالة تأكيد مع التفاصيل
        if not self.show_confirmation_message():
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            expense_type = self.expense_type_combo.currentText()

            # إعداد البيانات الأساسية
            expense_data = {
                'معرف_المشروع': self.project_id,
                'نوع_المصروف': expense_type,
                'معرف_العهدة': self.custody_combo.currentData() if expense_type == "مرتبط_بعهدة" else None,
                'المسؤول': self.responsible_edit.text().strip(),
                'وصف_المصروف': self.expense_description_edit.text().strip(),
                'المبلغ': self.expense_amount_spin.value(),
                'تاريخ_المصروف': self.expense_date_edit.date().toString("yyyy-MM-dd"),
                'طريقة_الدفع': self.payment_method_edit.currentText().strip(),
                'رقم_الفاتورة': self.invoice_number_edit.text().strip(),
                'المورد': self.supplier_edit.currentText().strip(),
                'ملاحظات': self.notes_edit.text().strip(),
                'المستخدم': 'النظام',  # يمكن تحديث هذا ليكون المستخدم الحالي
                'السنة': datetime.now().year,
                # الحقول الجديدة
                'متحمل_الخسائر': self.loss_bearer_combo.currentText() if expense_type == "خسائر" else None,
                'معرف_المتحمل': self.responsible_combo.currentData() if expense_type == "خسائر" and self.loss_bearer_combo.currentText() != "الشركة" else None,
                'معرف_العهدة_المردودة': self.return_custody_combo.currentData() if expense_type == "مردودات" else None
            }

            if self.is_edit_mode:
                # تحديث المصروف الموجود
                cursor.execute("""
                    UPDATE المقاولات_مصروفات_العهد
                    SET نوع_المصروف = %(نوع_المصروف)s, معرف_العهدة = %(معرف_العهدة)s,
                        المسؤول = %(المسؤول)s, وصف_المصروف = %(وصف_المصروف)s,
                        المبلغ = %(المبلغ)s, تاريخ_المصروف = %(تاريخ_المصروف)s,
                        طريقة_الدفع = %(طريقة_الدفع)s, رقم_الفاتورة = %(رقم_الفاتورة)s,
                        المورد = %(المورد)s, ملاحظات = %(ملاحظات)s,
                        متحمل_الخسائر = %(متحمل_الخسائر)s, معرف_المتحمل = %(معرف_المتحمل)s,
                        معرف_العهدة_المردودة = %(معرف_العهدة_المردودة)s
                    WHERE id = %s
                """, {**expense_data, 'id': self.expense_id})
            else:
                # إضافة مصروف جديد
                cursor.execute("""
                    INSERT INTO المقاولات_مصروفات_العهد
                    (معرف_المشروع, نوع_المصروف, معرف_العهدة, المسؤول, وصف_المصروف,
                     المبلغ, تاريخ_المصروف, طريقة_الدفع, رقم_الفاتورة, المورد,
                     ملاحظات, المستخدم, السنة, متحمل_الخسائر, معرف_المتحمل, معرف_العهدة_المردودة)
                    VALUES (%(معرف_المشروع)s, %(نوع_المصروف)s, %(معرف_العهدة)s, %(المسؤول)s,
                            %(وصف_المصروف)s, %(المبلغ)s, %(تاريخ_المصروف)s, %(طريقة_الدفع)s,
                            %(رقم_الفاتورة)s, %(المورد)s, %(ملاحظات)s, %(المستخدم)s, %(السنة)s,
                            %(متحمل_الخسائر)s, %(معرف_المتحمل)s, %(معرف_العهدة_المردودة)s)
                """, expense_data)

            # معالجة الخسائر والمردودات
            if not self.is_edit_mode:  # فقط للمصروفات الجديدة
                if expense_type == "خسائر":
                    self.handle_loss_expense(cursor, expense_data)
                elif expense_type == "مردودات":
                    self.handle_return_expense(cursor, expense_data)

            conn.commit()
            conn.close()

            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ مصروف العهد: {str(e)}")

    # معالجة مصروف الخسائر
    def handle_loss_expense(self, cursor, expense_data):
        try:
            bearer = expense_data['متحمل_الخسائر']
            amount = expense_data['المبلغ']
            description = expense_data['وصف_المصروف']
            expense_date = expense_data['تاريخ_المصروف']

            # الحصول على اسم المشروع من الجدول الصحيح
            cursor.execute("SELECT اسم_المشروع FROM المشاريع WHERE id = %s", (self.project_id,))
            project_result = cursor.fetchone()
            project_name = project_result[0] if project_result else "مشروع غير محدد"

            loss_description = f"خسائر - {project_name} - {description}"

            if bearer == "الشركة":
                # إضافة المصروف إلى جدول الحسابات (بدون السنة لأنها محسوبة تلقائياً)
                cursor.execute("""
                    INSERT INTO الحسابات
                    (التصنيف, المصروف, المبلغ, تاريخ_المصروف, المستلم, ملاحظات, المستخدم)
                    VALUES ('خسائر', %s, %s, %s, 'الشركة', %s, %s)
                """, (loss_description, amount, expense_date, f"خسائر من مشروع: {project_name}",
                      expense_data['المستخدم']))

            else:
                # إضافة خصم للموظف المسؤول
                employee_id = expense_data['معرف_المتحمل']
                if employee_id:
                    cursor.execute("""
                        INSERT INTO الموظفين_معاملات_مالية
                        (معرف_الموظف, نوع_العملية, نوع_المعاملة, المبلغ, التاريخ, الوصف, المستخدم)
                        VALUES (%s, 'خصم', 'خصم مبلغ', %s, %s, %s, %s)
                    """, (employee_id, amount, expense_date, loss_description,
                          expense_data['المستخدم']))

                    # تحديث رصيد الموظف
                    cursor.execute("""
                        UPDATE الموظفين
                        SET الرصيد = الرصيد - %s
                        WHERE id = %s
                    """, (amount, employee_id))

        except Exception as e:
            print(f"خطأ في معالجة الخسائر: {e}")

    # معالجة مصروف المردودات
    def handle_return_expense(self, cursor, expense_data):
        try:
            custody_id = expense_data['معرف_العهدة_المردودة']
            amount = expense_data['المبلغ']

            if custody_id:
                # طرح المبلغ من مصروفات العهدة المحددة فقط
                # المتبقي سيتم حسابه تلقائياً لأنه عمود محسوب
                cursor.execute("""
                    UPDATE المقاولات_العهد
                    SET المصروف = المصروف - %s
                    WHERE id = %s
                """, (amount, custody_id))

        except Exception as e:
            print(f"خطأ في معالجة المردودات: {e}")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# ==================== حوارات التفاصيل ====================

# حوار عرض تفاصيل العهدة
class CustodyDetailsDialog(QDialog):

    # init
    def __init__(self, parent=None, custody_id=None):
        super().__init__(parent)
        self.custody_id = custody_id

        self.setWindowTitle("تفاصيل العهدة المالية")
        self.setGeometry(200, 200, 500, 400)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

        self.create_interface()
        self.load_data()
        self.apply_styles()

    # إنشاء الواجهة
    def create_interface(self):
        layout = QVBoxLayout(self)

        # منطقة عرض التفاصيل
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        layout.addWidget(self.details_text)

        # زر الإغلاق
        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

    # تحميل تفاصيل العهدة
    def load_data(self):
        # يمكن تطوير هذه الدالة لعرض تفاصيل شاملة للعهدة
        self.details_text.setPlainText("تفاصيل العهدة ستظهر هنا...")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# حوار عرض تفاصيل دفعة العهد
class PaymentDetailsDialog(QDialog):

    # init
    def __init__(self, parent=None, payment_id=None):
        super().__init__(parent)
        self.payment_id = payment_id

        self.setWindowTitle("تفاصيل دفعة العهد")
        self.setGeometry(200, 200, 400, 300)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

        self.create_interface()
        self.load_data()
        self.apply_styles()

    # إنشاء الواجهة
    def create_interface(self):
        layout = QVBoxLayout(self)

        # منطقة عرض التفاصيل
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        layout.addWidget(self.details_text)

        # زر الإغلاق
        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

    # تحميل تفاصيل الدفعة
    def load_data(self):
        # يمكن تطوير هذه الدالة لعرض تفاصيل شاملة للدفعة
        self.details_text.setPlainText("تفاصيل الدفعة ستظهر هنا...")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# حوار عرض تفاصيل مصروف العهد
class ExpenseDetailsDialog(QDialog):

    # init
    def __init__(self, parent=None, expense_id=None):
        super().__init__(parent)
        self.expense_id = expense_id

        self.setWindowTitle("تفاصيل مصروف العهد")
        self.setGeometry(200, 200, 400, 300)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

        self.create_interface()
        self.load_data()
        self.apply_styles()

    # إنشاء الواجهة
    def create_interface(self):
        layout = QVBoxLayout(self)

        # منطقة عرض التفاصيل
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        layout.addWidget(self.details_text)

        # زر الإغلاق
        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

    # تحميل تفاصيل المصروف
    def load_data(self):
        # يمكن تطوير هذه الدالة لعرض تفاصيل شاملة للمصروف
        self.details_text.setPlainText("تفاصيل المصروف ستظهر هنا...")

    # تطبيق الأنماط
    def apply_styles(self):
        try:
            apply_stylesheet(self)
        except:
            pass


# ==================== تابات مدمجة لنافذة إدارة المشروع ====================

# تاب العهد المالية المدمج
class CustodyTabWidget(QWidget):

    # init
    def __init__(self, parent=None, project_id=None, project_data=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.project_data = project_data or {}

        self.create_interface()
        self.load_data()

    # إنشاء واجهة التاب
    def create_interface(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # تخطيط أفقي للأزرار والفلاتر
        top_layout = QHBoxLayout()

        # أزرار العمليات (في الجانب الأيمن)
        buttons_layout = QHBoxLayout()

        self.add_custody_btn = QPushButton("إضافة")
        self.add_custody_btn.setIcon(qta.icon('fa5s.plus'))
        self.add_custody_btn.clicked.connect(self.add_custody)
        self.add_custody_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        buttons_layout.addWidget(self.add_custody_btn)

        self.edit_custody_btn = QPushButton("تعديل")
        self.edit_custody_btn.setIcon(qta.icon('fa5s.edit'))
        self.edit_custody_btn.clicked.connect(self.edit_custody)
        self.edit_custody_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5dade2;
            }
        """)
        buttons_layout.addWidget(self.edit_custody_btn)

        self.close_custody_btn = QPushButton("إغلاق")
        self.close_custody_btn.setIcon(qta.icon('fa5s.lock'))
        self.close_custody_btn.clicked.connect(self.close_custody)
        self.close_custody_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #ec7063;
            }
        """)
        buttons_layout.addWidget(self.close_custody_btn)

        self.refresh_custody_btn = QPushButton("تحديث")
        self.refresh_custody_btn.setIcon(qta.icon('fa5s.sync'))
        self.refresh_custody_btn.clicked.connect(self.load_data)
        self.refresh_custody_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #aab7b8;
            }
        """)
        buttons_layout.addWidget(self.refresh_custody_btn)

        top_layout.addLayout(buttons_layout)

        # مساحة فارغة
        top_layout.addStretch()

        # الفلاتر (في الجانب الأيسر)
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("الحالة:"))
        self.custody_status_filter = QComboBox()
        self.custody_status_filter.addItems(["جميع الحالات", "مفتوحة", "مغلقة", "مرحلة"])
        self.custody_status_filter.currentTextChanged.connect(self.filter_custody_combined)
        filter_layout.addWidget(self.custody_status_filter)

        # شريط البحث
        filter_layout.addWidget(QLabel("البحث:"))
        self.custody_search_edit = QLineEdit()
        self.custody_search_edit.setPlaceholderText("ابحث في العهد المالية...")
        self.custody_search_edit.textChanged.connect(self.filter_custody_combined)
        filter_layout.addWidget(self.custody_search_edit)

        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: البطاقات الإحصائية (في صف منفصل)
        self.create_custody_statistics_cards(layout)

        # جدول العهد المالية
        self.custody_table = QTableWidget()
        self.setup_custody_table()
        layout.addWidget(self.custody_table)

    # إعداد جدول العهد المالية
    def setup_custody_table(self):
        headers = [
            "المعرف", "رقم العهدة", "وصف العهدة", "مبلغ العهدة", "نسبة المكتب %",
            "مبلغ المكتب", "المبلغ الصافي", "المصروف", "المتبقي",
            "تاريخ الاستلام", "الحالة", "ملاحظات"
        ]

        self.custody_table.setColumnCount(len(headers))
        self.custody_table.setHorizontalHeaderLabels(headers)

        # إعداد خصائص الجدول
        self.custody_table.setAlternatingRowColors(True)
        self.custody_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_table.setSortingEnabled(True)

        # إخفاء عمود المعرف
        self.custody_table.setColumnHidden(0, True)

        # تعديل عرض الأعمدة
        header = self.custody_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # ربط الأحداث
        self.custody_table.itemDoubleClicked.connect(self.edit_custody)
        self.custody_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_table.customContextMenuRequested.connect(self.show_custody_context_menu)

    # إنشاء البطاقات الإحصائية للعهد المالية
    def create_custody_statistics_cards(self, parent_layout):
        # إنشاء حاوية للبطاقات الإحصائية
        stats_container = QFrame()

        # تخطيط أفقي للبطاقات
        stats_layout = QHBoxLayout(stats_container)

        # إنشاء البطاقات الإحصائية
        self.custody_total_count_label = QLabel("0")
        self.custody_total_amount_label = QLabel("0")
        self.custody_total_expenses_label = QLabel("0")
        self.custody_total_remaining_label = QLabel("0")

        # إنشاء البطاقات
        stats = [
            ("إجمالي العهد", self.custody_total_count_label, "#3498db"),           # أزرق للعدد
            ("إجمالي المبالغ", self.custody_total_amount_label, "#8e44ad"),        # بنفسجي للمبالغ
            ("إجمالي المصروفات", self.custody_total_expenses_label, "#e74c3c"),   # أحمر للمصروفات
            ("إجمالي المتبقي", self.custody_total_remaining_label, "#27ae60"),     # أخضر للمتبقي
        ]

        # إنشاء البطاقات
        for title, label, color in stats:
            card = self.create_custody_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

        # تحميل الإحصائيات الأولية
        self.update_custody_statistics()

    # إنشاء بطاقة إحصائية للعهد المالية
    def create_custody_stat_card(self, title, value_label, color):
        card = QFrame()
        card.setFixedHeight(80)
        card.setMinimumWidth(130)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 12px;
                padding: 0px;
                border: none;
                margin: 0px;
            }}
            QFrame:hover {{
                background-color: {self.adjust_color_brightness(color, -20)};
            }}
            QLabel {{
                color: white;
                font-weight: bold;
                background: transparent;
                border: none;
            }}
        """)

        card_layout = QVBoxLayout(card)
        card_layout.setAlignment(Qt.AlignCenter)

        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 12px; font-weight: bold;")
        title_label.setWordWrap(True)

        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 16px; font-weight: bold;")

        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)

        return card

    # تعديل سطوع اللون للتأثير عند التمرير
    def adjust_color_brightness(self, hex_color, amount):
        try:
            # إزالة # من بداية اللون
            hex_color = hex_color.lstrip('#')

            # تحويل إلى RGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            # تعديل السطوع
            r = max(0, min(255, r + amount))
            g = max(0, min(255, g + amount))
            b = max(0, min(255, b + amount))

            # إرجاع اللون الجديد
            return f"#{r:02x}{g:02x}{b:02x}"
        except:
            return hex_color

    # فلترة العهد المدمجة حسب الحالة والبحث
    def filter_custody_combined(self):
        selected_status = self.custody_status_filter.currentText()
        search_text = self.custody_search_edit.text().lower()

        for row in range(self.custody_table.rowCount()):
            show_row = True

            # فلترة حسب الحالة
            if selected_status != "جميع الحالات":
                status_item = self.custody_table.item(row, 10)  # عمود الحالة
                if status_item:
                    if selected_status != status_item.text():
                        show_row = False
                else:
                    show_row = False

            # فلترة حسب البحث
            if show_row and search_text:
                row_match = False
                for col in range(self.custody_table.columnCount()):
                    item = self.custody_table.item(row, col)
                    if item and search_text in item.text().lower():
                        row_match = True
                        break
                if not row_match:
                    show_row = False

            self.custody_table.setRowHidden(row, not show_row)

        # تحديث إحصائيات العهد بعد الفلترة
        self.update_custody_statistics()

    # تحديث إحصائيات العهد بناءً على البيانات المعروضة في الجدول
    def update_custody_statistics(self):
        try:
            # التحقق من وجود البطاقات
            if not hasattr(self, 'custody_total_count_label'):
                return

            # التحقق من وجود الجدول
            if not hasattr(self, 'custody_table'):
                return

            total_count = 0
            total_amount = 0
            total_expenses = 0
            total_remaining = 0

            # حساب الإحصائيات من الصفوف المرئية في الجدول
            for row in range(self.custody_table.rowCount()):
                if not self.custody_table.isRowHidden(row):
                    total_count += 1

                    # الحصول على المبالغ
                    amount_item = self.custody_table.item(row, 3)  # عمود مبلغ العهدة
                    expenses_item = self.custody_table.item(row, 7)  # عمود المصروف
                    remaining_item = self.custody_table.item(row, 8)  # عمود المتبقي

                    if amount_item and amount_item.text():
                        try:
                            amount = float(amount_item.text().replace(',', ''))
                            total_amount += amount
                        except ValueError:
                            continue

                    if expenses_item and expenses_item.text():
                        try:
                            expenses = float(expenses_item.text().replace(',', ''))
                            total_expenses += expenses
                        except ValueError:
                            continue

                    if remaining_item and remaining_item.text():
                        try:
                            remaining = float(remaining_item.text().replace(',', ''))
                            total_remaining += remaining
                        except ValueError:
                            continue

            # تحديث البطاقات
            self.custody_total_count_label.setText(f"{total_count} عهدة")
            self.custody_total_amount_label.setText(f"{total_amount:,.0f} {Currency_type}")
            self.custody_total_expenses_label.setText(f"{total_expenses:,.0f} {Currency_type}")
            self.custody_total_remaining_label.setText(f"{total_remaining:,.0f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحديث إحصائيات العهد: {e}")

    # إنشاء قسم إحصائيات العهد المالية
    def create_custody_statistics(self, layout):
        stats_frame = QFrame()
        stats_frame.setFrameStyle(QFrame.StyledPanel)
        stats_layout = QHBoxLayout(stats_frame)

        # بطاقات الإحصائيات
        self.total_custody_card = self.create_stat_card("إجمالي العهد", "0", "#3498db")
        self.total_amount_card = self.create_stat_card(f"إجمالي المبالغ", f"0 {Currency_type}", "#27ae60")
        self.total_expenses_card = self.create_stat_card(f"إجمالي المصروفات", f"0 {Currency_type}", "#e74c3c")
        self.total_remaining_card = self.create_stat_card(f"إجمالي المتبقي", f"0 {Currency_type}", "#f39c12")

        stats_layout.addWidget(self.total_custody_card)
        stats_layout.addWidget(self.total_amount_card)
        stats_layout.addWidget(self.total_expenses_card)
        stats_layout.addWidget(self.total_remaining_card)

        layout.addWidget(stats_frame)

    # إنشاء بطاقة إحصائية
    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setFrameStyle(QFrame.StyledPanel)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 10px;
                padding: 10px;
                margin: 5px;
            }}
            QLabel {{
                color: white;
                font-weight: bold;
            }}
        """)

        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)

        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 12px;")

        value_label = QLabel(value)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        value_label.setObjectName(f"{title}_value")

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        return card

    # تحميل بيانات العهد المالية
    def load_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT id, رقم_العهدة, وصف_العهدة, مبلغ_العهدة, نسبة_المكتب,
                       مبلغ_نسبة_المكتب, المبلغ_الصافي, المصروف, المتبقي,
                       تاريخ_الإستلام, حالة_العهدة, ملاحظات
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
                ORDER BY تاريخ_الإستلام DESC
            """

            cursor.execute(query, (self.project_id,))
            custody_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_table.setRowCount(len(custody_data))

            for row, data in enumerate(custody_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)

                    # تلوين الحالات
                    if col == 10:  # عمود الحالة
                        if value == "مفتوحة":
                            item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                        elif value == "مغلقة":
                            item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر
                        elif value == "مرحلة":
                            item.setForeground(QBrush(QColor(243, 156, 18)))  # برتقالي

                    self.custody_table.setItem(row, col, item)

            conn.close()
            self.update_custody_statistics()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات العهد المالية: {str(e)}")

    # تحديث الإحصائيات
    def update_statistics(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # إحصائيات العهد المالية
            cursor.execute("""
                SELECT
                    COUNT(*) as count_custody,
                    COALESCE(SUM(مبلغ_العهدة), 0) as total_amount,
                    COALESCE(SUM(المصروف), 0) as total_expenses,
                    COALESCE(SUM(المتبقي), 0) as total_remaining
                FROM المقاولات_العهد
                WHERE معرف_المشروع = %s
            """, (self.project_id,))

            result = cursor.fetchone()
            if result:
                count_custody, total_amount, total_expenses, total_remaining = result

                # تحديث البطاقات
                self.total_custody_card.findChild(QLabel, "إجمالي العهد_value").setText(str(count_custody))
                self.total_amount_card.findChild(QLabel, "إجمالي المبالغ_value").setText(f"{float(total_amount):,.2f}  {Currency_type}")
                self.total_expenses_card.findChild(QLabel, "إجمالي المصروفات_value").setText(f"{float(total_expenses):,.2f}  {Currency_type}")
                self.total_remaining_card.findChild(QLabel, "إجمالي المتبقي_value").setText(f"{float(total_remaining):,.2f}  {Currency_type}")

            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث الإحصائيات: {e}")

    # فلترة بيانات العهد المالية
    def filter_custody_data(self):
        search_text = self.custody_search_edit.text().lower()
        status_filter = self.custody_status_filter.currentText()

        for row in range(self.custody_table.rowCount()):
            show_row = True

            # فلترة النص
            if search_text:
                row_text = ""
                for col in range(self.custody_table.columnCount()):
                    item = self.custody_table.item(row, col)
                    if item:
                        row_text += item.text().lower() + " "

                if search_text not in row_text:
                    show_row = False

            # فلترة الحالة
            if status_filter != "جميع الحالات":
                status_item = self.custody_table.item(row, 10)  # عمود الحالة
                if status_item and status_item.text() != status_filter:
                    show_row = False

            self.custody_table.setRowHidden(row, not show_row)

    # عرض القائمة السياقية للعهد المالية
    def show_custody_context_menu(self, position):
        if self.custody_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        close_action = menu.addAction(qta.icon('fa5s.lock', color='#e74c3c'), "إغلاق العهدة")

        action = menu.exec_(self.custody_table.mapToGlobal(position))

        if action == view_action:
            self.view_custody_details()
        elif action == edit_action:
            self.edit_custody()
        elif action == close_action:
            self.close_custody()

    # إضافة عهدة مالية جديدة
    def add_custody(self):
        dialog = CustodyDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            self.update_custody_statistics()
            QMessageBox.information(self, "نجح", "تم إضافة العهدة المالية بنجاح")

    # تعديل عهدة مالية
    def edit_custody(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار عهدة للتعديل")
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        if not custody_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف العهدة")
            return

        custody_id = int(custody_id_item.text())
        dialog = CustodyDialog(self, project_id=self.project_id, custody_id=custody_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            self.update_custody_statistics()
            QMessageBox.information(self, "نجح", "تم تعديل العهدة المالية بنجاح")

    # إغلاق عهدة مالية
    def close_custody(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار عهدة للإغلاق")
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        custody_number_item = self.custody_table.item(current_row, 1)

        if not custody_id_item or not custody_number_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات العهدة")
            return

        custody_id = int(custody_id_item.text())
        custody_number = custody_number_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الإغلاق",
            f"هل تريد إغلاق العهدة رقم {custody_number}؟\n"
            "لن تتمكن من إضافة مصروفات جديدة لهذه العهدة بعد الإغلاق.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("""
                    UPDATE المقاولات_العهد
                    SET حالة_العهدة = 'مغلقة', تاريخ_الإغلاق = CURDATE()
                    WHERE id = %s
                """, (custody_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم إغلاق العهدة بنجاح")
                self.load_data()
                self.update_custody_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في إغلاق العهدة: {str(e)}")

    # عرض تفاصيل العهدة
    def view_custody_details(self):
        current_row = self.custody_table.currentRow()
        if current_row < 0:
            return

        custody_id_item = self.custody_table.item(current_row, 0)
        if not custody_id_item:
            return

        custody_id = int(custody_id_item.text())
        dialog = CustodyDetailsDialog(self, custody_id)
        dialog.exec_()


# تاب دفعات العهد المدمج
class CustodyPaymentsTabWidget(QWidget):

    # init
    def __init__(self, parent=None, project_id=None, project_data=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.project_data = project_data or {}

        self.create_interface()
        self.load_data()

    # إنشاء واجهة التاب
    def create_interface(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # تخطيط أفقي للأزرار والفلاتر
        top_layout = QHBoxLayout()

        # أزرار العمليات (في الجانب الأيمن)
        buttons_layout = QHBoxLayout()

        self.add_payment_btn = QPushButton("إضافة")
        self.add_payment_btn.setIcon(qta.icon('fa5s.plus'))
        self.add_payment_btn.clicked.connect(self.add_custody_payment)
        self.add_payment_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        buttons_layout.addWidget(self.add_payment_btn)

        self.edit_payment_btn = QPushButton("تعديل")
        self.edit_payment_btn.setIcon(qta.icon('fa5s.edit'))
        self.edit_payment_btn.clicked.connect(self.edit_custody_payment)
        self.edit_payment_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5dade2;
            }
        """)
        buttons_layout.addWidget(self.edit_payment_btn)

        self.delete_payment_btn = QPushButton("حذف")
        self.delete_payment_btn.setIcon(qta.icon('fa5s.trash'))
        self.delete_payment_btn.clicked.connect(self.delete_custody_payment)
        self.delete_payment_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #ec7063;
            }
        """)
        buttons_layout.addWidget(self.delete_payment_btn)

        self.refresh_payments_btn = QPushButton("تحديث")
        self.refresh_payments_btn.setIcon(qta.icon('fa5s.sync'))
        self.refresh_payments_btn.clicked.connect(self.load_data)
        self.refresh_payments_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #aab7b8;
            }
        """)
        buttons_layout.addWidget(self.refresh_payments_btn)

        top_layout.addLayout(buttons_layout)

        # مساحة فارغة
        top_layout.addStretch()

        # الفلاتر (في الجانب الأيسر)
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("العهدة:"))
        self.custody_filter_combo = QComboBox()
        self.custody_filter_combo.addItem("جميع العهد")
        self.custody_filter_combo.currentTextChanged.connect(self.filter_payments_combined)
        filter_layout.addWidget(self.custody_filter_combo)

        # شريط البحث
        filter_layout.addWidget(QLabel("البحث:"))
        self.payments_search_edit = QLineEdit()
        self.payments_search_edit.setPlaceholderText("ابحث في دفعات العهد...")
        self.payments_search_edit.textChanged.connect(self.filter_payments_combined)
        filter_layout.addWidget(self.payments_search_edit)

        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: البطاقات الإحصائية (في صف منفصل)
        self.create_payments_statistics_cards(layout)

        # جدول دفعات العهد
        self.custody_payments_table = QTableWidget()
        self.setup_custody_payments_table()
        layout.addWidget(self.custody_payments_table)

    # إعداد جدول دفعات العهد
    def setup_custody_payments_table(self):
        headers = [
            "المعرف", "رقم العهدة", "وصف الدفعة", "المبلغ", "تاريخ الدفعة",
            "طريقة الدفع", "المستلم", "ملاحظات"
        ]

        self.custody_payments_table.setColumnCount(len(headers))
        self.custody_payments_table.setHorizontalHeaderLabels(headers)

        # إعداد خصائص الجدول
        self.custody_payments_table.setAlternatingRowColors(True)
        self.custody_payments_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_payments_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_payments_table.setSortingEnabled(True)

        # إخفاء عمود المعرف
        self.custody_payments_table.setColumnHidden(0, True)

        # تعديل عرض الأعمدة
        header = self.custody_payments_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # ربط الأحداث
        self.custody_payments_table.itemDoubleClicked.connect(self.edit_custody_payment)
        self.custody_payments_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_payments_table.customContextMenuRequested.connect(self.show_payments_context_menu)

    # إنشاء البطاقات الإحصائية لدفعات العهد
    def create_payments_statistics_cards(self, parent_layout):
        # إنشاء حاوية للبطاقات الإحصائية
        stats_container = QFrame()

        # تخطيط أفقي للبطاقات
        stats_layout = QHBoxLayout(stats_container)

        # إنشاء البطاقات الإحصائية
        self.payments_total_count_label = QLabel("0")
        self.payments_total_amount_label = QLabel("0")
        self.payments_avg_amount_label = QLabel("0")

        # إنشاء البطاقات
        stats = [
            ("إجمالي الدفعات", self.payments_total_count_label, "#3498db"),        # أزرق للعدد
            ("إجمالي المبالغ", self.payments_total_amount_label, "#27ae60"),       # أخضر للمبالغ
            ("متوسط الدفعة", self.payments_avg_amount_label, "#8e44ad"),           # بنفسجي للمتوسط
        ]

        # إنشاء البطاقات
        for title, label, color in stats:
            card = self.create_payments_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

        # تحميل الإحصائيات الأولية
        self.update_payments_statistics()

    # إنشاء بطاقة إحصائية لدفعات العهد
    def create_payments_stat_card(self, title, value_label, color):
        card = QFrame()
        card.setFixedHeight(80)
        card.setMinimumWidth(130)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 12px;
                padding: 0px;
                border: none;
                margin: 0px;
            }}
            QFrame:hover {{
                background-color: {self.adjust_color_brightness(color, -20)};
            }}
            QLabel {{
                color: white;
                font-weight: bold;
                background: transparent;
                border: none;
            }}
        """)

        card_layout = QVBoxLayout(card)
        card_layout.setAlignment(Qt.AlignCenter)

        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 12px; font-weight: bold;")
        title_label.setWordWrap(True)

        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 16px; font-weight: bold;")

        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)

        return card

    # تعديل سطوع اللون للتأثير عند التمرير
    def adjust_color_brightness(self, hex_color, amount):
        try:
            # إزالة # من بداية اللون
            hex_color = hex_color.lstrip('#')

            # تحويل إلى RGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            # تعديل السطوع
            r = max(0, min(255, r + amount))
            g = max(0, min(255, g + amount))
            b = max(0, min(255, b + amount))

            # إرجاع اللون الجديد
            return f"#{r:02x}{g:02x}{b:02x}"
        except:
            return hex_color

    # فلترة دفعات العهد المدمجة حسب العهدة والبحث
    def filter_payments_combined(self):
        selected_custody = self.custody_filter_combo.currentText()
        search_text = self.payments_search_edit.text().lower()

        for row in range(self.custody_payments_table.rowCount()):
            show_row = True

            # فلترة حسب العهدة
            if selected_custody != "جميع العهد":
                custody_item = self.custody_payments_table.item(row, 1)  # عمود رقم العهدة
                if custody_item:
                    if selected_custody != custody_item.text():
                        show_row = False
                else:
                    show_row = False

            # فلترة حسب البحث
            if show_row and search_text:
                row_match = False
                for col in range(self.custody_payments_table.columnCount()):
                    item = self.custody_payments_table.item(row, col)
                    if item and search_text in item.text().lower():
                        row_match = True
                        break
                if not row_match:
                    show_row = False

            self.custody_payments_table.setRowHidden(row, not show_row)

        # تحديث إحصائيات دفعات العهد بعد الفلترة
        self.update_payments_statistics()

    # تحديث إحصائيات دفعات العهد بناءً على البيانات المعروضة في الجدول
    def update_payments_statistics(self):
        try:
            # التحقق من وجود البطاقات
            if not hasattr(self, 'payments_total_count_label'):
                return

            # التحقق من وجود الجدول
            if not hasattr(self, 'custody_payments_table'):
                return

            total_count = 0
            total_amount = 0

            # حساب الإحصائيات من الصفوف المرئية في الجدول
            for row in range(self.custody_payments_table.rowCount()):
                if not self.custody_payments_table.isRowHidden(row):
                    total_count += 1

                    # الحصول على المبلغ
                    amount_item = self.custody_payments_table.item(row, 3)  # عمود المبلغ

                    if amount_item and amount_item.text():
                        try:
                            amount = float(amount_item.text().replace(',', ''))
                            total_amount += amount
                        except ValueError:
                            continue

            # حساب المتوسط
            avg_amount = total_amount / total_count if total_count > 0 else 0

            # تحديث البطاقات
            self.payments_total_count_label.setText(f"{total_count} دفعة")
            self.payments_total_amount_label.setText(f"{total_amount:,.0f} {Currency_type}")
            if hasattr(self, 'payments_avg_amount_label'):
                self.payments_avg_amount_label.setText(f"{avg_amount:,.0f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحديث إحصائيات دفعات العهد: {e}")

    # تحميل بيانات دفعات العهد
    def load_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT dp.id, c.رقم_العهدة, dp.وصف_الدفعة, dp.المبلغ, dp.تاريخ_الدفعة,
                       dp.طريقة_الدفع, dp.المستلم, dp.ملاحظات
                FROM المقاولات_دفعات_العهد dp
                JOIN المقاولات_العهد c ON dp.معرف_العهدة = c.id
                WHERE c.معرف_المشروع = %s
                ORDER BY dp.تاريخ_الدفعة DESC
            """

            cursor.execute(query, (self.project_id,))
            payments_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_payments_table.setRowCount(len(payments_data))

            for row, data in enumerate(payments_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.custody_payments_table.setItem(row, col, item)

            # تحديث فلتر العهد
            self.update_custody_filter()
            # تحديث الإحصائيات
            self.update_payments_statistics()
            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات دفعات العهد: {str(e)}")

    # تحديث فلتر العهد
    def update_custody_filter(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT رقم_العهدة FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND رقم_العهدة IS NOT NULL
                ORDER BY رقم_العهدة
            """, (self.project_id,))

            custodies = cursor.fetchall()

            # تحديث ComboBox
            current_text = self.custody_filter_combo.currentText()
            self.custody_filter_combo.clear()
            self.custody_filter_combo.addItem("جميع العهد")

            for custody in custodies:
                self.custody_filter_combo.addItem(custody[0])

            # استعادة الاختيار السابق إن أمكن
            index = self.custody_filter_combo.findText(current_text)
            if index >= 0:
                self.custody_filter_combo.setCurrentIndex(index)

            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث فلتر العهد: {e}")

    # فلترة بيانات دفعات العهد
    def filter_payments_data(self):
        search_text = self.payments_search_edit.text().lower()
        custody_filter = self.custody_filter_combo.currentText()

        for row in range(self.custody_payments_table.rowCount()):
            show_row = True

            # فلترة النص
            if search_text:
                row_text = ""
                for col in range(self.custody_payments_table.columnCount()):
                    item = self.custody_payments_table.item(row, col)
                    if item:
                        row_text += item.text().lower() + " "

                if search_text not in row_text:
                    show_row = False

            # فلترة العهدة
            if custody_filter != "جميع العهد":
                custody_item = self.custody_payments_table.item(row, 1)  # عمود رقم العهدة
                if custody_item and custody_item.text() != custody_filter:
                    show_row = False

            self.custody_payments_table.setRowHidden(row, not show_row)

    # عرض القائمة السياقية لدفعات العهد
    def show_payments_context_menu(self, position):
        if self.custody_payments_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        delete_action = menu.addAction(qta.icon('fa5s.trash', color='#e74c3c'), "حذف")

        action = menu.exec_(self.custody_payments_table.mapToGlobal(position))

        if action == view_action:
            self.view_payment_details()
        elif action == edit_action:
            self.edit_custody_payment()
        elif action == delete_action:
            self.delete_custody_payment()

    # إضافة دفعة عهد جديدة
    def add_custody_payment(self):
        dialog = CustodyPaymentDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # تحديث تاب العهد المالية إذا كان متاحاً
            if hasattr(self.parent, 'load_custody_data'):
                self.parent.load_custody_data()
            self.update_payments_statistics()
            QMessageBox.information(self, "نجح", "تم إضافة دفعة العهد بنجاح")

    # تعديل دفعة عهد
    def edit_custody_payment(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار دفعة للتعديل")
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        if not payment_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الدفعة")
            return

        payment_id = int(payment_id_item.text())
        dialog = CustodyPaymentDialog(self, project_id=self.project_id, payment_id=payment_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # تحديث تاب العهد المالية إذا كان متاحاً
            if hasattr(self.parent, 'load_custody_data'):
                self.parent.load_custody_data()
            self.update_payments_statistics()
            QMessageBox.information(self, "نجح", "تم تعديل دفعة العهد بنجاح")

    # حذف دفعة عهد
    def delete_custody_payment(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار دفعة للحذف")
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        payment_desc_item = self.custody_payments_table.item(current_row, 2)

        if not payment_id_item or not payment_desc_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات الدفعة")
            return

        payment_id = int(payment_id_item.text())
        payment_desc = payment_desc_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل تريد حذف الدفعة '{payment_desc}'؟\n"
            "سيتم تحديث مبلغ العهدة تلقائياً.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("DELETE FROM المقاولات_دفعات_العهد WHERE id = %s", (payment_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم حذف دفعة العهد بنجاح")
                self.load_data()
                # تحديث تاب العهد المالية إذا كان متاحاً
                if hasattr(self.parent, 'load_custody_data'):
                    self.parent.load_custody_data()
                self.update_payments_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف دفعة العهد: {str(e)}")

    # عرض تفاصيل دفعة العهد
    def view_payment_details(self):
        current_row = self.custody_payments_table.currentRow()
        if current_row < 0:
            return

        payment_id_item = self.custody_payments_table.item(current_row, 0)
        if not payment_id_item:
            return

        payment_id = int(payment_id_item.text())
        dialog = PaymentDetailsDialog(self, payment_id)
        dialog.exec_()


# تاب مصروفات العهد المدمج
class CustodyExpensesTabWidget(QWidget):

    # init
    def __init__(self, parent=None, project_id=None, project_data=None):
        super().__init__(parent)
        self.parent = parent
        self.project_id = project_id
        self.project_data = project_data or {}

        self.create_interface()
        self.load_data()

    # إنشاء واجهة التاب
    def create_interface(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # تخطيط أفقي للأزرار والفلاتر
        top_layout = QHBoxLayout()

        # أزرار العمليات (في الجانب الأيمن)
        buttons_layout = QHBoxLayout()

        self.add_expense_btn = QPushButton("إضافة")
        self.add_expense_btn.setIcon(qta.icon('fa5s.plus'))
        self.add_expense_btn.clicked.connect(self.add_custody_expense)
        self.add_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        buttons_layout.addWidget(self.add_expense_btn)

        self.edit_expense_btn = QPushButton("تعديل")
        self.edit_expense_btn.setIcon(qta.icon('fa5s.edit'))
        self.edit_expense_btn.clicked.connect(self.edit_custody_expense)
        self.edit_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5dade2;
            }
        """)
        buttons_layout.addWidget(self.edit_expense_btn)

        self.delete_expense_btn = QPushButton("حذف")
        self.delete_expense_btn.setIcon(qta.icon('fa5s.trash'))
        self.delete_expense_btn.clicked.connect(self.delete_custody_expense)
        self.delete_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #ec7063;
            }
        """)
        buttons_layout.addWidget(self.delete_expense_btn)

        self.refresh_expenses_btn = QPushButton("تحديث")
        self.refresh_expenses_btn.setIcon(qta.icon('fa5s.sync'))
        self.refresh_expenses_btn.clicked.connect(self.load_data)
        self.refresh_expenses_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #aab7b8;
            }
        """)
        buttons_layout.addWidget(self.refresh_expenses_btn)

        top_layout.addLayout(buttons_layout)

        # مساحة فارغة
        top_layout.addStretch()

        # الفلاتر (في الجانب الأيسر)
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("نوع المصروف:"))
        self.expense_type_filter = QComboBox()
        self.expense_type_filter.addItems(["جميع الأنواع", "مرتبط_بعهدة", "غير_مرتبط_بعهدة", "خسائر", "مردودات"])
        self.expense_type_filter.currentTextChanged.connect(self.filter_expenses_combined)
        filter_layout.addWidget(self.expense_type_filter)

        filter_layout.addWidget(QLabel("العهدة:"))
        self.expense_custody_filter = QComboBox()
        self.expense_custody_filter.addItem("جميع العهد")
        self.expense_custody_filter.currentTextChanged.connect(self.filter_expenses_combined)
        filter_layout.addWidget(self.expense_custody_filter)

        # شريط البحث
        filter_layout.addWidget(QLabel("البحث:"))
        self.expenses_search_edit = QLineEdit()
        self.expenses_search_edit.setPlaceholderText("ابحث في المصروفات...")
        self.expenses_search_edit.textChanged.connect(self.filter_expenses_combined)
        filter_layout.addWidget(self.expenses_search_edit)

        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: البطاقات الإحصائية (في صف منفصل)
        self.create_expenses_statistics_cards(layout)

        # جدول مصروفات العهد
        self.custody_expenses_table = QTableWidget()
        self.setup_custody_expenses_table()
        layout.addWidget(self.custody_expenses_table)

    # إعداد جدول مصروفات العهد
    def setup_custody_expenses_table(self):
        headers = [
            "المعرف", "نوع المصروف", "رقم العهدة", "المسؤول", "وصف المصروف",
            "المبلغ", "تاريخ المصروف", "طريقة الدفع", "رقم الفاتورة",
            "المورد", "ملاحظات"
        ]

        self.custody_expenses_table.setColumnCount(len(headers))
        self.custody_expenses_table.setHorizontalHeaderLabels(headers)

        # إعداد خصائص الجدول
        self.custody_expenses_table.setAlternatingRowColors(True)
        self.custody_expenses_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.custody_expenses_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.custody_expenses_table.setSortingEnabled(True)

        # إخفاء عمود المعرف
        self.custody_expenses_table.setColumnHidden(0, True)

        # تعديل عرض الأعمدة
        header = self.custody_expenses_table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # ربط الأحداث
        self.custody_expenses_table.itemDoubleClicked.connect(self.edit_custody_expense)
        self.custody_expenses_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.custody_expenses_table.customContextMenuRequested.connect(self.show_expenses_context_menu)

    # إنشاء البطاقات الإحصائية لمصروفات العهد
    def create_expenses_statistics_cards(self, parent_layout):
        # إنشاء حاوية للبطاقات الإحصائية
        stats_container = QFrame()

        # تخطيط أفقي للبطاقات
        stats_layout = QHBoxLayout(stats_container)

        # إنشاء البطاقات الإحصائية
        self.expenses_total_count_label = QLabel("0")
        self.expenses_total_amount_label = QLabel("0")
        self.expenses_custody_amount_label = QLabel("0")
        self.expenses_losses_amount_label = QLabel("0")
        self.expenses_returns_amount_label = QLabel("0")

        # إنشاء البطاقات
        stats = [
            ("إجمالي المصروفات", self.expenses_total_count_label, "#3498db"),        # أزرق للعدد
            ("إجمالي المبالغ", self.expenses_total_amount_label, "#8e44ad"),         # بنفسجي للمبالغ
            ("مصروفات العهد", self.expenses_custody_amount_label, "#27ae60"),        # أخضر للعهد
            ("الخسائر", self.expenses_losses_amount_label, "#e74c3c"),               # أحمر للخسائر
            ("المردودات", self.expenses_returns_amount_label, "#f39c12"),            # برتقالي للمردودات
        ]

        # إنشاء البطاقات
        for title, label, color in stats:
            card = self.create_expenses_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

        # تحميل الإحصائيات الأولية
        self.update_expenses_statistics()

    # إنشاء بطاقة إحصائية لمصروفات العهد
    def create_expenses_stat_card(self, title, value_label, color):
        card = QFrame()
        card.setFixedHeight(80)
        card.setMinimumWidth(130)
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 12px;
                padding: 0px;
                border: none;
                margin: 0px;
            }}
            QFrame:hover {{
                background-color: {self.adjust_color_brightness(color, -20)};
            }}
            QLabel {{
                color: white;
                font-weight: bold;
                background: transparent;
                border: none;
            }}
        """)

        card_layout = QVBoxLayout(card)
        card_layout.setAlignment(Qt.AlignCenter)

        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 12px; font-weight: bold;")
        title_label.setWordWrap(True)

        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet("font-size: 16px; font-weight: bold;")

        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)

        return card

    # تعديل سطوع اللون للتأثير عند التمرير
    def adjust_color_brightness(self, hex_color, amount):
        try:
            # إزالة # من بداية اللون
            hex_color = hex_color.lstrip('#')

            # تحويل إلى RGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            # تعديل السطوع
            r = max(0, min(255, r + amount))
            g = max(0, min(255, g + amount))
            b = max(0, min(255, b + amount))

            # إرجاع اللون الجديد
            return f"#{r:02x}{g:02x}{b:02x}"
        except:
            return hex_color

    # فلترة مصروفات العهد المدمجة حسب النوع والعهدة والبحث
    def filter_expenses_combined(self):
        selected_type = self.expense_type_filter.currentText()
        selected_custody = self.expense_custody_filter.currentText()
        search_text = self.expenses_search_edit.text().lower()

        for row in range(self.custody_expenses_table.rowCount()):
            show_row = True

            # فلترة حسب نوع المصروف
            if selected_type != "جميع الأنواع":
                type_item = self.custody_expenses_table.item(row, 1)  # عمود نوع المصروف
                if type_item:
                    if selected_type != type_item.text():
                        show_row = False
                else:
                    show_row = False

            # فلترة حسب العهدة
            if show_row and selected_custody != "جميع العهد":
                custody_item = self.custody_expenses_table.item(row, 2)  # عمود رقم العهدة
                if custody_item:
                    if selected_custody != custody_item.text():
                        show_row = False
                else:
                    show_row = False

            # فلترة حسب البحث
            if show_row and search_text:
                row_match = False
                for col in range(self.custody_expenses_table.columnCount()):
                    item = self.custody_expenses_table.item(row, col)
                    if item and search_text in item.text().lower():
                        row_match = True
                        break
                if not row_match:
                    show_row = False

            self.custody_expenses_table.setRowHidden(row, not show_row)

        # تحديث إحصائيات المصروفات بعد الفلترة
        self.update_expenses_statistics()

    # تحديث إحصائيات مصروفات العهد بناءً على البيانات المعروضة في الجدول
    def update_expenses_statistics(self):
        try:
            # التحقق من وجود البطاقات
            if not hasattr(self, 'expenses_total_count_label'):
                return

            # التحقق من وجود الجدول
            if not hasattr(self, 'custody_expenses_table'):
                return

            total_count = 0
            total_amount = 0
            custody_amount = 0
            losses_amount = 0
            returns_amount = 0

            # حساب الإحصائيات من الصفوف المرئية في الجدول
            for row in range(self.custody_expenses_table.rowCount()):
                if not self.custody_expenses_table.isRowHidden(row):
                    total_count += 1

                    # الحصول على المبلغ ونوع المصروف
                    amount_item = self.custody_expenses_table.item(row, 5)  # عمود المبلغ
                    type_item = self.custody_expenses_table.item(row, 1)    # عمود نوع المصروف

                    if amount_item and amount_item.text():
                        try:
                            amount = float(amount_item.text().replace(',', ''))
                            total_amount += amount

                            # تصنيف المبلغ حسب النوع
                            if type_item:
                                expense_type = type_item.text()
                                if expense_type == "مرتبط_بعهدة":
                                    custody_amount += amount
                                elif expense_type == "خسائر":
                                    losses_amount += amount
                                elif expense_type == "مردودات":
                                    returns_amount += amount
                        except ValueError:
                            continue

            # تحديث البطاقات
            self.expenses_total_count_label.setText(f"{total_count} مصروف")
            self.expenses_total_amount_label.setText(f"{total_amount:,.0f} {Currency_type}")
            if hasattr(self, 'expenses_custody_amount_label'):
                self.expenses_custody_amount_label.setText(f"{custody_amount:,.0f} {Currency_type}")
            if hasattr(self, 'expenses_losses_amount_label'):
                self.expenses_losses_amount_label.setText(f"{losses_amount:,.0f} {Currency_type}")
            if hasattr(self, 'expenses_returns_amount_label'):
                self.expenses_returns_amount_label.setText(f"{returns_amount:,.0f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحديث إحصائيات المصروفات: {e}")

    # تحميل بيانات مصروفات العهد
    def load_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            query = """
                SELECT me.id, me.نوع_المصروف, COALESCE(c.رقم_العهدة, 'غير مرتبط'),
                       me.المسؤول, me.وصف_المصروف, me.المبلغ, me.تاريخ_المصروف,
                       me.طريقة_الدفع, me.رقم_الفاتورة, me.المورد, me.ملاحظات
                FROM المقاولات_مصروفات_العهد me
                LEFT JOIN المقاولات_العهد c ON me.معرف_العهدة = c.id
                WHERE me.معرف_المشروع = %s
                ORDER BY me.تاريخ_المصروف DESC
            """

            cursor.execute(query, (self.project_id,))
            expenses_data = cursor.fetchall()

            # تحديث الجدول
            self.custody_expenses_table.setRowCount(len(expenses_data))

            for row, data in enumerate(expenses_data):
                for col, value in enumerate(data):
                    if value is None:
                        value = ""
                    elif isinstance(value, Decimal):
                        value = f"{float(value):,.2f}"
                    elif isinstance(value, date):
                        value = value.strftime("%Y-%m-%d")

                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignCenter)

                    # تلوين أنواع المصروفات
                    if col == 1:  # عمود نوع المصروف
                        if value == "مرتبط_بعهدة":
                            item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                        elif value == "غير_مرتبط_بعهدة":
                            item.setForeground(QBrush(QColor(52, 152, 219)))  # أزرق
                        elif value == "خسائر":
                            item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر

                    self.custody_expenses_table.setItem(row, col, item)

            # تحديث فلتر العهد
            self.update_custody_filter()
            # تحديث الإحصائيات
            self.update_expenses_statistics()
            conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات مصروفات العهد: {str(e)}")

    # تحديث فلتر العهد في تاب المصروفات
    def update_custody_filter(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT رقم_العهدة FROM المقاولات_العهد
                WHERE معرف_المشروع = %s AND رقم_العهدة IS NOT NULL
                ORDER BY رقم_العهدة
            """, (self.project_id,))

            custodies = cursor.fetchall()

            # تحديث ComboBox
            current_text = self.expense_custody_filter.currentText()
            self.expense_custody_filter.clear()
            self.expense_custody_filter.addItem("جميع العهد")

            for custody in custodies:
                self.expense_custody_filter.addItem(custody[0])

            # استعادة الاختيار السابق إن أمكن
            index = self.expense_custody_filter.findText(current_text)
            if index >= 0:
                self.expense_custody_filter.setCurrentIndex(index)

            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث فلتر العهد: {e}")

    # فلترة بيانات مصروفات العهد - استخدام الدالة المدمجة
    def filter_expenses_data(self):
        self.filter_expenses_combined()

    # عرض القائمة السياقية لمصروفات العهد
    def show_expenses_context_menu(self, position):
        if self.custody_expenses_table.itemAt(position) is None:
            return

        menu = QMenu(self)
        menu.setLayoutDirection(Qt.RightToLeft)

        view_action = menu.addAction(qta.icon('fa5s.eye', color='#3498db'), "عرض التفاصيل")
        edit_action = menu.addAction(qta.icon('fa5s.edit', color='#f39c12'), "تعديل")
        delete_action = menu.addAction(qta.icon('fa5s.trash', color='#e74c3c'), "حذف")

        action = menu.exec_(self.custody_expenses_table.mapToGlobal(position))

        if action == view_action:
            self.view_expense_details()
        elif action == edit_action:
            self.edit_custody_expense()
        elif action == delete_action:
            self.delete_custody_expense()

    # إضافة مصروف عهد جديد
    def add_custody_expense(self):
        dialog = CustodyExpenseDialog(self, project_id=self.project_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # تحديث تاب العهد المالية إذا كان متاحاً
            if hasattr(self.parent, 'load_custody_data'):
                self.parent.load_custody_data()
            self.update_expenses_statistics()
            QMessageBox.information(self, "نجح", "تم إضافة مصروف العهد بنجاح")

    # تعديل مصروف عهد
    def edit_custody_expense(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار مصروف للتعديل")
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        if not expense_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المصروف")
            return

        expense_id = int(expense_id_item.text())
        dialog = CustodyExpenseDialog(self, project_id=self.project_id, expense_id=expense_id)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # تحديث تاب العهد المالية إذا كان متاحاً
            if hasattr(self.parent, 'load_custody_data'):
                self.parent.load_custody_data()
            self.update_expenses_statistics()
            QMessageBox.information(self, "نجح", "تم تعديل مصروف العهد بنجاح")

    # حذف مصروف عهد
    def delete_custody_expense(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار مصروف للحذف")
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        expense_desc_item = self.custody_expenses_table.item(current_row, 4)

        if not expense_id_item or not expense_desc_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على بيانات المصروف")
            return

        expense_id = int(expense_id_item.text())
        expense_desc = expense_desc_item.text()

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل تريد حذف المصروف '{expense_desc}'؟\n"
            "سيتم تحديث مبلغ العهدة تلقائياً.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user_r, password=password_r,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("DELETE FROM المقاولات_مصروفات_العهد WHERE id = %s", (expense_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", "تم حذف مصروف العهد بنجاح")
                self.load_data()
                # تحديث تاب العهد المالية إذا كان متاحاً
                if hasattr(self.parent, 'load_custody_data'):
                    self.parent.load_custody_data()
                self.update_expenses_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف مصروف العهد: {str(e)}")

    # عرض تفاصيل مصروف العهد
    def view_expense_details(self):
        current_row = self.custody_expenses_table.currentRow()
        if current_row < 0:
            return

        expense_id_item = self.custody_expenses_table.item(current_row, 0)
        if not expense_id_item:
            return

        expense_id = int(expense_id_item.text())
        dialog = ExpenseDetailsDialog(self, expense_id)
        dialog.exec_()


# ==================== دوال المساعدة ====================

# فتح نظام إدارة العهد المالية
def open_custody_management_system(parent=None, project_id=None, project_data=None):
    try:
        if not project_id:
            QMessageBox.warning(parent, "خطأ", "لا يمكن فتح نظام العهد المالية بدون معرف المشروع")
            return None

        dialog = CustodyManagementSystem(parent, project_id, project_data)
        dialog.show()
        return dialog

    except Exception as e:
        QMessageBox.warning(parent, "خطأ", f"فشل في فتح نظام إدارة العهد المالية: {str(e)}")
        return None


# ==================== اختبار النظام ====================

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # بيانات تجريبية للاختبار
    test_project_data = {
        'id': 1,
        'اسم_المشروع': 'مشروع تجريبي',
        'اسم_العميل': 'عميل تجريبي'
    }

    # فتح النظام
    window = CustodyManagementSystem(project_id=1, project_data=test_project_data)
    window.show()

    sys.exit(app.exec_())
