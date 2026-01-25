#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
 نافذة إدارة التدريب التفصيلية 
تحتوي على جميع الوظائف المطلوبة لإدارة الدورات التدريبية بشكل كامل
"""

import sys
import os
from datetime import datetime, date, timedelta
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
import mysql.connector
import qtawesome as qta

# إضافة المسار الحالي
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)

from الإعدادات_العامة import *
from ستايل import apply_stylesheet
from قائمة_الجداول import setup_table_context_menu
from متغيرات import *
from مساعد_أزرار_الطباعة import quick_add_print_button
from نظام_البطاقات import ModernCard, ModernCardsContainer
from أزرار_الواجهة import table_setting
from الأدوات import setup_date_edit_format

# نافذة شاملة لإدارة التدريب
class TrainingManagementWindow(QDialog):
    
    # init
    def __init__(self, parent=None, training_data=None):
        super().__init__(parent)
        self.parent = parent
        self.training_data = training_data or {}
        self.training_id = self.training_data.get('id', None)

        # تهيئة متغيرات الجداول
        self.groups_table = None
        self.students_table = None
        self.student_payments_table = None
        self.trainers_table = None
        self.trainer_payments_table = None
        self.expenses_table = None
        self.certificates_table = None

        # إعداد النافذة الأساسية
        self.setup_window()

        # إنشاء التابات
        self.create_tabs()

        # التأكد من وجود الأعمدة المطلوبة في قاعدة البيانات
        self.ensure_training_table_columns()

        # تحميل البيانات
        self.load_training_info()

        # تحديث جميع الحسابات التلقائية
        self.update_student_balances()
        self.update_course_totals()
        self.calculate_trainer_amounts()

        # تطبيق الستايل
        apply_stylesheet(self)

        # تطبيق الأنماط المركزية
        self.apply_training_management_styles()

        # إضافة أزرار الطباعة لجميع التابات
        self.add_print_buttons()

    # إعداد النافذة الأساسية
    def setup_window(self):
        training_name = self.training_data.get('عنوان_الدورة', 'دورة جديدة')
        self.setWindowTitle(f"إدارة التدريب - {training_name}")
        self.setGeometry(100, 100, 1600, 900)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setWindowFlags(Qt.Window | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
        
        # التخطيط الرئيسي
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # العنوان الرئيسي
        self.title_label = QLabel()
        self.title_label.setObjectName("main_title")
        self.title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(self.title_label)
        
        # إنشاء التابات
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
    # تحديث العنوان الرئيسي ليعكس التاب الحالي
    def update_title(self):
        try:
            training_name = self.training_data.get('عنوان_الدورة', 'دورة جديدة')
            current_tab_index = self.tab_widget.currentIndex()
            
            if current_tab_index >= 0:
                tab_text = self.tab_widget.tabText(current_tab_index)
                # إزالة أيقونات HTML من نص التاب إذا كانت موجودة
                import re
                clean_tab_text = re.sub(r'<[^>]+>', '', tab_text)
                title_text = f"إدارة تدريب {training_name} - {clean_tab_text}"
            else:
                title_text = f"إدارة تدريب {training_name}"
            
            self.title_label.setText(title_text)
            
        except Exception as e:
            print(f"خطأ في تحديث العنوان: {e}")
            self.title_label.setText(f"إدارة تدريب {self.training_data.get('عنوان_الدورة', 'دورة جديدة')}")
        
    # إنشاء التابات
    def create_tabs(self):
        # تاب معلومات الدورة (التبويب الأول)
        self.create_training_info_tab()
        
        # تاب إدارة المجموعات
        self.create_groups_management_tab()
        
        # تاب إدارة الطلبة
        self.create_students_management_tab()
        
        # تاب دفعات الطلبة
        self.create_student_payments_tab()
        
        # تاب المدربين
        self.create_trainers_tab()
        
        # تاب دفعات المدربين
        self.create_trainer_payments_tab()
        
        # تاب المصروفات
        self.create_expenses_tab()
        
        # تاب الشهادات
        self.create_certificates_tab()
        
        # تاب التقارير المالية
        self.create_financial_reports_tab()

        # ربط إشارة تغيير التاب بدالة التحديث التلقائي
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
        # تحديث العنوان الأولي
        self.update_title()

    # إنشاء تاب معلومات الدورة
    def create_training_info_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول - البطاقات الإحصائية
        self.create_training_statistics_cards(layout)

        # الصف الثاني - الحاويات الثلاث
        containers_layout = QHBoxLayout()
        containers_layout.setSpacing(15)

        # الحاوية الأولى - المعلومات الأساسية
        self.create_basic_info_container(containers_layout)

        # الحاوية الثانية - المعلومات المالية
        self.create_financial_info_container(containers_layout)

        # الحاوية الثالثة - المعلومات الإضافية
        self.create_additional_info_container(containers_layout)

        layout.addLayout(containers_layout)

        # صفوف العمليات الأخيرة
        self.create_recent_operations_section(layout)

        self.tab_widget.addTab(tab, qta.icon('fa5s.info-circle', color='#3498db'), "معلومات الدورة")

    # إنشاء البطاقات الإحصائية للدورة
    def create_training_statistics_cards(self, parent_layout):
        # إنشاء حاوية للبطاقات الإحصائية
        stats_container = QFrame()
        stats_container.setObjectName("training_stats_container")

        # تخطيط أفقي للبطاقات
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.students_count_label = QLabel("0")
        self.trainers_count_label = QLabel("0")
        self.total_revenue_label = QLabel(f"0 {Currency_type}")
        self.total_expenses_label = QLabel(f"0 {Currency_type}")
        self.net_profit_label = QLabel(f"0 {Currency_type}")

        # تعريف البطاقات الإحصائية
        stats = [
            ("عدد الطلبة", self.students_count_label, "#3498db"),      # أزرق للطلبة
            ("عدد المدربين", self.trainers_count_label, "#9b59b6"),    # بنفسجي للمدربين
            ("إجمالي الإيرادات", self.total_revenue_label, "#27ae60"),  # أخضر للإيرادات
            ("المصروفات", self.total_expenses_label, "#e74c3c"),       # أحمر للمصروفات
            ("صافي الربح", self.net_profit_label, "#f39c12"),          # برتقالي للربح
        ]

        # إنشاء البطاقات
        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

        # تحميل الإحصائيات الأولية
        self.update_training_statistics()

    # إنشاء بطاقة إحصائية للتدريب
    def create_training_stat_card(self, title, value_label, color):
        card = QFrame()
        card.setFixedHeight(80)
        card.setMinimumWidth(130)
        card.setObjectName("training_stat_card")
        
        # تخطيط البطاقة
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(10, 10, 10, 10)
        card_layout.setSpacing(5)

        # عنوان البطاقة
        title_label = QLabel(title)
        title_label.setObjectName("stat_card_title")
        title_label.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(title_label)

        # قيمة البطاقة
        value_label.setObjectName("stat_card_value")
        value_label.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(value_label)

        # تطبيق اللون
        card.setProperty("card_color", color)

        return card

    # إنشاء حاوية المعلومات الأساسية
    def create_basic_info_container(self, parent_layout):
        container = QGroupBox("المعلومات الأساسية")
        container.setObjectName("basic_info_container")
        layout = QVBoxLayout(container)
        layout.setSpacing(10)

        # معلومات الدورة
        info_layout = QFormLayout()
        info_layout.setSpacing(8)

        # عرض البيانات الأساسية
        self.course_name_label = QLabel(self.training_data.get('عنوان_الدورة', 'غير محدد'))
        self.category_label = QLabel(self.training_data.get('التصنيف', 'غير محدد'))
        self.trainer_label = QLabel(self.training_data.get('المدرب', 'غير محدد'))
        self.duration_label = QLabel(str(self.training_data.get('المدة', 'غير محدد')))

        info_layout.addRow("اسم الدورة:", self.course_name_label)
        info_layout.addRow("التصنيف:", self.category_label)
        info_layout.addRow("المدرب الرئيسي:", self.trainer_label)
        info_layout.addRow("المدة:", self.duration_label)

        layout.addLayout(info_layout)

        # أزرار العمليات
        buttons_layout = QHBoxLayout()

        edit_button = QPushButton("تعديل البيانات")
        edit_button.setObjectName("edit_data_button")
        edit_button.clicked.connect(self.edit_training_data)

        validate_button = QPushButton("التحقق من البيانات")
        validate_button.setObjectName("validate_data_button")
        validate_button.clicked.connect(self.validate_all_data)
        validate_button.setStyleSheet("QPushButton { background-color: #17a2b8; color: white; }")

        summary_button = QPushButton("الملخص المالي")
        summary_button.setObjectName("financial_summary_button")
        summary_button.clicked.connect(self.show_financial_summary)
        summary_button.setStyleSheet("QPushButton { background-color: #6f42c1; color: white; }")

        export_button = QPushButton("تصدير البيانات")
        export_button.setObjectName("export_data_button")
        export_button.clicked.connect(self.export_training_data)
        export_button.setStyleSheet("QPushButton { background-color: #28a745; color: white; }")

        buttons_layout.addWidget(edit_button)
        buttons_layout.addWidget(validate_button)
        buttons_layout.addWidget(summary_button)
        buttons_layout.addWidget(export_button)
        layout.addLayout(buttons_layout)

        parent_layout.addWidget(container)

    # إنشاء حاوية المعلومات المالية
    def create_financial_info_container(self, parent_layout):
        container = QGroupBox("المعلومات المالية")
        container.setObjectName("financial_info_container")
        layout = QVBoxLayout(container)
        layout.setSpacing(10)

        # معلومات مالية
        financial_layout = QFormLayout()
        financial_layout.setSpacing(8)

        # عرض البيانات المالية
        self.cost_label = QLabel(f"{self.training_data.get('التكلفة', 0)} {Currency_type}")
        self.total_amount_label = QLabel(f"{self.training_data.get('إجمالي_المبلغ', 0)} {Currency_type}")
        self.revenue_label = QLabel(f"0 {Currency_type}")
        self.expenses_label = QLabel(f"0 {Currency_type}")

        financial_layout.addRow("التكلفة:", self.cost_label)
        financial_layout.addRow("إجمالي المبلغ:", self.total_amount_label)
        financial_layout.addRow("الإيرادات:", self.revenue_label)
        financial_layout.addRow("المصروفات:", self.expenses_label)

        layout.addLayout(financial_layout)

        # زر إضافة دفعة طالب
        add_payment_button = QPushButton("إضافة دفعة طالب")
        add_payment_button.setObjectName("add_payment_button")
        add_payment_button.clicked.connect(self.add_student_payment)
        layout.addWidget(add_payment_button)

        parent_layout.addWidget(container)

    # إنشاء حاوية المعلومات الإضافية
    def create_additional_info_container(self, parent_layout):
        container = QGroupBox("المعلومات الإضافية")
        container.setObjectName("additional_info_container")
        layout = QVBoxLayout(container)
        layout.setSpacing(10)

        # معلومات إضافية
        additional_layout = QFormLayout()
        additional_layout.setSpacing(8)

        # عرض البيانات الإضافية
        self.start_date_label = QLabel(str(self.training_data.get('تاريخ_البدء', 'غير محدد')))
        self.end_date_label = QLabel(str(self.training_data.get('تاريخ_الإنتهاء', 'غير محدد')))
        self.status_label = QLabel(self.training_data.get('الحالة', 'غير محدد'))
        self.participants_label = QLabel(str(self.training_data.get('عدد_المشاركين', 0)))

        additional_layout.addRow("تاريخ البداية:", self.start_date_label)
        additional_layout.addRow("تاريخ النهاية:", self.end_date_label)
        additional_layout.addRow("حالة الدورة:", self.status_label)
        additional_layout.addRow("عدد المشاركين:", self.participants_label)

        layout.addLayout(additional_layout)

        # زر تغيير حالة الدورة
        change_status_button = QPushButton("تغيير حالة الدورة")
        change_status_button.setObjectName("change_status_button")
        change_status_button.clicked.connect(self.change_training_status)
        layout.addWidget(change_status_button)

        parent_layout.addWidget(container)

    # إنشاء قسم العمليات الأخيرة
    def create_recent_operations_section(self, parent_layout):
        operations_container = QGroupBox("العمليات الأخيرة")
        operations_container.setObjectName("recent_operations_container")
        layout = QVBoxLayout(operations_container)

        # قائمة العمليات الأخيرة (سيتم تنفيذها لاحقاً)
        operations_label = QLabel("سيتم عرض آخر العمليات هنا...")
        operations_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(operations_label)

        parent_layout.addWidget(operations_container)

    # تحميل معلومات الدورة
    def load_training_info(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات الدورة
            cursor.execute("""
                SELECT * FROM التدريب WHERE id = %s
            """, (self.training_id,))

            result = cursor.fetchone()
            if result:
                # تحديث البيانات المعروضة
                self.update_displayed_info()

            conn.close()

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تحميل بيانات الدورة: {str(e)}")

    # تحديث الإحصائيات
    def update_training_statistics(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # حساب عدد الطلبة الفعلي
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            students_count = cursor.fetchone()[0]

            # حساب عدد المدربين الفعلي
            cursor.execute("""
                SELECT COUNT(DISTINCT معرف_الموظف) FROM المشاريع_مهام_الفريق
                WHERE معرف_القسم = %s AND نوع_المهمة = 'مهمة تدريب'
            """, (self.training_id,))
            trainers_count = cursor.fetchone()[0]

            # حساب إجمالي الإيرادات من دفعات الطلبة
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ_المدفوع), 0) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_revenue = cursor.fetchone()[0]

            # حساب إجمالي المصروفات
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ), 0) FROM التدريب_مصروفات
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_expenses = cursor.fetchone()[0]

            # حساب إجمالي مبالغ المدربين
            cursor.execute("""
                SELECT COALESCE(SUM(مبلغ_الموظف), 0) FROM المشاريع_مهام_الفريق
                WHERE معرف_القسم = %s AND نوع_المهمة = 'مهمة تدريب'
            """, (self.training_id,))
            trainers_cost = cursor.fetchone()[0]

            # حساب صافي الربح
            net_profit = total_revenue - total_expenses - trainers_cost

            conn.close()

            # تحديث العناصر في الواجهة إذا كانت موجودة
            if hasattr(self, 'students_count_label'):
                self.students_count_label.setText(str(students_count))
            if hasattr(self, 'trainers_count_label'):
                self.trainers_count_label.setText(str(trainers_count))
            if hasattr(self, 'total_revenue_label'):
                self.total_revenue_label.setText(f"{total_revenue:,.2f} {Currency_type}")
            if hasattr(self, 'total_expenses_label'):
                self.total_expenses_label.setText(f"{total_expenses:,.2f} {Currency_type}")
            if hasattr(self, 'net_profit_label'):
                self.net_profit_label.setText(f"{net_profit:,.2f} {Currency_type}")

            # تحديث البيانات المحلية
            self.training_data['إجمالي_المبلغ'] = total_revenue
            self.training_data['عدد_المشاركين'] = students_count

        except Exception as e:
            print(f"خطأ في تحديث الإحصائيات: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحديث الإحصائيات: {str(e)}")

    # تحديث المعلومات المعروضة
    def update_displayed_info(self):
        if not self.training_data:
            return

        # تحديث المعلومات الأساسية
        self.course_name_label.setText(self.training_data.get('عنوان_الدورة', 'غير محدد'))
        self.category_label.setText(self.training_data.get('التصنيف', 'غير محدد'))
        self.trainer_label.setText(self.training_data.get('المدرب', 'غير محدد'))

        # تحديث المعلومات المالية
        cost = self.training_data.get('التكلفة', 0) or 0
        total_amount = self.training_data.get('إجمالي_المبلغ', 0) or 0

        self.cost_label.setText(f"{cost:,.2f} {Currency_type}")
        self.total_amount_label.setText(f"{total_amount:,.2f} {Currency_type}")

        # تحديث المعلومات الإضافية
        start_date = self.training_data.get('تاريخ_البدء', 'غير محدد')
        end_date = self.training_data.get('تاريخ_الإنتهاء', 'غير محدد')
        status = self.training_data.get('الحالة', 'غير محدد')
        participants = self.training_data.get('عدد_المشاركين', 0) or 0

        self.start_date_label.setText(str(start_date))
        self.end_date_label.setText(str(end_date))
        self.status_label.setText(status)
        self.participants_label.setText(str(participants))

    # معالج تغيير التاب
    def on_tab_changed(self, index):
        # تحديث العنوان الرئيسي
        self.update_title()
        
        # تحديث البيانات عند تغيير التاب
        tab_names = [
            "معلومات الدورة", "إدارة المجموعات", "إدارة الطلبة", "دفعات الطلبة",
            "المدربين", "دفعات المدربين", "المصروفات", "الشهادات", "التقارير المالية"
        ]

        if index < len(tab_names):
            current_tab = tab_names[index]

            # تحديث البيانات حسب التاب المحدد
            if current_tab == "معلومات الدورة":
                self.update_training_statistics()
                self.update_displayed_info()
            elif current_tab == "إدارة المجموعات":
                self.load_groups_data()
            elif current_tab == "إدارة الطلبة":
                self.load_students_data()
            elif current_tab == "دفعات الطلبة":
                self.load_student_payments_data()
            elif current_tab == "المدربين":
                self.load_trainers_data()
            elif current_tab == "دفعات المدربين":
                self.load_trainer_payments_data()
            elif current_tab == "المصروفات":
                self.load_expenses_data()
            elif current_tab == "الشهادات":
                self.load_certificates_data()
            elif current_tab == "التقارير المالية":
                self.generate_financial_summary()

    # تحميل بيانات المجموعات
    def load_groups_data(self):
        if not self.training_id or not self.groups_table:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات المجموعات مع عدد الطلبة المسجلين
            cursor.execute("""
                SELECT g.id, g.`اسم المجموعه`, g.التوقيت,
                       COALESCE(g.عدد_المشتركين, 0) as current_count,
                       g.العدد_المطلوب, g.الحالة, g.ملاحظات
                FROM التدريب_المجموعات g
                WHERE g.معرف_الدورة = %s
                ORDER BY g.`اسم المجموعه`
            """, (self.training_id,))

            groups = cursor.fetchall()

            # تحديث عدد المشتركين الفعلي من جدول الطلبة
            for i, group in enumerate(groups):
                group_id = group[0]
                cursor.execute("""
                    SELECT COUNT(*) FROM التدريب_الطلاب
                    WHERE معرف_المجموعة = %s
                """, (group_id,))
                actual_count = cursor.fetchone()[0]

                # تحديث عدد المشتركين في قاعدة البيانات
                cursor.execute("""
                    UPDATE التدريب_المجموعات
                    SET عدد_المشتركين = %s
                    WHERE id = %s
                """, (actual_count, group_id))

                # تحديث البيانات في القائمة
                groups[i] = (group[0], group[1], group[2], actual_count, group[4], group[5], group[6])

            conn.commit()
            conn.close()

            # تحديث الجدول
            self.groups_table.setRowCount(len(groups))

            for row, group in enumerate(groups):
                group_id, name, timing, current_count, required_count, status, notes = group

                # إضافة البيانات إلى الجدول
                self.groups_table.setItem(row, 0, QTableWidgetItem(str(group_id)))
                self.groups_table.setItem(row, 1, QTableWidgetItem(name or ""))
                self.groups_table.setItem(row, 2, QTableWidgetItem(timing or ""))
                self.groups_table.setItem(row, 3, QTableWidgetItem(str(current_count)))
                self.groups_table.setItem(row, 4, QTableWidgetItem(str(required_count) if required_count else ""))
                self.groups_table.setItem(row, 5, QTableWidgetItem(status or ""))
                self.groups_table.setItem(row, 6, QTableWidgetItem(notes or ""))

                # تلوين الحالة
                status_item = self.groups_table.item(row, 5)
                if status == 'بدأت الدورة':
                    status_item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                elif status == 'ملغاة':
                    status_item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر
                elif status == 'مكتملة العدد':
                    status_item.setForeground(QBrush(QColor(255, 152, 0)))  # برتقالي

            # تحديث الإحصائيات
            total_groups = len(groups)
            active_groups = sum(1 for g in groups if g[5] in ['مفتوحة للتسجيل', 'بدأت الدورة'])
            completed_groups = sum(1 for g in groups if g[5] == 'منتهية')

            if hasattr(self, 'total_groups_label'):
                self.total_groups_label.setText(str(total_groups))
            if hasattr(self, 'active_groups_label'):
                self.active_groups_label.setText(str(active_groups))
            if hasattr(self, 'completed_groups_label'):
                self.completed_groups_label.setText(str(completed_groups))

        except Exception as e:
            print(f"خطأ في تحميل بيانات المجموعات: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المجموعات: {str(e)}")

    # تحميل بيانات الطلبة
    def load_students_data(self):
        if not self.training_id or not self.students_table:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات الطلبة مع معلومات المجموعة
            cursor.execute("""
                SELECT s.id, s.اسم_الطالب, s.رقم_الهاتف,
                       COALESCE(g.`اسم المجموعه`, 'غير محدد') as group_name,
                       s.المبلغ, s.المدفوع, s.الباقي,
                       CASE
                           WHEN s.الباقي = 0 THEN 'مكتمل الدفع'
                           WHEN s.المدفوع = 0 THEN 'لم يدفع'
                           ELSE 'دفع جزئي'
                       END as payment_status,
                       s.تاريخ_التسجيل
                FROM التدريب_الطلاب s
                LEFT JOIN التدريب_المجموعات g ON s.معرف_المجموعة = g.id
                WHERE s.معرف_الدورة = %s
                ORDER BY s.اسم_الطالب
            """, (self.training_id,))

            students = cursor.fetchall()
            conn.close()

            # تحديث الجدول
            self.students_table.setRowCount(len(students))

            for row, student in enumerate(students):
                student_id, name, phone, group_name, amount, paid, remaining, status, reg_date = student

                # إضافة البيانات إلى الجدول
                self.students_table.setItem(row, 0, QTableWidgetItem(str(student_id)))
                self.students_table.setItem(row, 1, QTableWidgetItem(name or ""))
                self.students_table.setItem(row, 2, QTableWidgetItem(phone or ""))
                self.students_table.setItem(row, 3, QTableWidgetItem(group_name))
                self.students_table.setItem(row, 4, QTableWidgetItem(f"{amount:,.2f}" if amount else "0.00"))
                self.students_table.setItem(row, 5, QTableWidgetItem(f"{paid:,.2f}" if paid else "0.00"))
                self.students_table.setItem(row, 6, QTableWidgetItem(f"{remaining:,.2f}" if remaining else "0.00"))
                self.students_table.setItem(row, 7, QTableWidgetItem(status))
                self.students_table.setItem(row, 8, QTableWidgetItem(str(reg_date) if reg_date else ""))

                # تلوين حالة الدفع
                status_item = self.students_table.item(row, 7)
                if status == 'مكتمل الدفع':
                    status_item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                elif status == 'لم يدفع':
                    status_item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر
                else:  # دفع جزئي
                    status_item.setForeground(QBrush(QColor(255, 152, 0)))  # برتقالي

            # تحديث الإحصائيات
            total_students = len(students)
            paid_students = sum(1 for s in students if s[7] == 'مكتمل الدفع')
            unpaid_students = sum(1 for s in students if s[7] == 'لم يدفع')
            partial_paid_students = sum(1 for s in students if s[7] == 'دفع جزئي')

            if hasattr(self, 'total_students_label'):
                self.total_students_label.setText(str(total_students))
            if hasattr(self, 'paid_students_label'):
                self.paid_students_label.setText(str(paid_students))
            if hasattr(self, 'unpaid_students_label'):
                self.unpaid_students_label.setText(str(unpaid_students))
            if hasattr(self, 'partial_paid_students_label'):
                self.partial_paid_students_label.setText(str(partial_paid_students))

        except Exception as e:
            print(f"خطأ في تحميل بيانات الطلبة: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات الطلبة: {str(e)}")

    # تحميل بيانات دفعات الطلبة
    def load_student_payments_data(self):
        if not self.training_id or not self.student_payments_table:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات دفعات الطلبة
            cursor.execute("""
                SELECT p.id, s.اسم_الطالب,
                       COALESCE(g.`اسم المجموعه`, 'غير محدد') as group_name,
                       p.وصف_المدفوع, p.المبلغ_المدفوع, p.تاريخ_الدفع,
                       p.طريقة_الدفع, p.المستلم
                FROM التدريب_دفعات_الطلاب p
                JOIN التدريب_الطلاب s ON p.معرف_الطالب = s.id
                LEFT JOIN التدريب_المجموعات g ON p.معرف_المجموعة = g.id
                WHERE p.معرف_الدورة = %s
                ORDER BY p.تاريخ_الدفع DESC, s.اسم_الطالب
            """, (self.training_id,))

            payments = cursor.fetchall()
            conn.close()

            # تحديث الجدول
            self.student_payments_table.setRowCount(len(payments))

            total_payments = 0
            for row, payment in enumerate(payments):
                payment_id, student_name, group_name, description, amount, payment_date, method, receiver = payment

                # إضافة البيانات إلى الجدول
                self.student_payments_table.setItem(row, 0, QTableWidgetItem(str(payment_id)))
                self.student_payments_table.setItem(row, 1, QTableWidgetItem(student_name or ""))
                self.student_payments_table.setItem(row, 2, QTableWidgetItem(group_name))
                self.student_payments_table.setItem(row, 3, QTableWidgetItem(description or ""))
                self.student_payments_table.setItem(row, 4, QTableWidgetItem(f"{amount:,.2f}" if amount else "0.00"))
                self.student_payments_table.setItem(row, 5, QTableWidgetItem(str(payment_date) if payment_date else ""))
                self.student_payments_table.setItem(row, 6, QTableWidgetItem(method or ""))
                self.student_payments_table.setItem(row, 7, QTableWidgetItem(receiver or ""))

                total_payments += amount if amount else 0

            # تحديث الإحصائيات
            if hasattr(self, 'total_payments_label'):
                self.total_payments_label.setText(f"{total_payments:,.2f} {Currency_type}")

            # حساب دفعات الشهر الحالي
            current_month = QDate.currentDate().month()
            current_year = QDate.currentDate().year()
            monthly_payments = sum(
                payment[4] for payment in payments
                if payment[5] and QDate.fromString(str(payment[5]), Qt.ISODate).month() == current_month
                and QDate.fromString(str(payment[5]), Qt.ISODate).year() == current_year
            )

            if hasattr(self, 'monthly_payments_label'):
                self.monthly_payments_label.setText(f"{monthly_payments:,.2f} {Currency_type}")

            # حساب دفعات اليوم
            today = QDate.currentDate()
            daily_payments = sum(
                payment[4] for payment in payments
                if payment[5] and QDate.fromString(str(payment[5]), Qt.ISODate) == today
            )

            if hasattr(self, 'daily_payments_label'):
                self.daily_payments_label.setText(f"{daily_payments:,.2f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحميل بيانات دفعات الطلبة: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات دفعات الطلبة: {str(e)}")

    # تحميل بيانات المدربين
    def load_trainers_data(self):
        if not self.training_id or not self.trainers_table:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات مهام المدربين من جدول المشاريع_مهام_الفريق
            cursor.execute("""
                SELECT t.id, e.اسم_الموظف, e.الوظيفة, t.عنوان_المهمة,
                       t.نسبة_الموظف, t.مبلغ_الموظف, t.حالة_مبلغ_الموظف,
                       t.الحالة, t.تاريخ_البدء, t.تاريخ_الانتهاء
                FROM المشاريع_مهام_الفريق t
                JOIN الموظفين e ON t.معرف_الموظف = e.id
                WHERE t.نوع_المهمة = 'مهمة تدريب'
                AND t.معرف_الدورة_التدريبية = %s
                ORDER BY t.تاريخ_البدء
            """, (self.training_id,))

            trainers = cursor.fetchall()
            conn.close()

            # تحديث الجدول
            self.trainers_table.setRowCount(len(trainers))

            total_amount = 0
            for row, trainer in enumerate(trainers):
                task_id, name, job_title, task_title, percentage, amount, amount_status, status, start_date, end_date = trainer

                # إضافة البيانات إلى الجدول
                self.trainers_table.setItem(row, 0, QTableWidgetItem(str(task_id)))
                self.trainers_table.setItem(row, 1, QTableWidgetItem(name or ""))
                self.trainers_table.setItem(row, 2, QTableWidgetItem(job_title or ""))
                self.trainers_table.setItem(row, 3, QTableWidgetItem(task_title or ""))
                self.trainers_table.setItem(row, 4, QTableWidgetItem(f"{percentage:.2f}" if percentage else "0.00"))
                self.trainers_table.setItem(row, 5, QTableWidgetItem(f"{amount:,.2f}" if amount else "0.00"))
                self.trainers_table.setItem(row, 6, QTableWidgetItem(amount_status or "غير مدرج"))
                self.trainers_table.setItem(row, 7, QTableWidgetItem(status or "لم يبدأ"))
                self.trainers_table.setItem(row, 8, QTableWidgetItem(str(start_date) if start_date else ""))
                self.trainers_table.setItem(row, 9, QTableWidgetItem(str(end_date) if end_date else ""))

                # تلوين حالة المبلغ
                amount_status_item = self.trainers_table.item(row, 6)
                if amount_status == 'تم الإدراج':
                    amount_status_item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                else:
                    amount_status_item.setForeground(QBrush(QColor(231, 76, 60)))  # أحمر

                # تلوين حالة المهمة
                task_status_item = self.trainers_table.item(row, 7)
                if status == 'مكتمل':
                    task_status_item.setForeground(QBrush(QColor(46, 125, 50)))  # أخضر
                elif status == 'قيد التنفيذ':
                    task_status_item.setForeground(QBrush(QColor(255, 152, 0)))  # برتقالي

                if amount:
                    total_amount += amount

            # تحديث الإحصائيات
            total_trainers = len(trainers)
            active_trainers = sum(1 for t in trainers if t[7] in ['قيد التنفيذ', 'مكتمل'])

            if hasattr(self, 'total_trainers_label'):
                self.total_trainers_label.setText(str(total_trainers))
            if hasattr(self, 'active_trainers_label'):
                self.active_trainers_label.setText(str(active_trainers))
            if hasattr(self, 'trainers_cost_label'):
                self.trainers_cost_label.setText(f"{total_amount:,.2f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحميل بيانات المدربين: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المدربين: {str(e)}")

    # تحميل بيانات دفعات المدربين
    def load_trainer_payments_data(self):
        if not self.training_id:
            return

        try:
            # في الوقت الحالي، سنستخدم بيانات افتراضية
            self.trainer_payments_table.setRowCount(0)

            # تحديث الإحصائيات
            self.total_trainer_payments_label.setText(f"0 {Currency_type}")
            self.monthly_trainer_payments_label.setText(f"0 {Currency_type}")
            self.pending_trainer_payments_label.setText(f"0 {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحميل بيانات دفعات المدربين: {e}")

    # تحميل بيانات المصروفات
    def load_expenses_data(self):
        if not self.training_id or not self.expenses_table:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب بيانات المصروفات
            cursor.execute("""
                SELECT id, وصف_المصروف, نوع_المصروف, المبلغ, تاريخ_المصروف,
                       المستلم, رقم_الفاتورة
                FROM التدريب_مصروفات
                WHERE معرف_الدورة = %s
                ORDER BY تاريخ_المصروف DESC
            """, (self.training_id,))

            expenses = cursor.fetchall()
            conn.close()

            # تحديث الجدول
            self.expenses_table.setRowCount(len(expenses))

            total_expenses = 0
            for row, expense in enumerate(expenses):
                expense_id, description, expense_type, amount, expense_date, receiver, invoice_num = expense

                # إضافة البيانات إلى الجدول
                self.expenses_table.setItem(row, 0, QTableWidgetItem(str(expense_id)))
                self.expenses_table.setItem(row, 1, QTableWidgetItem(description or ""))
                self.expenses_table.setItem(row, 2, QTableWidgetItem(expense_type or ""))
                self.expenses_table.setItem(row, 3, QTableWidgetItem(f"{amount:,.2f}" if amount else "0.00"))
                self.expenses_table.setItem(row, 4, QTableWidgetItem(str(expense_date) if expense_date else ""))
                self.expenses_table.setItem(row, 5, QTableWidgetItem("نقدي"))  # افتراضي
                self.expenses_table.setItem(row, 6, QTableWidgetItem(receiver or ""))
                self.expenses_table.setItem(row, 7, QTableWidgetItem(invoice_num or ""))

                total_expenses += amount if amount else 0

            # تحديث الإحصائيات
            if hasattr(self, 'total_expenses_stat_label'):
                self.total_expenses_stat_label.setText(f"{total_expenses:,.2f} {Currency_type}")

            # حساب مصروفات الشهر الحالي
            current_month = QDate.currentDate().month()
            current_year = QDate.currentDate().year()
            monthly_expenses = sum(
                expense[3] for expense in expenses
                if expense[4] and QDate.fromString(str(expense[4]), Qt.ISODate).month() == current_month
                and QDate.fromString(str(expense[4]), Qt.ISODate).year() == current_year
            )

            if hasattr(self, 'monthly_expenses_label'):
                self.monthly_expenses_label.setText(f"{monthly_expenses:,.2f} {Currency_type}")

            # حساب مصروفات اليوم
            today = QDate.currentDate()
            daily_expenses = sum(
                expense[3] for expense in expenses
                if expense[4] and QDate.fromString(str(expense[4]), Qt.ISODate) == today
            )

            if hasattr(self, 'daily_expenses_label'):
                self.daily_expenses_label.setText(f"{daily_expenses:,.2f} {Currency_type}")

        except Exception as e:
            print(f"خطأ في تحميل بيانات المصروفات: {e}")
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المصروفات: {str(e)}")

    # تحميل بيانات الشهادات
    def load_certificates_data(self):
        if not self.training_id:
            return

        try:
            # في الوقت الحالي، سنستخدم بيانات افتراضية
            self.certificates_table.setRowCount(0)

            # تحديث الإحصائيات
            participants = self.training_data.get('عدد_المشاركين', 0) or 0
            self.total_certificates_label.setText(str(participants))
            self.generated_certificates_label.setText("0")
            self.printed_certificates_label.setText("0")
            self.delivered_certificates_label.setText("0")

        except Exception as e:
            print(f"خطأ في تحميل بيانات الشهادات: {e}")

    # إنشاء الملخص المالي
    def generate_financial_summary(self):
        if not self.training_id:
            return

        try:
            # حساب الملخص المالي
            total_revenue = self.training_data.get('إجمالي_المبلغ', 0) or 0
            total_expenses = self.training_data.get('التكلفة', 0) or 0
            net_profit = total_revenue - total_expenses
            profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

            # تحديث التقرير
            report_text = f"""
تقرير مالي شامل للدورة التدريبية
=====================================

اسم الدورة: {self.training_data.get('عنوان_الدورة', 'غير محدد')}
التصنيف: {self.training_data.get('التصنيف', 'غير محدد')}
المدرب: {self.training_data.get('المدرب', 'غير محدد')}

الملخص المالي:
--------------
إجمالي الإيرادات: {total_revenue:,.2f} {Currency_type}
إجمالي المصروفات: {total_expenses:,.2f} {Currency_type}
صافي الربح: {net_profit:,.2f} {Currency_type}
هامش الربح: {profit_margin:.2f}%

تفاصيل إضافية:
--------------
عدد المشاركين: {self.training_data.get('عدد_المشاركين', 0)}
تاريخ البداية: {self.training_data.get('تاريخ_البدء', 'غير محدد')}
تاريخ النهاية: {self.training_data.get('تاريخ_الإنتهاء', 'غير محدد')}
الحالة: {self.training_data.get('الحالة', 'غير محدد')}

ملاحظات:
--------
{self.training_data.get('ملاحظات', 'لا توجد ملاحظات')}
            """

            self.report_display_area.setPlainText(report_text)

        except Exception as e:
            print(f"خطأ في إنشاء الملخص المالي: {e}")

    # إنشاء تاب إدارة المجموعات
    def create_groups_management_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم في الفلاتر والبحث
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_group_btn = QPushButton("إضافة مجموعة")
        add_group_btn.setObjectName("add_group_button")
        edit_group_btn = QPushButton("تعديل مجموعة")
        edit_group_btn.setObjectName("edit_group_button")
        delete_group_btn = QPushButton("حذف مجموعة")
        delete_group_btn.setObjectName("delete_group_button")

        actions_layout.addWidget(add_group_btn)
        actions_layout.addWidget(edit_group_btn)
        actions_layout.addWidget(delete_group_btn)
        actions_layout.addStretch()

        # ربط الأزرار بالدوال
        add_group_btn.clicked.connect(self.add_group)
        edit_group_btn.clicked.connect(self.edit_group)
        delete_group_btn.clicked.connect(self.delete_group)

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في المجموعات...")
        search_edit.setObjectName("groups_search_edit")
        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للمجموعات
        self.create_groups_statistics_cards(layout)

        # الصف الثالث: جدول المجموعات
        self.groups_table = QTableWidget()
        self.setup_groups_table()
        layout.addWidget(self.groups_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.users', color='#9b59b6'), "إدارة المجموعات")

    # إنشاء البطاقات الإحصائية للمجموعات
    def create_groups_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("groups_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_groups_label = QLabel("0")
        self.active_groups_label = QLabel("0")
        self.completed_groups_label = QLabel("0")

        # تعريف البطاقات
        stats = [
            ("إجمالي المجموعات", self.total_groups_label, "#3498db"),
            ("المجموعات النشطة", self.active_groups_label, "#27ae60"),
            ("المجموعات المكتملة", self.completed_groups_label, "#95a5a6"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول المجموعات
    def setup_groups_table(self):
        headers = ["المعرف", "اسم المجموعة", "التوقيت", "عدد المشتركين", "العدد المطلوب", "الحالة", "ملاحظات"]
        self.groups_table.setColumnCount(len(headers))
        self.groups_table.setHorizontalHeaderLabels(headers)
        self.groups_table.setLayoutDirection(Qt.RightToLeft)
        self.groups_table.hideColumn(0)  # إخفاء عمود المعرف
        table_setting(self.groups_table)

        # إضافة وظيفة النقر المزدوج لفتح حوار التعديل
        self.groups_table.itemDoubleClicked.connect(self.edit_group)

    # إنشاء تاب إدارة الطلبة
    def create_students_management_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_student_btn = QPushButton("إضافة طالب")
        add_student_btn.setObjectName("add_student_button")
        edit_student_btn = QPushButton("تعديل طالب")
        edit_student_btn.setObjectName("edit_student_button")
        delete_student_btn = QPushButton("حذف طالب")
        delete_student_btn.setObjectName("delete_student_button")

        actions_layout.addWidget(add_student_btn)
        actions_layout.addWidget(edit_student_btn)
        actions_layout.addWidget(delete_student_btn)
        actions_layout.addStretch()

        # ربط الأزرار بالدوال
        add_student_btn.clicked.connect(self.add_student)
        edit_student_btn.clicked.connect(self.edit_student)
        delete_student_btn.clicked.connect(self.delete_student)

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في الطلبة...")
        search_edit.setObjectName("students_search_edit")

        group_filter = QComboBox()
        group_filter.addItem("جميع المجموعات")
        group_filter.setObjectName("students_group_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("المجموعة:"))
        filter_layout.addWidget(group_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للطلبة
        self.create_students_statistics_cards(layout)

        # الصف الثالث: جدول الطلبة
        self.students_table = QTableWidget()
        self.setup_students_table()
        layout.addWidget(self.students_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.graduation-cap', color='#3498db'), "إدارة الطلبة")

    # إنشاء البطاقات الإحصائية للطلبة
    def create_students_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("students_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_students_label = QLabel("0")
        self.paid_students_label = QLabel("0")
        self.unpaid_students_label = QLabel("0")
        self.partial_paid_students_label = QLabel("0")

        # تعريف البطاقات
        stats = [
            ("إجمالي الطلبة", self.total_students_label, "#3498db"),
            ("خالص", self.paid_students_label, "#27ae60"),
            ("لم يدفع", self.unpaid_students_label, "#e74c3c"),
            ("دفع جزئي", self.partial_paid_students_label, "#f39c12"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول الطلبة
    def setup_students_table(self):
        headers = ["المعرف", "الاسم", "رقم الهاتف", "المجموعة", "المبلغ", "المدفوع", "الباقي", "الحالة", "تاريخ التسجيل"]
        self.students_table.setColumnCount(len(headers))
        self.students_table.setHorizontalHeaderLabels(headers)
        self.students_table.setLayoutDirection(Qt.RightToLeft)
        self.students_table.hideColumn(0)  # إخفاء عمود المعرف
        table_setting(self.students_table)

        # إضافة وظيفة النقر المزدوج لفتح حوار التعديل
        self.students_table.itemDoubleClicked.connect(self.edit_student)

    # إنشاء تاب دفعات الطلبة
    def create_student_payments_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_payment_btn = QPushButton("إضافة دفعة")
        add_payment_btn.setObjectName("add_student_payment_button")
        edit_payment_btn = QPushButton("تعديل دفعة")
        edit_payment_btn.setObjectName("edit_student_payment_button")
        delete_payment_btn = QPushButton("حذف دفعة")
        delete_payment_btn.setObjectName("delete_student_payment_button")

        actions_layout.addWidget(add_payment_btn)
        actions_layout.addWidget(edit_payment_btn)
        actions_layout.addWidget(delete_payment_btn)
        actions_layout.addStretch()

        # ربط الأزرار بالدوال
        add_payment_btn.clicked.connect(self.add_student_payment)
        edit_payment_btn.clicked.connect(self.edit_student_payment)
        delete_payment_btn.clicked.connect(self.delete_student_payment)

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في الدفعات...")
        search_edit.setObjectName("payments_search_edit")

        date_filter = QDateEdit()
        date_filter.setDate(QDate.currentDate())
        date_filter.setObjectName("payments_date_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("التاريخ:"))
        filter_layout.addWidget(date_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للدفعات
        self.create_payments_statistics_cards(layout)

        # الصف الثالث: جدول دفعات الطلبة
        self.student_payments_table = QTableWidget()
        self.setup_student_payments_table()
        layout.addWidget(self.student_payments_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.money-bill', color='#27ae60'), "دفعات الطلبة")

    # إنشاء البطاقات الإحصائية للدفعات
    def create_payments_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("payments_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_payments_label = QLabel(f"0 {Currency_type}")
        self.monthly_payments_label = QLabel(f"0 {Currency_type}")
        self.daily_payments_label = QLabel(f"0 {Currency_type}")

        # تعريف البطاقات
        stats = [
            ("إجمالي الدفعات", self.total_payments_label, "#27ae60"),
            ("دفعات الشهر", self.monthly_payments_label, "#3498db"),
            ("دفعات اليوم", self.daily_payments_label, "#f39c12"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول دفعات الطلبة
    def setup_student_payments_table(self):
        headers = ["المعرف", "اسم الطالب", "المجموعة", "وصف المدفوع", "المبلغ", "تاريخ الدفع", "طريقة الدفع", "المستلم"]
        self.student_payments_table.setColumnCount(len(headers))
        self.student_payments_table.setHorizontalHeaderLabels(headers)
        self.student_payments_table.setLayoutDirection(Qt.RightToLeft)
        self.student_payments_table.hideColumn(0)  # إخفاء عمود المعرف
        table_setting(self.student_payments_table)

        # إضافة وظيفة النقر المزدوج لفتح حوار التعديل
        self.student_payments_table.itemDoubleClicked.connect(self.edit_student_payment)

    # دوال معالجة الأحداث
    # تعديل بيانات الدورة
    def edit_training_data(self):
        try:
            # إنشاء نافذة تعديل مخصصة للتدريب
            dialog = QDialog(self)
            dialog.setWindowTitle("تعديل بيانات الدورة")
            dialog.setFixedSize(500, 600)
            dialog.setLayoutDirection(Qt.RightToLeft)
            dialog.setModal(True)

            layout = QVBoxLayout(dialog)
            layout.setSpacing(15)
            layout.setContentsMargins(20, 20, 20, 20)

            # عنوان النافذة
            title_label = QLabel("تعديل بيانات الدورة التدريبية")
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
            layout.addWidget(title_label)

            # نموذج الإدخال
            form_layout = QFormLayout()
            form_layout.setSpacing(10)

            # حقول الإدخال
            course_title_edit = QLineEdit()
            course_title_edit.setText(self.training_data.get('عنوان_الدورة', ''))
            course_title_edit.setPlaceholderText("عنوان الدورة")

            category_combo = QComboBox()
            category_combo.addItems(["دورة هندسية", "دورة تقنية", "ورشة عمل"])
            current_category = self.training_data.get('التصنيف', '')
            if current_category:
                index = category_combo.findText(current_category)
                if index >= 0:
                    category_combo.setCurrentIndex(index)

            trainer_edit = QLineEdit()
            trainer_edit.setText(self.training_data.get('المدرب', ''))
            trainer_edit.setPlaceholderText("اسم المدرب")

            cost_edit = QLineEdit()
            cost_edit.setText(str(self.training_data.get('التكلفة', 0)))
            cost_edit.setPlaceholderText("التكلفة")

            total_amount_edit = QLineEdit()
            total_amount_edit.setText(str(self.training_data.get('إجمالي_المبلغ', 0)))
            total_amount_edit.setPlaceholderText("إجمالي المبلغ")

            participants_edit = QLineEdit()
            participants_edit.setText(str(self.training_data.get('عدد_المشاركين', 0)))
            participants_edit.setPlaceholderText("عدد المشاركين")

            groups_edit = QLineEdit()
            groups_edit.setText(str(self.training_data.get('عدد_المجموعات', 0)))
            groups_edit.setPlaceholderText("عدد المجموعات")

            start_date_edit = QDateEdit()
            start_date_edit.setDate(QDate.currentDate())
            start_date_edit.setDisplayFormat("dd/MM/yyyy")
            start_date = self.training_data.get('تاريخ_البدء')
            if start_date:
                try:
                    if isinstance(start_date, str):
                        start_date_edit.setDate(QDate.fromString(start_date, "yyyy-MM-dd"))
                    else:
                        start_date_edit.setDate(QDate(start_date))
                except:
                    pass

            end_date_edit = QDateEdit()
            end_date_edit.setDate(QDate.currentDate())
            end_date_edit.setDisplayFormat("dd/MM/yyyy")
            end_date = self.training_data.get('تاريخ_الإنتهاء')
            if end_date:
                try:
                    if isinstance(end_date, str):
                        end_date_edit.setDate(QDate.fromString(end_date, "yyyy-MM-dd"))
                    else:
                        end_date_edit.setDate(QDate(end_date))
                except:
                    pass

            status_combo = QComboBox()
            status_combo.addItems(["قيد التسجيل", "جارية", "معلق", "منتهية", "ملغاه"])
            current_status = self.training_data.get('الحالة', '')
            if current_status:
                index = status_combo.findText(current_status)
                if index >= 0:
                    status_combo.setCurrentIndex(index)

            notes_edit = QTextEdit()
            notes_edit.setPlainText(self.training_data.get('ملاحظات', ''))
            notes_edit.setMaximumHeight(100)
            notes_edit.setPlaceholderText("ملاحظات")

            # إضافة الحقول إلى النموذج
            form_layout.addRow("عنوان الدورة:", course_title_edit)
            form_layout.addRow("التصنيف:", category_combo)
            form_layout.addRow("المدرب:", trainer_edit)
            form_layout.addRow("التكلفة:", cost_edit)
            form_layout.addRow("إجمالي المبلغ:", total_amount_edit)
            form_layout.addRow("عدد المشاركين:", participants_edit)
            form_layout.addRow("عدد المجموعات:", groups_edit)
            form_layout.addRow("تاريخ البدء:", start_date_edit)
            form_layout.addRow("تاريخ الانتهاء:", end_date_edit)
            form_layout.addRow("الحالة:", status_combo)
            form_layout.addRow("ملاحظات:", notes_edit)

            layout.addLayout(form_layout)

            # أزرار الحفظ والإلغاء
            buttons_layout = QHBoxLayout()
            save_btn = QPushButton("حفظ التعديلات")
            save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 6px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
            """)

            cancel_btn = QPushButton("إلغاء")
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #6c757d;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 6px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #5a6268;
                }
            """)

            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            buttons_layout.addWidget(save_btn)
            buttons_layout.addWidget(cancel_btn)
            layout.addLayout(buttons_layout)

            # تنفيذ النافذة
            if dialog.exec() == QDialog.Accepted:
                # جمع البيانات المحدثة
                updated_data = {
                    'عنوان_الدورة': course_title_edit.text().strip(),
                    'التصنيف': category_combo.currentText(),
                    'المدرب': trainer_edit.text().strip(),
                    'التكلفة': float(cost_edit.text()) if cost_edit.text().strip() else 0,
                    'إجمالي_المبلغ': float(total_amount_edit.text()) if total_amount_edit.text().strip() else 0,
                    'عدد_المشاركين': int(participants_edit.text()) if participants_edit.text().strip() else 0,
                    'عدد_المجموعات': int(groups_edit.text()) if groups_edit.text().strip() else 0,
                    'تاريخ_البدء': start_date_edit.date().toString("yyyy-MM-dd"),
                    'تاريخ_الإنتهاء': end_date_edit.date().toString("yyyy-MM-dd"),
                    'الحالة': status_combo.currentText(),
                    'ملاحظات': notes_edit.toPlainText().strip()
                }

                # تحديث قاعدة البيانات
                self.save_training_updates(updated_data)

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في فتح نافذة التعديل: {str(e)}")

    # حفظ التحديثات في قاعدة البيانات
    def save_training_updates(self, updated_data):
        try:
            conn = mysql.connector.connect(
                host=host, user=user_r, password=password_r,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # تحديث البيانات
            cursor.execute("""
                UPDATE التدريب SET
                    عنوان_الدورة = %s,
                    التصنيف = %s,
                    المدرب = %s,
                    التكلفة = %s,
                    إجمالي_المبلغ = %s,
                    عدد_المشاركين = %s,
                    عدد_المجموعات = %s,
                    تاريخ_البدء = %s,
                    تاريخ_الإنتهاء = %s,
                    الحالة = %s,
                    ملاحظات = %s,
                    تاريخ_التحديث = NOW()
                WHERE id = %s
            """, (
                updated_data['عنوان_الدورة'],
                updated_data['التصنيف'],
                updated_data['المدرب'],
                updated_data['التكلفة'],
                updated_data['إجمالي_المبلغ'],
                updated_data['عدد_المشاركين'],
                updated_data['عدد_المجموعات'],
                updated_data['تاريخ_البدء'],
                updated_data['تاريخ_الإنتهاء'],
                updated_data['الحالة'],
                updated_data['ملاحظات'],
                self.training_id
            ))

            conn.commit()
            conn.close()

            # تحديث البيانات المحلية
            self.training_data.update(updated_data)

            # تحديث العرض
            self.update_displayed_info()
            self.update_training_statistics()

            QMessageBox.information(self, "نجح", "تم تحديث بيانات الدورة بنجاح")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في حفظ التحديثات: {str(e)}")

    # إضافة دفعة طالب
    def add_student_payment(self):
        try:
            # نافذة إدخال بسيطة لإضافة دفعة طالب
            dialog = QDialog(self)
            dialog.setWindowTitle("إضافة دفعة طالب")
            dialog.setFixedSize(400, 300)
            dialog.setLayoutDirection(Qt.RightToLeft)

            layout = QVBoxLayout(dialog)

            # حقول الإدخال
            form_layout = QFormLayout()

            student_name_edit = QLineEdit()
            student_name_edit.setPlaceholderText("اسم الطالب")

            amount_edit = QLineEdit()
            amount_edit.setPlaceholderText("المبلغ")

            payment_method_combo = QComboBox()
            payment_method_combo.addItems(["نقدي", "تحويل بنكي", "شيك", "بطاقة ائتمان"])

            notes_edit = QTextEdit()
            notes_edit.setPlaceholderText("ملاحظات")
            notes_edit.setMaximumHeight(80)

            form_layout.addRow("اسم الطالب:", student_name_edit)
            form_layout.addRow("المبلغ:", amount_edit)
            form_layout.addRow("طريقة الدفع:", payment_method_combo)
            form_layout.addRow("ملاحظات:", notes_edit)

            layout.addLayout(form_layout)

            # أزرار
            buttons_layout = QHBoxLayout()
            save_btn = QPushButton("حفظ")
            cancel_btn = QPushButton("إلغاء")

            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            buttons_layout.addWidget(save_btn)
            buttons_layout.addWidget(cancel_btn)
            layout.addLayout(buttons_layout)

            if dialog.exec() == QDialog.Accepted:
                # هنا يمكن إضافة كود حفظ الدفعة في قاعدة البيانات
                QMessageBox.information(self, "نجح", "تم إضافة الدفعة بنجاح")
                self.load_student_payments_data()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في إضافة الدفعة: {str(e)}")

    # تغيير حالة الدورة
    def change_training_status(self):
        try:
            # نافذة تغيير الحالة
            dialog = QDialog(self)
            dialog.setWindowTitle("تغيير حالة الدورة")
            dialog.setFixedSize(350, 200)
            dialog.setLayoutDirection(Qt.RightToLeft)

            layout = QVBoxLayout(dialog)

            # عرض الحالة الحالية
            current_status = self.training_data.get('الحالة', 'غير محدد')
            current_label = QLabel(f"الحالة الحالية: {current_status}")
            layout.addWidget(current_label)

            # اختيار الحالة الجديدة
            status_combo = QComboBox()
            status_combo.addItems([
                "قيد التسجيل", "جارية", "معلق", "منتهية", "ملغاه"
            ])

            # تحديد الحالة الحالية
            current_index = status_combo.findText(current_status)
            if current_index >= 0:
                status_combo.setCurrentIndex(current_index)

            form_layout = QFormLayout()
            form_layout.addRow("الحالة الجديدة:", status_combo)
            layout.addLayout(form_layout)

            # أزرار
            buttons_layout = QHBoxLayout()
            save_btn = QPushButton("حفظ")
            cancel_btn = QPushButton("إلغاء")

            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            buttons_layout.addWidget(save_btn)
            buttons_layout.addWidget(cancel_btn)
            layout.addLayout(buttons_layout)

            if dialog.exec() == QDialog.Accepted:
                new_status = status_combo.currentText()

                # تحديث الحالة في قاعدة البيانات
                try:
                    conn = mysql.connector.connect(
                        host=host, user=user_r, password=password_r,
                        database="project_manager_V2"
                    )
                    cursor = conn.cursor()

                    cursor.execute("""
                        UPDATE التدريب SET الحالة = %s WHERE id = %s
                    """, (new_status, self.training_id))

                    conn.commit()
                    conn.close()

                    # تحديث البيانات المحلية
                    self.training_data['الحالة'] = new_status
                    self.update_displayed_info()

                    QMessageBox.information(self, "نجح", f"تم تغيير حالة الدورة إلى: {new_status}")

                except Exception as db_error:
                    QMessageBox.warning(self, "خطأ", f"فشل في تحديث قاعدة البيانات: {str(db_error)}")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تغيير الحالة: {str(e)}")

    # إنشاء تاب المدربين
    def create_trainers_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_trainer_btn = QPushButton("إضافة مدرب")
        add_trainer_btn.setObjectName("add_trainer_button")
        edit_trainer_btn = QPushButton("تعديل مدرب")
        edit_trainer_btn.setObjectName("edit_trainer_button")
        delete_trainer_btn = QPushButton("حذف مدرب")
        delete_trainer_btn.setObjectName("delete_trainer_button")

        actions_layout.addWidget(add_trainer_btn)
        actions_layout.addWidget(edit_trainer_btn)
        actions_layout.addWidget(delete_trainer_btn)

        # زر إدراج الأرصدة
        include_balances_btn = QPushButton("إدراج الأرصدة")
        include_balances_btn.setObjectName("include_trainer_balances_button")
        include_balances_btn.setStyleSheet("QPushButton { background-color: #28a745; color: white; }")
        actions_layout.addWidget(include_balances_btn)

        actions_layout.addStretch()

        # ربط الأزرار بالدوال
        add_trainer_btn.clicked.connect(self.add_trainer)
        edit_trainer_btn.clicked.connect(self.edit_trainer)
        delete_trainer_btn.clicked.connect(self.delete_trainer)
        include_balances_btn.clicked.connect(self.include_trainer_balances)

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في المدربين...")
        search_edit.setObjectName("trainers_search_edit")

        specialty_filter = QComboBox()
        specialty_filter.addItem("جميع التخصصات")
        specialty_filter.setObjectName("trainers_specialty_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("التخصص:"))
        filter_layout.addWidget(specialty_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للمدربين
        self.create_trainers_statistics_cards(layout)

        # الصف الثالث: جدول المدربين
        self.trainers_table = QTableWidget()
        self.setup_trainers_table()
        layout.addWidget(self.trainers_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.chalkboard-teacher', color='#9b59b6'), "المدربين")

    # إنشاء البطاقات الإحصائية للمدربين
    def create_trainers_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("trainers_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_trainers_label = QLabel("0")
        self.active_trainers_label = QLabel("0")
        self.trainers_cost_label = QLabel(f"0 {Currency_type}")

        # تعريف البطاقات
        stats = [
            ("إجمالي المدربين", self.total_trainers_label, "#9b59b6"),
            ("المدربين النشطين", self.active_trainers_label, "#27ae60"),
            ("إجمالي التكلفة", self.trainers_cost_label, "#e74c3c"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول المدربين
    def setup_trainers_table(self):
        headers = ["معرف المهمة", "اسم المدرب", "الوظيفة", "عنوان المهمة", "النسبة %", "مبلغ المدرب", "حالة المبلغ", "حالة المهمة", "تاريخ البدء", "تاريخ الانتهاء"]
        self.trainers_table.setColumnCount(len(headers))
        self.trainers_table.setHorizontalHeaderLabels(headers)
        self.trainers_table.setLayoutDirection(Qt.RightToLeft)
        self.trainers_table.hideColumn(0)  # إخفاء عمود معرف المهمة
        table_setting(self.trainers_table)

        # إضافة وظيفة النقر المزدوج لفتح حوار التعديل
        self.trainers_table.itemDoubleClicked.connect(self.edit_trainer)

    # إنشاء تاب دفعات المدربين
    def create_trainer_payments_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_trainer_payment_btn = QPushButton("إضافة دفعة مدرب")
        add_trainer_payment_btn.setObjectName("add_trainer_payment_button")
        edit_trainer_payment_btn = QPushButton("تعديل دفعة")
        edit_trainer_payment_btn.setObjectName("edit_trainer_payment_button")
        delete_trainer_payment_btn = QPushButton("حذف دفعة")
        delete_trainer_payment_btn.setObjectName("delete_trainer_payment_button")

        actions_layout.addWidget(add_trainer_payment_btn)
        actions_layout.addWidget(edit_trainer_payment_btn)
        actions_layout.addWidget(delete_trainer_payment_btn)
        actions_layout.addStretch()

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في دفعات المدربين...")
        search_edit.setObjectName("trainer_payments_search_edit")

        trainer_filter = QComboBox()
        trainer_filter.addItem("جميع المدربين")
        trainer_filter.setObjectName("trainer_payments_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("المدرب:"))
        filter_layout.addWidget(trainer_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية لدفعات المدربين
        self.create_trainer_payments_statistics_cards(layout)

        # الصف الثالث: جدول دفعات المدربين
        self.trainer_payments_table = QTableWidget()
        self.setup_trainer_payments_table()
        layout.addWidget(self.trainer_payments_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.hand-holding-usd', color='#f39c12'), "دفعات المدربين")

    # إنشاء البطاقات الإحصائية لدفعات المدربين
    def create_trainer_payments_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("trainer_payments_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_trainer_payments_label = QLabel(f"0 {Currency_type}")
        self.monthly_trainer_payments_label = QLabel(f"0 {Currency_type}")
        self.pending_trainer_payments_label = QLabel(f"0 {Currency_type}")

        # تعريف البطاقات
        stats = [
            ("إجمالي المدفوعات", self.total_trainer_payments_label, "#f39c12"),
            ("مدفوعات الشهر", self.monthly_trainer_payments_label, "#3498db"),
            ("مدفوعات معلقة", self.pending_trainer_payments_label, "#e74c3c"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول دفعات المدربين
    def setup_trainer_payments_table(self):
        headers = ["المعرف", "اسم المدرب", "المبلغ", "تاريخ الدفعة", "طريقة الدفع", "الحالة", "ملاحظات"]
        self.trainer_payments_table.setColumnCount(len(headers))
        self.trainer_payments_table.setHorizontalHeaderLabels(headers)
        self.trainer_payments_table.setLayoutDirection(Qt.RightToLeft)
        table_setting(self.trainer_payments_table)

    # إنشاء تاب المصروفات
    def create_expenses_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        add_expense_btn = QPushButton("إضافة مصروف")
        add_expense_btn.setObjectName("add_expense_button")
        edit_expense_btn = QPushButton("تعديل مصروف")
        edit_expense_btn.setObjectName("edit_expense_button")
        delete_expense_btn = QPushButton("حذف مصروف")
        delete_expense_btn.setObjectName("delete_expense_button")

        actions_layout.addWidget(add_expense_btn)
        actions_layout.addWidget(edit_expense_btn)
        actions_layout.addWidget(delete_expense_btn)
        actions_layout.addStretch()

        # ربط الأزرار بالدوال
        add_expense_btn.clicked.connect(self.add_expense)
        edit_expense_btn.clicked.connect(self.edit_expense)
        delete_expense_btn.clicked.connect(self.delete_expense)

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في المصروفات...")
        search_edit.setObjectName("expenses_search_edit")

        expense_type_filter = QComboBox()
        expense_type_filter.addItem("جميع الأنواع")
        expense_type_filter.setObjectName("expenses_type_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("نوع المصروف:"))
        filter_layout.addWidget(expense_type_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للمصروفات
        self.create_expenses_statistics_cards(layout)

        # الصف الثالث: جدول المصروفات
        self.expenses_table = QTableWidget()
        self.setup_expenses_table()
        layout.addWidget(self.expenses_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.receipt', color='#e74c3c'), "المصروفات")

    # إنشاء البطاقات الإحصائية للمصروفات
    def create_expenses_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("expenses_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_expenses_stat_label = QLabel(f"0 {Currency_type}")
        self.monthly_expenses_label = QLabel(f"0 {Currency_type}")
        self.daily_expenses_label = QLabel(f"0 {Currency_type}")

        # تعريف البطاقات
        stats = [
            ("إجمالي المصروفات", self.total_expenses_stat_label, "#e74c3c"),
            ("مصروفات الشهر", self.monthly_expenses_label, "#f39c12"),
            ("مصروفات اليوم", self.daily_expenses_label, "#95a5a6"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول المصروفات
    def setup_expenses_table(self):
        headers = ["المعرف", "وصف المصروف", "نوع المصروف", "المبلغ", "تاريخ المصروف", "طريقة الدفع", "المستلم", "رقم الفاتورة"]
        self.expenses_table.setColumnCount(len(headers))
        self.expenses_table.setHorizontalHeaderLabels(headers)
        self.expenses_table.setLayoutDirection(Qt.RightToLeft)
        self.expenses_table.hideColumn(0)  # إخفاء عمود المعرف
        table_setting(self.expenses_table)

        # إضافة وظيفة النقر المزدوج لفتح حوار التعديل
        self.expenses_table.itemDoubleClicked.connect(self.edit_expense)

    # إنشاء تاب الشهادات
    def create_certificates_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # الصف الأول: أزرار الإجراءات + عناصر التحكم
        top_layout = QHBoxLayout()

        # أزرار الإجراءات
        actions_layout = QHBoxLayout()
        generate_certificates_btn = QPushButton("إنشاء شهادات")
        generate_certificates_btn.setObjectName("generate_certificates_button")
        print_certificate_btn = QPushButton("طباعة شهادة")
        print_certificate_btn.setObjectName("print_certificate_button")
        export_certificates_btn = QPushButton("تصدير الشهادات")
        export_certificates_btn.setObjectName("export_certificates_button")

        actions_layout.addWidget(generate_certificates_btn)
        actions_layout.addWidget(print_certificate_btn)
        actions_layout.addWidget(export_certificates_btn)
        actions_layout.addStretch()

        # عناصر البحث والفلترة
        filter_layout = QHBoxLayout()
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("البحث في الشهادات...")
        search_edit.setObjectName("certificates_search_edit")

        status_filter = QComboBox()
        status_filter.addItem("جميع الحالات")
        status_filter.addItem("مُنشأة")
        status_filter.addItem("مطبوعة")
        status_filter.addItem("مُسلمة")
        status_filter.setObjectName("certificates_status_filter")

        filter_layout.addWidget(QLabel("البحث:"))
        filter_layout.addWidget(search_edit)
        filter_layout.addWidget(QLabel("الحالة:"))
        filter_layout.addWidget(status_filter)

        top_layout.addLayout(actions_layout)
        top_layout.addLayout(filter_layout)
        layout.addLayout(top_layout)

        # الصف الثاني: بطاقات إحصائية للشهادات
        self.create_certificates_statistics_cards(layout)

        # الصف الثالث: جدول الشهادات
        self.certificates_table = QTableWidget()
        self.setup_certificates_table()
        layout.addWidget(self.certificates_table)

        self.tab_widget.addTab(tab, qta.icon('fa5s.certificate', color='#f39c12'), "الشهادات")

    # إنشاء البطاقات الإحصائية للشهادات
    def create_certificates_statistics_cards(self, parent_layout):
        stats_container = QFrame()
        stats_container.setObjectName("certificates_stats_container")
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setSpacing(15)

        # تهيئة التسميات
        self.total_certificates_label = QLabel("0")
        self.generated_certificates_label = QLabel("0")
        self.printed_certificates_label = QLabel("0")
        self.delivered_certificates_label = QLabel("0")

        # تعريف البطاقات
        stats = [
            ("إجمالي الشهادات", self.total_certificates_label, "#f39c12"),
            ("مُنشأة", self.generated_certificates_label, "#3498db"),
            ("مطبوعة", self.printed_certificates_label, "#9b59b6"),
            ("مُسلمة", self.delivered_certificates_label, "#27ae60"),
        ]

        for title, label, color in stats:
            card = self.create_training_stat_card(title, label, color)
            stats_layout.addWidget(card)

        parent_layout.addWidget(stats_container)

    # إعداد جدول الشهادات
    def setup_certificates_table(self):
        headers = ["المعرف", "اسم الطالب", "المجموعة", "رقم الشهادة", "تاريخ الإنشاء", "الحالة", "ملاحظات"]
        self.certificates_table.setColumnCount(len(headers))
        self.certificates_table.setHorizontalHeaderLabels(headers)
        self.certificates_table.setLayoutDirection(Qt.RightToLeft)
        table_setting(self.certificates_table)

    # إنشاء تاب التقارير المالية
    def create_financial_reports_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)

        # عنوان التقارير المالية
        title_label = QLabel("التقارير المالية للدورة")
        title_label.setObjectName("financial_reports_title")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # قسم الملخص المالي
        summary_container = QGroupBox("الملخص المالي")
        summary_container.setObjectName("financial_summary_container")
        summary_layout = QVBoxLayout(summary_container)

        # بطاقات الملخص المالي
        summary_cards_layout = QHBoxLayout()

        # إجمالي الإيرادات
        revenue_card = self.create_financial_summary_card("إجمالي الإيرادات", f"0 {Currency_type}", "#27ae60")
        summary_cards_layout.addWidget(revenue_card)

        # إجمالي المصروفات
        expenses_card = self.create_financial_summary_card("إجمالي المصروفات", f"0 {Currency_type}", "#e74c3c")
        summary_cards_layout.addWidget(expenses_card)

        # صافي الربح
        profit_card = self.create_financial_summary_card("صافي الربح", f"0 {Currency_type}", "#f39c12")
        summary_cards_layout.addWidget(profit_card)

        # هامش الربح
        margin_card = self.create_financial_summary_card("هامش الربح", "0%", "#9b59b6")
        summary_cards_layout.addWidget(margin_card)

        summary_layout.addLayout(summary_cards_layout)
        layout.addWidget(summary_container)

        # قسم التقارير التفصيلية
        reports_container = QGroupBox("التقارير التفصيلية")
        reports_container.setObjectName("detailed_reports_container")
        reports_layout = QVBoxLayout(reports_container)

        # أزرار التقارير
        reports_buttons_layout = QHBoxLayout()

        revenue_report_btn = QPushButton("تقرير الإيرادات")
        revenue_report_btn.setObjectName("revenue_report_button")

        expenses_report_btn = QPushButton("تقرير المصروفات")
        expenses_report_btn.setObjectName("expenses_report_button")

        students_report_btn = QPushButton("تقرير الطلبة")
        students_report_btn.setObjectName("students_report_button")

        trainers_report_btn = QPushButton("تقرير المدربين")
        trainers_report_btn.setObjectName("trainers_report_button")

        export_report_btn = QPushButton("تصدير التقرير")
        export_report_btn.setObjectName("export_report_button")

        reports_buttons_layout.addWidget(revenue_report_btn)
        reports_buttons_layout.addWidget(expenses_report_btn)
        reports_buttons_layout.addWidget(students_report_btn)
        reports_buttons_layout.addWidget(trainers_report_btn)
        reports_buttons_layout.addWidget(export_report_btn)

        reports_layout.addLayout(reports_buttons_layout)

        # منطقة عرض التقرير
        self.report_display_area = QTextEdit()
        self.report_display_area.setObjectName("report_display_area")
        self.report_display_area.setReadOnly(True)
        self.report_display_area.setPlainText("اختر نوع التقرير لعرضه هنا...")
        reports_layout.addWidget(self.report_display_area)

        layout.addWidget(reports_container)

        self.tab_widget.addTab(tab, qta.icon('fa5s.chart-line', color='#3498db'), "التقارير المالية")

    # إنشاء بطاقة ملخص مالي
    def create_financial_summary_card(self, title, value, color):
        card = QFrame()
        card.setFixedHeight(100)
        card.setMinimumWidth(150)
        card.setObjectName("financial_summary_card")

        # تخطيط البطاقة
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(15, 15, 15, 15)
        card_layout.setSpacing(10)

        # عنوان البطاقة
        title_label = QLabel(title)
        title_label.setObjectName("financial_card_title")
        title_label.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(title_label)

        # قيمة البطاقة
        value_label = QLabel(value)
        value_label.setObjectName("financial_card_value")
        value_label.setAlignment(Qt.AlignCenter)
        card_layout.addWidget(value_label)

        # تطبيق اللون
        card.setProperty("card_color", color)

        return card

    # تطبيق أنماط نافذة إدارة التدريب
    def apply_training_management_styles(self):
        self.setStyleSheet("""
            /* نافذة إدارة التدريب الرئيسية */
            TrainingManagementWindow {
                background-color: #f8f9fa;
                font-family: 'Janna LT';
            }

            /* التابات */
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: white;
                border-radius: 8px;
            }

            QTabWidget::tab-bar {
               /* alignment: right;*/
            }

            QTabBar::tab {
                background-color: #e9ecef;
                color: #495057;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
            }

            QTabBar::tab:selected {
                background-color: white;
                color: #007bff;
                border-bottom: 2px solid #007bff;
            }

            QTabBar::tab:hover {
                background-color: #f8f9fa;
            }

            /* حاويات الإحصائيات */
            QFrame[objectName*="stats_container"] {
                background-color: transparent;
                border: none;
                margin: 10px 0;
            }

            /* بطاقات الإحصائيات */
            QFrame[objectName="training_stat_card"] {
                background-color: white;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                margin: 5px;
            }

            QFrame[objectName="training_stat_card"]:hover {
                border-color: #007bff;
                
            }

            QLabel[objectName="stat_card_title"] {
                font-size: 12px;
                font-weight: bold;
                color: #6c757d;
            }

            QLabel[objectName="stat_card_value"] {
                font-size: 18px;
                font-weight: bold;
                color: #212529;
            }

            /* حاويات المعلومات */
            QGroupBox[objectName*="info_container"] {
                font-size: 14px;
                font-weight: bold;
                color: #495057;
                border: 2px solid #e9ecef;
                border-radius: 8px;
                margin: 10px 5px;
                padding-top: 15px;
            }

            QGroupBox[objectName*="info_container"]::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
                background-color: white;
            }

            /* الأزرار */
            QPushButton[objectName*="button"] {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
            }

            QPushButton[objectName*="button"]:hover {
                background-color: #0056b3;
            }

            QPushButton[objectName*="button"]:pressed {
                background-color: #004085;
            }

            /* أزرار خاصة بالألوان */
            QPushButton[objectName*="add"] {
                background-color: #28a745;
            }

            QPushButton[objectName*="add"]:hover {
                background-color: #218838;
            }

            QPushButton[objectName*="edit"] {
                background-color: #ffc107;
                color: #212529;
            }

            QPushButton[objectName*="edit"]:hover {
                background-color: #e0a800;
            }

            QPushButton[objectName*="delete"] {
                background-color: #dc3545;
            }

            QPushButton[objectName*="delete"]:hover {
                background-color: #c82333;
            }

            /* حقول الإدخال */
            QLineEdit[objectName*="search"] {
                padding: 8px 12px;
                border: 1px solid #ced4da;
                border-radius: 6px;
                font-size: 12px;
            }

            QLineEdit[objectName*="search"]:focus {
                border-color: #007bff;
                outline: none;
            }

            /* القوائم المنسدلة */
            QComboBox[objectName*="filter"] {
                padding: 6px 12px;
                border: 1px solid #ced4da;
                border-radius: 6px;
                background-color: white;
                font-size: 12px;
            }

            QComboBox[objectName*="filter"]:focus {
                border-color: #007bff;
            }

            /* الجداول */
            QTableWidget {
                gridline-color: #e9ecef;
                background-color: white;
                alternate-background-color: #f8f9fa;
                selection-background-color: #007bff;
                border: 1px solid #dee2e6;
                border-radius: 6px;
            }

            QHeaderView::section {
                background-color: #e9ecef;
                color: #495057;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
        """)

    # دوال إدارة الطلبة
    # إضافة طالب جديد
    def add_student(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "يجب حفظ الدورة أولاً قبل إضافة الطلبة")
            return

        dialog = StudentDialog(self, training_id=self.training_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_students_data()
            self.update_training_statistics()

    # تعديل طالب
    def edit_student(self):
        if not self.students_table:
            return

        current_row = self.students_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد طالب للتعديل")
            return

        student_id_item = self.students_table.item(current_row, 0)
        if not student_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الطالب")
            return

        student_id = int(student_id_item.text())
        dialog = StudentDialog(self, training_id=self.training_id, student_id=student_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_students_data()
            self.update_training_statistics()

    # حذف طالب
    def delete_student(self):
        if not self.students_table:
            return

        current_row = self.students_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد طالب للحذف")
            return

        student_id_item = self.students_table.item(current_row, 0)
        if not student_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الطالب")
            return

        student_name_item = self.students_table.item(current_row, 1)
        student_name = student_name_item.text() if student_name_item else "غير محدد"

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل أنت متأكد من حذف الطالب '{student_name}'؟\n\nسيتم حذف جميع دفعات الطالب أيضاً.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user, password=password,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                student_id = int(student_id_item.text())

                # حذف دفعات الطالب أولاً
                cursor.execute("""
                    DELETE FROM التدريب_دفعات_الطلاب
                    WHERE معرف_الطالب = %s
                """, (student_id,))

                # حذف الطالب
                cursor.execute("""
                    DELETE FROM التدريب_الطلاب
                    WHERE id = %s
                """, (student_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", f"تم حذف الطالب '{student_name}' بنجاح")
                self.load_students_data()
                self.update_training_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف الطالب: {str(e)}")

    # دوال إدارة المجموعات
    # إضافة مجموعة جديدة
    def add_group(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "يجب حفظ الدورة أولاً قبل إضافة المجموعات")
            return

        dialog = GroupDialog(self, training_id=self.training_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_groups_data()
            self.update_training_statistics()

    # تعديل مجموعة
    def edit_group(self):
        if not self.groups_table:
            return

        current_row = self.groups_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مجموعة للتعديل")
            return

        group_id_item = self.groups_table.item(current_row, 0)
        if not group_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المجموعة")
            return

        group_id = int(group_id_item.text())
        dialog = GroupDialog(self, training_id=self.training_id, group_id=group_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_groups_data()
            self.update_training_statistics()

    # حذف مجموعة
    def delete_group(self):
        if not self.groups_table:
            return

        current_row = self.groups_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مجموعة للحذف")
            return

        group_id_item = self.groups_table.item(current_row, 0)
        if not group_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المجموعة")
            return

        group_name_item = self.groups_table.item(current_row, 1)
        group_name = group_name_item.text() if group_name_item else "غير محدد"

        # التحقق من وجود طلبة في المجموعة
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            group_id = int(group_id_item.text())
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_الطلاب
                WHERE معرف_المجموعة = %s
            """, (group_id,))

            student_count = cursor.fetchone()[0]
            conn.close()

            if student_count > 0:
                QMessageBox.warning(
                    self, "تحذير",
                    f"لا يمكن حذف المجموعة '{group_name}' لأنها تحتوي على {student_count} طالب.\n\nيرجى نقل الطلبة إلى مجموعة أخرى أو حذفهم أولاً."
                )
                return

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في التحقق من الطلبة: {str(e)}")
            return

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل أنت متأكد من حذف المجموعة '{group_name}'؟",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user, password=password,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                cursor.execute("""
                    DELETE FROM التدريب_المجموعات
                    WHERE id = %s
                """, (group_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", f"تم حذف المجموعة '{group_name}' بنجاح")
                self.load_groups_data()
                self.update_training_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف المجموعة: {str(e)}")

    # دوال إدارة دفعات الطلبة
    # إضافة دفعة طالب جديدة
    def add_student_payment(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "يجب حفظ الدورة أولاً قبل إضافة الدفعات")
            return

        dialog = StudentPaymentDialog(self, training_id=self.training_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_student_payments_data()
            self.load_students_data()  # تحديث جدول الطلبة لإظهار الأرصدة المحدثة
            self.update_training_statistics()

    # تعديل دفعة طالب
    def edit_student_payment(self):
        if not self.student_payments_table:
            return

        current_row = self.student_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد دفعة للتعديل")
            return

        payment_id_item = self.student_payments_table.item(current_row, 0)
        if not payment_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الدفعة")
            return

        payment_id = int(payment_id_item.text())
        dialog = StudentPaymentDialog(self, training_id=self.training_id, payment_id=payment_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_student_payments_data()
            self.load_students_data()  # تحديث جدول الطلبة لإظهار الأرصدة المحدثة
            self.update_training_statistics()

    # حذف دفعة طالب
    def delete_student_payment(self):
        if not self.student_payments_table:
            return

        current_row = self.student_payments_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد دفعة للحذف")
            return

        payment_id_item = self.student_payments_table.item(current_row, 0)
        if not payment_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف الدفعة")
            return

        student_name_item = self.student_payments_table.item(current_row, 1)
        amount_item = self.student_payments_table.item(current_row, 3)
        student_name = student_name_item.text() if student_name_item else "غير محدد"
        amount = amount_item.text() if amount_item else "0"

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل أنت متأكد من حذف دفعة الطالب '{student_name}' بمبلغ {amount}؟\n\nسيتم تحديث رصيد الطالب تلقائياً.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user, password=password,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                payment_id = int(payment_id_item.text())

                # الحصول على معرف الطالب والمبلغ قبل الحذف
                cursor.execute("""
                    SELECT معرف_الطالب, المبلغ_المدفوع
                    FROM التدريب_دفعات_الطلاب
                    WHERE id = %s
                """, (payment_id,))

                result = cursor.fetchone()
                if not result:
                    QMessageBox.warning(self, "خطأ", "لا يمكن العثور على الدفعة")
                    return

                student_id, payment_amount = result

                # حذف الدفعة
                cursor.execute("""
                    DELETE FROM التدريب_دفعات_الطلاب
                    WHERE id = %s
                """, (payment_id,))

                # تحديث رصيد الطالب
                cursor.execute("""
                    UPDATE التدريب_الطلاب
                    SET المدفوع = المدفوع - %s
                    WHERE id = %s
                """, (payment_amount, student_id))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", f"تم حذف الدفعة بنجاح وتحديث رصيد الطالب")
                self.load_student_payments_data()
                self.load_students_data()  # تحديث جدول الطلبة لإظهار الأرصدة المحدثة
                self.update_training_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف الدفعة: {str(e)}")

    # دوال إدارة المصروفات
    # إضافة مصروف جديد
    def add_expense(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "يجب حفظ الدورة أولاً قبل إضافة المصروفات")
            return

        dialog = ExpenseDialog(self, training_id=self.training_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_expenses_data()
            self.update_training_statistics()

    # تعديل مصروف
    def edit_expense(self):
        if not self.expenses_table:
            return

        current_row = self.expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مصروف للتعديل")
            return

        expense_id_item = self.expenses_table.item(current_row, 0)
        if not expense_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المصروف")
            return

        expense_id = int(expense_id_item.text())
        dialog = ExpenseDialog(self, training_id=self.training_id, expense_id=expense_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_expenses_data()
            self.update_training_statistics()

    # حذف مصروف
    def delete_expense(self):
        if not self.expenses_table:
            return

        current_row = self.expenses_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مصروف للحذف")
            return

        expense_id_item = self.expenses_table.item(current_row, 0)
        if not expense_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المصروف")
            return

        description_item = self.expenses_table.item(current_row, 1)
        amount_item = self.expenses_table.item(current_row, 3)
        description = description_item.text() if description_item else "غير محدد"
        amount = amount_item.text() if amount_item else "0"

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل أنت متأكد من حذف المصروف '{description}' بمبلغ {amount}؟",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user, password=password,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                expense_id = int(expense_id_item.text())

                cursor.execute("""
                    DELETE FROM التدريب_مصروفات
                    WHERE id = %s
                """, (expense_id,))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", f"تم حذف المصروف '{description}' بنجاح")
                self.load_expenses_data()
                self.update_training_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف المصروف: {str(e)}")

    # دوال إدارة المدربين
    # إضافة مدرب جديد للدورة
    def add_trainer(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "يجب حفظ الدورة أولاً قبل إضافة المدربين")
            return

        dialog = TrainerTaskDialog(self, training_id=self.training_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_trainers_data()
            self.update_training_statistics()

    # تعديل مهمة مدرب
    def edit_trainer(self):
        if not self.trainers_table:
            return

        current_row = self.trainers_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مدرب للتعديل")
            return

        task_id_item = self.trainers_table.item(current_row, 0)
        if not task_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المهمة")
            return

        task_id = int(task_id_item.text())
        dialog = TrainerTaskDialog(self, training_id=self.training_id, task_id=task_id)
        if dialog.exec() == QDialog.Accepted:
            self.load_trainers_data()
            self.update_training_statistics()

    # حذف مهمة مدرب
    def delete_trainer(self):
        if not self.trainers_table:
            return

        current_row = self.trainers_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "تحذير", "يرجى تحديد مدرب للحذف")
            return

        task_id_item = self.trainers_table.item(current_row, 0)
        if not task_id_item:
            QMessageBox.warning(self, "خطأ", "لا يمكن الحصول على معرف المهمة")
            return

        trainer_name_item = self.trainers_table.item(current_row, 1)
        trainer_name = trainer_name_item.text() if trainer_name_item else "غير محدد"

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            f"هل أنت متأكد من حذف المدرب '{trainer_name}' من الدورة؟\n\nسيتم حذف جميع المهام المرتبطة بهذا المدرب.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                conn = mysql.connector.connect(
                    host=host, user=user, password=password,
                    database="project_manager_V2"
                )
                cursor = conn.cursor()

                task_id = int(task_id_item.text())

                cursor.execute("""
                    DELETE FROM المشاريع_مهام_الفريق
                    WHERE id = %s AND نوع_المهمة = 'مهمة تدريب' AND معرف_الدورة_التدريبية = %s
                """, (task_id, self.training_id))

                conn.commit()
                conn.close()

                QMessageBox.information(self, "نجح", f"تم حذف المدرب '{trainer_name}' بنجاح")
                self.load_trainers_data()
                self.update_training_statistics()

            except Exception as e:
                QMessageBox.warning(self, "خطأ", f"فشل في حذف المدرب: {str(e)}")

    # إدراج أرصدة المدربين الذين لم يتم إدراج أرصدتهم
    def include_trainer_balances(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "لا توجد دورة محددة")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب المدربين الذين لم يتم إدراج أرصدتهم
            cursor.execute("""
                SELECT t.id, e.اسم_الموظف, t.مبلغ_الموظف, t.معرف_الموظف
                FROM المشاريع_مهام_الفريق t
                JOIN الموظفين e ON t.معرف_الموظف = e.id
                WHERE t.معرف_الدورة_التدريبية = %s
                AND t.نوع_المهمة = 'مهمة تدريب'
                AND t.حالة_مبلغ_الموظف = 'غير مدرج'
                AND t.مبلغ_الموظف > 0
            """, (self.training_id,))

            pending_trainers = cursor.fetchall()

            if not pending_trainers:
                QMessageBox.information(self, "معلومات", "لا توجد أرصدة مدربين في انتظار الإدراج")
                conn.close()
                return

            # حساب إجمالي المبلغ
            total_amount = sum(trainer[2] for trainer in pending_trainers)
            trainers_list = "\n".join([f"- {trainer[1]}: {trainer[2]:,.2f}" for trainer in pending_trainers])

            reply = QMessageBox.question(
                self, "تأكيد إدراج الأرصدة",
                f"سيتم إدراج الأرصدة التالية:\n\n{trainers_list}\n\nإجمالي المبلغ: {total_amount:,.2f} {Currency_type}\n\nهل تريد المتابعة؟",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # تحديث حالة جميع المهام وإضافة المبالغ لأرصدة المدربين
                for task_id, trainer_name, amount, employee_id in pending_trainers:
                    # تحديث حالة المبلغ
                    cursor.execute("""
                        UPDATE المشاريع_مهام_الفريق
                        SET حالة_مبلغ_الموظف = 'تم الإدراج'
                        WHERE id = %s
                    """, (task_id,))

                    # إضافة المبلغ إلى رصيد المدرب
                    cursor.execute("""
                        UPDATE الموظفين
                        SET الرصيد = الرصيد + %s
                        WHERE id = %s
                    """, (amount, employee_id))

                conn.commit()
                conn.close()

                QMessageBox.information(
                    self, "نجح",
                    f"تم إدراج إجمالي مبلغ {total_amount:,.2f} {Currency_type} لأرصدة {len(pending_trainers)} مدرب بنجاح"
                )
                self.load_trainers_data()
                self.update_training_statistics()
            else:
                conn.close()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في إدراج الأرصدة: {str(e)}")

    # إنشاء مهام جديدة للمدربين عند إضافة دفعات جديدة
    def create_trainer_tasks_for_new_payments(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # جلب المدربين الذين تم إدراج أرصدتهم السابقة
            cursor.execute("""
                SELECT DISTINCT t.معرف_الموظف, e.اسم_الموظف, t.نسبة_الموظف
                FROM المشاريع_مهام_الفريق t
                JOIN الموظفين e ON t.معرف_الموظف = e.id
                WHERE t.معرف_الدورة_التدريبية = %s
                AND t.نوع_المهمة = 'مهمة تدريب'
                AND t.حالة_مبلغ_الموظف = 'تم الإدراج'
            """, (self.training_id,))

            trainers_with_included_balances = cursor.fetchall()

            if not trainers_with_included_balances:
                conn.close()
                return

            # حساب إجمالي الدفعات الجديدة منذ آخر إدراج
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ_المدفوع), 0) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s
                AND تاريخ_الدفع > (
                    SELECT MAX(تاريخ_التحديث) FROM المشاريع_مهام_الفريق
                    WHERE معرف_الدورة_التدريبية = %s AND نوع_المهمة = 'مهمة تدريب'
                    AND حالة_مبلغ_الموظف = 'تم الإدراج'
                )
            """, (self.training_id, self.training_id))

            new_payments_total = cursor.fetchone()[0]

            if new_payments_total > 0:
                # إنشاء مهام جديدة للمدربين بناءً على النسب
                for trainer_id, trainer_name, percentage in trainers_with_included_balances:
                    if percentage and percentage > 0:
                        trainer_amount = (new_payments_total * percentage) / 100

                        # إنشاء مهمة جديدة
                        cursor.execute("""
                            INSERT INTO المشاريع_مهام_الفريق
                            (معرف_الموظف, نوع_المهمة, معرف_القسم, معرف_الدورة_التدريبية,
                             عنوان_المهمة, وصف_المهمة, نوع_دور_المهمة, نسبة_الموظف, مبلغ_الموظف,
                             حالة_مبلغ_الموظف, تاريخ_البدء, تاريخ_الانتهاء, الحالة, المستخدم)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (trainer_id, 'مهمة تدريب', None, self.training_id, f"دفعة إضافية - {trainer_name}",f"مبلغ إضافي من الدفعات الجديدة بقيمة {new_payments_total:,.2f}",'دور_عام', percentage, trainer_amount, 'غير مدرج', QDate.currentDate().toString(Qt.ISODate),QDate.currentDate().toString(Qt.ISODate),'قيد التنفيذ', 'admin'))

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"خطأ في إنشاء مهام المدربين الجديدة: {e}")

    

    # التحقق من صحة بيانات الدورة
    def validate_training_data(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "لا توجد دورة محددة للتحقق من بياناتها")
            return False

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # التحقق من وجود الدورة
            cursor.execute("SELECT COUNT(*) FROM التدريب WHERE id = %s", (self.training_id,))
            if cursor.fetchone()[0] == 0:
                QMessageBox.critical(self, "خطأ", "الدورة غير موجودة في قاعدة البيانات")
                conn.close()
                return False

            # التحقق من صحة التواريخ
            cursor.execute("""
                SELECT تاريخ_البدء, تاريخ_الإنتهاء FROM التدريب WHERE id = %s
            """, (self.training_id,))

            result = cursor.fetchone()
            if result:
                start_date, end_date = result
                if start_date and end_date and start_date > end_date:
                    QMessageBox.warning(self, "تحذير", "تاريخ البدء لا يمكن أن يكون بعد تاريخ الانتهاء")
                    conn.close()
                    return False

            # التحقق من وجود مجموعات بدون طلبة
            cursor.execute("""
                SELECT g.`اسم المجموعه`
                FROM التدريب_المجموعات g
                LEFT JOIN التدريب_الطلاب s ON g.id = s.معرف_المجموعة
                WHERE g.معرف_الدورة = %s
                GROUP BY g.id, g.`اسم المجموعه`
                HAVING COUNT(s.id) = 0
            """, (self.training_id,))

            empty_groups = cursor.fetchall()
            if empty_groups:
                group_names = ", ".join([group[0] for group in empty_groups])
                QMessageBox.information(
                    self, "تنبيه",
                    f"توجد مجموعات فارغة (بدون طلبة): {group_names}"
                )

            conn.close()
            return True

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في التحقق من صحة البيانات: {str(e)}")
            return False

    # التحقق من صحة البيانات المالية
    def validate_financial_data(self):
        if not self.training_id:
            return False

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # التحقق من وجود دفعات بمبالغ سالبة
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s AND المبلغ_المدفوع < 0
            """, (self.training_id,))

            negative_payments = cursor.fetchone()[0]
            if negative_payments > 0:
                QMessageBox.warning(
                    self, "تحذير",
                    f"توجد {negative_payments} دفعة بمبالغ سالبة"
                )

            # التحقق من وجود مصروفات بمبالغ سالبة
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_مصروفات
                WHERE معرف_الدورة = %s AND المبلغ < 0
            """, (self.training_id,))

            negative_expenses = cursor.fetchone()[0]
            if negative_expenses > 0:
                QMessageBox.warning(
                    self, "تحذير",
                    f"توجد {negative_expenses} مصروف بمبالغ سالبة"
                )

            # التحقق من طلبة دفعوا أكثر من المطلوب
            cursor.execute("""
                SELECT اسم_الطالب, المبلغ, المدفوع
                FROM التدريب_الطلاب
                WHERE معرف_الدورة = %s AND المدفوع > المبلغ
            """, (self.training_id,))

            overpaid_students = cursor.fetchall()
            if overpaid_students:
                student_names = ", ".join([student[0] for student in overpaid_students])
                QMessageBox.information(
                    self, "تنبيه",
                    f"طلبة دفعوا أكثر من المطلوب: {student_names}"
                )

            conn.close()
            return True

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في التحقق من البيانات المالية: {str(e)}")
            return False

    # معالجة أخطاء قاعدة البيانات بشكل موحد
    def handle_database_error(self, operation, error):
        error_message = str(error)

        # رسائل خطأ مخصصة حسب نوع الخطأ
        if "Connection" in error_message or "connect" in error_message.lower():
            QMessageBox.critical(
                self, "خطأ في الاتصال",
                f"فشل في الاتصال بقاعدة البيانات أثناء {operation}.\n\nيرجى التحقق من إعدادات الاتصال والمحاولة مرة أخرى."
            )
        elif "Duplicate entry" in error_message:
            QMessageBox.warning(
                self, "بيانات مكررة",
                f"البيانات المدخلة موجودة مسبقاً أثناء {operation}.\n\nيرجى التحقق من البيانات والمحاولة مرة أخرى."
            )
        elif "Foreign key constraint" in error_message:
            QMessageBox.warning(
                self, "خطأ في الربط",
                f"لا يمكن تنفيذ العملية {operation} بسبب وجود بيانات مرتبطة.\n\nيرجى حذف البيانات المرتبطة أولاً."
            )
        elif "Data too long" in error_message:
            QMessageBox.warning(
                self, "بيانات طويلة",
                f"البيانات المدخلة طويلة جداً أثناء {operation}.\n\nيرجى تقليل حجم البيانات والمحاولة مرة أخرى."
            )
        else:
            QMessageBox.critical(
                self, "خطأ في قاعدة البيانات",
                f"حدث خطأ أثناء {operation}:\n\n{error_message}\n\nيرجى المحاولة مرة أخرى أو الاتصال بالدعم الفني."
            )

    # عرض رسالة نجاح موحدة
    def show_success_message(self, operation, details=""):
        message = f"تم {operation} بنجاح"
        if details:
            message += f"\n\n{details}"

        QMessageBox.information(self, "نجح", message)

    # تأكيد عملية الحذف
    def confirm_delete_operation(self, item_type, item_name, additional_info=""):
        message = f"هل أنت متأكد من حذف {item_type} '{item_name}'؟"
        if additional_info:
            message += f"\n\n{additional_info}"

        reply = QMessageBox.question(
            self, "تأكيد الحذف",
            message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        return reply == QMessageBox.Yes

    # التحقق من صحة جميع البيانات
    def validate_all_data(self):
        try:
            # التحقق من البيانات الأساسية
            if not self.validate_training_data():
                return

            # التحقق من البيانات المالية
            if not self.validate_financial_data():
                return

            # إذا نجحت جميع عمليات التحقق
            self.show_success_message(
                "التحقق من البيانات",
                "جميع البيانات صحيحة ومتسقة"
            )

        except Exception as e:
            self.handle_database_error("التحقق من صحة البيانات", e)

    # تحديث إجماليات الدورة في قاعدة البيانات
    def update_course_totals(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # حساب إجمالي الإيرادات
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ_المدفوع), 0) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_revenue = cursor.fetchone()[0]

            # حساب عدد الطلبة
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            students_count = cursor.fetchone()[0]

            # تحديث جدول التدريب
            cursor.execute("""
                UPDATE التدريب
                SET إجمالي_المبلغ = %s, عدد_المشاركين = %s
                WHERE id = %s
            """, (total_revenue, students_count, self.training_id))

            conn.commit()
            conn.close()

            # تحديث البيانات المحلية
            self.training_data['إجمالي_المبلغ'] = total_revenue
            self.training_data['عدد_المشاركين'] = students_count

            print(f"تم تحديث الإجماليات: الإيرادات = {total_revenue:,.2f}, عدد الطلبة = {students_count}")

        except Exception as e:
            print(f"خطأ في تحديث إجماليات الدورة: {e}")

    # تحديث أرصدة جميع الطلبة تلقائياً
    def update_student_balances(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # تحديث المدفوع والباقي لكل طالب
            cursor.execute("""
                UPDATE التدريب_الطلاب s
                SET المدفوع = (
                    SELECT COALESCE(SUM(المبلغ_المدفوع), 0)
                    FROM التدريب_دفعات_الطلاب p
                    WHERE p.معرف_الطالب = s.id
                ),
                الباقي = s.المبلغ - (
                    SELECT COALESCE(SUM(المبلغ_المدفوع), 0)
                    FROM التدريب_دفعات_الطلاب p
                    WHERE p.معرف_الطالب = s.id
                )
                WHERE s.معرف_الدورة = %s
            """, (self.training_id,))

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"خطأ في تحديث أرصدة الطلبة: {e}")

    # حساب مبالغ المدربين تلقائياً بناءً على النسب
    def calculate_trainer_amounts(self):
        if not self.training_id:
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # حساب إجمالي الإيرادات
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ_المدفوع), 0) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_revenue = cursor.fetchone()[0]

            # تحديث مبالغ المدربين بناءً على النسب
            cursor.execute("""
                UPDATE المشاريع_مهام_الفريق
                SET مبلغ_الموظف = (نسبة_الموظف * %s) / 100
                WHERE معرف_الدورة_التدريبية = %s
                AND نوع_المهمة = 'مهمة تدريب'
                AND نسبة_الموظف > 0
                AND حالة_مبلغ_الموظف = 'غير مدرج'
            """, (total_revenue, self.training_id))

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"خطأ في حساب مبالغ المدربين: {e}")

    # عرض ملخص مالي شامل للدورة
    def show_financial_summary(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "لا توجد دورة محددة")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # حساب الإيرادات
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ_المدفوع), 0) FROM التدريب_دفعات_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_revenue = cursor.fetchone()[0]

            # حساب المصروفات
            cursor.execute("""
                SELECT COALESCE(SUM(المبلغ), 0) FROM التدريب_مصروفات
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            total_expenses = cursor.fetchone()[0]

            # حساب مبالغ المدربين
            cursor.execute("""
                SELECT COALESCE(SUM(مبلغ_الموظف), 0) FROM المشاريع_مهام_الفريق
                WHERE معرف_الدورة_التدريبية = %s AND نوع_المهمة = 'مهمة تدريب'
            """, (self.training_id,))
            trainers_cost = cursor.fetchone()[0]

            # حساب عدد الطلبة
            cursor.execute("""
                SELECT COUNT(*) FROM التدريب_الطلاب
                WHERE معرف_الدورة = %s
            """, (self.training_id,))
            students_count = cursor.fetchone()[0]

            conn.close()

            # حساب صافي الربح
            net_profit = total_revenue - total_expenses - trainers_cost

            # إنشاء رسالة الملخص
            summary = f"""
الملخص المالي للدورة التدريبية

📊 الإحصائيات العامة:
• عدد الطلبة: {students_count}
• إجمالي الإيرادات: {total_revenue:,.2f} {Currency_type}

💰 التكاليف:
• المصروفات: {total_expenses:,.2f} {Currency_type}
• مبالغ المدربين: {trainers_cost:,.2f} {Currency_type}
• إجمالي التكاليف: {(total_expenses + trainers_cost):,.2f} {Currency_type}

📈 النتيجة النهائية:
• صافي الربح: {net_profit:,.2f} {Currency_type}
• هامش الربح: {(net_profit/total_revenue*100 if total_revenue > 0 else 0):.2f}%
            """

            QMessageBox.information(self, "الملخص المالي", summary.strip())

        except Exception as e:
            self.handle_database_error("عرض الملخص المالي", e)

    # تصدير بيانات الدورة إلى ملف Excel
    def export_training_data(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "لا توجد دورة محددة للتصدير")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            # إنشاء ملف Excel جديد
            wb = Workbook()

            # حذف الورقة الافتراضية
            wb.remove(wb.active)

            # إنشاء أوراق مختلفة للبيانات
            self.export_students_sheet(wb)
            self.export_payments_sheet(wb)
            self.export_expenses_sheet(wb)
            self.export_trainers_sheet(wb)

            # حفظ الملف
            course_title = self.training_data.get('عنوان_الدورة', 'دورة_تدريبية')
            filename = f"تقرير_{course_title}_{QDate.currentDate().toString('yyyy-MM-dd')}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ تقرير الدورة", filename, "Excel Files (*.xlsx)"
            )

            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "نجح", f"تم تصدير البيانات بنجاح إلى:\n{file_path}")

        except ImportError:
            QMessageBox.warning(
                self, "خطأ",
                "مكتبة openpyxl غير مثبتة.\nيرجى تثبيتها باستخدام: pip install openpyxl"
            )
        except Exception as e:
            self.handle_database_error("تصدير البيانات", e)

    # تصدير بيانات الطلبة إلى ورقة Excel
    def export_students_sheet(self, workbook):
        ws = workbook.create_sheet("الطلبة")

        # العناوين
        headers = ["اسم الطالب", "رقم الهاتف", "المجموعة", "المبلغ", "المدفوع", "الباقي", "تاريخ التسجيل"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # البيانات
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT s.اسم_الطالب, s.رقم_الهاتف,
                       COALESCE(g.`اسم المجموعه`, 'غير محدد'),
                       s.المبلغ, s.المدفوع, s.الباقي, s.تاريخ_التسجيل
                FROM التدريب_الطلاب s
                LEFT JOIN التدريب_المجموعات g ON s.معرف_المجموعة = g.id
                WHERE s.معرف_الدورة = %s
                ORDER BY s.اسم_الطالب
            """, (self.training_id,))

            students = cursor.fetchall()
            conn.close()

            for row, student in enumerate(students, 2):
                for col, value in enumerate(student, 1):
                    ws.cell(row=row, column=col, value=value)

        except Exception as e:
            print(f"خطأ في تصدير بيانات الطلبة: {e}")

    # تصدير بيانات المدربين إلى ورقة Excel
    def export_trainers_sheet(self, workbook):
        ws = workbook.create_sheet("المدربين")

        # العناوين
        headers = ["اسم المدرب", "الوظيفة", "عنوان المهمة", "النسبة", "المبلغ", "حالة المبلغ", "الحالة"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # البيانات
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT e.اسم_الموظف, e.الوظيفة, t.عنوان_المهمة,
                       t.نسبة_الموظف, t.مبلغ_الموظف, t.حالة_مبلغ_الموظف, t.الحالة
                FROM المشاريع_مهام_الفريق t
                JOIN الموظفين e ON t.معرف_الموظف = e.id
                WHERE t.معرف_الدورة_التدريبية = %s AND t.نوع_المهمة = 'مهمة تدريب'
                ORDER BY e.اسم_الموظف
            """, (self.training_id,))

            trainers = cursor.fetchall()
            conn.close()

            for row, trainer in enumerate(trainers, 2):
                for col, value in enumerate(trainer, 1):
                    ws.cell(row=row, column=col, value=value)

        except Exception as e:
            print(f"خطأ في تصدير بيانات المدربين: {e}")

    # إنشاء نسخة احتياطية من بيانات الدورة
    def backup_training_data(self):
        if not self.training_id:
            QMessageBox.warning(self, "تحذير", "لا توجد دورة محددة للنسخ الاحتياطي")
            return

        try:
            import json
            from datetime import datetime

            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            backup_data = {
                "backup_date": datetime.now().isoformat(),
                "training_id": self.training_id,
                "training_info": {},
                "students": [],
                "groups": [],
                "payments": [],
                "expenses": [],
                "trainers": []
            }

            # نسخ معلومات الدورة
            cursor.execute("SELECT * FROM التدريب WHERE id = %s", (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            training_row = cursor.fetchone()
            if training_row:
                backup_data["training_info"] = dict(zip(columns, training_row))

            # نسخ بيانات الطلبة
            cursor.execute("SELECT * FROM التدريب_الطلاب WHERE معرف_الدورة = %s", (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                backup_data["students"].append(dict(zip(columns, row)))

            # نسخ بيانات المجموعات
            cursor.execute("SELECT * FROM التدريب_المجموعات WHERE معرف_الدورة = %s", (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                backup_data["groups"].append(dict(zip(columns, row)))

            # نسخ بيانات الدفعات
            cursor.execute("SELECT * FROM التدريب_دفعات_الطلاب WHERE معرف_الدورة = %s", (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                backup_data["payments"].append(dict(zip(columns, row)))

            # نسخ بيانات المصروفات
            cursor.execute("SELECT * FROM التدريب_مصروفات WHERE معرف_الدورة = %s", (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                backup_data["expenses"].append(dict(zip(columns, row)))

            # نسخ بيانات المدربين
            cursor.execute("""
                SELECT * FROM المشاريع_مهام_الفريق
                WHERE معرف_القسم = %s AND نوع_المهمة = 'مهمة تدريب'
            """, (self.training_id,))
            columns = [desc[0] for desc in cursor.description]
            for row in cursor.fetchall():
                backup_data["trainers"].append(dict(zip(columns, row)))

            conn.close()

            # حفظ النسخة الاحتياطية
            course_title = self.training_data.get('عنوان_الدورة', 'دورة_تدريبية')
            filename = f"نسخة_احتياطية_{course_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ النسخة الاحتياطية", filename, "JSON Files (*.json)"
            )

            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(backup_data, f, ensure_ascii=False, indent=2, default=str)

                QMessageBox.information(
                    self, "نجح",
                    f"تم إنشاء النسخة الاحتياطية بنجاح:\n{file_path}"
                )

        except Exception as e:
            self.handle_database_error("إنشاء النسخة الاحتياطية", e)

    # إضافة أزرار الطباعة لجميع التابات
    def add_print_buttons(self):
        try:
            # إضافة أزرار الطباعة تلقائياً لجميع التابات
            quick_add_print_button(self, self.tab_widget)

        except Exception as e:
            print(f"خطأ في إضافة أزرار الطباعة: {e}")

    # التأكد من وجود الأعمدة المطلوبة في جداول التدريب
    def ensure_training_table_columns(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            # التحقق من وجود عمود إجمالي_المبلغ في جدول التدريب
            cursor.execute("""
                SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = 'project_manager_V2'
                AND TABLE_NAME = 'التدريب'
                AND COLUMN_NAME = 'إجمالي_المبلغ'
            """)

            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    ALTER TABLE التدريب
                    ADD COLUMN إجمالي_المبلغ DECIMAL(10,2) DEFAULT 0
                """)
                print("تم إضافة عمود إجمالي_المبلغ إلى جدول التدريب")

            # التحقق من وجود عمود عدد_المشاركين في جدول التدريب
            cursor.execute("""
                SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = 'project_manager_V2'
                AND TABLE_NAME = 'التدريب'
                AND COLUMN_NAME = 'عدد_المشاركين'
            """)

            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    ALTER TABLE التدريب
                    ADD COLUMN عدد_المشاركين INT DEFAULT 0
                """)
                print("تم إضافة عمود عدد_المشاركين إلى جدول التدريب")

            # التحقق من وجود عمود معرف_الدورة_التدريبية في جدول المشاريع_مهام_الفريق
            cursor.execute("""
                SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = 'project_manager_V2'
                AND TABLE_NAME = 'المشاريع_مهام_الفريق'
                AND COLUMN_NAME = 'معرف_الدورة_التدريبية'
            """)

            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    ALTER TABLE المشاريع_مهام_الفريق
                    ADD COLUMN معرف_الدورة_التدريبية INT NULL,
                    ADD INDEX idx_معرف_الدورة_التدريبية (معرف_الدورة_التدريبية)
                """)
                print("تم إضافة عمود معرف_الدورة_التدريبية إلى جدول المشاريع_مهام_الفريق")

            conn.commit()
            conn.close()

        except Exception as e:
            print(f"خطأ في التحقق من أعمدة جداول التدريب: {e}")

# حوار إضافة/تعديل مهمة مدرب
class TrainerTaskDialog(QDialog):

    # init
    def __init__(self, parent=None, training_id=None, task_id=None):
        super().__init__(parent)
        self.training_id = training_id
        self.task_id = task_id
        self.is_edit_mode = task_id is not None

        self.setup_dialog()
        self.create_ui()
        self.load_trainers()

        if self.is_edit_mode:
            self.load_task_data()

        apply_stylesheet(self)

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل مهمة مدرب" if self.is_edit_mode else "إضافة مدرب جديد للدورة"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 550, 450)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء واجهة المستخدم
    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان النافذة
        title_label = QLabel("تعديل مهمة المدرب" if self.is_edit_mode else "إضافة مدرب جديد للدورة")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # نموذج الإدخال
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # المدرب
        self.trainer_combo = QComboBox()
        self.trainer_combo.addItem("اختر المدرب", None)
        form_layout.addRow("المدرب:", self.trainer_combo)

        # عنوان المهمة
        self.task_title_edit = QLineEdit()
        self.task_title_edit.setPlaceholderText("مثال: تدريب دورة البرمجة")
        form_layout.addRow("عنوان المهمة:", self.task_title_edit)

        # وصف المهمة
        self.task_description_edit = QTextEdit()
        self.task_description_edit.setPlaceholderText("وصف تفصيلي للمهمة")
        self.task_description_edit.setMaximumHeight(80)
        form_layout.addRow("وصف المهمة:", self.task_description_edit)

        # نسبة المدرب
        self.percentage_edit = QLineEdit()
        self.percentage_edit.setPlaceholderText("نسبة المدرب من إجمالي الدفعات")
        percentage_validator = QDoubleValidator(0.0, 100.0, 2)
        percentage_validator.setNotation(QDoubleValidator.StandardNotation)
        self.percentage_edit.setValidator(percentage_validator)
        form_layout.addRow("نسبة المدرب (%):", self.percentage_edit)

        # مبلغ المدرب
        self.amount_edit = QLineEdit()
        self.amount_edit.setPlaceholderText("مبلغ ثابت للمدرب (اختياري)")
        amount_validator = QDoubleValidator(0.0, 999999.99, 2)
        amount_validator.setNotation(QDoubleValidator.StandardNotation)
        self.amount_edit.setValidator(amount_validator)
        form_layout.addRow("مبلغ المدرب:", self.amount_edit)

        # حالة المبلغ
        self.amount_status_combo = QComboBox()
        self.amount_status_combo.addItems(["غير مدرج", "تم الإدراج"])
        form_layout.addRow("حالة المبلغ:", self.amount_status_combo)

        # تاريخ البدء
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate())
        setup_date_edit_format(self.start_date_edit)
        form_layout.addRow("تاريخ البدء:", self.start_date_edit)

        # تاريخ الانتهاء
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate().addDays(30))
        setup_date_edit_format(self.end_date_edit)
        form_layout.addRow("تاريخ الانتهاء:", self.end_date_edit)

        # حالة المهمة
        self.task_status_combo = QComboBox()
        self.task_status_combo.addItems(["لم يبدأ", "قيد التنفيذ", "مكتمل", "متوقف"])
        form_layout.addRow("حالة المهمة:", self.task_status_combo)

        # ملاحظات
        self.notes_edit = QTextEdit()
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات")
        self.notes_edit.setMaximumHeight(80)
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("حفظ")
        save_button.clicked.connect(self.save_task)
        cancel_button = QPushButton("إلغاء")
        cancel_button.clicked.connect(self.reject)

        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)

        # ربط تغيير المدرب بتحديث النسبة
        self.trainer_combo.currentIndexChanged.connect(self.update_trainer_percentage)

    # تحميل المدربين المتاحين من جدول الموظفين
    def load_trainers(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, اسم_الموظف, الوظيفة, النسبة
                FROM الموظفين
                WHERE التصنيف = 'مدرب'
                ORDER BY اسم_الموظف
            """)

            trainers = cursor.fetchall()
            conn.close()

            for trainer_id, name, job_title, percentage in trainers:
                display_text = f"{name} - {job_title}"
                self.trainer_combo.addItem(display_text, {
                    'id': trainer_id,
                    'name': name,
                    'job_title': job_title,
                    'percentage': percentage
                })

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل المدربين: {str(e)}")

    # تحديث نسبة المدرب تلقائياً عند اختياره
    def update_trainer_percentage(self):
        trainer_data = self.trainer_combo.currentData()
        if trainer_data and trainer_data['percentage']:
            self.percentage_edit.setText(str(trainer_data['percentage']))

    # تحميل بيانات المهمة للتعديل
    def load_task_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT معرف_الموظف, عنوان_المهمة, وصف_المهمة, نسبة_الموظف,
                       مبلغ_الموظف, حالة_مبلغ_الموظف, تاريخ_البدء, تاريخ_الانتهاء,
                       الحالة, ملاحظات
                FROM المشاريع_مهام_الفريق
                WHERE id = %s AND نوع_المهمة = 'مهمة تدريب' AND معرف_الدورة_التدريبية = %s
            """, (self.task_id, self.training_id))

            result = cursor.fetchone()
            conn.close()

            if result:
                employee_id, title, description, percentage, amount, amount_status, start_date, end_date, status, notes = result

                # تحديد المدرب
                for i in range(self.trainer_combo.count()):
                    trainer_data = self.trainer_combo.itemData(i)
                    if trainer_data and trainer_data['id'] == employee_id:
                        self.trainer_combo.setCurrentIndex(i)
                        break

                self.task_title_edit.setText(title or "")
                self.task_description_edit.setPlainText(description or "")
                self.percentage_edit.setText(str(percentage) if percentage else "")
                self.amount_edit.setText(str(amount) if amount else "")

                # تحديد حالة المبلغ
                if amount_status:
                    index = self.amount_status_combo.findText(amount_status)
                    if index >= 0:
                        self.amount_status_combo.setCurrentIndex(index)

                if start_date:
                    self.start_date_edit.setDate(QDate.fromString(str(start_date), Qt.ISODate))

                if end_date:
                    self.end_date_edit.setDate(QDate.fromString(str(end_date), Qt.ISODate))

                # تحديد حالة المهمة
                if status:
                    index = self.task_status_combo.findText(status)
                    if index >= 0:
                        self.task_status_combo.setCurrentIndex(index)

                self.notes_edit.setPlainText(notes or "")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المهمة: {str(e)}")

    # حفظ بيانات المهمة
    def save_task(self):
        # التحقق من صحة البيانات
        trainer_data = self.trainer_combo.currentData()
        if not trainer_data:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار المدرب")
            return

        if not self.task_title_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال عنوان المهمة")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            employee_id = trainer_data['id']
            title = self.task_title_edit.text().strip()
            description = self.task_description_edit.toPlainText().strip()
            percentage = float(self.percentage_edit.text()) if self.percentage_edit.text().strip() else 0.0
            amount = float(self.amount_edit.text()) if self.amount_edit.text().strip() else 0.0
            amount_status = self.amount_status_combo.currentText()
            start_date = self.start_date_edit.date().toString(Qt.ISODate)
            end_date = self.end_date_edit.date().toString(Qt.ISODate)
            task_status = self.task_status_combo.currentText()
            notes = self.notes_edit.toPlainText().strip()

            if self.is_edit_mode:
                # تحديث المهمة الموجودة
                cursor.execute("""
                    UPDATE المشاريع_مهام_الفريق
                    SET معرف_الموظف = %s, عنوان_المهمة = %s, وصف_المهمة = %s,
                        نسبة_الموظف = %s, مبلغ_الموظف = %s, حالة_مبلغ_الموظف = %s,
                        تاريخ_البدء = %s, تاريخ_الانتهاء = %s, الحالة = %s, ملاحظات = %s
                    WHERE id = %s AND معرف_الدورة_التدريبية = %s
                """, (employee_id, title, description, percentage, amount, amount_status,
                      start_date, end_date, task_status, notes, self.task_id, self.training_id))

                message = "تم تحديث مهمة المدرب بنجاح"
            else:
                # إضافة مهمة جديدة (معرف_القسم = NULL، معرف_الدورة_التدريبية = training_id)
                cursor.execute("""
                    INSERT INTO المشاريع_مهام_الفريق
                    (معرف_الموظف, نوع_المهمة, معرف_القسم, معرف_الدورة_التدريبية,
                     عنوان_المهمة, وصف_المهمة, نوع_دور_المهمة, نسبة_الموظف, مبلغ_الموظف,
                     حالة_مبلغ_الموظف, تاريخ_البدء, تاريخ_الانتهاء, الحالة, ملاحظات, المستخدم)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (employee_id, 'مهمة تدريب', None, self.training_id, title, description,
                      'دور_عام', percentage, amount, amount_status,
                      start_date, end_date, task_status, notes, 'admin'))

                message = "تم إضافة المدرب بنجاح"

            conn.commit()
            conn.close()

            QMessageBox.information(self, "نجح", message)
            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ مهمة المدرب: {str(e)}")


# حوار إضافة/تعديل مصروف
class ExpenseDialog(QDialog):

    # init
    def __init__(self, parent=None, training_id=None, expense_id=None):
        super().__init__(parent)
        self.training_id = training_id
        self.expense_id = expense_id
        self.is_edit_mode = expense_id is not None

        self.setup_dialog()
        self.create_ui()

        if self.is_edit_mode:
            self.load_expense_data()

        apply_stylesheet(self)

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل مصروف" if self.is_edit_mode else "إضافة مصروف جديد"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 500, 400)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء واجهة المستخدم
    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان النافذة
        title_label = QLabel("تعديل بيانات المصروف" if self.is_edit_mode else "إضافة مصروف جديد")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # نموذج الإدخال
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # وصف المصروف
        self.description_edit = QLineEdit()
        self.description_edit.setPlaceholderText("أدخل وصف المصروف")
        form_layout.addRow("وصف المصروف:", self.description_edit)

        # نوع المصروف
        self.expense_type_combo = QComboBox()
        self.expense_type_combo.addItems([
            "مواد تدريبية",
            "أجهزة ومعدات",
            "إيجار قاعة",
            "مواصلات",
            "ضيافة",
            "طباعة وتصوير",
            "أخرى"
        ])
        form_layout.addRow("نوع المصروف:", self.expense_type_combo)

        # المبلغ
        self.amount_edit = QLineEdit()
        self.amount_edit.setPlaceholderText("أدخل مبلغ المصروف")
        amount_validator = QDoubleValidator(0.0, 999999.99, 2)
        amount_validator.setNotation(QDoubleValidator.StandardNotation)
        self.amount_edit.setValidator(amount_validator)
        form_layout.addRow("المبلغ:", self.amount_edit)

        # تاريخ المصروف
        self.expense_date_edit = QDateEdit()
        self.expense_date_edit.setCalendarPopup(True)
        self.expense_date_edit.setDate(QDate.currentDate())
        setup_date_edit_format(self.expense_date_edit)
        form_layout.addRow("تاريخ المصروف:", self.expense_date_edit)

        # المستلم
        self.receiver_edit = QLineEdit()
        self.receiver_edit.setPlaceholderText("اسم الشخص الذي استلم المبلغ")
        form_layout.addRow("المستلم:", self.receiver_edit)

        # رقم الفاتورة
        self.invoice_number_edit = QLineEdit()
        self.invoice_number_edit.setPlaceholderText("رقم الفاتورة (اختياري)")
        form_layout.addRow("رقم الفاتورة:", self.invoice_number_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("حفظ")
        save_button.clicked.connect(self.save_expense)
        cancel_button = QPushButton("إلغاء")
        cancel_button.clicked.connect(self.reject)

        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)

    # تحميل بيانات المصروف للتعديل
    def load_expense_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT وصف_المصروف, نوع_المصروف, المبلغ, تاريخ_المصروف,
                       المستلم, رقم_الفاتورة
                FROM التدريب_مصروفات
                WHERE id = %s
            """, (self.expense_id,))

            result = cursor.fetchone()
            conn.close()

            if result:
                description, expense_type, amount, expense_date, receiver, invoice_num = result

                self.description_edit.setText(description or "")

                # تحديد نوع المصروف
                if expense_type:
                    index = self.expense_type_combo.findText(expense_type)
                    if index >= 0:
                        self.expense_type_combo.setCurrentIndex(index)

                self.amount_edit.setText(str(amount) if amount else "")

                if expense_date:
                    self.expense_date_edit.setDate(QDate.fromString(str(expense_date), Qt.ISODate))

                self.receiver_edit.setText(receiver or "")
                self.invoice_number_edit.setText(invoice_num or "")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المصروف: {str(e)}")

    # حفظ بيانات المصروف
    def save_expense(self):
        # التحقق من صحة البيانات
        if not self.description_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال وصف المصروف")
            return

        if not self.amount_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال المبلغ")
            return

        try:
            amount = float(self.amount_edit.text())
            if amount <= 0:
                QMessageBox.warning(self, "تحذير", "المبلغ يجب أن يكون أكبر من صفر")
                return
        except ValueError:
            QMessageBox.warning(self, "تحذير", "يرجى إدخال مبلغ صحيح")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            description = self.description_edit.text().strip()
            expense_type = self.expense_type_combo.currentText()
            amount = float(self.amount_edit.text())
            expense_date = self.expense_date_edit.date().toString(Qt.ISODate)
            receiver = self.receiver_edit.text().strip()
            invoice_num = self.invoice_number_edit.text().strip()

            if self.is_edit_mode:
                # تحديث المصروف الموجود
                cursor.execute("""
                    UPDATE التدريب_مصروفات
                    SET وصف_المصروف = %s, نوع_المصروف = %s, المبلغ = %s,
                        تاريخ_المصروف = %s, المستلم = %s, رقم_الفاتورة = %s
                    WHERE id = %s
                """, (description, expense_type, amount, expense_date, receiver,
                      invoice_num, self.expense_id))

                message = "تم تحديث المصروف بنجاح"
            else:
                # إضافة مصروف جديد
                cursor.execute("""
                    INSERT INTO التدريب_مصروفات
                    (معرف_الدورة, وصف_المصروف, نوع_المصروف, المبلغ,
                     تاريخ_المصروف, المستلم, رقم_الفاتورة)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (self.training_id, description, expense_type, amount, expense_date,
                      receiver, invoice_num))

                message = "تم إضافة المصروف بنجاح"

            conn.commit()
            conn.close()

            QMessageBox.information(self, "نجح", message)
            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ المصروف: {str(e)}")


# حوار إضافة/تعديل دفعة طالب
class StudentPaymentDialog(QDialog):

    # init
    def __init__(self, parent=None, training_id=None, payment_id=None):
        super().__init__(parent)
        self.training_id = training_id
        self.payment_id = payment_id
        self.is_edit_mode = payment_id is not None

        self.setup_dialog()
        self.create_ui()
        self.load_students()

        if self.is_edit_mode:
            self.load_payment_data()

        apply_stylesheet(self)

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل دفعة طالب" if self.is_edit_mode else "إضافة دفعة طالب جديدة"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 500, 400)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء واجهة المستخدم
    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان النافذة
        title_label = QLabel("تعديل دفعة الطالب" if self.is_edit_mode else "إضافة دفعة طالب جديدة")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # نموذج الإدخال
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # الطالب
        self.student_combo = QComboBox()
        self.student_combo.addItem("اختر الطالب", None)
        form_layout.addRow("الطالب:", self.student_combo)

        # المجموعة (للعرض فقط)
        self.group_combo = QComboBox()
        self.group_combo.addItem("اختر المجموعة", None)
        form_layout.addRow("المجموعة:", self.group_combo)

        # وصف المدفوع
        self.description_edit = QLineEdit()
        self.description_edit.setPlaceholderText("مثال: دفعة أولى، دفعة ثانية، إلخ")
        form_layout.addRow("وصف المدفوع:", self.description_edit)

        # المبلغ المدفوع
        self.amount_edit = QLineEdit()
        self.amount_edit.setPlaceholderText("أدخل المبلغ المدفوع")
        amount_validator = QDoubleValidator(0.0, 999999.99, 2)
        amount_validator.setNotation(QDoubleValidator.StandardNotation)
        self.amount_edit.setValidator(amount_validator)
        form_layout.addRow("المبلغ المدفوع:", self.amount_edit)

        # تاريخ الدفع
        self.payment_date_edit = QDateEdit()
        self.payment_date_edit.setCalendarPopup(True)
        self.payment_date_edit.setDate(QDate.currentDate())
        setup_date_edit_format(self.payment_date_edit)
        form_layout.addRow("تاريخ الدفع:", self.payment_date_edit)

        # طريقة الدفع
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItems(["نقدي", "شيك", "تحويل بنكي"])
        form_layout.addRow("طريقة الدفع:", self.payment_method_combo)

        # المستلم
        self.receiver_edit = QLineEdit()
        self.receiver_edit.setPlaceholderText("اسم الشخص الذي استلم المبلغ")
        form_layout.addRow("المستلم:", self.receiver_edit)

        layout.addLayout(form_layout)

        # معلومات الطالب (للعرض)
        info_frame = QFrame()
        info_frame.setStyleSheet("QFrame { border: 1px solid #ddd; border-radius: 5px; padding: 10px; background-color: #f8f9fa; }")
        info_layout = QFormLayout(info_frame)

        self.student_total_label = QLabel("0.00")
        self.student_paid_label = QLabel("0.00")
        self.student_remaining_label = QLabel("0.00")

        info_layout.addRow("إجمالي مبلغ الطالب:", self.student_total_label)
        info_layout.addRow("المدفوع حالياً:", self.student_paid_label)
        info_layout.addRow("المتبقي:", self.student_remaining_label)

        layout.addWidget(info_frame)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("حفظ")
        save_button.clicked.connect(self.save_payment)
        cancel_button = QPushButton("إلغاء")
        cancel_button.clicked.connect(self.reject)

        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)

        # ربط تغيير الطالب بتحديث المعلومات
        self.student_combo.currentIndexChanged.connect(self.update_student_info)

    # تحميل الطلبة المتاحين للدورة
    def load_students(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT s.id, s.اسم_الطالب, s.المبلغ, s.المدفوع, s.الباقي,
                       COALESCE(g.`اسم المجموعه`, 'غير محدد') as group_name,
                       s.معرف_المجموعة
                FROM التدريب_الطلاب s
                LEFT JOIN التدريب_المجموعات g ON s.معرف_المجموعة = g.id
                WHERE s.معرف_الدورة = %s
                ORDER BY s.اسم_الطالب
            """, (self.training_id,))

            students = cursor.fetchall()
            conn.close()

            for student_id, name, total, paid, remaining, group_name, group_id in students:
                display_text = f"{name} - {group_name} (متبقي: {remaining:,.2f})"
                self.student_combo.addItem(display_text, {
                    'id': student_id,
                    'name': name,
                    'total': total,
                    'paid': paid,
                    'remaining': remaining,
                    'group_name': group_name,
                    'group_id': group_id
                })

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل الطلبة: {str(e)}")

    # تحديث معلومات الطالب المحدد
    def update_student_info(self):
        student_data = self.student_combo.currentData()
        if student_data:
            self.student_total_label.setText(f"{student_data['total']:,.2f}")
            self.student_paid_label.setText(f"{student_data['paid']:,.2f}")
            self.student_remaining_label.setText(f"{student_data['remaining']:,.2f}")

            # تحديث المجموعة
            self.group_combo.clear()
            self.group_combo.addItem(student_data['group_name'], student_data['group_id'])
        else:
            self.student_total_label.setText("0.00")
            self.student_paid_label.setText("0.00")
            self.student_remaining_label.setText("0.00")
            self.group_combo.clear()

    # تحميل بيانات الدفعة للتعديل
    def load_payment_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT معرف_الطالب, معرف_المجموعة, وصف_المدفوع, المبلغ_المدفوع,
                       تاريخ_الدفع, طريقة_الدفع, المستلم
                FROM التدريب_دفعات_الطلاب
                WHERE id = %s
            """, (self.payment_id,))

            result = cursor.fetchone()
            conn.close()

            if result:
                student_id, group_id, description, amount, payment_date, method, receiver = result

                # تحديد الطالب
                for i in range(self.student_combo.count()):
                    student_data = self.student_combo.itemData(i)
                    if student_data and student_data['id'] == student_id:
                        self.student_combo.setCurrentIndex(i)
                        break

                self.description_edit.setText(description or "")
                self.amount_edit.setText(str(amount) if amount else "")

                if payment_date:
                    self.payment_date_edit.setDate(QDate.fromString(str(payment_date), Qt.ISODate))

                # تحديد طريقة الدفع
                if method:
                    index = self.payment_method_combo.findText(method)
                    if index >= 0:
                        self.payment_method_combo.setCurrentIndex(index)

                self.receiver_edit.setText(receiver or "")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات الدفعة: {str(e)}")

    # حفظ بيانات الدفعة
    def save_payment(self):
        # التحقق من صحة البيانات
        student_data = self.student_combo.currentData()
        if not student_data:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار الطالب")
            return

        if not self.description_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال وصف المدفوع")
            return

        if not self.amount_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال المبلغ المدفوع")
            return

        try:
            amount = float(self.amount_edit.text())
            if amount <= 0:
                QMessageBox.warning(self, "تحذير", "المبلغ يجب أن يكون أكبر من صفر")
                return
        except ValueError:
            QMessageBox.warning(self, "تحذير", "يرجى إدخال مبلغ صحيح")
            return

        # التحقق من عدم تجاوز المبلغ المتبقي (في حالة الإضافة)
        if not self.is_edit_mode:
            if amount > student_data['remaining']:
                reply = QMessageBox.question(
                    self, "تحذير",
                    f"المبلغ المدخل ({amount:,.2f}) أكبر من المبلغ المتبقي ({student_data['remaining']:,.2f}).\n\nهل تريد المتابعة؟",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if reply == QMessageBox.No:
                    return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            student_id = student_data['id']
            group_id = student_data['group_id']
            description = self.description_edit.text().strip()
            amount = float(self.amount_edit.text())
            payment_date = self.payment_date_edit.date().toString(Qt.ISODate)
            method = self.payment_method_combo.currentText()
            receiver = self.receiver_edit.text().strip()

            if self.is_edit_mode:
                # الحصول على المبلغ القديم لتحديث رصيد الطالب
                cursor.execute("""
                    SELECT المبلغ_المدفوع FROM التدريب_دفعات_الطلاب
                    WHERE id = %s
                """, (self.payment_id,))
                old_amount = cursor.fetchone()[0]

                # تحديث الدفعة الموجودة
                cursor.execute("""
                    UPDATE التدريب_دفعات_الطلاب
                    SET معرف_الطالب = %s, معرف_المجموعة = %s, وصف_المدفوع = %s,
                        المبلغ_المدفوع = %s, تاريخ_الدفع = %s, طريقة_الدفع = %s, المستلم = %s
                    WHERE id = %s
                """, (student_id, group_id, description, amount, payment_date, method, receiver, self.payment_id))

                # تحديث رصيد الطالب (إزالة المبلغ القديم وإضافة الجديد)
                cursor.execute("""
                    UPDATE التدريب_الطلاب
                    SET المدفوع = المدفوع - %s + %s
                    WHERE id = %s
                """, (old_amount, amount, student_id))

                message = "تم تحديث الدفعة بنجاح"
            else:
                # إضافة دفعة جديدة
                cursor.execute("""
                    INSERT INTO التدريب_دفعات_الطلاب
                    (معرف_الدورة, معرف_المجموعة, معرف_الطالب, وصف_المدفوع,
                     المبلغ_المدفوع, تاريخ_الدفع, طريقة_الدفع, المستلم)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (self.training_id, group_id, student_id, description, amount, payment_date, method, receiver))

                # تحديث رصيد الطالب
                cursor.execute("""
                    UPDATE التدريب_الطلاب
                    SET المدفوع = المدفوع + %s
                    WHERE id = %s
                """, (amount, student_id))

                message = "تم إضافة الدفعة بنجاح"

            conn.commit()
            conn.close()

            # تحديث أرصدة الطلبة وإجماليات الدورة ومبالغ المدربين
            if hasattr(self.parent(), 'update_student_balances'):
                self.parent().update_student_balances()
            if hasattr(self.parent(), 'update_course_totals'):
                self.parent().update_course_totals()
            if hasattr(self.parent(), 'calculate_trainer_amounts'):
                self.parent().calculate_trainer_amounts()

            # إنشاء مهام جديدة للمدربين إذا كانت دفعة جديدة
            if not self.is_edit_mode and hasattr(self.parent(), 'create_trainer_tasks_for_new_payments'):
                self.parent().create_trainer_tasks_for_new_payments()

            QMessageBox.information(self, "نجح", message)
            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ الدفعة: {str(e)}")


# حوار إضافة/تعديل مجموعة
class GroupDialog(QDialog):

    # init
    def __init__(self, parent=None, training_id=None, group_id=None):
        super().__init__(parent)
        self.training_id = training_id
        self.group_id = group_id
        self.is_edit_mode = group_id is not None

        self.setup_dialog()
        self.create_ui()

        if self.is_edit_mode:
            self.load_group_data()

        apply_stylesheet(self)

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل مجموعة" if self.is_edit_mode else "إضافة مجموعة جديدة"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 450, 350)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء واجهة المستخدم
    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان النافذة
        title_label = QLabel("تعديل بيانات المجموعة" if self.is_edit_mode else "إضافة مجموعة جديدة")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # نموذج الإدخال
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # اسم المجموعة
        self.group_name_edit = QLineEdit()
        self.group_name_edit.setPlaceholderText("أدخل اسم المجموعة")
        form_layout.addRow("اسم المجموعة:", self.group_name_edit)

        # التوقيت
        self.timing_edit = QLineEdit()
        self.timing_edit.setPlaceholderText("مثال: السبت والاثنين 10:00 ص - 12:00 م")
        form_layout.addRow("التوقيت:", self.timing_edit)

        # العدد المطلوب
        self.required_count_edit = QLineEdit()
        self.required_count_edit.setPlaceholderText("أدخل العدد المطلوب للمجموعة")
        required_validator = QIntValidator(1, 999)
        self.required_count_edit.setValidator(required_validator)
        form_layout.addRow("العدد المطلوب:", self.required_count_edit)

        # الحالة
        self.status_combo = QComboBox()
        self.status_combo.addItems([
            "مفتوحة للتسجيل",
            "مكتملة العدد",
            "بدأت الدورة",
            "منتهية",
            "ملغاة"
        ])
        form_layout.addRow("الحالة:", self.status_combo)

        # ملاحظات
        self.notes_edit = QTextEdit()
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات")
        self.notes_edit.setMaximumHeight(80)
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("حفظ")
        save_button.clicked.connect(self.save_group)
        cancel_button = QPushButton("إلغاء")
        cancel_button.clicked.connect(self.reject)

        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)

    # تحميل بيانات المجموعة للتعديل
    def load_group_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT `اسم المجموعه`, التوقيت, العدد_المطلوب, الحالة, ملاحظات
                FROM التدريب_المجموعات
                WHERE id = %s
            """, (self.group_id,))

            result = cursor.fetchone()
            conn.close()

            if result:
                group_name, timing, required_count, status, notes = result

                self.group_name_edit.setText(group_name or "")
                self.timing_edit.setText(timing or "")
                self.required_count_edit.setText(str(required_count) if required_count else "")

                # تحديد الحالة
                if status:
                    index = self.status_combo.findText(status)
                    if index >= 0:
                        self.status_combo.setCurrentIndex(index)

                self.notes_edit.setPlainText(notes or "")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات المجموعة: {str(e)}")

    # حفظ بيانات المجموعة
    def save_group(self):
        # التحقق من صحة البيانات
        if not self.group_name_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال اسم المجموعة")
            return

        if not self.required_count_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال العدد المطلوب")
            return

        try:
            required_count = int(self.required_count_edit.text())
            if required_count <= 0:
                QMessageBox.warning(self, "تحذير", "العدد المطلوب يجب أن يكون أكبر من صفر")
                return
        except ValueError:
            QMessageBox.warning(self, "تحذير", "يرجى إدخال عدد صحيح")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            group_name = self.group_name_edit.text().strip()
            timing = self.timing_edit.text().strip()
            required_count = int(self.required_count_edit.text())
            status = self.status_combo.currentText()
            notes = self.notes_edit.toPlainText().strip()

            if self.is_edit_mode:
                # تحديث المجموعة الموجودة
                cursor.execute("""
                    UPDATE التدريب_المجموعات
                    SET `اسم المجموعه` = %s, التوقيت = %s, العدد_المطلوب = %s,
                        الحالة = %s, ملاحظات = %s
                    WHERE id = %s
                """, (group_name, timing, required_count, status, notes, self.group_id))

                message = "تم تحديث بيانات المجموعة بنجاح"
            else:
                # إضافة مجموعة جديدة
                cursor.execute("""
                    INSERT INTO التدريب_المجموعات
                    (معرف_الدورة, `اسم المجموعه`, التوقيت, عدد_المشتركين, العدد_المطلوب, الحالة, ملاحظات)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (self.training_id, group_name, timing, 0, required_count, status, notes))

                message = "تم إضافة المجموعة بنجاح"

            conn.commit()
            conn.close()

            QMessageBox.information(self, "نجح", message)
            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ بيانات المجموعة: {str(e)}")


# حوار إضافة/تعديل طالب
class StudentDialog(QDialog):

    # init
    def __init__(self, parent=None, training_id=None, student_id=None):
        super().__init__(parent)
        self.training_id = training_id
        self.student_id = student_id
        self.is_edit_mode = student_id is not None

        self.setup_dialog()
        self.create_ui()
        self.load_groups()

        if self.is_edit_mode:
            self.load_student_data()

        apply_stylesheet(self)

    # إعداد الحوار
    def setup_dialog(self):
        title = "تعديل طالب" if self.is_edit_mode else "إضافة طالب جديد"
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 500, 400)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setModal(True)

    # إنشاء واجهة المستخدم
    def create_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان النافذة
        title_label = QLabel("تعديل بيانات الطالب" if self.is_edit_mode else "إضافة طالب جديد")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # نموذج الإدخال
        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # اسم الطالب
        self.student_name_edit = QLineEdit()
        self.student_name_edit.setPlaceholderText("أدخل اسم الطالب")
        form_layout.addRow("اسم الطالب:", self.student_name_edit)

        # رقم الهاتف
        self.phone_edit = QLineEdit()
        self.phone_edit.setPlaceholderText("أدخل رقم الهاتف")
        form_layout.addRow("رقم الهاتف:", self.phone_edit)

        # العنوان
        self.address_edit = QLineEdit()
        self.address_edit.setPlaceholderText("أدخل العنوان")
        form_layout.addRow("العنوان:", self.address_edit)

        # المجموعة
        self.group_combo = QComboBox()
        self.group_combo.addItem("اختر المجموعة", None)
        form_layout.addRow("المجموعة:", self.group_combo)

        # المبلغ
        self.amount_edit = QLineEdit()
        self.amount_edit.setPlaceholderText("أدخل مبلغ الدورة للطالب")
        amount_validator = QDoubleValidator(0.0, 999999.99, 2)
        amount_validator.setNotation(QDoubleValidator.StandardNotation)
        self.amount_edit.setValidator(amount_validator)
        form_layout.addRow("المبلغ:", self.amount_edit)

        # تاريخ التسجيل
        self.registration_date_edit = QDateEdit()
        self.registration_date_edit.setCalendarPopup(True)
        self.registration_date_edit.setDate(QDate.currentDate())
        setup_date_edit_format(self.registration_date_edit)
        form_layout.addRow("تاريخ التسجيل:", self.registration_date_edit)

        # ملاحظات
        self.notes_edit = QTextEdit()
        self.notes_edit.setPlaceholderText("أدخل أي ملاحظات")
        self.notes_edit.setMaximumHeight(80)
        form_layout.addRow("ملاحظات:", self.notes_edit)

        layout.addLayout(form_layout)

        # أزرار الحفظ والإلغاء
        buttons_layout = QHBoxLayout()
        save_button = QPushButton("حفظ")
        save_button.clicked.connect(self.save_student)
        cancel_button = QPushButton("إلغاء")
        cancel_button.clicked.connect(self.reject)

        buttons_layout.addWidget(save_button)
        buttons_layout.addWidget(cancel_button)
        layout.addLayout(buttons_layout)

    # تحميل المجموعات المتاحة للدورة
    def load_groups(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT id, `اسم المجموعه`, التوقيت
                FROM التدريب_المجموعات
                WHERE معرف_الدورة = %s
                ORDER BY `اسم المجموعه`
            """, (self.training_id,))

            groups = cursor.fetchall()
            conn.close()

            for group_id, group_name, timing in groups:
                display_text = f"{group_name}"
                if timing:
                    display_text += f" - {timing}"
                self.group_combo.addItem(display_text, group_id)

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل المجموعات: {str(e)}")

    # تحميل بيانات الطالب للتعديل
    def load_student_data(self):
        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            cursor.execute("""
                SELECT اسم_الطالب, رقم_الهاتف, العنوان, معرف_المجموعة,
                       المبلغ, تاريخ_التسجيل, ملاحظات
                FROM التدريب_الطلاب
                WHERE id = %s
            """, (self.student_id,))

            result = cursor.fetchone()
            conn.close()

            if result:
                name, phone, address, group_id, amount, reg_date, notes = result

                self.student_name_edit.setText(name or "")
                self.phone_edit.setText(phone or "")
                self.address_edit.setText(address or "")

                # تحديد المجموعة
                if group_id:
                    for i in range(self.group_combo.count()):
                        if self.group_combo.itemData(i) == group_id:
                            self.group_combo.setCurrentIndex(i)
                            break

                self.amount_edit.setText(str(amount) if amount else "")

                if reg_date:
                    self.registration_date_edit.setDate(QDate.fromString(str(reg_date), Qt.ISODate))

                self.notes_edit.setPlainText(notes or "")

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في تحميل بيانات الطالب: {str(e)}")

    # حفظ بيانات الطالب
    def save_student(self):
        # التحقق من صحة البيانات
        if not self.student_name_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال اسم الطالب")
            return

        if self.group_combo.currentData() is None:
            QMessageBox.warning(self, "تحذير", "يرجى اختيار المجموعة")
            return

        if not self.amount_edit.text().strip():
            QMessageBox.warning(self, "تحذير", "يرجى إدخال المبلغ")
            return

        try:
            amount = float(self.amount_edit.text())
            if amount < 0:
                QMessageBox.warning(self, "تحذير", "المبلغ يجب أن يكون أكبر من أو يساوي صفر")
                return
        except ValueError:
            QMessageBox.warning(self, "تحذير", "يرجى إدخال مبلغ صحيح")
            return

        try:
            conn = mysql.connector.connect(
                host=host, user=user, password=password,
                database="project_manager_V2"
            )
            cursor = conn.cursor()

            student_name = self.student_name_edit.text().strip()
            phone = self.phone_edit.text().strip()
            address = self.address_edit.text().strip()
            group_id = self.group_combo.currentData()
            amount = float(self.amount_edit.text())
            reg_date = self.registration_date_edit.date().toString(Qt.ISODate)
            notes = self.notes_edit.toPlainText().strip()

            if self.is_edit_mode:
                # تحديث الطالب الموجود
                cursor.execute("""
                    UPDATE التدريب_الطلاب
                    SET اسم_الطالب = %s, رقم_الهاتف = %s, العنوان = %s,
                        معرف_المجموعة = %s, المبلغ = %s, تاريخ_التسجيل = %s, ملاحظات = %s
                    WHERE id = %s
                """, (student_name, phone, address, group_id, amount, reg_date, notes, self.student_id))

                message = "تم تحديث بيانات الطالب بنجاح"
            else:
                # إضافة طالب جديد
                cursor.execute("""
                    INSERT INTO التدريب_الطلاب
                    (معرف_الدورة, معرف_المجموعة, اسم_الطالب, رقم_الهاتف, العنوان,
                     المبلغ, تاريخ_التسجيل, ملاحظات)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (self.training_id, group_id, student_name, phone, address, amount, reg_date, notes))

                message = "تم إضافة الطالب بنجاح"

            conn.commit()
            conn.close()

            QMessageBox.information(self, "نجح", message)
            self.accept()

        except Exception as e:
            QMessageBox.warning(self, "خطأ", f"فشل في حفظ بيانات الطالب: {str(e)}")

# دالة فتح نافذة إدارة التدريب
# فتح نافذة إدارة التدريب
def open_training_management_window(parent, training_data):
    window = TrainingManagementWindow(parent, training_data)
    window.show()
    return window
