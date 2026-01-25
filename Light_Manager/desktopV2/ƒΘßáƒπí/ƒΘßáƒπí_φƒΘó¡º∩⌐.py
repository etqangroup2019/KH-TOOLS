from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtPrintSupport import *
from PySide6.QtWebEngineWidgets import QWebEngineView
import pandas as pd
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from datetime import datetime
import tempfile
import subprocess
from الإعدادات_العامة import *

# تحويل الأرقام إلى كلمات عربية
def number_to_arabic_words(number):
    try:
        # تحويل الرقم إلى عدد صحيح
        num = int(float(str(number)))

        if num == 0:
            return "صفر"

        # قوائم الأرقام العربية
        ones = ["", "واحد", "اثنان", "ثلاثة", "أربعة", "خمسة", "ستة", "سبعة", "ثمانية", "تسعة"]
        tens = ["", "", "عشرون", "ثلاثون", "أربعون", "خمسون", "ستون", "سبعون", "ثمانون", "تسعون"]
        hundreds = ["", "مائة", "مائتان", "ثلاثمائة", "أربعمائة", "خمسمائة", "ستمائة", "سبعمائة", "ثمانمائة", "تسعمائة"]

        # تحويل المئات
        def convert_hundreds(n):
            result = ""

            # المئات
            if n >= 100:
                h = n // 100
                result += hundreds[h]
                n %= 100
                if n > 0:
                    result += " و"

            # العشرات والآحاد
            if n >= 20:
                t = n // 10
                result += tens[t]
                n %= 10
                if n > 0:
                    result += " و" + ones[n]
            elif n >= 11:
                teens = ["", "أحد عشر", "اثنا عشر", "ثلاثة عشر", "أربعة عشر", "خمسة عشر",
                        "ستة عشر", "سبعة عشر", "ثمانية عشر", "تسعة عشر"]
                result += teens[n - 10]
            elif n == 10:
                result += "عشرة"
            elif n > 0:
                result += ones[n]

            return result

        if num < 1000:
            return convert_hundreds(num)
        elif num < 1000000:
            thousands = num // 1000
            remainder = num % 1000
            result = convert_hundreds(thousands) + " ألف"
            if remainder > 0:
                result += " و" + convert_hundreds(remainder)
            return result
        else:
            return str(num)  # للأرقام الكبيرة جداً

    except:
        return str(number)

# نافذة سند صرف المصروفات
class ExpenseVoucherDialog(QDialog):
    # init
    def __init__(self, expense_data, parent=None):
        super().__init__(parent)
        self.expense_data = expense_data
        self.parent_window = parent

        # مرجع لنافذة المعاينة لإدارتها
        self.preview_dialog = None

        # إعدادات الشركة
        self.company_name = company_name
        self.logo_path = logo_path
        self.company_phone = company_phone
        self.company_address = company_address
        self.company_email = company_email
        self.currency_type = Currency_type

        self.setup_ui()
        self.apply_styles()

    # معالجة إغلاق النافذة الرئيسية
    def closeEvent(self, event):
        try:
            # إغلاق نافذة المعاينة إذا كانت مفتوحة
            if self.preview_dialog is not None:
                try:
                    self.preview_dialog.close()
                    self.preview_dialog = None
                except:
                    pass

            # قبول الإغلاق
            event.accept()
        except Exception as e:
            print(f"خطأ في إغلاق النافذة: {e}")
            event.accept()

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("سند صرف مصروفات")
        self.setFixedSize(800, 600)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout()

        # عنوان رئيسي
        title = QLabel("سند صرف مصروفات")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50; margin: 15px;")
        layout.addWidget(title)

        # معلومات السند
        info_group = QGroupBox("معلومات السند")
        info_layout = QFormLayout()

        # عرض بيانات المصروف
        self.expense_type_label = QLabel(str(self.expense_data.get('التصنيف', '')))
        self.expense_desc_label = QLabel(str(self.expense_data.get('المصروف', '')))
        self.amount_label = QLabel(f"{self.expense_data.get('المبلغ', 0)} {self.currency_type}")
        self.date_label = QLabel(str(self.expense_data.get('تاريخ_المصروف', '')))
        self.recipient_label = QLabel(str(self.expense_data.get('المستلم', '')))
        self.phone_label = QLabel(str(self.expense_data.get('رقم_الهاتف', '')))
        self.invoice_label = QLabel(str(self.expense_data.get('رقم_الفاتورة', '')))

        # تحويل المبلغ إلى كلمات
        amount_in_words = number_to_arabic_words(self.expense_data.get('المبلغ', 0))
        self.amount_words_label = QLabel(f"{amount_in_words} {self.currency_type} فقط لا غير")

        info_layout.addRow("نوع المصروف:", self.expense_type_label)
        info_layout.addRow("وصف المصروف:", self.expense_desc_label)
        info_layout.addRow("المبلغ:", self.amount_label)
        info_layout.addRow("المبلغ بالكلمات:", self.amount_words_label)
        info_layout.addRow("تاريخ المصروف:", self.date_label)
        info_layout.addRow("المستلم:", self.recipient_label)
        info_layout.addRow("رقم الهاتف:", self.phone_label)
        info_layout.addRow("رقم الفاتورة:", self.invoice_label)

        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        # أزرار العمليات
        buttons_layout = QHBoxLayout()

        self.preview_btn = QPushButton("معاينة السند")
        self.preview_btn.clicked.connect(self.preview_voucher)
        buttons_layout.addWidget(self.preview_btn)

        self.print_btn = QPushButton("طباعة السند")
        self.print_btn.clicked.connect(self.print_voucher)
        buttons_layout.addWidget(self.print_btn)

        self.pdf_btn = QPushButton("حفظ PDF")
        self.pdf_btn.clicked.connect(self.save_pdf)
        buttons_layout.addWidget(self.pdf_btn)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    # تطبيق الأنماط على النافذة
    def apply_styles(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                border-radius: 10px;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin: 10px 0;
                padding-top: 15px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #2c3e50;
                background-color: #f8f9fa;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 120px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
                padding: 5px;
            }
            QFormLayout QLabel {
                font-weight: bold;
            }
        """)

    # إنشاء محتوى HTML لسند الصرف
    def create_voucher_html(self):
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M")

        # تحويل المبلغ إلى كلمات
        amount_in_words = number_to_arabic_words(self.expense_data.get('المبلغ', 0))

        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="ar">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>سند صرف مصروفات</title>
            <style>
                @page {{
                    size: A4;
                    margin: 2.5cm;
                }}

                @media print {{
                    body {{
                        -webkit-print-color-adjust: exact;
                        print-color-adjust: exact;
                    }}
                }}

                body {{
                    font-family: 'Arial', 'Tahoma', sans-serif;
                    direction: rtl;
                    text-align: right;
                    margin: 0;
                    padding: 15px;
                    background-color: white;
                    color: #333;
                    line-height: 1.4;
                    height: 100vh;
                    overflow: hidden;
                    box-sizing: border-box;
                }}

                .header {{
                    text-align: center;
                    margin-bottom: 20px;
                    border-bottom: 2px solid #2c3e50;
                    padding-bottom: 15px;
                }}

                .company-logo {{
                    max-width: 80px;
                    max-height: 60px;
                    margin-bottom: 8px;
                }}

                .company-name {{
                    font-size: 22px;
                    font-weight: bold;
                    color: #2c3e50;
                    margin: 8px 0;
                }}

                .voucher-title {{
                    font-size: 20px;
                    font-weight: bold;
                    color: #e74c3c;
                    margin: 15px 0;
                    text-decoration: underline;
                }}

                .voucher-number {{
                    font-size: 14px;
                    color: #7f8c8d;
                    margin-bottom: 8px;
                }}

                .voucher-content {{
                    background-color: #f8f9fa;
                    border: 2px solid #3498db;
                    border-radius: 8px;
                    padding: 20px;
                    margin: 15px 0;
                }}

                .voucher-row {{
                    display: flex;
                    justify-content: space-between;
                    margin: 10px 0;
                    padding: 8px;
                    background-color: white;
                    border-radius: 4px;
                    border-right: 3px solid #3498db;
                }}

                .voucher-label {{
                    font-weight: bold;
                    color: #2c3e50;
                    min-width: 130px;
                    font-size: 14px;
                }}

                .voucher-value {{
                    color: #34495e;
                    flex: 1;
                    text-align: center;
                    font-size: 14px;
                    padding: 0 10px;
                }}

                .amount-section {{
                    background-color: #ecf0f1;
                    border: 2px solid #e74c3c;
                    border-radius: 8px;
                    padding: 15px;
                    margin: 15px 0;
                    text-align: center;
                }}

                .amount-number {{
                    font-size: 20px;
                    font-weight: bold;
                    color: #e74c3c;
                    margin: 8px 0;
                }}

                .amount-words {{
                    font-size: 16px;
                    font-weight: bold;
                    color: #2c3e50;
                    background-color: white;
                    padding: 12px;
                    border-radius: 4px;
                    border: 1px solid #bdc3c7;
                }}

                .signatures {{
                    display: flex;
                    justify-content: space-between;
                    margin-top: 30px;
                    padding-top: 20px;
                }}

                .signature-box {{
                    text-align: center;
                    width: 150px;
                    border-top: 2px solid #2c3e50;
                    padding-top: 8px;
                }}

                .signature-label {{
                    font-weight: bold;
                    color: #2c3e50;
                    font-size: 14px;
                }}

                .footer {{
                    position: fixed;
                    bottom: 0;
                    left: 0;
                    right: 0;
                    background-color: #587ca0;
                    color: white;
                    text-align: center;
                    padding: 15px;
                    font-size: 11px;
                    border-top: 2px solid #3498db;
                    height: calc(100vh - 99vh);
                    min-height: 25px;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }}

                .content-wrapper {{
                    margin-bottom: 0;
                    padding-bottom: 20px;
                    min-height: 92vh;
                    display: flex;
                    flex-direction: column;
                    justify-content: space-between;
                }}

                .footer-item {{
                    display: inline-block;
                    margin: 0 15px;
                }}

                .print-info {{
                    position: fixed;
                    top: 15px;
                    left: 15px;
                    font-size: 11px;
                    color: #7f8c8d;
                }}
            </style>
        </head>
        <body>
            <div class="print-info">
                تاريخ الطباعة: {current_date} | الوقت: {current_time}
            </div>

            <div class="content-wrapper">
                <div class="header">
                    {f'<img src="{self.logo_path}" class="company-logo" alt="شعار الشركة">' if self.logo_path and os.path.exists(self.logo_path) else ''}
                    <div class="company-name">{self.company_name}</div>
                    <div class="voucher-title">سند صرف مصروفات</div>
                    <div class="voucher-number">رقم السند: {self.expense_data.get('id', 'غير محدد')}</div>
                </div>

            <div class="voucher-content">
                <div class="voucher-row">
                    <span class="voucher-label">نوع المصروف:</span>
                    <span class="voucher-value">{self.expense_data.get('التصنيف', '')}</span>
                </div>

                <div class="voucher-row">
                    <span class="voucher-label">وصف المصروف:</span>
                    <span class="voucher-value">{self.expense_data.get('المصروف', '')}</span>
                </div>

                <div class="voucher-row">
                    <span class="voucher-label">تاريخ المصروف:</span>
                    <span class="voucher-value">{self.expense_data.get('تاريخ_المصروف', '')}</span>
                </div>

                <div class="voucher-row">
                    <span class="voucher-label">المستلم:</span>
                    <span class="voucher-value">{self.expense_data.get('المستلم', '')}</span>
                </div>

                <div class="voucher-row">
                    <span class="voucher-label">رقم الهاتف:</span>
                    <span class="voucher-value">{self.expense_data.get('رقم_الهاتف', '')}</span>
                </div>

                <div class="voucher-row">
                    <span class="voucher-label">رقم الفاتورة:</span>
                    <span class="voucher-value">{self.expense_data.get('رقم_الفاتورة', '')}</span>
                </div>

                {f'<div class="voucher-row"><span class="voucher-label">ملاحظات:</span><span class="voucher-value">{self.expense_data.get("ملاحظات", "")}</span></div>' if self.expense_data.get('ملاحظات') else ''}
            </div>

            <div class="amount-section">
                <div class="amount-number">
                    المبلغ: {self.expense_data.get('المبلغ', 0)} {self.currency_type}
                </div>
                <div class="amount-words">
                    {amount_in_words} {self.currency_type} فقط لا غير
                </div>
            </div>

            <div class="signatures">
                <div class="signature-box">
                    <div class="signature-label">توقيع المستلم</div>
                </div>
                <div class="signature-box">
                    <div class="signature-label">توقيع المحاسب</div>
                </div>
                <div class="signature-box">
                    <div class="signature-label">توقيع المدير</div>
                </div>
            </div>
            </div>

            <div class="footer">
                <span class="footer-item">📍 العنوان: {self.company_address}</span>
                <span class="footer-item">📞 الهاتف: {self.company_phone}</span>
                <span class="footer-item">✉ البريد الإلكتروني: {self.company_email}</span>
            </div>
        </body>
        </html>
        """

        return html_content

    # معاينة سند الصرف
    def preview_voucher(self):
        try:
            # إغلاق نافذة المعاينة السابقة إذا كانت موجودة
            if self.preview_dialog is not None:
                try:
                    self.preview_dialog.close()
                    self.preview_dialog = None
                except:
                    pass

            html_content = self.create_voucher_html()

            # إنشاء ملف مؤقت
            temp_dir = os.path.join(documents_folder, "Reports")
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            temp_file = os.path.join(temp_dir, f"expense_voucher_{self.expense_data.get('id', 'temp')}.html")

            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # فتح نافذة المعاينة باستخدام دالة وحفظ المرجع
            self.preview_dialog = preview_expense_voucher_dialog(temp_file, self)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في معاينة السند:\n{str(e)}")

    # طباعة سند الصرف
    def print_voucher(self):
        try:
            html_content = self.create_voucher_html()

            # إنشاء ملف مؤقت
            temp_dir = os.path.join(documents_folder, "Reports")
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            temp_file = os.path.join(temp_dir, f"expense_voucher_{self.expense_data.get('id', 'temp')}.html")

            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # فتح نافذة الطباعة باستخدام دالة
            preview_expense_voucher_dialog(temp_file, self, print_mode=True)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في طباعة السند:\n{str(e)}")

    # حفظ سند الصرف كملف PDF
    def save_pdf(self):
        try:
            # اختيار مكان الحفظ
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "حفظ سند الصرف",
                f"سند_صرف_{self.expense_data.get('id', 'temp')}.pdf",
                "PDF Files (*.pdf)"
            )

            if not file_path:
                return

            html_content = self.create_voucher_html()

            # إنشاء ملف HTML مؤقت
            temp_dir = os.path.join(documents_folder, "Reports")
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            temp_file = os.path.join(temp_dir, f"temp_voucher_{self.expense_data.get('id', 'temp')}.html")

            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # تحويل إلى PDF
            web_view = QWebEngineView()
            web_view.load(QUrl.fromLocalFile(temp_file))

            # عند الانتهاء من الحمل
            def on_load_finished():
                web_view.page().printToPdf(file_path)

            # على PDF الانتهاء
            def on_pdf_finished(pdf_file_path, success):
                if success:
                    QMessageBox.information(self, "تم الحفظ", f"تم حفظ سند الصرف بنجاح في:\n{pdf_file_path}")
                else:
                    QMessageBox.critical(self, "خطأ", "فشل في حفظ ملف PDF")

                # تنظيف الملف المؤقت
                try:
                    os.remove(temp_file)
                except:
                    pass

            web_view.loadFinished.connect(on_load_finished)
            web_view.page().pdfPrintingFinished.connect(on_pdf_finished)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في حفظ ملف PDF:\n{str(e)}")

# نافذة معاينة سند الصرف
class ExpenseVoucherPreviewDialog(QDialog):
    # init
    def __init__(self, html_file_path, parent=None, print_mode=False):
        super().__init__(parent)
        self.html_file_path = html_file_path
        self.print_mode = print_mode

        self.setWindowTitle("معاينة سند الصرف")
        self.resize(1000, 700)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout()

        # عرض HTML
        self.web_view = QWebEngineView()
        self.web_view.load(QUrl.fromLocalFile(html_file_path))
        layout.addWidget(self.web_view)

        # أزرار
        buttons_layout = QHBoxLayout()

        if not print_mode:
            print_btn = QPushButton("طباعة")
            print_btn.clicked.connect(self.print_document)
            buttons_layout.addWidget(print_btn)

        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(self.accept)
        buttons_layout.addWidget(close_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        # إذا كان في وضع الطباعة، افتح نافذة الطباعة تلقائياً
        if print_mode:
            QTimer.singleShot(1000, self.print_document)

    # طباعة المستند
    def print_document(self):
        try:
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPageSize(QPageSize.A4))
            printer.setPageOrientation(QPageLayout.Portrait)
            printer.setPageMargins(QMarginsF(25, 25, 25, 25), QPageLayout.Millimeter)

            dialog = QPrintDialog(printer, self)
            if dialog.exec() == QPrintDialog.Accepted:
                self.web_view.page().print(printer, lambda success: None)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")

# دالة معاينة سند الصرف
def preview_expense_voucher_dialog(html_file_path, parent=None, print_mode=False):
    dialog = None
    try:
        # إنشاء نافذة المعاينة كنافذة مستقلة
        dialog = QMainWindow(parent)  # ربط بالوالد لمنع الإغلاق التلقائي
        dialog.setWindowTitle("معاينة سند الصرف")
        dialog.resize(1000, 700)
        dialog.setLayoutDirection(Qt.RightToLeft)

        # جعل النافذة مستقلة ولكن مرتبطة بالوالد
        dialog.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMinMaxButtonsHint)

        # إنشاء widget مركزي
        central_widget = QWidget()
        dialog.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # عرض HTML
        web_view = QWebEngineView()

        # التأكد من تحميل الملف بشكل صحيح
        if not os.path.exists(html_file_path):
            raise Exception(f"الملف غير موجود: {html_file_path}")

        web_view.load(QUrl.fromLocalFile(html_file_path))
        layout.addWidget(web_view)

        # أزرار
        buttons_layout = QHBoxLayout()

        if not print_mode:
            print_btn = QPushButton("طباعة")
            print_btn.clicked.connect(lambda: print_expense_voucher_document(web_view, dialog, auto_close=False))
            buttons_layout.addWidget(print_btn)

        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(dialog.close)
        buttons_layout.addWidget(close_btn)

        layout.addLayout(buttons_layout)

        # عرض النافذة
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

        # إذا كان في وضع الطباعة، افتح نافذة الطباعة بعد تحميل الصفحة
        if print_mode:
            # عند الانتهاء من الحمل
            def on_load_finished():
                QTimer.singleShot(500, lambda: print_expense_voucher_document(web_view, dialog, auto_close=True))
            web_view.loadFinished.connect(on_load_finished)

        # إرجاع مرجع للنافذة
        return dialog

    except Exception as e:
        if dialog:
            dialog.close()
        QMessageBox.critical(parent, "خطأ", f"فشل في معاينة السند:\n{str(e)}")
        return None


# طباعة مستند سند الصرف
def print_expense_voucher_document(web_view, parent, auto_close=False):
    try:
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPageSize(QPageSize.A4))
        printer.setPageOrientation(QPageLayout.Portrait)
        printer.setPageMargins(QMarginsF(25, 25, 25, 25), QPageLayout.Millimeter)

        dialog = QPrintDialog(printer, parent)
        if dialog.exec() == QPrintDialog.Accepted:
            # التأكد من وجود مجلد Reports
            temp_dir = os.path.join(documents_folder, "Reports")
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)

            # استخدام printToPdf ثم طباعة الملف
            temp_pdf = os.path.join(temp_dir, "temp_print.pdf")

            # دالة للتعامل مع انتهاء إنشاء PDF
            # على PDF الانتهاء
            def on_pdf_finished(file_path, success):
                if success and os.path.exists(file_path):
                    try:
                        # طباعة الملف باستخدام النظام
                        if os.name == 'nt':  # Windows
                            os.startfile(file_path, "print")
                        else:
                            import subprocess
                            subprocess.run(['lpr', file_path])

                        # عرض رسالة نجاح
                        if auto_close:
                            # في حالة الإغلاق التلقائي، عرض رسالة سريعة ثم إغلاق النافذة
                            QMessageBox.information(parent, "نجح", "تم إرسال السند للطباعة")
                            QTimer.singleShot(1000, parent.close)  # إغلاق النافذة بعد ثانية واحدة
                        else:
                            QMessageBox.information(parent, "نجح", "تم إرسال السند للطباعة")

                    except Exception as print_error:
                        QMessageBox.warning(parent, "تحذير", f"فشل في الطباعة المباشرة: {print_error}\nيمكنك فتح الملف وطباعته يدوياً")
                        if auto_close:
                            QTimer.singleShot(2000, parent.close)  # إغلاق النافذة بعد ثانيتين في حالة الخطأ
                else:
                    QMessageBox.critical(parent, "خطأ", f"فشل في إنشاء ملف PDF للطباعة")
                    if auto_close:
                        QTimer.singleShot(2000, parent.close)

            # ربط الإشارة وبدء إنشاء PDF
            web_view.page().pdfPrintingFinished.connect(on_pdf_finished)
            web_view.page().printToPdf(temp_pdf)

        else:
            # إذا ألغى المستخدم الطباعة وكان في وضع الإغلاق التلقائي
            if auto_close:
                QTimer.singleShot(500, parent.close)

    except Exception as e:
        QMessageBox.critical(parent, "خطأ", f"فشل في الطباعة:\n{str(e)}")
        if auto_close:
            QTimer.singleShot(2000, parent.close)  # إغلاق النافذة بعد ثانيتين في حالة الخطأ

# AdvancedPrintExportDialog
class AdvancedPrintExportDialog(QDialog):
    # init
    def __init__(self, table_data, headers, section_name, parent=None):
        super().__init__(parent)
        self.table_data = table_data
        self.headers = headers
        self.section_name = section_name
        self.parent_window = parent
        self.preview_dialog = None  # مرجع لنافذة المعاينة

        # إعدادات الشركة
        self.company_name = company_name
        self.logo_path = logo_path
        self.company_phone = company_phone
        self.company_address = company_address
        self.company_email = company_email
        self.currency_type = Currency_type
        self.account_type = account_type

        self.setup_ui()
        self.apply_styles()

        # إعدادات آمنة للنافذة
        self.setAttribute(Qt.WA_DeleteOnClose, True)
        self.setAttribute(Qt.WA_QuitOnClose, False)

    # تسجيل خط عربي للـ PDF
    def register_arabic_font(self):
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # قائمة بالخطوط العربية المحتملة
            arabic_fonts = [
                "fonts/NotoSansArabic-Regular.ttf",
                "fonts/Arial-Unicode.ttf",
                "fonts/Tahoma.ttf",
                "fonts/DejaVuSans.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/tahoma.ttf",
                "C:/Windows/Fonts/calibri.ttf"
            ]

            # محاولة تسجيل أول خط متاح
            for font_path in arabic_fonts:
                try:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('Arabic', font_path))
                        print(f"تم تسجيل الخط العربي: {font_path}")
                        return 'Arabic'
                except Exception as e:
                    print(f"فشل في تسجيل الخط {font_path}: {e}")
                    continue

            # إذا لم يتم العثور على خط عربي، استخدم خط النظام
            try:
                # محاولة استخدام خط النظام الافتراضي
                pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
                return 'Arabic'
            except:
                pass

            # كحل أخير، استخدم Helvetica مع تحذير
            print("تحذير: لم يتم العثور على خط عربي مناسب، سيتم استخدام Helvetica")
            return 'Helvetica'

        except Exception as e:
            print(f"خطأ في تسجيل الخط العربي: {e}")
            return 'Helvetica'

    # إصلاح النص العربي للعرض الصحيح في HTML
    def fix_arabic_text(self, text):
        if not text or not isinstance(text, str):
            return text

        # للمتصفح، نحتاج فقط لإضافة علامات الاتجاه بدون تغيير شكل النص
        # لأن CSS و JavaScript سيتولى الباقي
        if any('\u0600' <= char <= '\u06FF' for char in text):  # فحص وجود أحرف عربية
            # إضافة علامة اتجاه من اليمين إلى اليسار للمتصفح
            return f"&rlm;{text}&rlm;"  # Right-to-Left Mark for HTML

        return text

    # إصلاح النص العربي للعرض الصحيح في PDF
    def fix_arabic_text_for_pdf(self, text):
        if not text or not isinstance(text, str):
            return text

        try:
            # محاولة استخدام مكتبات دعم النص العربي
            import arabic_reshaper
            from bidi.algorithm import get_display

            # فحص وجود أحرف عربية
            if any('\u0600' <= char <= '\u06FF' for char in text):
                # إعادة تشكيل النص العربي
                reshaped_text = arabic_reshaper.reshape(text)
                # تطبيق خوارزمية الاتجاه الثنائي
                bidi_text = get_display(reshaped_text)
                return bidi_text

            return text

        except ImportError:
            # إذا لم تكن المكتبات متاحة، استخدم النص كما هو مع تحذير
            print("تحذير: مكتبات دعم النص العربي غير متاحة. قم بتشغيل install_arabic_support.bat")
            return text
        except Exception as e:
            print(f"خطأ في معالجة النص العربي: {e}")
            return text

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("طباعة وتصدير متقدم")
        self.setFixedSize(600, 700)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout()

        # عنوان رئيسي
        title = QLabel(f"طباعة وتصدير بيانات {self.section_name}")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50; margin: 10px;")
        layout.addWidget(title)

        # مجموعة إعدادات التصدير
        export_group = QGroupBox("إعدادات التصدير")
        export_layout = QVBoxLayout()

        # نوع التصدير
        export_type_layout = QHBoxLayout()
        export_type_layout.addWidget(QLabel("نوع التصدير:"))

        self.export_type = QComboBox()
        self.export_type.addItems(["PDF احترافي", "Excel متقدم", "طباعة مباشرة"])
        export_type_layout.addWidget(self.export_type)
        export_layout.addLayout(export_type_layout)

        # إعدادات الصفحة
        page_layout = QHBoxLayout()
        page_layout.addWidget(QLabel("اتجاه الصفحة:"))

        self.orientation = QComboBox()
        self.orientation.addItems(["عمودي", "أفقي"])
        page_layout.addWidget(self.orientation)
        export_layout.addLayout(page_layout)

        # حجم الخط
        font_layout = QHBoxLayout()
        font_layout.addWidget(QLabel("حجم الخط:"))

        self.font_size = QSpinBox()
        self.font_size.setRange(8, 16)
        self.font_size.setValue(10)
        font_layout.addWidget(self.font_size)
        export_layout.addLayout(font_layout)

        export_group.setLayout(export_layout)
        layout.addWidget(export_group)

        # مجموعة إعدادات المحتوى
        content_group = QGroupBox("إعدادات المحتوى")
        content_layout = QVBoxLayout()

        # خيارات المحتوى
        self.include_logo = QCheckBox("تضمين شعار الشركة")
        self.include_logo.setChecked(True)
        content_layout.addWidget(self.include_logo)

        self.include_company_info = QCheckBox("تضمين معلومات الشركة")
        self.include_company_info.setChecked(True)
        content_layout.addWidget(self.include_company_info)

        self.include_date = QCheckBox("تضمين التاريخ والوقت")
        self.include_date.setChecked(True)
        content_layout.addWidget(self.include_date)

        self.include_user = QCheckBox("تضمين اسم المستخدم")
        self.include_user.setChecked(True)
        content_layout.addWidget(self.include_user)

        self.include_page_numbers = QCheckBox("ترقيم الصفحات")
        self.include_page_numbers.setChecked(True)
        content_layout.addWidget(self.include_page_numbers)

        content_group.setLayout(content_layout)
        layout.addWidget(content_group)

        # مجموعة فلترة البيانات
        filter_group = QGroupBox("فلترة البيانات")
        filter_layout = QVBoxLayout()

        # اختيار الأعمدة
        columns_layout = QVBoxLayout()
        columns_layout.addWidget(QLabel("اختر الأعمدة للتصدير:"))

        self.columns_list = QListWidget()
        self.columns_list.setSelectionMode(QAbstractItemView.MultiSelection)

        for header in self.headers:
            item = QListWidgetItem(header)
            # تحديد جميع الأعمدة تلقائياً ما عدا أعمدة id
            should_select = not self.is_معرف_column(header)
            item.setSelected(should_select)
            self.columns_list.addItem(item)

        columns_layout.addWidget(self.columns_list)

        # أزرار تحديد الأعمدة
        columns_buttons = QHBoxLayout()
        select_all_btn = QPushButton("تحديد الكل")
        select_all_btn.clicked.connect(self.select_all_columns)
        columns_buttons.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("إلغاء تحديد الكل")
        deselect_all_btn.clicked.connect(self.deselect_all_columns)
        columns_buttons.addWidget(deselect_all_btn)

        columns_layout.addLayout(columns_buttons)
        filter_layout.addLayout(columns_layout)

        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)

        # أزرار العمليات
        buttons_layout = QHBoxLayout()

        self.preview_btn = QPushButton("معاينة")
        self.preview_btn.clicked.connect(self.preview_document)
        buttons_layout.addWidget(self.preview_btn)

        self.export_btn = QPushButton("تصدير")
        self.export_btn.clicked.connect(self.export_document)
        buttons_layout.addWidget(self.export_btn)

        self.print_btn = QPushButton("طباعة")
        self.print_btn.clicked.connect(self.print_document)
        buttons_layout.addWidget(self.print_btn)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        self.setLayout(layout)

    # معالجة إغلاق النافذة الرئيسية
    def closeEvent(self, event):
        try:
            # إغلاق نافذة المعاينة إذا كانت مفتوحة
            if self.preview_dialog is not None:
                try:
                    self.preview_dialog.close()
                    self.preview_dialog = None
                except:
                    pass

            # قبول الإغلاق
            event.accept()
        except Exception as e:
            print(f"خطأ في إغلاق النافذة الرئيسية: {e}")
            event.accept()

    # تطبيق الأنماط على النافذة
    def apply_styles(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                border-radius: 10px;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin: 10px 0;
                padding-top: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #2c3e50;
                background-color: #f8f9fa;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QCheckBox {
                spacing: 5px;
                color: #2c3e50;
                font-weight: normal;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #3498db;
                border-radius: 3px;
                background-color: #3498db;
                image: url(icons/check.png);
            }
            QComboBox, QSpinBox {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                padding: 5px;
                background-color: white;
                min-width: 100px;
            }
            QComboBox:focus, QSpinBox:focus {
                border-color: #3498db;
            }
            QListWidget {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                background-color: white;
                selection-background-color: #3498db;
            }
        """)

    # تحديد جميع الأعمدة
    def select_all_columns(self):
        for i in range(self.columns_list.count()):
            self.columns_list.item(i).setSelected(True)

    # إلغاء تحديد جميع الأعمدة
    def deselect_all_columns(self):
        for i in range(self.columns_list.count()):
            self.columns_list.item(i).setSelected(False)

    # الحصول على الأعمدة المحددة
    def get_selected_columns(self):
        selected_items = self.columns_list.selectedItems()
        selected_headers = [item.text() for item in selected_items]
        selected_indices = [self.headers.index(header) for header in selected_headers]
        return selected_indices, selected_headers

    # معاينة المستند
    def preview_document(self):
        export_type = self.export_type.currentText()

        if export_type == "PDF احترافي":
            self.preview_pdf()
        elif export_type == "Excel متقدم":
            self.preview_excel()
        else:
            self.preview_print()

    # تصدير المستند
    def export_document(self):
        export_type = self.export_type.currentText()

        if export_type == "PDF احترافي":
            self.export_pdf()
        elif export_type == "Excel متقدم":
            self.export_excel()
        else:
            self.print_document()

    # إنشاء محتوى HTML للطباعة - نسخة مبسطة للاختبار
    def create_html_content(self):
        return self.create_html_document()

    # طباعة المستند مباشرة
    def print_document(self):
        try:
            # إنشاء HTML للطباعة
            html_content = self.create_html_content()

            # إغلاق نافذة المعاينة السابقة إذا كانت موجودة
            if self.preview_dialog is not None:
                try:
                    self.preview_dialog.close()
                    self.preview_dialog = None
                except:
                    pass

            # إنشاء نافذة معاينة للطباعة مع الاحتفاظ بمرجع
            self.preview_dialog = PrintPreviewDialog(html_content, self)
            self.preview_dialog.setModal(False)  # جعل النافذة غير modal
            self.preview_dialog.setAttribute(Qt.WA_DeleteOnClose, True)  # حذف النافذة عند الإغلاق

            # ربط إشارة الإغلاق لتنظيف المرجع
            self.preview_dialog.finished.connect(self.on_preview_closed)
            self.preview_dialog.destroyed.connect(self.on_preview_closed)  # إضافة إشارة إضافية

            self.preview_dialog.show()  # استخدام show بدلاً من exec

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")

    # تنظيف مرجع نافذة المعاينة عند إغلاقها
    def on_preview_closed(self):
        try:
            if self.preview_dialog is not None:
                # قطع الاتصال بالإشارات لتجنب استدعاءات متكررة
                try:
                    self.preview_dialog.finished.disconnect()
                    self.preview_dialog.destroyed.disconnect()
                except:
                    pass

                # تنظيف المرجع
                self.preview_dialog = None
        except Exception as e:
            print(f"خطأ في تنظيف مرجع نافذة المعاينة: {e}")
            # تنظيف قسري للمرجع
            self.preview_dialog = None

    # فتح ملف PDF بطريقة آمنة ومتقدمة
    def open_pdf_file(self, file_path):
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"الملف غير موجود: {file_path}")

            # التحقق من حجم الملف
            if os.path.getsize(file_path) == 0:
                raise ValueError("الملف فارغ أو تالف")

            # محاولة استخدام قارئ PDF مخصص أولاً
            try:
                # البحث عن PDF.exe في مجلد البرامج
                if hasattr(self, 'parent_window') and hasattr(self.parent_window, 'Programs_dir'):
                    programs_dir = self.parent_window.Programs_dir
                elif 'Programs_dir' in globals():
                    programs_dir = Programs_dir
                else:
                    programs_dir = os.path.join(os.path.dirname(__file__), "Programs")

                custom_pdf_reader_path = os.path.join(programs_dir, "PDF.exe")

                if os.path.exists(custom_pdf_reader_path):
                    subprocess.run([custom_pdf_reader_path, file_path], check=True)
                    return True
            except Exception as e:
                print(f"فشل في استخدام قارئ PDF المخصص: {e}")

            # محاولة استخدام البرنامج الافتراضي
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                    return True
                elif os.name == 'posix':  # Linux/Mac
                    subprocess.run(['xdg-open', file_path], check=True)
                    return True
                else:
                    subprocess.run(['open', file_path], check=True)  # macOS
                    return True
            except Exception as e:
                print(f"فشل في فتح الملف بالبرنامج الافتراضي: {e}")

            # محاولة أخيرة باستخدام متصفح الويب
            try:
                import webbrowser
                webbrowser.open(f'file://{file_path}')
                return True
            except Exception as e:
                print(f"فشل في فتح الملف بالمتصفح: {e}")

            return False

        except Exception as e:
            print(f"خطأ في فتح ملف PDF: {e}")
            return False

    # معاينة PDF
    def preview_pdf(self):
        try:
            # إنشاء ملف مؤقت للمعاينة
            temp_file = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            temp_file.close()

            # إنشاء PDF
            self.create_pdf(temp_file.name)

            # التحقق من إنشاء الملف بنجاح
            if not os.path.exists(temp_file.name):
                raise FileNotFoundError("فشل في إنشاء ملف PDF المؤقت")

            if os.path.getsize(temp_file.name) == 0:
                raise ValueError("ملف PDF المؤقت فارغ")

            # فتح الملف للمعاينة
            if not self.open_pdf_file(temp_file.name):
                # إذا فشل فتح الملف، اعرض رسالة مع مسار الملف
                QMessageBox.information(
                    self,
                    "معلومات",
                    f"تم إنشاء ملف PDF بنجاح ولكن فشل في فتحه تلقائياً.\n"
                    f"يمكنك فتح الملف يدوياً من:\n{temp_file.name}\n\n"
                    f"تأكد من وجود برنامج قارئ PDF على النظام."
                )

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في إنشاء معاينة PDF:\n{str(e)}")

    # تصدير PDF
    def export_pdf(self):
        try:
            # اختيار مكان الحفظ
            documents_folder = os.path.expanduser("~\\Documents")
            default_name = f"{self.section_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "حفظ ملف PDF",
                os.path.join(documents_folder, default_name),
                "PDF Files (*.pdf)"
            )

            if file_path:
                self.create_pdf(file_path)
                QMessageBox.information(self, "نجح التصدير", f"تم حفظ الملف بنجاح:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير PDF:\n{str(e)}")

    # إنشاء ملف PDF احترافي
    def create_pdf(self, file_path):
        # تسجيل خط عربي
        arabic_font = self.register_arabic_font()

        # إعداد الصفحة
        page_size = A4 if self.orientation.currentText() == "عمودي" else (A4[1], A4[0])

        # إنشاء المستند
        doc = SimpleDocTemplate(
            file_path,
            pagesize=page_size,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # قائمة العناصر
        story = []

        # إضافة الهيدر
        story.extend(self.create_pdf_header(arabic_font))

        # إضافة الجدول
        story.extend(self.create_pdf_table(arabic_font))

        # إضافة الفوتر
        story.extend(self.create_pdf_footer(arabic_font))

        # بناء المستند
        doc.build(story, onFirstPage=self.add_page_info, onLaterPages=self.add_page_info)

    # إنشاء هيدر PDF
    def create_pdf_header(self, font_name):
        elements = []

        if self.include_logo.isChecked() and self.logo_path and os.path.exists(self.logo_path):
            try:
                # إضافة الشعار
                logo = Image(self.logo_path, width=2*inch, height=1*inch)
                logo.hAlign = 'CENTER'
                elements.append(logo)
                elements.append(Spacer(1, 0.2*inch))
            except:
                pass

        # اسم الشركة
        company_style = ParagraphStyle(
            'CompanyName',
            parent=getSampleStyleSheet()['Title'],
            fontName=font_name,
            fontSize=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=0.1*inch,
            rightIndent=0,
            leftIndent=0,
            wordWrap='RTL'  # اتجاه من اليمين إلى اليسار
        )

        company_para = Paragraph(self.fix_arabic_text_for_pdf(self.company_name), company_style)
        elements.append(company_para)

        # عنوان التقرير
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=getSampleStyleSheet()['Heading1'],
            fontName=font_name,
            fontSize=14,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=0.2*inch,
            rightIndent=0,
            leftIndent=0,
            wordWrap='RTL'  # اتجاه من اليمين إلى اليسار
        )

        title_para = Paragraph(self.fix_arabic_text_for_pdf(f"تقرير {self.section_name}"), title_style)
        elements.append(title_para)

        if self.include_date.isChecked():
            # التاريخ والوقت
            date_style = ParagraphStyle(
                'DateStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontName=font_name,
                fontSize=10,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=0.3*inch,
                rightIndent=0,
                leftIndent=0,
                wordWrap='RTL'  # اتجاه من اليمين إلى اليسار
            )

            current_date = datetime.now().strftime("%Y/%m/%d - %H:%M")
            date_para = Paragraph(self.fix_arabic_text_for_pdf(f"تاريخ الطباعة: {current_date}"), date_style)
            elements.append(date_para)

        elements.append(Spacer(1, 0.2*inch))
        return elements

    # إنشاء جدول PDF
    def create_pdf_table(self, font_name):
        elements = []

        # الحصول على البيانات المفلترة
        selected_indices, selected_headers = self.get_selected_columns()

        if not selected_indices:
            return elements

        # إعداد بيانات الجدول
        table_data = []

        # إضافة الهيدر مع إصلاح النص العربي وعكس الترتيب للاتجاه العربي
        fixed_headers = [self.fix_arabic_text_for_pdf(header) for header in selected_headers]
        # عكس ترتيب الأعمدة للاتجاه العربي (من اليمين إلى اليسار)
        fixed_headers.reverse()
        table_data.append(fixed_headers)

        # إضافة البيانات
        for row in self.table_data:
            filtered_row = [str(row[i]) if row[i] is not None else "" for i in selected_indices]
            # تحويل الأرقام المالية وإصلاح النصوص العربية
            for j, cell in enumerate(filtered_row):
                if self.is_currency_column(selected_headers[j]):
                    try:
                        amount = float(cell)
                        filtered_row[j] = f"{amount:,.2f} {self.currency_type}"
                    except:
                        pass
                else:
                    # إصلاح النص العربي للخلايا النصية
                    filtered_row[j] = self.fix_arabic_text_for_pdf(filtered_row[j])

            # عكس ترتيب البيانات لتتماشى مع الهيدر المعكوس
            filtered_row.reverse()
            table_data.append(filtered_row)

        # حساب عرض الأعمدة بناءً على عرض الصفحة
        page_width = A4[0] if self.orientation.currentText() == "عمودي" else A4[1]
        available_width = page_width - 2*cm  # طرح الهوامش

        # توزيع العرض بالتساوي على الأعمدة
        num_columns = len(selected_headers)
        col_width = available_width / num_columns
        col_widths = [col_width] * num_columns

        # إنشاء الجدول مع عرض الأعمدة المحدد
        table = Table(table_data, colWidths=col_widths)

        # تنسيق الجدول مع دعم الاتجاه العربي
        table_style = TableStyle([
            # تنسيق الهيدر
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), self.font_size.value() + 1),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), self.font_size.value()),

            # حدود الجدول
            ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2c3e50')),

            # تنسيق الصفوف
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MidDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ])

        table.setStyle(table_style)

        # محاذاة الجدول إلى اليمين (للاتجاه العربي)
        table.hAlign = 'RIGHT'

        elements.append(table)

        # إضافة إحصائيات حسب نوع القسم
        if self.section_name == "المشاريع":
            elements.extend(self.create_project_statistics(selected_headers, font_name))
        elif self.section_name == "العملاء":
            elements.extend(self.create_client_statistics(selected_headers, font_name))
        elif self.section_name == "الحسابات":
            elements.extend(self.create_expense_statistics(selected_headers, font_name))
        elif self.section_name == "الموظفين":
            elements.extend(self.create_employee_statistics(selected_headers, font_name))
        elif self.section_name == "التدريب":
            elements.extend(self.create_training_statistics(selected_headers, font_name))

        return elements

    # إنشاء فوتر PDF
    def create_pdf_footer(self, font_name):
        elements = []

        if self.include_company_info.isChecked():
            elements.append(Spacer(1, 0.3*inch))

            # معلومات الشركة
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontName=font_name,
                fontSize=9,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=0.1*inch
            )

            company_info = f"""
            <b>{self.fix_arabic_text_for_pdf(self.company_name)}</b><br/>
            {self.fix_arabic_text_for_pdf(f"العنوان: {self.company_address}")}<br/>
            {self.fix_arabic_text_for_pdf(f"الهاتف: {self.company_phone}")} | {self.fix_arabic_text_for_pdf(f"البريد الإلكتروني: {self.company_email}")}
            """

            if self.include_user.isChecked():
                company_info += f"<br/>{self.fix_arabic_text_for_pdf(f'طُبع بواسطة: {self.account_type}')}"

            footer_para = Paragraph(company_info, footer_style)
            elements.append(footer_para)

        return elements

    # إضافة معلومات الصفحة
    def add_page_info(self, canvas, doc):
        if self.include_page_numbers.isChecked():
            canvas.saveState()

            # محاولة استخدام خط عربي إذا كان متاحاً
            try:
                arabic_font = self.register_arabic_font()
                if arabic_font != 'Helvetica':
                    canvas.setFont(arabic_font, 9)
                else:
                    canvas.setFont('Helvetica', 9)
            except:
                canvas.setFont('Helvetica', 9)

            # رقم الصفحة مع إصلاح النص العربي
            page_num = canvas.getPageNumber()
            text = self.fix_arabic_text_for_pdf(f"صفحة {page_num}")
            canvas.drawRightString(doc.pagesize[0] - 1*cm, 1*cm, text)

            # التاريخ في الزاوية اليسرى
            if self.include_date.isChecked():
                date_text = datetime.now().strftime("%Y/%m/%d")
                canvas.drawString(1*cm, 1*cm, date_text)

            canvas.restoreState()

    # تحديد ما إذا كان العمود يحتوي على قيم مالية
    def is_currency_column(self, column_name):
        currency_keywords = ["مبلغ", "المبلغ", "مدفوع", "المدفوع", "باقي", "الباقي", "سعر", "السعر", "تكلفة", "التكلفة", "راتب", "المرتب"]
        return any(keyword in column_name for keyword in currency_keywords)

    # تحديد ما إذا كان العمود يحتوي على id
    def is_معرف_column(self, column_name):
        معرف_keywords = ["id", "id", "id", "id", "id", "الid", "رقم_", "الرقم_"]
        return any(keyword in column_name for keyword in معرف_keywords)

    # إنشاء إحصائيات المشاريع
    def create_project_statistics(self, headers, font_name):
        elements = []

        try:
            # البحث عن أعمدة المبالغ
            amount_col = None
            paid_col = None
            remaining_col = None

            for i, header in enumerate(headers):
                if "المبلغ" in header and "مدفوع" not in header and "باقي" not in header:
                    amount_col = i
                elif "مدفوع" in header:
                    paid_col = i
                elif "باقي" in header:
                    remaining_col = i

            if amount_col is not None:
                elements.append(Spacer(1, 0.2*inch))

                # حساب الإحصائيات
                total_amount = 0
                total_paid = 0
                total_remaining = 0
                project_count = len(self.table_data)

                selected_indices, _ = self.get_selected_columns()

                for row in self.table_data:
                    try:
                        if amount_col < len(selected_indices):
                            amount = float(row[selected_indices[amount_col]] or 0)
                            total_amount += amount

                        if paid_col is not None and paid_col < len(selected_indices):
                            paid = float(row[selected_indices[paid_col]] or 0)
                            total_paid += paid

                        if remaining_col is not None and remaining_col < len(selected_indices):
                            remaining = float(row[selected_indices[remaining_col]] or 0)
                            total_remaining += remaining
                    except:
                        continue

                # إنشاء جدول الإحصائيات مع إصلاح النص العربي وعكس الترتيب
                stats_data = [
                    [self.fix_arabic_text_for_pdf("القيمة"), self.fix_arabic_text_for_pdf("الإحصائيات")],  # عكس الترتيب
                    [str(project_count), self.fix_arabic_text_for_pdf("عدد المشاريع")],
                    [f"{total_amount:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي المبالغ")],
                    [f"{total_paid:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي المدفوع")],
                    [f"{total_remaining:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي المتبقي")]
                ]

                stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])  # عكس عرض الأعمدة
                stats_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fff5f5'), colors.white]),
                ])

                stats_table.setStyle(stats_style)
                # محاذاة جدول الإحصائيات إلى اليمين
                stats_table.hAlign = 'RIGHT'
                elements.append(stats_table)

        except Exception as e:
            print(f"خطأ في إنشاء الإحصائيات: {e}")

        return elements

    # إنشاء إحصائيات العملاء
    def create_client_statistics(self, headers, font_name):
        elements = []

        try:
            elements.append(Spacer(1, 0.2*inch))

            # حساب الإحصائيات
            total_clients = len(self.table_data)

            # تصنيف العملاء حسب النوع
            client_types = {}
            classification_col = None

            for i, header in enumerate(headers):
                if "التصنيف" in header or "نوع" in header:
                    classification_col = i
                    break

            if classification_col is not None:
                selected_indices, _ = self.get_selected_columns()
                for row in self.table_data:
                    try:
                        if classification_col < len(selected_indices):
                            client_type = str(row[selected_indices[classification_col]] or "غير محدد")
                            client_types[client_type] = client_types.get(client_type, 0) + 1
                    except:
                        continue

            # إنشاء جدول الإحصائيات مع عكس الترتيب وإصلاح النص العربي
            stats_data = [
                [self.fix_arabic_text_for_pdf("القيمة"), self.fix_arabic_text_for_pdf("الإحصائيات")],  # عكس الترتيب
                [str(total_clients), self.fix_arabic_text_for_pdf("إجمالي العملاء")]
            ]

            # إضافة تصنيف العملاء
            for client_type, count in client_types.items():
                stats_data.append([str(count), self.fix_arabic_text_for_pdf(f"عملاء {client_type}")])

            stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])  # عكس عرض الأعمدة
            stats_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f0fff4'), colors.white]),
            ])

            stats_table.setStyle(stats_style)
            stats_table.hAlign = 'RIGHT'  # محاذاة إلى اليمين
            elements.append(stats_table)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات العملاء: {e}")

        return elements

    # إنشاء إحصائيات الحسابات
    def create_expense_statistics(self, headers, font_name):
        elements = []

        try:
            elements.append(Spacer(1, 0.2*inch))

            # البحث عن عمود المبلغ والتصنيف
            amount_col = None
            classification_col = None

            selected_indices, _ = self.get_selected_columns()

            for i, header in enumerate(headers):
                if "المبلغ" in header:
                    amount_col = i
                elif "التصنيف" in header or "نوع" in header:
                    classification_col = i

            if amount_col is not None:
                # حساب الإحصائيات
                total_amount = 0
                expense_count = len(self.table_data)
                expense_types = {}

                for row in self.table_data:
                    try:
                        if amount_col < len(selected_indices):
                            amount = float(row[selected_indices[amount_col]] or 0)
                            total_amount += amount

                        # تصنيف المصروفات
                        if classification_col is not None and classification_col < len(selected_indices):
                            expense_type = str(row[selected_indices[classification_col]] or "غير محدد")
                            if expense_type not in expense_types:
                                expense_types[expense_type] = {'count': 0, 'amount': 0}
                            expense_types[expense_type]['count'] += 1
                            expense_types[expense_type]['amount'] += amount
                    except:
                        continue

                # إنشاء جدول الإحصائيات مع عكس الترتيب وإصلاح النص العربي
                stats_data = [
                    [self.fix_arabic_text_for_pdf("القيمة"), self.fix_arabic_text_for_pdf("الإحصائيات")],  # عكس الترتيب
                    [str(expense_count), self.fix_arabic_text_for_pdf("عدد المصروفات")],
                    [f"{total_amount:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي المصروفات")],
                    [f"{total_amount/expense_count if expense_count > 0 else 0:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("متوسط المصروف")]
                ]

                # إضافة تصنيف المصروفات
                for expense_type, data in expense_types.items():
                    stats_data.append([f"{data['amount']:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf(f"مصروفات {expense_type}")])

                stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])  # عكس عرض الأعمدة
                stats_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e67e22')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fef9e7'), colors.white]),
                ])

                stats_table.setStyle(stats_style)
                stats_table.hAlign = 'RIGHT'  # محاذاة إلى اليمين
                elements.append(stats_table)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات الحسابات: {e}")

        return elements

    # إنشاء إحصائيات الموظفين
    def create_employee_statistics(self, headers, font_name):
        elements = []

        try:
            elements.append(Spacer(1, 0.2*inch))

            # البحث عن الأعمدة المهمة
            salary_col = None
            balance_col = None
            withdrawal_col = None
            classification_col = None

            selected_indices, _ = self.get_selected_columns()

            for i, header in enumerate(headers):
                if "المرتب" in header or "راتب" in header:
                    salary_col = i
                elif "الرصيد" in header or "رصيد" in header:
                    balance_col = i

                elif "التصنيف" in header or "الوظيفة" in header:
                    classification_col = i

            # حساب الإحصائيات
            total_employees = len(self.table_data)
            total_salaries = 0
            total_balance = 0

            employee_types = {}

            for row in self.table_data:
                try:
                    # حساب المرتبات
                    if salary_col is not None and salary_col < len(selected_indices):
                        salary = float(row[selected_indices[salary_col]] or 0)
                        total_salaries += salary

                    # حساب الأرصدة
                    if balance_col is not None and balance_col < len(selected_indices):
                        balance = float(row[selected_indices[balance_col]] or 0)
                        total_balance += balance



                    # تصنيف الموظفين
                    if classification_col is not None and classification_col < len(selected_indices):
                        emp_type = str(row[selected_indices[classification_col]] or "غير محدد")
                        employee_types[emp_type] = employee_types.get(emp_type, 0) + 1
                except:
                    continue

            # إنشاء جدول الإحصائيات مع عكس الترتيب وإصلاح النص العربي
            stats_data = [
                [self.fix_arabic_text_for_pdf("القيمة"), self.fix_arabic_text_for_pdf("الإحصائيات")],  # عكس الترتيب
                [str(total_employees), self.fix_arabic_text_for_pdf("عدد الموظفين")],
                [f"{total_salaries:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي المرتبات")],
                [f"{total_balance:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي الأرصدة")],

                [f"{total_salaries/total_employees if total_employees > 0 else 0:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("متوسط المرتب")]
            ]

            # إضافة تصنيف الموظفين
            for emp_type, count in employee_types.items():
                stats_data.append([str(count), self.fix_arabic_text_for_pdf(f"{emp_type}")])

            stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])  # عكس عرض الأعمدة
            stats_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8e44ad')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f4f1f8'), colors.white]),
            ])

            stats_table.setStyle(stats_style)
            stats_table.hAlign = 'RIGHT'  # محاذاة إلى اليمين
            elements.append(stats_table)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات الموظفين: {e}")

        return elements

    # إنشاء إحصائيات التدريب
    def create_training_statistics(self, headers, font_name):
        elements = []

        try:
            elements.append(Spacer(1, 0.2*inch))

            # البحث عن الأعمدة المهمة
            cost_col = None
            total_amount_col = None
            participants_col = None
            groups_col = None
            status_col = None
            classification_col = None

            selected_indices, _ = self.get_selected_columns()

            for i, header in enumerate(headers):
                if "التكلفة" in header:
                    cost_col = i
                elif "إجمالي" in header and "المبلغ" in header:
                    total_amount_col = i
                elif "المشاركين" in header or "عدد_المشاركين" in header:
                    participants_col = i
                elif "المجموعات" in header or "عدد_المجموعات" in header:
                    groups_col = i
                elif "الحالة" in header:
                    status_col = i
                elif "التصنيف" in header or "نوع" in header:
                    classification_col = i

            # حساب الإحصائيات
            total_courses = len(self.table_data)
            total_revenue = 0
            total_participants = 0
            total_groups = 0
            course_status = {}
            course_types = {}

            for row in self.table_data:
                try:
                    # حساب الإيرادات
                    if total_amount_col is not None and total_amount_col < len(selected_indices):
                        amount = float(row[selected_indices[total_amount_col]] or 0)
                        total_revenue += amount
                    elif cost_col is not None and cost_col < len(selected_indices):
                        cost = float(row[selected_indices[cost_col]] or 0)
                        # إذا لم يكن هناك عمود إجمالي، استخدم التكلفة
                        if participants_col is not None and participants_col < len(selected_indices):
                            participants = int(row[selected_indices[participants_col]] or 0)
                            total_revenue += cost * participants
                        else:
                            total_revenue += cost

                    # حساب المشاركين
                    if participants_col is not None and participants_col < len(selected_indices):
                        participants = int(row[selected_indices[participants_col]] or 0)
                        total_participants += participants

                    # حساب المجموعات
                    if groups_col is not None and groups_col < len(selected_indices):
                        groups = int(row[selected_indices[groups_col]] or 0)
                        total_groups += groups

                    # تصنيف حسب الحالة
                    if status_col is not None and status_col < len(selected_indices):
                        status = str(row[selected_indices[status_col]] or "غير محدد")
                        course_status[status] = course_status.get(status, 0) + 1

                    # تصنيف حسب النوع
                    if classification_col is not None and classification_col < len(selected_indices):
                        course_type = str(row[selected_indices[classification_col]] or "غير محدد")
                        course_types[course_type] = course_types.get(course_type, 0) + 1
                except:
                    continue

            # إنشاء جدول الإحصائيات مع عكس الترتيب وإصلاح النص العربي
            stats_data = [
                [self.fix_arabic_text_for_pdf("القيمة"), self.fix_arabic_text_for_pdf("الإحصائيات")],  # عكس الترتيب
                [str(total_courses), self.fix_arabic_text_for_pdf("عدد الدورات")],
                [f"{total_revenue:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("إجمالي الإيرادات")],
                [str(total_participants), self.fix_arabic_text_for_pdf("إجمالي المشاركين")],
                [str(total_groups), self.fix_arabic_text_for_pdf("إجمالي المجموعات")],
                [f"{total_participants/total_courses if total_courses > 0 else 0:.1f}", self.fix_arabic_text_for_pdf("متوسط المشاركين/دورة")],
                [f"{total_revenue/total_courses if total_courses > 0 else 0:,.2f} {self.currency_type}", self.fix_arabic_text_for_pdf("متوسط الإيراد/دورة")]
            ]

            # إضافة تصنيف حسب الحالة
            for status, count in course_status.items():
                stats_data.append([str(count), self.fix_arabic_text_for_pdf(f"دورات {status}")])

            # إضافة تصنيف حسب النوع
            for course_type, count in course_types.items():
                stats_data.append([str(count), self.fix_arabic_text_for_pdf(f"{course_type}")])

            stats_table = Table(stats_data, colWidths=[2*inch, 3*inch])  # عكس عرض الأعمدة
            stats_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f39c12')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRid', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fef5e7'), colors.white]),
            ])

            stats_table.setStyle(stats_style)
            stats_table.hAlign = 'RIGHT'  # محاذاة إلى اليمين
            elements.append(stats_table)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات التدريب: {e}")

        return elements

    # فتح ملف Excel بطريقة آمنة
    def open_excel_file(self, file_path):
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"الملف غير موجود: {file_path}")

            # التحقق من حجم الملف
            if os.path.getsize(file_path) == 0:
                raise ValueError("الملف فارغ أو تالف")

            # محاولة فتح الملف بالبرنامج الافتراضي
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                    return True
                elif os.name == 'posix':  # Linux/Mac
                    subprocess.run(['xdg-open', file_path], check=True)
                    return True
                else:
                    subprocess.run(['open', file_path], check=True)  # macOS
                    return True
            except Exception as e:
                print(f"فشل في فتح ملف Excel: {e}")
                return False

        except Exception as e:
            print(f"خطأ في فتح ملف Excel: {e}")
            return False

    # معاينة Excel
    def preview_excel(self):
        try:
            # إنشاء ملف مؤقت للمعاينة
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.close()

            # إنشاء Excel
            self.create_excel(temp_file.name)

            # التحقق من إنشاء الملف بنجاح
            if not os.path.exists(temp_file.name):
                raise FileNotFoundError("فشل في إنشاء ملف Excel المؤقت")

            if os.path.getsize(temp_file.name) == 0:
                raise ValueError("ملف Excel المؤقت فارغ")

            # فتح الملف للمعاينة
            if not self.open_excel_file(temp_file.name):
                # إذا فشل فتح الملف، اعرض رسالة مع مسار الملف
                QMessageBox.information(
                    self,
                    "معلومات",
                    f"تم إنشاء ملف Excel بنجاح ولكن فشل في فتحه تلقائياً.\n"
                    f"يمكنك فتح الملف يدوياً من:\n{temp_file.name}\n\n"
                    f"تأكد من وجود برنامج Excel أو LibreOffice على النظام."
                )

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في إنشاء معاينة Excel:\n{str(e)}")

    # تصدير Excel
    def export_excel(self):
        try:
            # اختيار مكان الحفظ
            documents_folder = os.path.expanduser("~\\Documents")
            default_name = f"{self.section_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "حفظ ملف Excel",
                os.path.join(documents_folder, default_name),
                "Excel Files (*.xlsx)"
            )

            if file_path:
                self.create_excel(file_path)
                QMessageBox.information(self, "نجح التصدير", f"تم حفظ الملف بنجاح:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير Excel:\n{str(e)}")

    # إنشاء ملف Excel احترافي
    def create_excel(self, file_path):
        try:
            # الحصول على البيانات المفلترة
            selected_indices, selected_headers = self.get_selected_columns()

            if not selected_indices:
                QMessageBox.warning(self, "تحذير", "لم يتم تحديد أي أعمدة للتصدير")
                return

            # إعداد البيانات
            filtered_data = []
            for row in self.table_data:
                filtered_row = [row[i] if row[i] is not None else "" for i in selected_indices]
                # تحويل الأرقام المالية
                for j, cell in enumerate(filtered_row):
                    if self.is_currency_column(selected_headers[j]):
                        try:
                            filtered_row[j] = float(cell)
                        except:
                            pass
                filtered_data.append(filtered_row)

            # إنشاء DataFrame
            df = pd.DataFrame(filtered_data, columns=selected_headers)

            # إنشاء ملف Excel مع تنسيق متقدم
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # كتابة البيانات الرئيسية
                df.to_excel(writer, sheet_name='البيانات', index=False, startrow=8)

                # الحصول على الورقة والكتاب
                workbook = writer.book
                worksheet = writer.sheets['البيانات']

                # تطبيق التنسيق
                self.format_excel_sheet(workbook, worksheet, df, selected_headers)

                # إضافة ورقة الإحصائيات حسب نوع القسم
                if self.section_name == "المشاريع":
                    self.create_excel_statistics(writer, df, selected_headers)
                elif self.section_name in ["العملاء", "الحسابات", "الموظفين", "التدريب"]:
                    self.create_excel_general_statistics(writer, df, selected_headers)

        except Exception as e:
            raise Exception(f"خطأ في إنشاء ملف Excel: {str(e)}")

    # تنسيق ورقة Excel
    def format_excel_sheet(self, workbook, worksheet, df, headers):
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing import image

        # ألوان الشركة
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        company_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        # خطوط
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        company_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)

        # حدود
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # إضافة معلومات الشركة في الأعلى
        if self.include_company_info.isChecked():
            # اسم الشركة
            worksheet['A1'] = self.company_name
            worksheet['A1'].font = company_font
            worksheet['A1'].fill = company_fill
            worksheet['A1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:' + get_column_letter(len(headers)) + '1')

            # عنوان التقرير
            worksheet['A2'] = f"تقرير {self.section_name}"
            worksheet['A2'].font = Font(name="Arial", size=14, bold=True)
            worksheet['A2'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A2:' + get_column_letter(len(headers)) + '2')

            # التاريخ والمستخدم
            if self.include_date.isChecked():
                worksheet['A3'] = f"تاريخ الطباعة: {datetime.now().strftime('%Y/%m/%d - %H:%M')}"
                worksheet['A3'].font = Font(name="Arial", size=10, italic=True)

            if self.include_user.isChecked():
                worksheet['A4'] = f"طُبع بواسطة: {self.account_type}"
                worksheet['A4'].font = Font(name="Arial", size=10, italic=True)

            # معلومات الشركة
            worksheet['A5'] = f"العنوان: {self.company_address}"
            worksheet['A5'].font = Font(name="Arial", size=9)

            worksheet['A6'] = f"الهاتف: {self.company_phone} | البريد الإلكتروني: {self.company_email}"
            worksheet['A6'].font = Font(name="Arial", size=9)

        # تنسيق الهيدر
        header_row = 9
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # تنسيق البيانات
        for row_num in range(header_row + 1, header_row + 1 + len(df)):
            for col_num in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # تنسيق الأعمدة المالية
                if self.is_currency_column(headers[col_num - 1]):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = f'#,##0.00 "{self.currency_type}"'

        # ضبط عرض الأعمدة
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(header)), 15)

            # حساب أقصى طول في العمود
            for row_num in range(header_row + 1, header_row + 1 + len(df)):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # إضافة الشعار إذا كان متوفراً
        if self.include_logo.isChecked() and self.logo_path and os.path.exists(self.logo_path):
            try:
                img = image.Image(self.logo_path)
                img.height = 60
                img.width = 120
                worksheet.add_image(img, f'{get_column_letter(len(headers) - 1)}1')
            except:
                pass

    # إنشاء ورقة إحصائيات Excel
    def create_excel_statistics(self, writer, df, headers):
        try:
            # البحث عن أعمدة المبالغ
            amount_col = None
            paid_col = None
            remaining_col = None

            for header in headers:
                if "المبلغ" in header and "مدفوع" not in header and "باقي" not in header:
                    amount_col = header
                elif "مدفوع" in header:
                    paid_col = header
                elif "باقي" in header:
                    remaining_col = header

            if amount_col:
                # إنشاء بيانات الإحصائيات
                stats_data = {
                    'الإحصائية': ['عدد المشاريع', 'إجمالي المبالغ', 'إجمالي المدفوع', 'إجمالي المتبقي', 'متوسط قيمة المشروع'],
                    'القيمة': [
                        len(df),
                        df[amount_col].sum() if amount_col in df.columns else 0,
                        df[paid_col].sum() if paid_col and paid_col in df.columns else 0,
                        df[remaining_col].sum() if remaining_col and remaining_col in df.columns else 0,
                        df[amount_col].mean() if amount_col in df.columns else 0
                    ]
                }

                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='الإحصائيات', index=False)

                # تنسيق ورقة الإحصائيات
                stats_worksheet = writer.sheets['الإحصائيات']
                self.format_stats_sheet(writer.book, stats_worksheet)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات Excel: {e}")

    # إنشاء ورقة إحصائيات عامة Excel
    def create_excel_general_statistics(self, writer, df, headers):
        try:
            # إنشاء بيانات الإحصائيات حسب نوع القسم
            if self.section_name == "العملاء":
                stats_data = self.get_client_excel_stats(df, headers)
            elif self.section_name == "الحسابات":
                stats_data = self.get_expense_excel_stats(df, headers)
            elif self.section_name == "الموظفين":
                stats_data = self.get_employee_excel_stats(df, headers)
            
            elif self.section_name == "التدريب":
                stats_data = self.get_training_excel_stats(df, headers)
            else:
                return

            if stats_data:
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='الإحصائيات', index=False)

                # تنسيق ورقة الإحصائيات
                stats_worksheet = writer.sheets['الإحصائيات']
                self.format_stats_sheet(writer.book, stats_worksheet)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات عامة Excel: {e}")

    # الحصول على إحصائيات العملاء لـ Excel
    def get_client_excel_stats(self, df, headers):
        try:
            total_clients = len(df)
            stats_data = {
                'الإحصائية': ['إجمالي العملاء'],
                'القيمة': [total_clients]
            }

            # تصنيف العملاء حسب النوع
            if 'التصنيف' in df.columns:
                client_types = df['التصنيف'].value_counts()
                for client_type, count in client_types.items():
                    stats_data['الإحصائية'].append(f'عملاء {client_type}')
                    stats_data['القيمة'].append(count)

            return stats_data
        except:
            return None

    # الحصول على إحصائيات الحسابات لـ Excel
    def get_expense_excel_stats(self, df, headers):
        try:
            total_expenses = len(df)
            total_amount = 0

            # البحث عن عمود المبلغ
            amount_col = None
            for col in df.columns:
                if 'المبلغ' in col:
                    amount_col = col
                    break

            if amount_col:
                total_amount = df[amount_col].sum()
                avg_amount = df[amount_col].mean()
            else:
                avg_amount = 0

            stats_data = {
                'الإحصائية': [
                    'عدد المصروفات',
                    'إجمالي المصروفات',
                    'متوسط المصروف'
                ],
                'القيمة': [
                    total_expenses,
                    f"{total_amount:,.2f} {self.currency_type}",
                    f"{avg_amount:,.2f} {self.currency_type}"
                ]
            }

            # تصنيف المصروفات حسب النوع
            if 'التصنيف' in df.columns:
                expense_types = df['التصنيف'].value_counts()
                for expense_type, count in expense_types.items():
                    stats_data['الإحصائية'].append(f'مصروفات {expense_type}')
                    stats_data['القيمة'].append(count)

            return stats_data
        except:
            return None

    # الحصول على إحصائيات الموظفين لـ Excel
    def get_employee_excel_stats(self, df, headers):
        try:
            total_employees = len(df)
            total_salaries = 0
            total_balance = 0

            # البحث عن الأعمدة المهمة
            salary_col = None
            balance_col = None

            for col in df.columns:
                if 'المرتب' in col or 'راتب' in col:
                    salary_col = col
                elif 'الرصيد' in col or 'رصيد' in col:
                    balance_col = col

            if salary_col:
                total_salaries = df[salary_col].sum()
                avg_salary = df[salary_col].mean()
            else:
                avg_salary = 0

            if balance_col:
                total_balance = df[balance_col].sum()

            stats_data = {
                'الإحصائية': [
                    'عدد الموظفين',
                    'إجمالي المرتبات',
                    'إجمالي الأرصدة',
                    'متوسط المرتب'
                ],
                'القيمة': [
                    total_employees,
                    f"{total_salaries:,.2f} {self.currency_type}",
                    f"{total_balance:,.2f} {self.currency_type}",
                    f"{avg_salary:,.2f} {self.currency_type}"
                ]
            }

            # تصنيف الموظفين حسب النوع
            if 'التصنيف' in df.columns:
                employee_types = df['التصنيف'].value_counts()
                for emp_type, count in employee_types.items():
                    stats_data['الإحصائية'].append(f'{emp_type}')
                    stats_data['القيمة'].append(count)
            elif 'الوظيفة' in df.columns:
                job_types = df['الوظيفة'].value_counts()
                for job_type, count in job_types.items():
                    stats_data['الإحصائية'].append(f'{job_type}')
                    stats_data['القيمة'].append(count)

            return stats_data
        except:
            return None

    

    # الحصول على إحصائيات التدريب لـ Excel
    def get_training_excel_stats(self, df, headers):
        try:
            total_courses = len(df)
            total_revenue = 0
            total_participants = 0

            # البحث عن الأعمدة المهمة
            amount_col = None
            participants_col = None

            for col in df.columns:
                if 'إجمالي' in col and 'المبلغ' in col:
                    amount_col = col
                elif 'المشاركين' in col:
                    participants_col = col

            if amount_col:
                total_revenue = df[amount_col].sum()
                avg_revenue = df[amount_col].mean()
            else:
                avg_revenue = 0

            if participants_col:
                total_participants = df[participants_col].sum()
                avg_participants = df[participants_col].mean()
            else:
                avg_participants = 0

            stats_data = {
                'الإحصائية': [
                    'عدد الدورات',
                    'إجمالي الإيرادات',
                    'إجمالي المشاركين',
                    'متوسط الإيراد/دورة',
                    'متوسط المشاركين/دورة'
                ],
                'القيمة': [
                    total_courses,
                    f"{total_revenue:,.2f} {self.currency_type}",
                    int(total_participants),
                    f"{avg_revenue:,.2f} {self.currency_type}",
                    f"{avg_participants:.1f}"
                ]
            }

            # تصنيف الدورات حسب الحالة
            if 'الحالة' in df.columns:
                course_status = df['الحالة'].value_counts()
                for status, count in course_status.items():
                    stats_data['الإحصائية'].append(f'دورات {status}')
                    stats_data['القيمة'].append(count)

            # تصنيف الدورات حسب النوع
            if 'التصنيف' in df.columns:
                course_types = df['التصنيف'].value_counts()
                for course_type, count in course_types.items():
                    stats_data['الإحصائية'].append(f'{course_type}')
                    stats_data['القيمة'].append(count)

            return stats_data
        except:
            return None

    # تنسيق ورقة الإحصائيات
    def format_stats_sheet(self, workbook, worksheet):
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # ألوان وخطوط
        header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=11)

        # تنسيق الهيدر
        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # تنسيق البيانات
        for row in range(2, 7):
            for col in range(1, 3):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # تنسيق الأرقام المالية
                if col == 2 and row > 2:  # عمود القيم (عدا عدد المشاريع)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = f'#,##0.00 "{self.currency_type}"'

        # ضبط عرض الأعمدة
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20

    # معاينة الطباعة
    def preview_print(self):
        try:
            # إنشاء مستند HTML للطباعة
            html_content = self.create_html_document()

            # إنشاء نافذة معاينة غير modal
            preview_dialog = PrintPreviewDialog(html_content, self)
            preview_dialog.setModal(False)  # جعل النافذة غير modal
            preview_dialog.show()  # استخدام show بدلاً من exec

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في معاينة الطباعة:\n{str(e)}")

    # طباعة المستند
    def print_document(self):
        try:
            # إنشاء مستند HTML للطباعة
            html_content = self.create_html_document()

            # إنشاء نافذة الطباعة غير modal
            print_dialog = PrintDialog(html_content, self)
            print_dialog.setModal(False)  # جعل النافذة غير modal
            print_dialog.show()  # استخدام show بدلاً من exec

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")

    # إنشاء مستند HTML محسن للطباعة مع دعم كامل للعربية
    def create_enhanced_html_document(self):
        # الحصول على البيانات المفلترة
        selected_indices, selected_headers = self.get_selected_columns()

        if not selected_indices:
            return self.create_simple_error_html("لم يتم تحديد أي أعمدة للطباعة")

        # إصلاح النصوص العربية مسبقاً
        fixed_headers = [self.fix_arabic_text(header) for header in selected_headers]

        # بناء HTML محسن
        html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.fix_arabic_text(f'تقرير {self.section_name}')}</title>
    <style>
        @page {{
            size: {'A4 landscape' if self.orientation.currentText() == 'أفقي' else 'A4 portrait'};
            margin: 1.5cm;
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}

        html {{
            direction: rtl !important;
            text-align: right !important;
            font-family: 'Tahoma', 'Arial Unicode MS', 'Segoe UI', 'DejaVu Sans', sans-serif;
        }}

        body {{
            font-family: 'Tahoma', 'Arial Unicode MS', 'Segoe UI', 'DejaVu Sans', sans-serif;
            font-size: {self.font_size.value()}px;
            direction: rtl !important;
            text-align: right !important;
            unicode-bidi: bidi-override !important;
            line-height: 1.6;
            color: #333;
            background: white;
            margin: 0;
            padding: 20px;
        }}

        .document-container {{
            max-width: 100%;
            margin: 0 auto;
            direction: rtl !important;
            text-align: right !important;
        }}

        .header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 3px solid #3498db;
            direction: rtl !important;
        }}

        .company-name {{
            font-size: {self.font_size.value() + 6}px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 10px;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
            text-align: center;
        }}

        .report-title {{
            font-size: {self.font_size.value() + 3}px;
            color: #34495e;
            margin-bottom: 15px;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
            text-align: center;
        }}

        .date-info {{
            font-size: {self.font_size.value() - 1}px;
            color: #7f8c8d;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
            text-align: center;
        }}

        .table-container {{
            width: 100%;
            overflow-x: auto;
            margin: 20px 0;
            direction: rtl !important;
        }}

        table {{
            width: 100% !important;
            border-collapse: collapse;
            direction: rtl !important;
            table-layout: fixed !important;
            font-size: {self.font_size.value()}px;
            background: white;
            
        }}

        th {{
            background: linear-gradient(135deg, #3498db, #2980b9);
            color: white !important;
            padding: 12px 8px;
            text-align: center !important;
            border: 1px solid #2980b9;
            font-weight: bold;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
            font-size: {self.font_size.value() + 1}px;
            white-space: nowrap;
        }}

        td {{
            padding: 10px 8px;
            text-align: center !important;
            border: 1px solid #bdc3c7;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
            vertical-align: middle;
            background: white;
            word-wrap: break-word;
        }}

        tr:nth-child(even) td {{
            background-color: #f8f9fa;
        }}

        tr:hover td {{
            background-color: #e3f2fd;
        }}

        .statistics {{
            margin-top: 30px;
            background: #fff5f5;
            border: 2px solid #e74c3c;
            border-radius: 8px;
            padding: 20px;
            direction: rtl !important;
        }}

        .statistics h3 {{
            color: #e74c3c;
            margin-bottom: 15px;
            text-align: center;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
        }}

        .stats-table {{
            width: 100%;
            margin: 15px 0;
            direction: rtl !important;
        }}

        .stats-table th {{
            background: #e74c3c;
            color: white;
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
        }}

        .stats-table td {{
            direction: rtl !important;
            unicode-bidi: bidi-override !important;
        }}

        .footer {{
            margin-top: 40px;
            text-align: center;
            font-size: {max(8, self.font_size.value() - 3)}px;
            color: #7f8c8d;
            border-top: 2px solid #bdc3c7;
            padding-top: 15px;
            direction: rtl !important;
        }}

        @media print {{
            body {{
                margin: 0 !important;
                padding: 10px !important;
                font-size: {max(8, self.font_size.value() - 1)}px !important;
                direction: rtl !important;
            }}

            .header {{
                margin-bottom: 20px !important;
                padding-bottom: 15px !important;
            }}

            table {{
                page-break-inside: avoid;
                font-size: {max(7, self.font_size.value() - 2)}px !important;
                direction: rtl !important;
            }}

            th, td {{
                padding: 6px 4px !important;
                font-size: {max(7, self.font_size.value() - 2)}px !important;
                direction: rtl !important;
                unicode-bidi: bidi-override !important;
            }}

            .statistics {{
                page-break-before: avoid;
                margin-top: 20px !important;
                direction: rtl !important;
            }}
        }}
    </style>
    <script>
        // تحسين الطباعة المباشرة ودعم اللغة العربية
        window.onload = function() {{
            console.log('تحميل الصفحة - بدء إعداد الاتجاه العربي');

            // ضبط الاتجاه العام للصفحة
            document.documentElement.dir = 'rtl';
            document.documentElement.lang = 'ar';
            document.body.dir = 'rtl';
            document.body.style.direction = 'rtl';
            document.body.style.unicodeBidi = 'bidi-override';
            document.body.style.textAlign = 'right';

            // تطبيق الاتجاه على جميع العناصر النصية
            var textElements = document.querySelectorAll('div, span, p, h1, h2, h3, h4, h5, h6, th, td');
            textElements.forEach(function(element) {{
                element.style.direction = 'rtl';
                element.style.unicodeBidi = 'bidi-override';
                element.style.textAlign = 'center';

                // إضافة علامة الاتجاه للنصوص العربية
                var text = element.textContent || element.innerText;
                if (text && /[\u0600-\u06FF]/.test(text)) {{
                    element.style.unicodeBidi = 'bidi-override';
                    element.dir = 'rtl';
                }}
            }});

            // تطبيق خاص على الجداول
            var tables = document.querySelectorAll('table');
            tables.forEach(function(table) {{
                table.style.direction = 'rtl';
                table.style.tableLayout = 'fixed';
                table.style.width = '100%';
                table.dir = 'rtl';

                var cells = table.querySelectorAll('th, td');
                cells.forEach(function(cell) {{
                    cell.style.direction = 'rtl';
                    cell.style.unicodeBidi = 'bidi-override';
                    cell.style.textAlign = 'center';
                    cell.dir = 'rtl';

                    // معالجة خاصة للنصوص العربية في الخلايا
                    var cellText = cell.textContent || cell.innerText;
                    if (cellText && /[\u0600-\u06FF]/.test(cellText)) {{
                        cell.innerHTML = '&rlm;' + cellText + '&rlm;';
                    }}
                }});
            }});

            // تطبيق على العناوين والنصوص
            var headers = document.querySelectorAll('.company-name, .report-title, .date-info');
            headers.forEach(function(header) {{
                header.style.direction = 'rtl';
                header.style.unicodeBidi = 'bidi-override';
                header.style.textAlign = 'center';
                header.dir = 'rtl';
            }});

            console.log('تم إعداد الاتجاه العربي بنجاح');
        }};

        // تحسين الطباعة
        window.onbeforeprint = function() {{
            console.log('بدء الطباعة - تطبيق إعدادات الاتجاه');

            document.body.style.direction = 'rtl';
            document.body.style.unicodeBidi = 'bidi-override';
            document.body.style.textAlign = 'right';

            var tables = document.querySelectorAll('table');
            tables.forEach(function(table) {{
                table.style.direction = 'rtl';
                table.style.width = '100%';
                table.style.tableLayout = 'fixed';
                table.dir = 'rtl';
            }});

            var cells = document.querySelectorAll('th, td');
            cells.forEach(function(cell) {{
                cell.style.direction = 'rtl';
                cell.style.unicodeBidi = 'bidi-override';
                cell.style.textAlign = 'center';
                cell.dir = 'rtl';
            }});
        }};
    </script>
</head>
<body>
    <div class="document-container">"""

        return html

    # إنشاء HTML بسيط لرسالة خطأ
    def create_simple_error_html(self, error_message):
        return f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="UTF-8">
    <title>خطأ</title>
    <style>
        body {{
            font-family: 'Tahoma', 'Arial Unicode MS', sans-serif;
            direction: rtl;
            text-align: center;
            padding: 50px;
            color: #e74c3c;
        }}
    </style>
</head>
<body>
    <h1>{self.fix_arabic_text(error_message)}</h1>
</body>
</html>"""

    # إنشاء مستند HTML للطباعة
    def create_html_document(self):
        # الحصول على البيانات المفلترة
        selected_indices, selected_headers = self.get_selected_columns()

        if not selected_indices:
            return self.create_simple_error_html("لم يتم تحديد أي أعمدة للطباعة")

        # بناء HTML محسن
        html = self.create_enhanced_html_document()

        # إضافة الهيدر
        html += self.create_html_header()

        # إضافة الجدول
        html += self.create_html_table(selected_indices, selected_headers)

        # إضافة الإحصائيات
        if self.section_name == "المشاريع":
            stats_html = self.create_html_statistics(selected_headers)
            if stats_html:
                html += stats_html
        elif self.section_name in ["العملاء", "الحسابات", "الموظفين", "التدريب"]:
            stats_html = self.create_html_general_statistics(selected_headers)
            if stats_html:
                html += stats_html

        # إضافة الفوتر
        html += self.create_html_footer()

        # إغلاق HTML
        html += """
    </div>
</body>
</html>"""

        return html

    # إنشاء هيدر HTML
    def create_html_header(self):
        html = ""
        if self.include_company_info.isChecked() or self.include_date.isChecked():
            html += '<div class="header">'

            if self.include_company_info.isChecked():
                company_name_fixed = self.fix_arabic_text(self.company_name)
                report_title_fixed = self.fix_arabic_text(f"تقرير {self.section_name}")
                html += f'<div class="company-name" dir="rtl">{company_name_fixed}</div>'
                html += f'<div class="report-title" dir="rtl">{report_title_fixed}</div>'

            if self.include_date.isChecked():
                current_date = datetime.now().strftime("%Y/%m/%d - %H:%M")
                date_text_fixed = self.fix_arabic_text(f"تاريخ الطباعة: {current_date}")
                html += f'<div class="date-info" dir="rtl">{date_text_fixed}'

                if self.include_user.isChecked():
                    user_text_fixed = self.fix_arabic_text(f"طُبع بواسطة: {self.account_type}")
                    html += f' | {user_text_fixed}'

                html += '</div>'

            html += '</div>'
        return html

    # إنشاء جدول HTML
    def create_html_table(self, selected_indices, selected_headers):
        column_count = len(selected_headers)
        column_width = f"{100/column_count:.1f}%" if column_count > 0 else "auto"

        # إنشاء الجدول مع تحديد عرض الأعمدة بوضوح
        html = f'''<div class="table-container">
        <table style="width: 100% !important; table-layout: fixed !important; border-collapse: collapse; margin: 15px 0; direction: rtl; box-sizing: border-box;">
        <colgroup>'''

        # إضافة تعريف عرض الأعمدة
        for _ in range(column_count):
            html += f'<col style="width: {column_width};">'

        html += '</colgroup>'

        # هيدر الجدول مع عكس الترتيب للاتجاه العربي وضبط العرض
        html += '<thead><tr>'
        # عكس ترتيب الهيدر للاتجاه العربي
        reversed_headers = list(reversed(selected_headers))
        for header in reversed_headers:
            header_fixed = self.fix_arabic_text(header)
            html += f'<th style="width: {column_width};" dir="rtl">{header_fixed}</th>'
        html += '</tr></thead>'

        # بيانات الجدول
        html += '<tbody>'
        for row in self.table_data:
            html += '<tr>'
            # عكس ترتيب البيانات لتتماشى مع الهيدر المعكوس
            reversed_indices = list(reversed(selected_indices))
            reversed_selected_headers = list(reversed(selected_headers))

            for i, original_index in enumerate(reversed_indices):
                cell_value = str(row[original_index]) if row[original_index] is not None else ""

                # تنسيق الأرقام المالية
                if self.is_currency_column(reversed_selected_headers[i]):
                    try:
                        amount = float(cell_value)
                        cell_value = f"{amount:,.2f} {self.currency_type}"
                    except:
                        pass
                else:
                    # إصلاح النص العربي للخلايا النصية
                    cell_value = self.fix_arabic_text(cell_value)

                html += f'<td style="width: {column_width}; box-sizing: border-box;" dir="rtl">{cell_value}</td>'
            html += '</tr>'
        html += '</tbody></table></div>'
        return html

    # إنشاء فوتر HTML
    def create_html_footer(self):
        html = ""
        if self.include_company_info.isChecked():
            company_name_fixed = self.fix_arabic_text(self.company_name)
            address_fixed = self.fix_arabic_text(f"العنوان: {self.company_address}")
            phone_fixed = self.fix_arabic_text(f"الهاتف: {self.company_phone}")
            email_fixed = self.fix_arabic_text(f"البريد الإلكتروني: {self.company_email}")

            html += f'''
            <div class="footer" dir="rtl">
                <strong>{company_name_fixed}</strong><br>
                {address_fixed}<br>
                {phone_fixed} | {email_fixed}
            </div>
            '''
        return html

    # إنشاء إحصائيات HTML
    def create_html_statistics(self, headers):
        try:
            # البحث عن أعمدة المبالغ
            amount_col = None
            paid_col = None
            remaining_col = None

            selected_indices, _ = self.get_selected_columns()

            for i, header in enumerate(headers):
                if "المبلغ" in header and "مدفوع" not in header and "باقي" not in header:
                    amount_col = i
                elif "مدفوع" in header:
                    paid_col = i
                elif "باقي" in header:
                    remaining_col = i

            if amount_col is not None:
                # حساب الإحصائيات
                total_amount = 0
                total_paid = 0
                total_remaining = 0
                project_count = len(self.table_data)

                for row in self.table_data:
                    try:
                        if amount_col < len(selected_indices):
                            amount = float(row[selected_indices[amount_col]] or 0)
                            total_amount += amount

                        if paid_col is not None and paid_col < len(selected_indices):
                            paid = float(row[selected_indices[paid_col]] or 0)
                            total_paid += paid

                        if remaining_col is not None and remaining_col < len(selected_indices):
                            remaining = float(row[selected_indices[remaining_col]] or 0)
                            total_remaining += remaining
                    except:
                        continue

                # إنشاء HTML للإحصائيات مع إصلاح النص العربي وعكس الترتيب
                stats_html = f'''
                <div class="statistics">
                    <h3>{self.fix_arabic_text("إحصائيات المشاريع")}</h3>
                    <table class="stats-table">
                        <tr><th>{self.fix_arabic_text("القيمة")}</th><th>{self.fix_arabic_text("الإحصائية")}</th></tr>
                        <tr><td>{project_count}</td><td>{self.fix_arabic_text("عدد المشاريع")}</td></tr>
                        <tr><td>{total_amount:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي المبالغ")}</td></tr>
                        <tr><td>{total_paid:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي المدفوع")}</td></tr>
                        <tr><td>{total_remaining:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي المتبقي")}</td></tr>
                    </table>
                </div>
                '''
                return stats_html

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات HTML: {e}")

        return ""

    # إنشاء إحصائيات HTML عامة
    def create_html_general_statistics(self, headers):
        try:
            selected_indices, _ = self.get_selected_columns()

            # تحديد لون الإحصائيات حسب القسم
            section_colors = {
                "العملاء": "#27ae60",
                "الحسابات": "#e67e22",
                "الموظفين": "#8e44ad",
                "التدريب": "#f39c12"
            }

            color = section_colors.get(self.section_name, "#3498db")

            if self.section_name == "العملاء":
                return self.create_client_html_stats(headers, selected_indices, color)
            elif self.section_name == "الحسابات":
                return self.create_expense_html_stats(headers, selected_indices, color)
            elif self.section_name == "الموظفين":
                return self.create_employee_html_stats(headers, selected_indices, color)
            
            elif self.section_name == "التدريب":
                return self.create_training_html_stats(headers, selected_indices, color)

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات HTML عامة: {e}")

        return ""

    # إنشاء إحصائيات العملاء HTML
    def create_client_html_stats(self, headers, selected_indices, color):
        try:
            total_clients = len(self.table_data)
            client_types = {}

            # البحث عن عمود التصنيف
            classification_col = None
            for i, header in enumerate(headers):
                if "التصنيف" in header or "نوع" in header:
                    classification_col = i
                    break

            # تصنيف العملاء
            if classification_col is not None:
                for row in self.table_data:
                    try:
                        if classification_col < len(selected_indices):
                            client_type = str(row[selected_indices[classification_col]] or "غير محدد")
                            client_types[client_type] = client_types.get(client_type, 0) + 1
                    except:
                        continue

            # إنشاء HTML مع عكس الترتيب وإصلاح النص العربي
            stats_html = f'''
            <div class="statistics" style="border-color: {color};">
                <h3 style="color: {color};">{self.fix_arabic_text("إحصائيات العملاء")}</h3>
                <table class="stats-table">
                    <tr><th style="background-color: {color};">{self.fix_arabic_text("القيمة")}</th><th style="background-color: {color};">{self.fix_arabic_text("الإحصائية")}</th></tr>
                    <tr><td>{total_clients}</td><td>{self.fix_arabic_text("إجمالي العملاء")}</td></tr>
            '''

            # إضافة تصنيف العملاء
            for client_type, count in client_types.items():
                stats_html += f'<tr><td>{count}</td><td>{self.fix_arabic_text(f"عملاء {client_type}")}</td></tr>'

            stats_html += '</table></div>'
            return stats_html

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات العملاء HTML: {e}")
            return ""

    # إنشاء إحصائيات الحسابات HTML
    def create_expense_html_stats(self, headers, selected_indices, color):
        try:
            total_expenses = len(self.table_data)
            total_amount = 0
            expense_types = {}

            # البحث عن الأعمدة
            amount_col = None
            classification_col = None

            for i, header in enumerate(headers):
                if "المبلغ" in header:
                    amount_col = i
                elif "التصنيف" in header or "نوع" in header:
                    classification_col = i

            # حساب الإحصائيات
            for row in self.table_data:
                try:
                    if amount_col is not None and amount_col < len(selected_indices):
                        amount = float(row[selected_indices[amount_col]] or 0)
                        total_amount += amount

                    if classification_col is not None and classification_col < len(selected_indices):
                        expense_type = str(row[selected_indices[classification_col]] or "غير محدد")
                        if expense_type not in expense_types:
                            expense_types[expense_type] = {'count': 0, 'amount': 0}
                        expense_types[expense_type]['count'] += 1
                        expense_types[expense_type]['amount'] += amount
                except:
                    continue

            avg_amount = total_amount / total_expenses if total_expenses > 0 else 0

            # إنشاء HTML مع عكس الترتيب وإصلاح النص العربي
            stats_html = f'''
            <div class="statistics" style="border-color: {color};">
                <h3 style="color: {color};">{self.fix_arabic_text("إحصائيات الحسابات")}</h3>
                <table class="stats-table">
                    <tr><th style="background-color: {color};">{self.fix_arabic_text("القيمة")}</th><th style="background-color: {color};">{self.fix_arabic_text("الإحصائية")}</th></tr>
                    <tr><td>{total_expenses}</td><td>{self.fix_arabic_text("عدد المصروفات")}</td></tr>
                    <tr><td>{total_amount:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي المصروفات")}</td></tr>
                    <tr><td>{avg_amount:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("متوسط المصروف")}</td></tr>
            '''

            # إضافة تصنيف المصروفات
            for expense_type, data in expense_types.items():
                stats_html += f'<tr><td>{data["amount"]:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text(f"مصروفات {expense_type}")}</td></tr>'

            stats_html += '</table></div>'
            return stats_html

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات الحسابات HTML: {e}")
            return ""

    # إنشاء إحصائيات الموظفين HTML
    def create_employee_html_stats(self, headers, selected_indices, color):
        try:
            total_employees = len(self.table_data)
            total_salaries = 0
            total_balance = 0
            employee_types = {}

            # البحث عن الأعمدة
            salary_col = None
            balance_col = None
            classification_col = None

            for i, header in enumerate(headers):
                if "المرتب" in header or "راتب" in header:
                    salary_col = i
                elif "الرصيد" in header or "رصيد" in header:
                    balance_col = i
                elif "التصنيف" in header or "الوظيفة" in header:
                    classification_col = i

            # حساب الإحصائيات
            for row in self.table_data:
                try:
                    if salary_col is not None and salary_col < len(selected_indices):
                        salary = float(row[selected_indices[salary_col]] or 0)
                        total_salaries += salary

                    if balance_col is not None and balance_col < len(selected_indices):
                        balance = float(row[selected_indices[balance_col]] or 0)
                        total_balance += balance

                    if classification_col is not None and classification_col < len(selected_indices):
                        emp_type = str(row[selected_indices[classification_col]] or "غير محدد")
                        employee_types[emp_type] = employee_types.get(emp_type, 0) + 1
                except:
                    continue

            avg_salary = total_salaries / total_employees if total_employees > 0 else 0

            # إنشاء HTML مع عكس الترتيب وإصلاح النص العربي
            stats_html = f'''
            <div class="statistics" style="border-color: {color};">
                <h3 style="color: {color};">{self.fix_arabic_text("إحصائيات الموظفين")}</h3>
                <table class="stats-table">
                    <tr><th style="background-color: {color};">{self.fix_arabic_text("القيمة")}</th><th style="background-color: {color};">{self.fix_arabic_text("الإحصائية")}</th></tr>
                    <tr><td>{total_employees}</td><td>{self.fix_arabic_text("عدد الموظفين")}</td></tr>
                    <tr><td>{total_salaries:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي المرتبات")}</td></tr>
                    <tr><td>{total_balance:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي الأرصدة")}</td></tr>
                    <tr><td>{avg_salary:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("متوسط المرتب")}</td></tr>
            '''

            # إضافة تصنيف الموظفين
            for emp_type, count in employee_types.items():
                stats_html += f'<tr><td>{count}</td><td>{self.fix_arabic_text(emp_type)}</td></tr>'

            stats_html += '</table></div>'
            return stats_html

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات الموظفين HTML: {e}")
            return ""

    
    # إنشاء إحصائيات التدريب HTML
    def create_training_html_stats(self, headers, selected_indices, color):
        try:
            total_courses = len(self.table_data)
            total_revenue = 0
            total_participants = 0
            course_status = {}
            course_types = {}

            # البحث عن الأعمدة
            amount_col = None
            participants_col = None
            status_col = None
            classification_col = None

            for i, header in enumerate(headers):
                if "إجمالي" in header and "المبلغ" in header:
                    amount_col = i
                elif "المشاركين" in header:
                    participants_col = i
                elif "الحالة" in header:
                    status_col = i
                elif "التصنيف" in header or "نوع" in header:
                    classification_col = i

            # حساب الإحصائيات
            for row in self.table_data:
                try:
                    if amount_col is not None and amount_col < len(selected_indices):
                        amount = float(row[selected_indices[amount_col]] or 0)
                        total_revenue += amount

                    if participants_col is not None and participants_col < len(selected_indices):
                        participants = int(row[selected_indices[participants_col]] or 0)
                        total_participants += participants

                    if status_col is not None and status_col < len(selected_indices):
                        status = str(row[selected_indices[status_col]] or "غير محدد")
                        course_status[status] = course_status.get(status, 0) + 1

                    if classification_col is not None and classification_col < len(selected_indices):
                        course_type = str(row[selected_indices[classification_col]] or "غير محدد")
                        course_types[course_type] = course_types.get(course_type, 0) + 1
                except:
                    continue

            avg_revenue = total_revenue / total_courses if total_courses > 0 else 0
            avg_participants = total_participants / total_courses if total_courses > 0 else 0

            # إنشاء HTML مع عكس الترتيب وإصلاح النص العربي
            stats_html = f'''
            <div class="statistics" style="border-color: {color};">
                <h3 style="color: {color};">{self.fix_arabic_text("إحصائيات التدريب")}</h3>
                <table class="stats-table">
                    <tr><th style="background-color: {color};">{self.fix_arabic_text("القيمة")}</th><th style="background-color: {color};">{self.fix_arabic_text("الإحصائية")}</th></tr>
                    <tr><td>{total_courses}</td><td>{self.fix_arabic_text("عدد الدورات")}</td></tr>
                    <tr><td>{total_revenue:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("إجمالي الإيرادات")}</td></tr>
                    <tr><td>{total_participants}</td><td>{self.fix_arabic_text("إجمالي المشاركين")}</td></tr>
                    <tr><td>{avg_revenue:,.2f} {self.currency_type}</td><td>{self.fix_arabic_text("متوسط الإيراد/دورة")}</td></tr>
                    <tr><td>{avg_participants:.1f}</td><td>{self.fix_arabic_text("متوسط المشاركين/دورة")}</td></tr>
            '''

            # إضافة تصنيف حسب الحالة
            for status, count in course_status.items():
                stats_html += f'<tr><td>{count}</td><td>{self.fix_arabic_text(f"دورات {status}")}</td></tr>'

            # إضافة تصنيف حسب النوع
            for course_type, count in course_types.items():
                stats_html += f'<tr><td>{count}</td><td>{self.fix_arabic_text(course_type)}</td></tr>'

            stats_html += '</table></div>'
            return stats_html

        except Exception as e:
            print(f"خطأ في إنشاء إحصائيات التدريب HTML: {e}")
            return ""


# نافذة معاينة الطباعة
class PrintPreviewDialog(QDialog):
    # init
    def __init__(self, html_content, parent=None):
        super().__init__(parent)
        self.html_content = html_content
        self.parent_window = parent
        self.web_view = None
        self.text_view = None
        self.setup_ui()

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("معاينة الطباعة")
        self.setGeometry(100, 100, 1000, 700)
        self.setLayoutDirection(Qt.RightToLeft)

        # جعل النافذة مستقلة وغير modal مع إعدادات آمنة
        self.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint | Qt.WindowMinMaxButtonsHint)
        self.setAttribute(Qt.WA_DeleteOnClose, True)  # حذف النافذة عند الإغلاق
        self.setAttribute(Qt.WA_QuitOnClose, False)  # عدم إنهاء التطبيق عند إغلاق النافذة

        layout = QVBoxLayout()

        # شريط الأدوات
        toolbar = QHBoxLayout()

        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_document)
        toolbar.addWidget(print_btn)

        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(self.close_window)  # استخدام دالة مخصصة للإغلاق
        toolbar.addWidget(close_btn)

        toolbar.addStretch()
        layout.addLayout(toolbar)

        # عارض المحتوى
        try:
            from PySide6.QtWebEngineWidgets import QWebEngineView
            self.web_view = QWebEngineView()
            self.web_view.setHtml(self.html_content)
            layout.addWidget(self.web_view)
        except ImportError:
            # إذا لم يكن WebEngine متاحاً، استخدم QTextEdit
            from PySide6.QtWidgets import QTextEdit
            self.text_view = QTextEdit()
            self.text_view.setHtml(self.html_content)
            self.text_view.setReadOnly(True)
            layout.addWidget(self.text_view)

        self.setLayout(layout)

    # إغلاق النافذة بطريقة آمنة
    def close_window(self):
        try:
            # تنظيف الموارد قبل الإغلاق
            self.cleanup_resources()
            # إغلاق النافذة
            self.close()
        except Exception as e:
            print(f"خطأ في إغلاق نافذة المعاينة: {e}")
            # محاولة إغلاق قسري
            try:
                self.close()
            except:
                pass

    # تنظيف الموارد
    def cleanup_resources(self):
        try:
            # تنظيف محتوى WebView
            if self.web_view is not None:
                self.web_view.setHtml("")
                self.web_view.deleteLater()
                self.web_view = None

            # تنظيف محتوى TextView
            if self.text_view is not None:
                self.text_view.clear()
                self.text_view.deleteLater()
                self.text_view = None

            # تنظيف المحتوى
            self.html_content = None

        except Exception as e:
            print(f"خطأ في تنظيف الموارد: {e}")

    # معالجة إغلاق النافذة
    def closeEvent(self, event):
        try:
            # تنظيف الموارد
            self.cleanup_resources()
            # قبول الإغلاق
            event.accept()
        except Exception as e:
            print(f"خطأ في معالجة إغلاق النافذة: {e}")
            # قبول الإغلاق حتى لو حدث خطأ
            event.accept()

    # طباعة المستند
    def print_document(self):
        try:
            printer = QPrinter(QPrinter.HighResolution)
            print_dialog = QPrintDialog(printer, self)

            if print_dialog.exec() == QPrintDialog.Accepted:
                # استخدام QTextDocument للطباعة مع QPainter
                document = QTextDocument()
                document.setHtml(self.html_content)

                # تعيين حجم الصفحة
                try:
                    page_rect = printer.pageRect(QPrinter.DevicePixel)
                    document.setPageSize(page_rect.size())
                except:
                    from PySide6.QtCore import QSizeF
                    document.setPageSize(QSizeF(595, 842))  # A4 size

                # الطباعة باستخدام QPainter
                painter = QPainter()
                if painter.begin(printer):
                    try:
                        document.drawContents(painter)
                        painter.end()
                        QMessageBox.information(self, "طباعة", "تمت الطباعة بنجاح")
                    except Exception as draw_error:
                        painter.end()
                        QMessageBox.critical(self, "خطأ", f"فشل في رسم المحتوى:\n{str(draw_error)}")
                else:
                    QMessageBox.critical(self, "خطأ", "فشل في بدء عملية الطباعة")
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")


# نافذة الطباعة المباشرة
class PrintDialog(QDialog):
    # init
    def __init__(self, html_content, parent=None):
        super().__init__(parent)
        self.html_content = html_content
        self.parent_window = parent
        self.setup_ui()

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("طباعة")
        self.setFixedSize(400, 200)
        self.setLayoutDirection(Qt.RightToLeft)

        # جعل النافذة مستقلة وغير modal مع إعدادات آمنة
        self.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint)
        self.setAttribute(Qt.WA_DeleteOnClose, True)  # حذف النافذة عند الإغلاق
        self.setAttribute(Qt.WA_QuitOnClose, False)  # عدم إنهاء التطبيق عند إغلاق النافذة

        layout = QVBoxLayout()

        # رسالة
        label = QLabel("هل تريد طباعة المستند؟")
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("font-size: 14px; margin: 20px;")
        layout.addWidget(label)

        # أزرار
        buttons_layout = QHBoxLayout()

        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_document)
        buttons_layout.addWidget(print_btn)

        cancel_btn = QPushButton("إلغاء")
        cancel_btn.clicked.connect(self.close_window)  # استخدام دالة مخصصة للإغلاق
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    # إغلاق النافذة بطريقة آمنة
    def close_window(self):
        try:
            # تنظيف الموارد قبل الإغلاق
            self.html_content = None
            # إغلاق النافذة
            self.close()
        except Exception as e:
            print(f"خطأ في إغلاق نافذة الطباعة: {e}")
            # محاولة إغلاق قسري
            try:
                self.close()
            except:
                pass

    # معالجة إغلاق النافذة
    def closeEvent(self, event):
        try:
            # تنظيف الموارد
            self.html_content = None
            # قبول الإغلاق
            event.accept()
        except Exception as e:
            print(f"خطأ في معالجة إغلاق النافذة: {e}")
            # قبول الإغلاق حتى لو حدث خطأ
            event.accept()

    # طباعة المستند
    def print_document(self):
        try:
            printer = QPrinter(QPrinter.HighResolution)
            print_dialog = QPrintDialog(printer, self)

            if print_dialog.exec() == QPrintDialog.Accepted:
                # إنشاء مستند نصي للطباعة
                document = QTextDocument()
                document.setHtml(self.html_content)

                # تعيين حجم الصفحة بطريقة آمنة
                try:
                    page_rect = printer.pageRect(QPrinter.DevicePixel)
                    document.setPageSize(page_rect.size())
                except:
                    # استخدام حجم افتراضي إذا فشل
                    from PySide6.QtCore import QSizeF
                    document.setPageSize(QSizeF(595, 842))  # A4 size

                # استخدام QPainter للطباعة
                painter = QPainter()
                if painter.begin(printer):
                    try:
                        # رسم المحتوى
                        document.drawContents(painter)
                        painter.end()
                        QMessageBox.information(self, "طباعة", "تمت الطباعة بنجاح")
                        self.close_window()  # إغلاق النافذة بطريقة آمنة
                    except Exception as draw_error:
                        painter.end()
                        QMessageBox.critical(self, "خطأ", f"فشل في رسم المحتوى:\n{str(draw_error)}")
                else:
                    QMessageBox.critical(self, "خطأ", "فشل في بدء عملية الطباعة")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")


# دالة لفتح نافذة الطباعة والتصدير
# فتح نافذة الطباعة والتصدير المتقدمة
def open_print_export_dialog(table_widget, section_name, parent=None):
    try:
        # استخراج البيانات من الجدول
        table_data = []
        headers = []

        # الحصول على الهيدرز
        for col in range(table_widget.columnCount()):
            header_item = table_widget.horizontalHeaderItem(col)
            headers.append(header_item.text() if header_item else f"عمود {col + 1}")

        # الحصول على البيانات
        for row in range(table_widget.rowCount()):
            row_data = []
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                row_data.append(item.text() if item else "")
            table_data.append(row_data)

        # التحقق من وجود نافذة طباعة مفتوحة مسبقاً في النافذة الأب
        if hasattr(parent, 'print_export_dialog') and parent.print_export_dialog is not None:
            try:
                parent.print_export_dialog.close()
                parent.print_export_dialog = None
            except:
                pass

        # فتح النافذة بطريقة non-modal لتجنب التعارض
        dialog = AdvancedPrintExportDialog(table_data, headers, section_name, parent)

        # إضافة إعدادات آمنة للنافذة
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        dialog.setAttribute(Qt.WA_QuitOnClose, False)
        dialog.setModal(False)  # جعل النافذة غير modal

        # حفظ مرجع في النافذة الأب لإدارة أفضل
        if hasattr(parent, '__dict__'):
            parent.print_export_dialog = dialog
            # ربط إشارة الإغلاق لتنظيف المرجع
            dialog.finished.connect(lambda: setattr(parent, 'print_export_dialog', None))
            dialog.destroyed.connect(lambda: setattr(parent, 'print_export_dialog', None))

        # عرض النافذة
        dialog.show()
        return dialog

    except Exception as e:
        QMessageBox.critical(parent, "خطأ", f"فشل في فتح نافذة الطباعة والتصدير:\n{str(e)}")
        return None


# ==================== نظام الطباعة الموحد ====================

import sys
import qtawesome as qta
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# مدير الطباعة الموحد لجميع نوافذ الإدارة
class UnifiedPrintManager:

    # init
    def __init__(self):
        # إعدادات الشركة
        self.company_name = company_name
        self.logo_path = logo_path
        self.company_phone = company_phone
        self.company_address = company_address
        self.company_email = company_email
        self.currency_type = Currency_type
        self.account_type = account_type

        # إعدادات الطباعة الافتراضية
        self.default_settings = {
            'orientation': 'portrait',  # portrait or landscape
            'font_size': 12,
            'include_logo': True,
            'include_company_info': True,
            'include_user_info': True,
            'include_date': True,
            'include_page_numbers': True,
            'page_margins': {'top': 25, 'bottom': 25, 'left': 25, 'right': 25}
        }

    # فتح حوار الطباعة الموحد
    def open_print_dialog(self, parent_window, tab_widget=None, current_tab_index=None):
        try:
            # استخراج بيانات التاب الحالي
            tab_data = self.extract_tab_data(parent_window, tab_widget, current_tab_index)

            if not tab_data:
                QMessageBox.warning(parent_window, "تحذير", "لا توجد بيانات للطباعة")
                return

            # فتح حوار الطباعة
            dialog = UnifiedPrintDialog(tab_data, parent_window)
            dialog.exec()

        except Exception as e:
            QMessageBox.critical(parent_window, "خطأ", f"فشل في فتح حوار الطباعة:\n{str(e)}")

    # استخراج بيانات التاب الحالي أو النافذة الكاملة
    def extract_tab_data(self, parent_window, tab_widget, current_tab_index):
        try:
            # إذا كانت النافذة تحتوي على تابات
            if tab_widget and current_tab_index is not None:
                current_tab = tab_widget.widget(current_tab_index)
                tab_name = tab_widget.tabText(current_tab_index)

                # البحث عن الجداول في التاب الحالي
                tables = self.find_tables_in_widget(current_tab)

                # البحث عن البطاقات الإحصائية
                statistics_cards = self.find_statistics_cards(current_tab)

            else:
                # النافذة لا تحتوي على تابات - استخراج البيانات من النافذة الكاملة
                tab_name = self.get_window_title(parent_window)

                # البحث عن الجداول في النافذة الكاملة
                tables = self.find_tables_in_widget(parent_window)

                # البحث عن البطاقات الإحصائية في النافذة الكاملة
                statistics_cards = self.find_statistics_cards(parent_window)

            # تحديد نوع النافذة
            window_type = self.detect_window_type(parent_window)

            return {
                'window_type': window_type,
                'tab_name': tab_name,
                'tables': tables,
                'statistics': statistics_cards,
                'parent_window': parent_window
            }

        except Exception as e:
            print(f"خطأ في استخراج بيانات التاب: {e}")
            return None

    # الحصول على عنوان النافذة
    def get_window_title(self, window):
        try:
            if hasattr(window, 'windowTitle') and window.windowTitle():
                return window.windowTitle()
            elif hasattr(window, 'project_data') and window.project_data:
                # للنوافذ التي تحتوي على بيانات المشروع
                project_name = window.project_data.get('اسم_المشروع', 'مشروع غير محدد')
                return f"إدارة دفعات - {project_name}"
            else:
                # عنوان افتراضي حسب نوع النافذة
                window_type = self.detect_window_type(window)
                return window_type
        except Exception as e:
            print(f"خطأ في الحصول على عنوان النافذة: {e}")
            return "نافذة البيانات"

    # البحث عن جميع الجداول في الويدجت
    def find_tables_in_widget(self, widget):
        tables = []

        # ابحث عن العودية
        def search_recursive(w):
            if isinstance(w, QTableWidget):
                table_data = self.extract_table_data(w)
                if table_data:
                    tables.append(table_data)

            # البحث في الويدجتات الفرعية
            for child in w.findChildren(QWidget):
                if isinstance(child, QTableWidget):
                    table_data = self.extract_table_data(child)
                    if table_data:
                        tables.append(table_data)

        search_recursive(widget)
        return tables

    # استخراج بيانات الجدول
    def extract_table_data(self, table_widget):
        try:
            if not table_widget or table_widget.rowCount() == 0:
                return None

            # استخراج العناوين
            headers = []
            for col in range(table_widget.columnCount()):
                if not table_widget.isColumnHidden(col):
                    header_item = table_widget.horizontalHeaderItem(col)
                    headers.append(header_item.text() if header_item else f"عمود {col + 1}")

            # استخراج البيانات
            data = []
            for row in range(table_widget.rowCount()):
                row_data = []
                for col in range(table_widget.columnCount()):
                    if not table_widget.isColumnHidden(col):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                data.append(row_data)

            return {
                'headers': headers,
                'data': data,
                'name': table_widget.objectName() or "جدول البيانات"
            }

        except Exception as e:
            print(f"خطأ في استخراج بيانات الجدول: {e}")
            return None

    # البحث عن البطاقات الإحصائية
    def find_statistics_cards(self, widget):
        statistics = []

        # البحث عن عناصر البطاقات بطرق مختلفة
        for child in widget.findChildren(QWidget):
            card_data = None

            # البحث بناءً على objectName
            if hasattr(child, 'objectName') and child.objectName():
                obj_name = child.objectName().lower()
                if any(keyword in obj_name for keyword in ['card', 'stat', 'info', 'summary']):
                    card_data = self.extract_card_data(child)

            # البحث بناءً على نوع الويدجت
            if not card_data and isinstance(child, (QFrame, QGroupBox)):
                # التحقق من وجود نصوص تبدو كإحصائيات
                labels = child.findChildren(QLabel)
                if len(labels) >= 2:  # على الأقل عنوان وقيمة
                    card_data = self.extract_card_data(child)

            # البحث في الويدجتات التي تحتوي على أرقام أو عملات
            if not card_data:
                labels = child.findChildren(QLabel)
                for label in labels:
                    text = label.text()
                    # البحث عن نصوص تحتوي على أرقام أو رموز عملات
                    if any(char.isdigit() for char in text) and any(keyword in text for keyword in ['ريال', 'دولار', 'دينار', '$', '€', '£']):
                        card_data = self.extract_card_data(child)
                        break

            if card_data:
                statistics.append(card_data)

        # إضافة إحصائيات خاصة بنافذة الدفعات
        if hasattr(widget, 'project_data') and widget.project_data:
            project_stats = self.extract_project_statistics(widget.project_data)
            statistics.extend(project_stats)

        return statistics

    # استخراج إحصائيات المشروع من البيانات
    def extract_project_statistics(self, project_data):
        statistics = []

        try:
            if project_data:
                # إحصائيات أساسية للمشروع
                if 'اسم_المشروع' in project_data:
                    statistics.append({
                        'title': 'اسم المشروع',
                        'value': project_data['اسم_المشروع']
                    })

                if 'المبلغ' in project_data:
                    statistics.append({
                        'title': 'إجمالي المبلغ',
                        'value': f"{project_data['المبلغ']:,} {self.currency_type}"
                    })

                if 'المدفوع' in project_data:
                    statistics.append({
                        'title': 'المبلغ المدفوع',
                        'value': f"{project_data['المدفوع']:,} {self.currency_type}"
                    })

                if 'الباقي' in project_data:
                    statistics.append({
                        'title': 'المبلغ المتبقي',
                        'value': f"{project_data['الباقي']:,} {self.currency_type}"
                    })

                # حساب النسبة المئوية للإنجاز
                if 'المبلغ' in project_data and 'المدفوع' in project_data and project_data['المبلغ'] > 0:
                    percentage = (project_data['المدفوع'] / project_data['المبلغ']) * 100
                    statistics.append({
                        'title': 'نسبة الإنجاز',
                        'value': f"{percentage:.1f}%"
                    })

                if 'اسم_العميل' in project_data:
                    statistics.append({
                        'title': 'العميل',
                        'value': project_data['اسم_العميل']
                    })

        except Exception as e:
            print(f"خطأ في استخراج إحصائيات المشروع: {e}")

        return statistics

    # استخراج بيانات البطاقة
    def extract_card_data(self, card_widget):
        try:
            # البحث عن النصوص في البطاقة
            labels = card_widget.findChildren(QLabel)

            title = ""
            value = ""

            for label in labels:
                text = label.text()
                if text and text.strip():
                    if not title:
                        title = text
                    elif not value and text != title:
                        value = text

            if title:
                return {'title': title, 'value': value}

        except Exception as e:
            print(f"خطأ في استخراج بيانات البطاقة: {e}")

        return None

    # تحديد نوع النافذة
    def detect_window_type(self, window):
        window_class = window.__class__.__name__

        if 'Project' in window_class or 'مشروع' in window_class:
            return "إدارة المشاريع"
        elif 'Client' in window_class or 'عميل' in window_class:
            return "إدارة العملاء"
        elif 'Training' in window_class or 'تدريب' in window_class:
            return "إدارة التدريب"
        elif 'Employee' in window_class or 'موظف' in window_class:
            return "إدارة الموظفين"
        elif 'Payment' in window_class or 'دفعات' in window_class:
            return "إدارة الدفعات"
        elif 'Debt' in window_class or 'ديون' in window_class:
            return "إدارة الديون"
        elif 'Supplier' in window_class or 'مورد' in window_class:
            return "إدارة الموردين"
        else:
            return "نظام الإدارة"


# حوار الطباعة الموحد
class UnifiedPrintDialog(QDialog):

    # init
    def __init__(self, tab_data, parent=None):
        super().__init__(parent)
        self.tab_data = tab_data
        self.parent_window = parent
        self.print_manager = UnifiedPrintManager()

        self.setup_ui()
        self.apply_styles()

        # إعدادات النافذة
        self.setModal(True)
        self.setAttribute(Qt.WA_DeleteOnClose, True)

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("طباعة وتصدير متقدم")
        self.setFixedSize(700, 800)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان رئيسي
        title_label = QLabel(f"طباعة وتصدير بيانات {self.tab_data['tab_name']}")
        title_label.setObjectName("dialog_title")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # إنشاء التاب الموحد
        self.create_tabs(layout)

        # أزرار التحكم
        self.create_control_buttons(layout)

    # إنشاء تاب موحد للطباعة والإعدادات
    def create_tabs(self, layout):
        # إنشاء تاب واحد موحد بدلاً من عدة تابات
        self.create_unified_tab(layout)

    # إنشاء تاب موحد للطباعة والإعدادات
    def create_unified_tab(self, layout):
        # إنشاء مجموعة إعدادات التصدير (مثل AdvancedPrintExportDialog)
        export_group = QGroupBox("إعدادات التصدير")
        export_group.setObjectName("settings_group")
        export_layout = QVBoxLayout()

        # نوع التصدير
        export_type_layout = QHBoxLayout()
        export_type_layout.addWidget(QLabel("نوع التصدير:"))

        self.export_type = QComboBox()
        self.export_type.setObjectName("settings_combo")
        self.export_type.addItems(["PDF احترافي", "Excel متقدم", "طباعة مباشرة"])
        export_type_layout.addWidget(self.export_type)
        export_layout.addLayout(export_type_layout)

        # اتجاه الصفحة
        orientation_layout = QHBoxLayout()
        orientation_layout.addWidget(QLabel("اتجاه الصفحة:"))

        self.orientation_combo = QComboBox()
        self.orientation_combo.setObjectName("settings_combo")
        self.orientation_combo.addItems(["عمودي (Portrait)", "أفقي (Landscape)"])
        orientation_layout.addWidget(self.orientation_combo)
        export_layout.addLayout(orientation_layout)

        # حجم الخط
        font_layout = QHBoxLayout()
        font_layout.addWidget(QLabel("حجم الخط:"))

        self.font_size_spin = QSpinBox()
        self.font_size_spin.setObjectName("settings_spin")
        self.font_size_spin.setRange(8, 24)
        self.font_size_spin.setValue(12)
        font_layout.addWidget(self.font_size_spin)
        export_layout.addLayout(font_layout)

        export_group.setLayout(export_layout)
        layout.addWidget(export_group)

        # مجموعة إعدادات المحتوى
        content_group = QGroupBox("إعدادات المحتوى")
        content_group.setObjectName("settings_group")
        content_group_layout = QGridLayout()

        # إنشاء مربعات الاختيار
        self.include_logo_cb = QCheckBox("تضمين شعار الشركة")
        self.include_logo_cb.setObjectName("settings_checkbox")
        self.include_logo_cb.setChecked(True)

        self.include_company_info_cb = QCheckBox("تضمين معلومات الشركة")
        self.include_company_info_cb.setObjectName("settings_checkbox")
        self.include_company_info_cb.setChecked(True)

        self.include_user_info_cb = QCheckBox("تضمين اسم المستخدم")
        self.include_user_info_cb.setObjectName("settings_checkbox")
        self.include_user_info_cb.setChecked(True)

        self.include_date_cb = QCheckBox("تضمين تاريخ الطباعة")
        self.include_date_cb.setObjectName("settings_checkbox")
        self.include_date_cb.setChecked(True)

        self.include_page_numbers_cb = QCheckBox("تضمين أرقام الصفحات")
        self.include_page_numbers_cb.setObjectName("settings_checkbox")
        self.include_page_numbers_cb.setChecked(True)

        self.include_statistics_cb = QCheckBox("تضمين الإحصائيات")
        self.include_statistics_cb.setObjectName("settings_checkbox")
        self.include_statistics_cb.setChecked(True)

        # ترتيب مربعات الاختيار في تخطيط شبكي (3 صفوف × 2 أعمدة)
        # الصف الأول: شعار الشركة + معلومات الشركة
        content_group_layout.addWidget(self.include_logo_cb, 0, 0)
        content_group_layout.addWidget(self.include_company_info_cb, 0, 1)

        # الصف الثاني: اسم المستخدم + تاريخ الطباعة
        content_group_layout.addWidget(self.include_user_info_cb, 1, 0)
        content_group_layout.addWidget(self.include_date_cb, 1, 1)

        # الصف الثالث: أرقام الصفحات + الإحصائيات
        content_group_layout.addWidget(self.include_page_numbers_cb, 2, 0)
        content_group_layout.addWidget(self.include_statistics_cb, 2, 1)

        # تعيين محاذاة للتخطيط الشبكي لضمان التوافق مع RTL
        content_group_layout.setAlignment(Qt.AlignTop)
        content_group_layout.setSpacing(10)
        content_group_layout.setContentsMargins(10, 10, 10, 10)

        content_group.setLayout(content_group_layout)
        layout.addWidget(content_group)

        # مجموعة فلترة البيانات (اختيار الأعمدة)
        filter_group = QGroupBox("فلترة البيانات")
        filter_group.setObjectName("settings_group")
        filter_layout = QVBoxLayout()

        # عنوان اختيار الأعمدة
        columns_title = QLabel("اختر الأعمدة للتصدير:")
        columns_title.setObjectName("section_title")
        filter_layout.addWidget(columns_title)

        # إنشاء تخطيط شبكي للأعمدة (3 أعمدة في كل صف)
        self.columns_grid_layout = QGridLayout()
        self.columns_checkboxes = []

        # إضافة الأعمدة من الجداول في تخطيط شبكي
        row = 0
        col = 0
        for table in self.tab_data.get('tables', []):
            for i, header in enumerate(table['headers']):
                checkbox = QCheckBox(header)
                checkbox.setObjectName("column_checkbox")
                checkbox.setChecked(True)  # محدد افتراضياً
                checkbox.setProperty('table_name', table['name'])
                checkbox.setProperty('column_index', i)

                self.columns_checkboxes.append(checkbox)
                self.columns_grid_layout.addWidget(checkbox, row, col)

                col += 1
                if col >= 3:  # 3 أعمدة في كل صف
                    col = 0
                    row += 1

        filter_layout.addLayout(self.columns_grid_layout)

        # أزرار تحديد الأعمدة
        columns_buttons = QHBoxLayout()
        select_all_btn = QPushButton("تحديد الكل")
        select_all_btn.setObjectName("action_button")
        select_all_btn.clicked.connect(self.select_all_columns)
        columns_buttons.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("إلغاء تحديد الكل")
        deselect_all_btn.setObjectName("action_button")
        deselect_all_btn.clicked.connect(self.deselect_all_columns)
        columns_buttons.addWidget(deselect_all_btn)

        filter_layout.addLayout(columns_buttons)
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)





    # إنشاء أزرار التحكم
    def create_control_buttons(self, layout):
        buttons_layout = QHBoxLayout()

        # زر المعاينة
        self.preview_btn = QPushButton("معاينة")
        self.preview_btn.setObjectName("action_button")
        try:
            self.preview_btn.setIcon(qta.icon('fa5s.eye', color='white'))
        except:
            pass
        self.preview_btn.clicked.connect(self.show_print_preview)

        # زر التصدير (يعتمد على نوع التصدير المحدد)
        self.export_btn = QPushButton("تصدير")
        self.export_btn.setObjectName("primary_button")
        try:
            self.export_btn.setIcon(qta.icon('fa5s.download', color='white'))
        except:
            pass
        self.export_btn.clicked.connect(self.export_document)

        # زر الطباعة المباشرة
        self.print_btn = QPushButton("طباعة")
        self.print_btn.setObjectName("secondary_button")
        try:
            self.print_btn.setIcon(qta.icon('fa5s.print', color='white'))
        except:
            pass
        self.print_btn.clicked.connect(self.print_document)

        # زر الإلغاء
        cancel_btn = QPushButton("إلغاء")
        cancel_btn.setObjectName("cancel_button")
        try:
            cancel_btn.setIcon(qta.icon('fa5s.times', color='white'))
        except:
            pass
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addWidget(self.preview_btn)
        buttons_layout.addWidget(self.export_btn)
        buttons_layout.addWidget(self.print_btn)
        buttons_layout.addStretch()
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

    # تحديد جميع الأعمدة
    def select_all_columns(self):
        for checkbox in self.columns_checkboxes:
            checkbox.setChecked(True)

    # إلغاء تحديد جميع الأعمدة
    def deselect_all_columns(self):
        for checkbox in self.columns_checkboxes:
            checkbox.setChecked(False)



    # الحصول على الأعمدة المحددة
    def get_selected_columns(self):
        selected = []
        for checkbox in self.columns_checkboxes:
            if checkbox.isChecked():
                selected.append({
                    'table_name': checkbox.property('table_name'),
                    'column_index': checkbox.property('column_index'),
                    'header': checkbox.text()
                })
        return selected

    # تصدير المستند حسب النوع المحدد
    def export_document(self):
        export_type = self.export_type.currentText()

        if export_type == "PDF احترافي":
            self.export_pdf()
        elif export_type == "Excel متقدم":
            self.export_excel()
        else:  # طباعة مباشرة
            self.print_document()

    # الحصول على إعدادات الطباعة
    def get_print_settings(self):
        return {
            'orientation': 'landscape' if self.orientation_combo.currentIndex() == 1 else 'portrait',
            'font_size': self.font_size_spin.value(),
            'include_logo': self.include_logo_cb.isChecked(),
            'include_company_info': self.include_company_info_cb.isChecked(),
            'include_user_info': self.include_user_info_cb.isChecked(),
            'include_date': self.include_date_cb.isChecked(),
            'include_page_numbers': self.include_page_numbers_cb.isChecked(),
            'include_statistics': self.include_statistics_cb.isChecked()
        }

    # طباعة المستند
    def print_document(self):
        try:
            selected_columns = self.get_selected_columns()
            if not selected_columns:
                QMessageBox.warning(self, "تحذير", "يرجى اختيار عمود واحد على الأقل للطباعة")
                return

            settings = self.get_print_settings()

            # إنشاء مستند HTML للطباعة
            html_content = self.generate_html_document(selected_columns, settings)

            # طباعة المستند
            self.print_html_document(html_content)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")

    # تصدير إلى PDF
    def export_pdf(self):
        try:
            selected_columns = self.get_selected_columns()
            if not selected_columns:
                QMessageBox.warning(self, "تحذير", "يرجى اختيار عمود واحد على الأقل للتصدير")
                return

            # اختيار مكان الحفظ
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ ملف PDF",
                f"{self.tab_data['tab_name']}.pdf",
                "PDF Files (*.pdf)"
            )

            if file_path:
                settings = self.get_print_settings()
                html_content = self.generate_html_document(selected_columns, settings)

                # تصدير إلى PDF
                self.export_html_to_pdf(html_content, file_path)

                QMessageBox.information(self, "نجح", f"تم تصدير الملف بنجاح:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير PDF:\n{str(e)}")

    # تصدير إلى Excel
    def export_excel(self):
        try:
            selected_columns = self.get_selected_columns()
            if not selected_columns:
                QMessageBox.warning(self, "تحذير", "يرجى اختيار عمود واحد على الأقل للتصدير")
                return

            # اختيار مكان الحفظ
            file_path, _ = QFileDialog.getSaveFileName(
                self, "حفظ ملف Excel",
                f"{self.tab_data['tab_name']}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                self.export_to_excel(selected_columns, file_path)
                QMessageBox.information(self, "نجح", f"تم تصدير الملف بنجاح:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير Excel:\n{str(e)}")

    # إنشاء مستند HTML للطباعة
    def generate_html_document(self, selected_columns, settings):
        try:
            html = f"""
            <!DOCTYPE html>
            <html dir="rtl" lang="ar">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>{self.tab_data['tab_name']}</title>
                <style>
                    @page {{
                        size: {'A4 landscape' if settings['orientation'] == 'landscape' else 'A4 portrait'};
                        margin: 2cm;
                    }}
                    body {{
                        font-family: 'Arial', 'Tahoma', sans-serif;
                        font-size: {settings['font_size']}px;
                        direction: rtl;
                        text-align: right;
                        margin: 0;
                        padding: 0;
                    }}
                    .header {{
                        text-align: center;
                        margin-bottom: 30px;
                        border-bottom: 2px solid #3498db;
                        padding-bottom: 20px;
                    }}
                    .company-info {{
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                        margin-bottom: 20px;
                    }}
                    .company-logo {{
                        max-width: 100px;
                        max-height: 80px;
                    }}
                    .company-name {{
                        font-size: {settings['font_size'] + 8}px;
                        font-weight: bold;
                        color: #2c3e50;
                    }}
                    .document-title {{
                        font-size: {settings['font_size'] + 4}px;
                        font-weight: bold;
                        color: #3498db;
                        margin: 20px 0;
                    }}
                    .statistics {{
                        background-color: #f8f9fa;
                        padding: 15px;
                        border-radius: 5px;
                        margin-bottom: 20px;
                        border: 1px solid #dee2e6;
                    }}
                    .statistics h3 {{
                        color: #495057;
                        margin-top: 0;
                    }}
                    .stats-grid {{
                        display: grid;
                        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                        gap: 10px;
                    }}
                    .stat-item {{
                        background: white;
                        padding: 10px;
                        border-radius: 3px;
                        border: 1px solid #e9ecef;
                    }}
                    .stat-title {{
                        font-weight: bold;
                        color: #6c757d;
                    }}
                    .stat-value {{
                        font-size: {settings['font_size'] + 2}px;
                        color: #495057;
                        margin-top: 5px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-top: 20px;
                        background: white;
                    }}
                    th, td {{
                        border: 1px solid #dee2e6;
                        padding: 8px;
                        text-align: right;
                    }}
                    th {{
                        background-color: #3498db;
                        color: white;
                        font-weight: bold;
                    }}
                    tr:nth-child(even) {{
                        background-color: #f8f9fa;
                    }}
                    .footer {{
                        margin-top: 30px;
                        padding-top: 20px;
                        border-top: 1px solid #dee2e6;
                        font-size: {settings['font_size'] - 2}px;
                        color: #6c757d;
                        text-align: center;
                    }}
                    .print-info {{
                        display: flex;
                        justify-content: space-between;
                        margin-top: 20px;
                        font-size: {settings['font_size'] - 2}px;
                        color: #6c757d;
                    }}
                </style>
            </head>
            <body>
            """

            # إضافة الهيدر
            if settings['include_logo'] or settings['include_company_info']:
                html += '<div class="header">'

                if settings['include_company_info']:
                    html += '<div class="company-info">'

                    if settings['include_logo'] and self.print_manager.logo_path and os.path.exists(self.print_manager.logo_path):
                        html += f'<img src="file:///{self.print_manager.logo_path}" class="company-logo" alt="شعار الشركة">'

                    html += f'<div class="company-name">{self.print_manager.company_name}</div>'
                    html += '</div>'

                html += f'<div class="document-title">{self.tab_data["window_type"]} - {self.tab_data["tab_name"]}</div>'
                html += '</div>'

            # إضافة الإحصائيات
            if settings['include_statistics'] and self.tab_data.get('statistics'):
                html += '<div class="statistics">'
                html += '<h3>الإحصائيات</h3>'
                html += '<div class="stats-grid">'

                for stat in self.tab_data['statistics']:
                    html += f'''
                    <div class="stat-item">
                        <div class="stat-title">{stat['title']}</div>
                        <div class="stat-value">{stat['value']}</div>
                    </div>
                    '''

                html += '</div></div>'

            # إضافة الجداول
            for table in self.tab_data.get('tables', []):
                # تصفية الأعمدة المحددة
                filtered_headers = []
                filtered_indices = []

                for col_info in selected_columns:
                    if col_info['table_name'] == table['name']:
                        col_index = col_info['column_index']
                        if col_index < len(table['headers']):
                            filtered_headers.append(table['headers'][col_index])
                            filtered_indices.append(col_index)

                if filtered_headers:
                    html += f'<h3>{table["name"]}</h3>'
                    html += '<table>'

                    # إضافة العناوين
                    html += '<thead><tr>'
                    for header in filtered_headers:
                        html += f'<th>{header}</th>'
                    html += '</tr></thead>'

                    # إضافة البيانات
                    html += '<tbody>'
                    for row in table['data']:
                        html += '<tr>'
                        for col_index in filtered_indices:
                            cell_data = row[col_index] if col_index < len(row) else ""
                            html += f'<td>{cell_data}</td>'
                        html += '</tr>'
                    html += '</tbody>'

                    html += '</table>'

            # إضافة الفوتر
            html += '<div class="footer">'

            if settings['include_company_info']:
                html += f'''
                <div>
                    {self.print_manager.company_name}<br>
                    الهاتف: {self.print_manager.company_phone}<br>
                    العنوان: {self.print_manager.company_address}<br>
                    البريد الإلكتروني: {self.print_manager.company_email}
                </div>
                '''

            if settings['include_date'] or settings['include_user_info']:
                html += '<div class="print-info">'

                if settings['include_date']:
                    html += f'<span>تاريخ الطباعة: {datetime.now().strftime("%Y-%m-%d %H:%M")}</span>'

                if settings['include_user_info']:
                    html += f'<span>المستخدم: {self.print_manager.account_type}</span>'

                html += '</div>'

            html += '</div>'
            html += '</body></html>'

            return html

        except Exception as e:
            print(f"خطأ في إنشاء مستند HTML: {e}")
            return ""

    # طباعة مستند HTML
    def print_html_document(self, html_content):
        try:
            # إنشاء ملف مؤقت
            temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8')
            temp_file.write(html_content)
            temp_file.close()

            # إنشاء QWebEngineView للطباعة
            web_view = QWebEngineView()
            web_view.setHtml(html_content)

            # إعداد الطابعة
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPageSize(QPageSize.A4))

            settings = self.get_print_settings()
            if settings['orientation'] == 'landscape':
                printer.setPageOrientation(QPageLayout.Landscape)
            else:
                printer.setPageOrientation(QPageLayout.Portrait)

            # فتح حوار الطباعة
            print_dialog = QPrintDialog(printer, self)
            if print_dialog.exec() == QPrintDialog.Accepted:
                # طباعة الصفحة
                # طباعة منتهية
                def print_finished(success):
                    if success:
                        QMessageBox.information(self, "نجح", "تمت الطباعة بنجاح")
                    else:
                        QMessageBox.warning(self, "تحذير", "فشلت عملية الطباعة")

                    # تنظيف الملف المؤقت
                    try:
                        os.unlink(temp_file.name)
                    except:
                        pass

                web_view.page().print(printer, print_finished)
            else:
                # تنظيف الملف المؤقت في حالة الإلغاء
                try:
                    os.unlink(temp_file.name)
                except:
                    pass

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في الطباعة:\n{str(e)}")

    # تصدير HTML إلى PDF
    def export_html_to_pdf(self, html_content, file_path):
        try:
            # إنشاء QWebEngineView
            web_view = QWebEngineView()
            web_view.setHtml(html_content)

            # إعداد تخطيط الصفحة
            page_layout = QPageLayout()
            page_layout.setPageSize(QPageSize(QPageSize.A4))

            settings = self.get_print_settings()
            if settings['orientation'] == 'landscape':
                page_layout.setOrientation(QPageLayout.Landscape)
            else:
                page_layout.setOrientation(QPageLayout.Portrait)

            page_layout.setMargins(QMarginsF(25, 25, 25, 25), QPageLayout.Millimeter)

            # تصدير إلى PDF
            # انتهى PDF
            def pdf_finished(success):
                if not success:
                    QMessageBox.critical(self, "خطأ", "فشل في تصدير PDF")

            web_view.page().printToPdf(file_path, page_layout)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير PDF:\n{str(e)}")

    # تصدير البيانات إلى Excel
    def export_to_excel(self, selected_columns, file_path):
        try:
            # إنشاء مصنف Excel جديد
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = self.tab_data['tab_name']

            # تنسيق الخط العربي
            arabic_font = Font(name='Arial', size=12)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')

            # تنسيق الخلايا
            header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            current_row = 1

            # إضافة معلومات الشركة
            settings = self.get_print_settings()
            if settings['include_company_info']:
                worksheet.cell(row=current_row, column=1, value=self.print_manager.company_name)
                worksheet.cell(row=current_row, column=1).font = Font(name='Arial', size=16, bold=True)
                current_row += 1

                worksheet.cell(row=current_row, column=1, value=f"التقرير: {self.tab_data['tab_name']}")
                worksheet.cell(row=current_row, column=1).font = Font(name='Arial', size=14, bold=True)
                current_row += 2

            # إضافة الإحصائيات
            if settings['include_statistics'] and self.tab_data.get('statistics'):
                worksheet.cell(row=current_row, column=1, value="الإحصائيات")
                worksheet.cell(row=current_row, column=1).font = Font(name='Arial', size=14, bold=True)
                current_row += 1

                for stat in self.tab_data['statistics']:
                    worksheet.cell(row=current_row, column=1, value=stat['title'])
                    worksheet.cell(row=current_row, column=2, value=stat['value'])
                    current_row += 1

                current_row += 1

            # إضافة الجداول
            for table in self.tab_data.get('tables', []):
                # تصفية الأعمدة المحددة
                filtered_headers = []
                filtered_indices = []

                for col_info in selected_columns:
                    if col_info['table_name'] == table['name']:
                        col_index = col_info['column_index']
                        if col_index < len(table['headers']):
                            filtered_headers.append(table['headers'][col_index])
                            filtered_indices.append(col_index)

                if filtered_headers:
                    # إضافة عنوان الجدول
                    worksheet.cell(row=current_row, column=1, value=table['name'])
                    worksheet.cell(row=current_row, column=1).font = Font(name='Arial', size=14, bold=True)
                    current_row += 1

                    # إضافة عناوين الأعمدة
                    for col_index, header in enumerate(filtered_headers, 1):
                        cell = worksheet.cell(row=current_row, column=col_index, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                    current_row += 1

                    # إضافة البيانات
                    for row_data in table['data']:
                        for col_index, table_col_index in enumerate(filtered_indices, 1):
                            cell_value = row_data[table_col_index] if table_col_index < len(row_data) else ""
                            cell = worksheet.cell(row=current_row, column=col_index, value=cell_value)
                            cell.font = arabic_font
                            cell.border = border
                            cell.alignment = Alignment(horizontal='right')

                        current_row += 1

                    current_row += 1

            # إضافة معلومات الطباعة
            if settings['include_date'] or settings['include_user_info']:
                current_row += 1

                if settings['include_date']:
                    worksheet.cell(row=current_row, column=1, value=f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                    current_row += 1

                if settings['include_user_info']:
                    worksheet.cell(row=current_row, column=1, value=f"المستخدم: {self.print_manager.account_type}")
                    current_row += 1

            # تعديل عرض الأعمدة
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # حفظ الملف
            workbook.save(file_path)

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في تصدير Excel:\n{str(e)}")

    # عرض معاينة الطباعة
    def show_print_preview(self):
        try:
            selected_columns = self.get_selected_columns()
            if not selected_columns:
                QMessageBox.warning(self, "تحذير", "يرجى اختيار عمود واحد على الأقل للمعاينة")
                return

            settings = self.get_print_settings()
            html_content = self.generate_html_document(selected_columns, settings)

            # إنشاء نافذة معاينة
            preview_dialog = PrintPreviewDialog(html_content, self)
            preview_dialog.exec()

        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"فشل في عرض المعاينة:\n{str(e)}")

    # تطبيق الأنماط المركزية
    def apply_styles(self):
        try:
            from ستايل import apply_stylesheet
            apply_stylesheet(self)

            # تطبيق ستايل إضافي مشابه لـ AdvancedPrintExportDialog
            additional_style = """
                QDialog {
                    background-color: #f8f9fa;
                    border-radius: 10px;
                }
                QGroupBox {
                    font-weight: bold;
                    border: 2px solid #3498db;
                    border-radius: 8px;
                    margin: 10px 0;
                    padding-top: 15px;
                    background-color: white;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px 0 5px;
                    color: #2c3e50;
                    background-color: #f8f9fa;
                }
                QPushButton#primary_button {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                    min-width: 120px;
                    font-size: 14px;
                }
                QPushButton#primary_button:hover {
                    background-color: #2ecc71;
                }
                QPushButton#secondary_button {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                    min-width: 120px;
                    font-size: 14px;
                }
                QPushButton#secondary_button:hover {
                    background-color: #2980b9;
                }
                QPushButton#action_button {
                    background-color: #f39c12;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 5px;
                    font-weight: bold;
                    min-width: 100px;
                    font-size: 13px;
                }
                QPushButton#action_button:hover {
                    background-color: #e67e22;
                }
                QPushButton#cancel_button {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    font-weight: bold;
                    min-width: 120px;
                    font-size: 14px;
                }
                QPushButton#cancel_button:hover {
                    background-color: #c0392b;
                }
                QCheckBox#settings_checkbox {
                    spacing: 5px;
                    color: #2c3e50;
                    font-weight: normal;
                    font-size: 13px;
                }
                QCheckBox#column_checkbox {
                    spacing: 5px;
                    color: #2c3e50;
                    font-weight: normal;
                    font-size: 12px;
                }
                QLabel#section_title {
                    color: #2c3e50;
                    font-size: 16px;
                    font-weight: bold;
                    padding: 5px;
                }
                QLabel#dialog_title {
                    color: #2c3e50;
                    font-size: 18px;
                    font-weight: bold;
                    padding: 10px;
                }
                QComboBox#settings_combo {
                    padding: 5px;
                    border: 1px solid #bdc3c7;
                    border-radius: 3px;
                    background-color: white;
                }
                QSpinBox#settings_spin {
                    padding: 5px;
                    border: 1px solid #bdc3c7;
                    border-radius: 3px;
                    background-color: white;
                }
            """
            self.setStyleSheet(self.styleSheet() + additional_style)
        except Exception as e:
            print(f"خطأ في تطبيق الأنماط: {e}")


# نافذة معاينة الطباعة
class PrintPreviewDialog(QDialog):

    # init
    def __init__(self, html_content, parent=None):
        super().__init__(parent)
        self.html_content = html_content
        self.setup_ui()

    # إعداد واجهة المستخدم
    def setup_ui(self):
        self.setWindowTitle("معاينة الطباعة")
        self.setGeometry(100, 100, 1000, 700)
        self.setLayoutDirection(Qt.RightToLeft)

        layout = QVBoxLayout(self)

        # شريط الأدوات
        toolbar = QHBoxLayout()

        close_btn = QPushButton("إغلاق")
        close_btn.setIcon(qta.icon('fa5s.times'))
        close_btn.clicked.connect(self.close)

        toolbar.addStretch()
        toolbar.addWidget(close_btn)

        layout.addLayout(toolbar)

        # عرض المحتوى
        self.web_view = QWebEngineView()
        self.web_view.setHtml(self.html_content)
        layout.addWidget(self.web_view)


# إنشاء مثيل عام من مدير الطباعة
print_manager = UnifiedPrintManager()
